from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import text
from functools import wraps
from datetime import datetime, timedelta
import os
import csv
import io
import base64
import uuid
import json
import re
import unicodedata
import difflib
import tempfile
import openpyxl
from decimal import Decimal
import qrcode
from openpyxl import Workbook

APP_VERSION = "20.11.1"

JOURNAL_ACCOUNT_TYPES = [
    "", "عميل", "مورد", "موظف", "مندوب مبيعات", "بنك", "صندوق",
    "مصروف", "إيراد", "أصل", "التزام", "جهة حكومية", "أخرى"
]

ITEM_UNITS = [
    "وحدة", "قطعة", "متر", "متر مربع", "متر مكعب", "كجم", "طن",
    "لتر", "رول", "بكرة", "صندوق", "كرتون", "حزمة", "طقم",
    "ساعة", "يوم", "شهر", "خدمة", "مقطوعية"
]

app = Flask(__name__, template_folder=".")
app.secret_key = os.environ.get("SECRET_KEY", "development-only-change-me")

database_url = os.environ.get("DATABASE_URL", "sqlite:///erp.db")

if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql+psycopg://", 1)
elif database_url.startswith("postgresql://") and not database_url.startswith("postgresql+psycopg://"):
    database_url = database_url.replace("postgresql://", "postgresql+psycopg://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}

db = SQLAlchemy(app)

def rows(query, params=None):
    result = db.session.execute(text(query), params or {})
    return result.mappings().all()

def row(query, params=None):
    result = db.session.execute(text(query), params or {})
    return result.mappings().first()

def execute(query, params=None):
    db.session.execute(text(query), params or {})
    db.session.commit()

def audit(action, entity, details=""):
    try:
        db.session.execute(
            text("""INSERT INTO audit_logs(user_id, username, action, entity, details, created_at)
                    VALUES(:user_id, :username, :action, :entity, :details, :created_at)"""),
            {
                "user_id": session.get("user_id"),
                "username": session.get("username", "system"),
                "action": action,
                "entity": entity,
                "details": details,
                "created_at": datetime.now(),
            },
        )
        db.session.commit()
    except Exception:
        db.session.rollback()



def next_invoice_draft_number():
    """
    Create the same date/time based number used in the original invoice model.
    No DR, DRAFT, TEMP or other word is included.
    Example: 20260716145837
    """
    current = datetime.now()
    base_number = current.strftime("%Y%m%d%H%M%S")
    candidate = base_number
    suffix = 0

    # Protect against two invoices created during the same second.
    while row("""SELECT id FROM invoices
                 WHERE invoice_no=:number OR draft_number=:number
                 LIMIT 1""", {"number": candidate}):
        suffix += 1
        candidate = f"{base_number}{suffix:02d}"

    return candidate


def invoice_approval_preflight(invoice_id):
    """
    Validate the accounting requirements before consuming the official sequence.
    """
    invoice = row("""SELECT * FROM invoices WHERE id=:id""", {"id": invoice_id})
    if not invoice:
        raise ValueError("الفاتورة غير موجودة.")

    ensure_open_period(invoice["invoice_date"])
    require_accounts([
        "customer_account_id",
        "sales_account_id",
        "vat_output_account_id",
    ])

    if sales_invoice_cost(invoice_id) > 0:
        require_accounts([
            "inventory_account_id",
            "cost_of_sales_account_id",
        ])

    return invoice


def assign_official_invoice_number(invoice_id):
    """
    Replace only the visible serial number at final approval.
    All other invoice data remains unchanged.
    """
    invoice = invoice_approval_preflight(invoice_id)

    original_number = invoice.get("draft_number") or invoice["invoice_no"]
    official_number = next_invoice_number(invoice["invoice_date"])

    execute("""UPDATE invoices
               SET draft_number=COALESCE(draft_number,:draft_number),
                   official_invoice_no=:official_number,
                   invoice_no=:official_number,
                   updated_at=:updated_at
               WHERE id=:id""",
            {
                "draft_number": original_number,
                "official_number": official_number,
                "updated_at": datetime.now(),
                "id": invoice_id,
            })

    return original_number, official_number


def next_invoice_number(invoice_date_value):
    """Create a yearly sequential invoice number such as INV-2026-000001."""
    if isinstance(invoice_date_value, str):
        invoice_year = datetime.strptime(invoice_date_value, "%Y-%m-%d").year
    else:
        invoice_year = invoice_date_value.year

    result = db.session.execute(
        text("""
            INSERT INTO invoice_sequences(sequence_year, last_number)
            VALUES(:year, 1)
            ON CONFLICT (sequence_year)
            DO UPDATE SET last_number = invoice_sequences.last_number + 1
            RETURNING last_number
        """),
        {"year": invoice_year},
    )
    sequence_number = result.scalar_one()
    db.session.commit()
    return f"INV-{invoice_year}-{sequence_number:06d}"



def next_journal_number(journal_date_value):
    if isinstance(journal_date_value, str):
        journal_year = datetime.strptime(journal_date_value, "%Y-%m-%d").year
    else:
        journal_year = journal_date_value.year

    result = db.session.execute(
        text("""
            INSERT INTO journal_sequences(sequence_year, last_number)
            VALUES(:year, 1)
            ON CONFLICT (sequence_year)
            DO UPDATE SET last_number = journal_sequences.last_number + 1
            RETURNING last_number
        """),
        {"year": journal_year},
    )
    number = result.scalar_one()
    db.session.commit()
    return f"JV-{journal_year}-{number:06d}"



def next_batch_number(batch_date_value):
    if isinstance(batch_date_value, str):
        batch_year = datetime.strptime(batch_date_value, "%Y-%m-%d").year
    else:
        batch_year = batch_date_value.year
    latest = db.session.execute(
        text("""SELECT COUNT(*) FROM journal_batches
                WHERE EXTRACT(YEAR FROM batch_date)=:year"""),
        {"year": batch_year}
    ).scalar() or 0
    return f"JB-{batch_year}-{latest + 1:06d}"



def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

def get_default_accounts():
    return row("""SELECT customer_account_id,supplier_account_id,sales_account_id,
                         purchases_account_id,vat_input_account_id,vat_output_account_id,
                         cash_account_id,bank_account_id,inventory_account_id,
                         cost_of_sales_account_id,retained_earnings_account_id
                  FROM settings WHERE id=1""")

def require_accounts(names):
    acc = get_default_accounts()
    missing = [n for n in names if not acc or not acc.get(n)]
    if missing:
        raise ValueError("أكمل ربط الحسابات الافتراضية من الإعدادات.")
    return acc

def ensure_open_period(document_date):
    p = row("""SELECT id,name,status FROM fiscal_periods
               WHERE :d BETWEEN start_date AND end_date
               ORDER BY start_date DESC LIMIT 1""", {"d":document_date})
    if not p:
        raise ValueError("لا توجد فترة مالية تغطي تاريخ المستند.")
    if p["status"] != "مفتوحة":
        raise ValueError(f"الفترة المالية {p['name']} مغلقة.")
    return p

def create_system_journal(journal_date, description, reference, source_type, source_id, lines):
    ensure_open_period(journal_date)
    td = round(sum(float(x.get("debit",0)) for x in lines),2)
    tc = round(sum(float(x.get("credit",0)) for x in lines),2)
    if td <= 0 or td != tc:
        raise ValueError("القيد التلقائي غير متوازن.")
    journal_no = next_journal_number(journal_date)
    db.session.execute(text("""INSERT INTO journal_entries(
        journal_no,journal_date,reference,description,status,total_debit,total_credit,
        created_by,created_at,source_type,source_id,posted_at)
        VALUES(:no,:dt,:ref,:des,'مرحّل',:td,:tc,:uid,:created,:stype,:sid,:posted)"""),
        {"no":journal_no,"dt":journal_date,"ref":reference,"des":description,
         "td":td,"tc":tc,"uid":session.get("user_id"),"created":datetime.now(),
         "stype":source_type,"sid":source_id,"posted":datetime.now()})
    journal_id = db.session.execute(text("SELECT id FROM journal_entries WHERE journal_no=:n"),
                                    {"n":journal_no}).scalar_one()
    for x in lines:
        db.session.execute(text("""INSERT INTO journal_entry_lines(
            journal_id,account_id,debit,credit,taxable,tax_direction,supplier_id,
            customer_id,party_type,tax_number,invoice_number,invoice_date,
            line_description,cost_center_id)
            VALUES(:jid,:aid,:debit,:credit,:taxable,:direction,:sid,:cid,:ptype,
                   :tax,:inv,:invdt,:des,:cc)"""),
            {"jid":journal_id,"aid":x["account_id"],"debit":round(float(x.get("debit",0)),2),
             "credit":round(float(x.get("credit",0)),2),"taxable":1 if x.get("taxable") else 0,
             "direction":x.get("tax_direction","غير مطبق"),"sid":x.get("supplier_id"),
             "cid":x.get("customer_id"),"ptype":x.get("party_type",""),
             "tax":x.get("tax_number",""),"inv":x.get("invoice_number",""),
             "invdt":x.get("invoice_date"),"des":x.get("line_description",""),
             "cc":x.get("cost_center_id")})
    db.session.commit()
    audit("POST",source_type,f"ترحيل تلقائي بالقيد {journal_no}")
    return journal_id

def post_invoice_to_ledger(invoice_id):
    inv = row("""SELECT i.*,c.vat_number customer_vat FROM invoices i
                 JOIN customers c ON c.id=i.customer_id WHERE i.id=:id""",{"id":invoice_id})
    if not inv: raise ValueError("الفاتورة غير موجودة.")
    if inv.get("journal_id"): return inv["journal_id"]
    a = require_accounts(["customer_account_id","sales_account_id","vat_output_account_id"])
    lines = [
      {"account_id":a["customer_account_id"],"debit":inv["total"],"customer_id":inv["customer_id"],
       "party_type":"عميل","tax_number":inv["customer_vat"] or "","invoice_number":inv["invoice_no"],
       "invoice_date":inv["invoice_date"],"line_description":"إجمالي فاتورة المبيعات"},
      {"account_id":a["sales_account_id"],"credit":inv["subtotal"],"customer_id":inv["customer_id"],
       "party_type":"عميل","tax_number":inv["customer_vat"] or "","invoice_number":inv["invoice_no"],
       "invoice_date":inv["invoice_date"],"line_description":"إيرادات المبيعات"},
    ]
    if float(inv["vat"] or 0):
        lines.append({"account_id":a["vat_output_account_id"],"credit":inv["vat"],"taxable":1,
          "tax_direction":"مخرجات","customer_id":inv["customer_id"],"party_type":"عميل",
          "tax_number":inv["customer_vat"] or "","invoice_number":inv["invoice_no"],
          "invoice_date":inv["invoice_date"],"line_description":"ضريبة المخرجات"})
    cost_total = sales_invoice_cost(invoice_id)
    if cost_total > 0:
        stock_accounts = require_accounts(["inventory_account_id","cost_of_sales_account_id"])
        lines.extend([
            {"account_id":stock_accounts["cost_of_sales_account_id"],"debit":cost_total,
             "invoice_number":inv["invoice_no"],"invoice_date":inv["invoice_date"],
             "line_description":"تكلفة البضاعة المباعة"},
            {"account_id":stock_accounts["inventory_account_id"],"credit":cost_total,
             "invoice_number":inv["invoice_no"],"invoice_date":inv["invoice_date"],
             "line_description":"إخراج المخزون المباع"},
        ])
    jid = create_system_journal(inv["invoice_date"],f"فاتورة مبيعات {inv['invoice_no']}",
                                inv["invoice_no"],"INVOICE",invoice_id,lines)
    execute("UPDATE invoices SET journal_id=:j,posting_status='مرحّل' WHERE id=:id",
            {"j":jid,"id":invoice_id})
    return jid

def post_expense_to_ledger(expense_id):
    e = row("""SELECT e.*,s.vat_number supplier_vat FROM expenses e
               LEFT JOIN suppliers s ON s.id=e.supplier_id WHERE e.id=:id""",{"id":expense_id})
    if not e: raise ValueError("المصروف غير موجود.")
    if e.get("journal_id"): return e["journal_id"]
    if not e["expense_account_id"] or not e["payment_account_id"]:
        raise ValueError("حدد حساب المصروف وحساب السداد.")
    a = require_accounts(["vat_input_account_id"] if float(e["vat"] or 0) else [])
    common = {"supplier_id":e["supplier_id"],"party_type":"مورد" if e["supplier_id"] else "",
              "tax_number":e["supplier_vat"] or "","invoice_number":e["invoice_number"] or "",
              "invoice_date":e["invoice_date"],"cost_center_id":e["cost_center_id"]}
    lines = [{"account_id":e["expense_account_id"],"debit":e["amount"],
              "line_description":e["description"] or e["category"],**common}]
    if float(e["vat"] or 0):
        lines.append({"account_id":a["vat_input_account_id"],"debit":e["vat"],"taxable":1,
                      "tax_direction":"مدخلات","line_description":"ضريبة المدخلات",**common})
    lines.append({"account_id":e["payment_account_id"],"credit":e["total"],
                  "line_description":"سداد المصروف",**common})
    jid = create_system_journal(e["expense_date"],f"مصروف {e['category']}",
                                e["invoice_number"] or f"EXP-{expense_id}",
                                "EXPENSE",expense_id,lines)
    execute("UPDATE expenses SET journal_id=:j,posting_status='مرحّل' WHERE id=:id",
            {"j":jid,"id":expense_id})
    return jid



ARABIC_TRANSLITERATION = {
    "ا":"a","أ":"a","إ":"i","آ":"aa","ب":"b","ت":"t","ث":"th","ج":"j","ح":"h",
    "خ":"kh","د":"d","ذ":"dh","ر":"r","ز":"z","س":"s","ش":"sh","ص":"s","ض":"d",
    "ط":"t","ظ":"z","ع":"a","غ":"gh","ف":"f","ق":"q","ك":"k","ل":"l","م":"m",
    "ن":"n","ه":"h","ة":"a","و":"w","ؤ":"o","ي":"y","ى":"a","ئ":"e","ء":"",
    "َ":"a","ُ":"u","ِ":"i","ّ":"","ْ":"","ً":"","ٌ":"","ٍ":"","ـ":""
}
BUSINESS_WORDS = {
    "شركة":"Company","الشركة":"Company","مؤسسة":"Establishment","المؤسسة":"Establishment",
    "محدودة":"Limited","المحدودة":"Limited","ذمم":"LLC","للمقاولات":"Contracting",
    "مقاولات":"Contracting","تجارة":"Trading","للتجارة":"Trading","صناعة":"Industries",
    "للصناعة":"Industries","خدمات":"Services","للخدمات":"Services","مجموعة":"Group",
    "مصنع":"Factory","مكتب":"Office","مركز":"Center","الدولية":"International",
    "العالمية":"International","العربية":"Arabian","السعودية":"Saudi","الخليج":"Gulf"
}

def transliterate_arabic_name(value):
    value = (value or "").strip()
    if not value:
        return ""
    words = value.split()
    result = []
    for word in words:
        normalized = word.strip("،,.-_/()")
        if normalized in BUSINESS_WORDS:
            result.append(BUSINESS_WORDS[normalized])
            continue
        latin = "".join(ARABIC_TRANSLITERATION.get(ch, ch) for ch in normalized)
        latin = re.sub(r"[^A-Za-z0-9]+", "", latin)
        if latin:
            result.append(latin[:1].upper() + latin[1:])
    # Improve common company suffix order.
    if result and result[0] == "Company":
        result = result[1:] + ["Company"]
    return " ".join(result)

@app.route("/api/transliterate", methods=["POST"])
@login_required
def transliterate_name_api():
    payload = request.get_json(silent=True) or {}
    return {"english_name": transliterate_arabic_name(payload.get("name",""))}



def next_treasury_number(voucher_type, voucher_date):
    prefix = {"قبض":"RV","صرف":"PV","تحويل":"TV"}.get(voucher_type,"TV")
    year = datetime.strptime(voucher_date,"%Y-%m-%d").year if isinstance(voucher_date,str) else voucher_date.year
    count = db.session.execute(text("""SELECT COUNT(*) FROM treasury_vouchers
        WHERE voucher_type=:t AND EXTRACT(YEAR FROM voucher_date)=:y"""),
        {"t":voucher_type,"y":year}).scalar() or 0
    return f"{prefix}-{year}-{count+1:06d}"

def post_treasury_voucher(voucher_id):
    v = row("""SELECT tv.*,c.name customer_name,c.vat_number customer_vat,
                     s.name supplier_name,s.vat_number supplier_vat
              FROM treasury_vouchers tv
              LEFT JOIN customers c ON c.id=tv.customer_id
              LEFT JOIN suppliers s ON s.id=tv.supplier_id
              WHERE tv.id=:id""", {"id":voucher_id})
    if not v:
        raise ValueError("السند غير موجود.")
    if v.get("journal_id"):
        return v["journal_id"]
    ensure_open_period(v["voucher_date"])

    common = {
        "customer_id":v["customer_id"],"supplier_id":v["supplier_id"],
        "party_type":v["party_type"],"tax_number":v["customer_vat"] or v["supplier_vat"] or "",
        "line_description":v["description"] or v["voucher_type"],
        "cost_center_id":v["cost_center_id"]
    }
    amount = float(v["amount"])
    if v["voucher_type"] == "قبض":
        lines = [
            {"account_id":v["cash_bank_account_id"],"debit":amount,**common},
            {"account_id":v["counter_account_id"],"credit":amount,**common},
        ]
    elif v["voucher_type"] == "صرف":
        lines = [
            {"account_id":v["counter_account_id"],"debit":amount,**common},
            {"account_id":v["cash_bank_account_id"],"credit":amount,**common},
        ]
    else:
        lines = [
            {"account_id":v["counter_account_id"],"debit":amount,**common},
            {"account_id":v["cash_bank_account_id"],"credit":amount,**common},
        ]

    jid = create_system_journal(
        v["voucher_date"], f"سند {v['voucher_type']} {v['voucher_no']}",
        v["reference"] or v["voucher_no"], "TREASURY", voucher_id, lines
    )
    execute("""UPDATE treasury_vouchers
               SET journal_id=:j,posting_status='مرحّل',status='معتمد'
               WHERE id=:id""", {"j":jid,"id":voucher_id})
    return jid



def next_document_number(table_name, date_field, prefix, document_date):
    year = datetime.strptime(document_date,"%Y-%m-%d").year if isinstance(document_date,str) else document_date.year
    allowed = {
        "purchase_requisitions": "requisition_date",
        "purchase_orders": "po_date",
        "goods_receipts": "grn_date",
        "supplier_invoices": "invoice_date",
    }
    if table_name not in allowed or allowed[table_name] != date_field:
        raise ValueError("Invalid document sequence.")
    count = db.session.execute(
        text(f"SELECT COUNT(*) FROM {table_name} WHERE EXTRACT(YEAR FROM {date_field})=:y"),
        {"y":year}
    ).scalar() or 0
    return f"{prefix}-{year}-{count+1:06d}"

def get_required_approver(document_type, amount):
    rule = row("""SELECT approver_role FROM approval_rules
                  WHERE document_type=:t AND active=1
                    AND :amount >= min_amount
                    AND (max_amount IS NULL OR :amount <= max_amount)
                  ORDER BY min_amount DESC LIMIT 1""",
               {"t":document_type,"amount":amount})
    return rule["approver_role"] if rule else "المدير المالي"

def post_supplier_invoice(invoice_id):
    inv = row("""SELECT si.*,s.vat_number supplier_vat
                 FROM supplier_invoices si
                 JOIN suppliers s ON s.id=si.supplier_id
                 WHERE si.id=:id""", {"id":invoice_id})
    if not inv:
        raise ValueError("فاتورة المورد غير موجودة.")
    if inv.get("journal_id"):
        return inv["journal_id"]
    acc = require_accounts(["supplier_account_id","vat_input_account_id"])
    common = {
        "supplier_id":inv["supplier_id"],"party_type":"مورد",
        "tax_number":inv["supplier_vat"] or "",
        "invoice_number":inv["supplier_invoice_no"],
        "invoice_date":inv["invoice_date"],
        "cost_center_id":inv["cost_center_id"],
    }
    lines = [
        {"account_id":inv["expense_or_inventory_account_id"],"debit":inv["subtotal"],
         "line_description":"فاتورة مورد - قيمة قبل الضريبة",**common}
    ]
    if float(inv["vat"] or 0) > 0:
        lines.append({"account_id":acc["vat_input_account_id"],"debit":inv["vat"],
                      "taxable":1,"tax_direction":"مدخلات",
                      "line_description":"ضريبة القيمة المضافة - مدخلات",**common})
    lines.append({"account_id":acc["supplier_account_id"],"credit":inv["total"],
                  "line_description":"ذمم المورد",**common})
    jid = create_system_journal(
        inv["invoice_date"], f"فاتورة مورد {inv['supplier_invoice_no']}",
        inv["supplier_invoice_no"], "SUPPLIER_INVOICE", invoice_id, lines
    )
    execute("""UPDATE supplier_invoices
               SET journal_id=:j,posting_status='مرحّل',status='معتمدة'
               WHERE id=:id""", {"j":jid,"id":invoice_id})
    return jid



def next_inventory_movement_number(movement_date):
    year = datetime.strptime(movement_date,"%Y-%m-%d").year if isinstance(movement_date,str) else movement_date.year
    count = db.session.execute(text("""SELECT COUNT(*) FROM inventory_movements
        WHERE EXTRACT(YEAR FROM movement_date)=:y"""),{"y":year}).scalar() or 0
    return f"IM-{year}-{count+1:07d}"

def next_inventory_count_number(count_date):
    year = datetime.strptime(count_date,"%Y-%m-%d").year if isinstance(count_date,str) else count_date.year
    count = db.session.execute(text("""SELECT COUNT(*) FROM inventory_counts
        WHERE EXTRACT(YEAR FROM count_date)=:y"""),{"y":year}).scalar() or 0
    return f"IC-{year}-{count+1:06d}"

def warehouse_stock(item_id, warehouse_id):
    value = db.session.execute(text("""
        SELECT COALESCE(SUM(
          CASE
            WHEN movement_type IN ('رصيد افتتاحي','استلام','تسوية زيادة','تحويل وارد','مرتجع مبيعات') THEN quantity
            WHEN movement_type IN ('صرف','تسوية نقص','تحويل صادر','بيع','مرتجع مشتريات') THEN -quantity
            ELSE 0 END
        ),0)
        FROM inventory_movements
        WHERE item_id=:item AND warehouse_id=:warehouse
    """),{"item":item_id,"warehouse":warehouse_id}).scalar()
    return float(value or 0)

def item_total_stock(item_id):
    value = db.session.execute(text("""
        SELECT COALESCE(SUM(
          CASE
            WHEN movement_type IN ('رصيد افتتاحي','استلام','تسوية زيادة','تحويل وارد','مرتجع مبيعات') THEN quantity
            WHEN movement_type IN ('صرف','تسوية نقص','تحويل صادر','بيع','مرتجع مشتريات') THEN -quantity
            ELSE 0 END
        ),0)
        FROM inventory_movements WHERE item_id=:item
    """),{"item":item_id}).scalar()
    return float(value or 0)

def record_inventory_movement(movement_date,movement_type,item_id,warehouse_id,quantity,
                              unit_cost=0,destination_warehouse_id=None,reference_type="",
                              reference_id=None,reference_no="",batch_no="",serial_no="",
                              production_date=None,expiry_date=None,notes=""):
    quantity=round(float(quantity),3)
    unit_cost=round(float(unit_cost or 0),4)
    if quantity<=0:
        raise ValueError("الكمية يجب أن تكون أكبر من صفر.")
    if movement_type in ("صرف","بيع","تحويل صادر","مرتجع مشتريات","تسوية نقص"):
        available=warehouse_stock(item_id,warehouse_id)
        if quantity>available+0.0001:
            raise ValueError(f"الرصيد غير كافٍ. المتاح {available:.3f}")
    no=next_inventory_movement_number(movement_date)
    execute("""INSERT INTO inventory_movements(
      movement_no,movement_date,movement_type,item_id,warehouse_id,destination_warehouse_id,
      quantity,unit_cost,total_cost,reference_type,reference_id,reference_no,batch_no,
      serial_no,production_date,expiry_date,notes,created_by,created_at)
      VALUES(:no,:dt,:type,:item,:warehouse,:dest,:qty,:cost,:total,:rtype,:rid,:rno,
      :batch,:serial,:prod,:expiry,:notes,:uid,:created)""",
      {"no":no,"dt":movement_date,"type":movement_type,"item":item_id,
       "warehouse":warehouse_id,"dest":destination_warehouse_id,"qty":quantity,
       "cost":unit_cost,"total":round(quantity*unit_cost,2),"rtype":reference_type,
       "rid":reference_id,"rno":reference_no,"batch":batch_no,"serial":serial_no,
       "prod":production_date,"expiry":expiry_date,"notes":notes,
       "uid":session.get("user_id"),"created":datetime.now()})
    # Moving-average cost on incoming movements.
    if movement_type in ("رصيد افتتاحي","استلام","تسوية زيادة","تحويل وارد","مرتجع مبيعات"):
        item=row("SELECT cost FROM inventory WHERE id=:id",{"id":item_id})
        previous_qty=max(item_total_stock(item_id)-quantity,0)
        old_cost=float(item["cost"] or 0)
        new_cost=((previous_qty*old_cost)+(quantity*unit_cost))/(previous_qty+quantity) if previous_qty+quantity else unit_cost
        execute("UPDATE inventory SET quantity=:q,cost=:c WHERE id=:id",
                {"q":item_total_stock(item_id),"c":round(new_cost,4),"id":item_id})
    else:
        execute("UPDATE inventory SET quantity=:q WHERE id=:id",
                {"q":item_total_stock(item_id),"id":item_id})
    return no



def next_sales_no(table_name, date_field, prefix, doc_date):
    allowed={"sales_quotations":"quotation_date","sales_orders":"order_date","sales_deliveries":"delivery_date"}
    if allowed.get(table_name)!=date_field: raise ValueError("Invalid sequence")
    y=datetime.strptime(doc_date,"%Y-%m-%d").year if isinstance(doc_date,str) else doc_date.year
    n=db.session.execute(text(f"SELECT COUNT(*) FROM {table_name} WHERE EXTRACT(YEAR FROM {date_field})=:y"),{"y":y}).scalar() or 0
    return f"{prefix}-{y}-{n+1:06d}"

def parse_sales_lines(form):
    ids=form.getlist("item_id[]"); qtys=form.getlist("quantity[]"); prices=form.getlist("unit_price[]")
    discounts=form.getlist("discount_rate[]"); vats=form.getlist("vat_rate[]")
    lines=[]; subtotal=discount_total=vat_total=total=0
    for i,item_id in enumerate(ids):
        if not item_id: continue
        item=row("SELECT id,sku,name,unit,sale_price FROM inventory WHERE id=:id",{"id":item_id})
        qty=float(qtys[i] or 0); price=float(prices[i] or item["sale_price"] or 0)
        dr=float(discounts[i] or 0); vr=float(vats[i] or 0)
        if qty<=0: raise ValueError("الكمية يجب أن تكون أكبر من صفر")
        base=round(qty*price,2); disc=round(base*dr/100,2); taxable=base-disc
        tax=round(taxable*vr/100,2); line_total=round(taxable+tax,2)
        lines.append({"item_id":item["id"],"item_name":item["name"],"quantity":qty,"unit":item["unit"],
                      "unit_price":price,"discount_rate":dr,"vat_rate":vr,
                      "line_subtotal":base,"line_discount":disc,"line_vat":tax,"line_total":line_total})
        subtotal+=base; discount_total+=disc; vat_total+=tax; total+=line_total
    if not lines: raise ValueError("أدخل صنفًا واحدًا على الأقل")
    return lines,round(subtotal,2),round(discount_total,2),round(vat_total,2),round(total,2)


def create_invoice_from_sales_delivery(delivery_id):
    delivery = row("""SELECT d.*,o.order_no,o.branch_id,o.cost_center_id,o.subtotal,o.discount,
                             o.vat,o.total
                      FROM sales_deliveries d
                      JOIN sales_orders o ON o.id=d.order_id
                      WHERE d.id=:id""", {"id":delivery_id})
    if not delivery:
        raise ValueError("إذن التسليم غير موجود.")
    if delivery.get("invoice_id"):
        return delivery["invoice_id"]

    delivery_items = rows("""SELECT di.*,oi.item_name,oi.unit,oi.unit_price,oi.vat_rate,
                                    oi.discount_rate
                             FROM sales_delivery_items di
                             JOIN sales_order_items oi ON oi.id=di.order_item_id
                             WHERE di.delivery_id=:id ORDER BY di.id""", {"id":delivery_id})
    if not delivery_items:
        raise ValueError("لا توجد بنود في إذن التسليم.")

    subtotal = vat_total = grand_total = 0.0
    prepared = []
    for x in delivery_items:
        base = round(float(x["quantity"]) * float(x["unit_price"]), 2)
        discount = round(base * float(x["discount_rate"] or 0) / 100, 2)
        taxable = round(base - discount, 2)
        tax = round(taxable * float(x["vat_rate"] or 0) / 100, 2)
        total = round(taxable + tax, 2)
        prepared.append({
            "item_id": x["item_id"], "item_name": x["item_name"], "quantity": x["quantity"],
            "unit": x["unit"], "unit_price": x["unit_price"], "vat_rate": x["vat_rate"],
            "line_subtotal": taxable, "line_vat": tax, "line_total": total
        })
        subtotal += taxable
        vat_total += tax
        grand_total += total

    invoice_no = next_invoice_number(delivery["delivery_date"])
    invoice_uuid = str(uuid.uuid4())
    execute("""INSERT INTO invoices(
        invoice_no,invoice_uuid,customer_id,invoice_date,subtotal,vat,total,status,
        branch_id,notes,created_at,sales_order_id,delivery_id,cost_center_id,warehouse_id
        ) VALUES(
        :no,:uuid,:customer,:dt,:sub,:vat,:total,'مسودة',:branch,:notes,:created,
        :order_id,:delivery_id,:cc,:warehouse
        )""", {
        "no":invoice_no,"uuid":invoice_uuid,"customer":delivery["customer_id"],
        "dt":delivery["delivery_date"],"sub":round(subtotal,2),"vat":round(vat_total,2),
        "total":round(grand_total,2),"branch":delivery["branch_id"],
        "notes":f"فاتورة ناتجة عن إذن التسليم {delivery['delivery_no']}",
        "created":datetime.now(),"order_id":delivery["order_id"],"delivery_id":delivery_id,
        "cc":delivery["cost_center_id"],"warehouse":delivery["warehouse_id"]
    })
    order_meta = row("""SELECT payment_terms,sales_person FROM sales_orders WHERE id=:id""",
                     {"id":delivery["order_id"]})
    due_date = delivery["delivery_date"]
    if order_meta and order_meta.get("payment_terms"):
        match = re.search(r"(\d+)", order_meta["payment_terms"])
        if match:
            due_date = delivery["delivery_date"] + timedelta(days=int(match.group(1)))
    execute("""UPDATE invoices SET due_date=:due,payment_terms=:terms,
               sales_person=:sales_person,payment_method='آجل'
               WHERE id=:id""",
            {"due":due_date,
             "terms":order_meta["payment_terms"] if order_meta else "",
             "sales_person":order_meta["sales_person"] if order_meta else "",
             "id":invoice_id})
    invoice_id = row("SELECT id FROM invoices WHERE invoice_no=:no",{"no":invoice_no})["id"]

    for x in prepared:
        execute("""INSERT INTO invoice_items(
            invoice_id,item_id,item_name,quantity,unit,unit_price,vat_rate,
            line_subtotal,line_vat,line_total
            ) VALUES(
            :invoice_id,:item_id,:item_name,:quantity,:unit,:unit_price,:vat_rate,
            :line_subtotal,:line_vat,:line_total
            )""", {"invoice_id":invoice_id, **x})

    execute("UPDATE sales_deliveries SET invoice_id=:invoice WHERE id=:id",
            {"invoice":invoice_id,"id":delivery_id})
    audit("CREATE","INVOICE",f"إنشاء فاتورة {invoice_no} من إذن التسليم {delivery['delivery_no']}")
    return invoice_id

def sales_invoice_cost(invoice_id):
    value = db.session.execute(text("""
        SELECT COALESCE(SUM(di.quantity * di.unit_cost),0)
        FROM sales_delivery_items di
        JOIN sales_deliveries d ON d.id=di.delivery_id
        JOIN invoices i ON i.delivery_id=d.id
        WHERE i.id=:invoice_id
    """), {"invoice_id":invoice_id}).scalar()
    if value:
        return float(value)
    value = db.session.execute(text("""
        SELECT COALESCE(SUM(ii.quantity * COALESCE(inv.cost,0)),0)
        FROM invoice_items ii LEFT JOIN inventory inv ON inv.id=ii.item_id
        WHERE ii.invoice_id=:invoice_id
    """), {"invoice_id":invoice_id}).scalar()
    return float(value or 0)

def post_sales_return_to_ledger(return_id):
    ret = row("""SELECT r.*,c.vat_number customer_vat
                 FROM sales_returns r JOIN customers c ON c.id=r.customer_id
                 WHERE r.id=:id""", {"id":return_id})
    if not ret:
        raise ValueError("مرتجع المبيعات غير موجود.")
    if ret.get("journal_id"):
        return ret["journal_id"]

    acc = require_accounts(["customer_account_id","sales_account_id","vat_output_account_id",
                            "inventory_account_id","cost_of_sales_account_id"])
    common = {
        "customer_id":ret["customer_id"],"party_type":"عميل",
        "tax_number":ret["customer_vat"] or "","invoice_number":ret["return_no"],
        "invoice_date":ret["return_date"],"line_description":"مرتجع مبيعات"
    }
    lines = [
        {"account_id":acc["sales_account_id"],"debit":ret["subtotal"],**common},
        {"account_id":acc["customer_account_id"],"credit":ret["total"],**common},
    ]
    if float(ret["vat"] or 0) > 0:
        lines.append({"account_id":acc["vat_output_account_id"],"debit":ret["vat"],
                      "taxable":1,"tax_direction":"مخرجات",**common})
    if float(ret["cost_total"] or 0) > 0:
        lines.extend([
            {"account_id":acc["inventory_account_id"],"debit":ret["cost_total"],**common},
            {"account_id":acc["cost_of_sales_account_id"],"credit":ret["cost_total"],**common},
        ])

    jid = create_system_journal(
        ret["return_date"],f"مرتجع مبيعات {ret['return_no']}",
        ret["return_no"],"SALES_RETURN",return_id,lines
    )
    execute("UPDATE sales_returns SET journal_id=:jid WHERE id=:id",
            {"jid":jid,"id":return_id})
    return jid



def invoice_paid_amount(invoice_id):
    value = db.session.execute(text("""
        SELECT COALESCE(SUM(allocated_amount),0)
        FROM invoice_payment_allocations WHERE invoice_id=:id
    """), {"id":invoice_id}).scalar()
    return round(float(value or 0),2)

def refresh_invoice_payment_status(invoice_id):
    inv = row("SELECT total FROM invoices WHERE id=:id",{"id":invoice_id})
    if not inv:
        return
    paid = invoice_paid_amount(invoice_id)
    total = round(float(inv["total"] or 0),2)
    if paid <= 0:
        status = "غير مسددة"
    elif paid + 0.005 >= total:
        status = "مسددة"
    else:
        status = "مسددة جزئيًا"
    execute("UPDATE invoices SET payment_status=:status WHERE id=:id",
            {"status":status,"id":invoice_id})

def allocate_customer_receipt(voucher_id, allocations):
    voucher = row("""SELECT * FROM treasury_vouchers
                     WHERE id=:id AND voucher_type='قبض' AND customer_id IS NOT NULL""",
                  {"id":voucher_id})
    if not voucher:
        raise ValueError("سند القبض غير صالح للربط بالفواتير.")
    allocation_total = round(sum(float(x["amount"]) for x in allocations),2)
    if allocation_total <= 0:
        raise ValueError("يجب تخصيص مبلغ لفاتورة واحدة على الأقل.")
    if allocation_total > round(float(voucher["amount"]),2) + 0.005:
        raise ValueError("إجمالي التخصيص أكبر من مبلغ سند القبض.")

    for allocation in allocations:
        invoice_id = int(allocation["invoice_id"])
        amount = round(float(allocation["amount"]),2)
        if amount <= 0:
            continue
        inv = row("""SELECT id,total,customer_id FROM invoices
                     WHERE id=:id AND customer_id=:customer""",
                  {"id":invoice_id,"customer":voucher["customer_id"]})
        if not inv:
            raise ValueError("إحدى الفواتير لا تخص العميل المحدد.")
        outstanding = round(float(inv["total"]) - invoice_paid_amount(invoice_id),2)
        if amount > outstanding + 0.005:
            raise ValueError(f"المبلغ المخصص أكبر من رصيد الفاتورة المتبقي {outstanding:.2f}.")
        execute("""INSERT INTO invoice_payment_allocations(
                    invoice_id,voucher_id,allocated_amount,allocation_date,created_by,created_at)
                   VALUES(:invoice,:voucher,:amount,:dt,:uid,:created)""",
                {"invoice":invoice_id,"voucher":voucher_id,"amount":amount,
                 "dt":voucher["voucher_date"],"uid":session.get("user_id"),
                 "created":datetime.now()})
        refresh_invoice_payment_status(invoice_id)



def sales_document_timeline(invoice_id=None, delivery_id=None, order_id=None, quotation_id=None):
    context = {
        "quotation": None, "order": None, "delivery": None,
        "invoice": None, "receipt_allocations": [], "return_docs": []
    }

    if invoice_id:
        context["invoice"] = row("""SELECT i.*,c.name customer_name
                                    FROM invoices i JOIN customers c ON c.id=i.customer_id
                                    WHERE i.id=:id""", {"id":invoice_id})
        if context["invoice"]:
            delivery_id = delivery_id or context["invoice"].get("delivery_id")
            order_id = order_id or context["invoice"].get("sales_order_id")
            context["receipt_allocations"] = rows("""SELECT a.*,tv.voucher_no,tv.voucher_date,
                tv.payment_method,tv.reference
                FROM invoice_payment_allocations a
                JOIN treasury_vouchers tv ON tv.id=a.voucher_id
                WHERE a.invoice_id=:id ORDER BY tv.voucher_date,tv.id""", {"id":invoice_id})
            context["return_docs"] = rows("""SELECT id,return_no,return_date,total,status
                                             FROM sales_returns WHERE invoice_id=:id
                                             ORDER BY return_date,id""", {"id":invoice_id})

    if delivery_id:
        context["delivery"] = row("""SELECT d.*,o.order_no
                                     FROM sales_deliveries d
                                     JOIN sales_orders o ON o.id=d.order_id
                                     WHERE d.id=:id""", {"id":delivery_id})
        if context["delivery"]:
            order_id = order_id or context["delivery"]["order_id"]

    if order_id:
        context["order"] = row("""SELECT o.*,q.quotation_no
                                  FROM sales_orders o
                                  LEFT JOIN sales_quotations q ON q.id=o.quotation_id
                                  WHERE o.id=:id""", {"id":order_id})
        if context["order"]:
            quotation_id = quotation_id or context["order"].get("quotation_id")

    if quotation_id:
        context["quotation"] = row("SELECT * FROM sales_quotations WHERE id=:id",
                                   {"id":quotation_id})

    return context



def financial_date_filters():
    return {
        "date_from": request.args.get("date_from","").strip(),
        "date_to": request.args.get("date_to","").strip(),
        "branch_id": request.args.get("branch_id","").strip(),
        "cost_center_id": request.args.get("cost_center_id","").strip(),
    }

def financial_where(filters, journal_alias="j", line_alias="l"):
    conditions=[f"{journal_alias}.status='مرحّل'"]
    params={}
    if filters.get("date_from"):
        conditions.append(f"{journal_alias}.journal_date>=:date_from")
        params["date_from"]=filters["date_from"]
    if filters.get("date_to"):
        conditions.append(f"{journal_alias}.journal_date<=:date_to")
        params["date_to"]=filters["date_to"]
    if filters.get("cost_center_id"):
        conditions.append(f"{line_alias}.cost_center_id=:cost_center_id")
        params["cost_center_id"]=filters["cost_center_id"]
    return " AND ".join(conditions),params

def account_signed_balance(account_type, debit, credit):
    debit=float(debit or 0); credit=float(credit or 0)
    if account_type in ("التزام","حقوق ملكية","إيراد"):
        return credit-debit
    return debit-credit

def financial_statement_data(filters):
    where,params=financial_where(filters)
    data=rows(f"""SELECT a.id,a.account_code,a.account_name_ar,a.account_name_en,
      a.account_type,a.statement_type,a.normal_balance,
      COALESCE(SUM(l.debit),0) total_debit,COALESCE(SUM(l.credit),0) total_credit
      FROM chart_of_accounts a
      LEFT JOIN journal_entry_lines l ON l.account_id=a.id
      LEFT JOIN journal_entries j ON j.id=l.journal_id AND {where}
      WHERE a.active=1
      GROUP BY a.id,a.account_code,a.account_name_ar,a.account_name_en,
               a.account_type,a.statement_type,a.normal_balance
      ORDER BY a.account_code""",params)
    result=[]
    for r in data:
        item=dict(r)
        item["balance"]=round(account_signed_balance(
            item["account_type"],item["total_debit"],item["total_credit"]),2)
        result.append(item)
    return result

def party_opening_balance(party_type,party_id,date_from=""):
    if not party_id or not date_from:
        return 0.0
    field="customer_id" if party_type=="customer" else "supplier_id"
    result=row(f"""SELECT COALESCE(SUM(l.debit-l.credit),0) balance
                   FROM journal_entry_lines l JOIN journal_entries j ON j.id=l.journal_id
                   WHERE l.{field}=:party AND j.status='مرحّل' AND j.journal_date<:date_from""",
               {"party":party_id,"date_from":date_from})
    return round(float(result["balance"] or 0),2)

def party_statement_rows(party_type, party_id, date_from="", date_to="", opening_balance=0):
    field="customer_id" if party_type=="customer" else "supplier_id"
    conditions=[f"l.{field}=:party","j.status='مرحّل'"]
    params={"party":party_id}
    if date_from:
        conditions.append("j.journal_date>=:date_from");params["date_from"]=date_from
    if date_to:
        conditions.append("j.journal_date<=:date_to");params["date_to"]=date_to
    data=rows(f"""SELECT j.id journal_id,j.journal_no,j.journal_date,j.reference,
      j.description,l.invoice_number,l.invoice_date,l.line_description,
      l.debit,l.credit,a.account_code,a.account_name_ar
      FROM journal_entry_lines l
      JOIN journal_entries j ON j.id=l.journal_id
      JOIN chart_of_accounts a ON a.id=l.account_id
      WHERE {' AND '.join(conditions)}
      ORDER BY j.journal_date,j.id,l.id""",params)
    running=float(opening_balance or 0)
    output=[]
    for r in data:
        item=dict(r)
        running += float(item["debit"] or 0)-float(item["credit"] or 0)
        item["running_balance"]=round(running,2)
        output.append(item)
    return output



def next_asset_number():
    count = db.session.execute(text("SELECT COUNT(*) FROM fixed_assets")).scalar() or 0
    return f"FA-{count+1:06d}"

def next_asset_run_number(period_date):
    year = period_date.year if hasattr(period_date,"year") else datetime.strptime(period_date,"%Y-%m-%d").year
    count = db.session.execute(text("""SELECT COUNT(*) FROM asset_depreciation_runs
        WHERE EXTRACT(YEAR FROM period_date)=:y"""),{"y":year}).scalar() or 0
    return f"DEP-{year}-{count+1:05d}"

def calculate_monthly_depreciation(asset):
    cost=float(asset["purchase_cost"] or 0)
    residual=float(asset["residual_value"] or 0)
    life=int(asset["useful_life_months"] or 1)
    accumulated=float(asset["accumulated_depreciation"] or 0)
    depreciable=max(cost-residual,0)
    remaining=max(depreciable-accumulated,0)
    if remaining<=0:
        return 0
    if asset["depreciation_method"]=="القسط الثابت":
        return round(min(depreciable/life,remaining),2)
    return round(min(depreciable/life,remaining),2)

def post_depreciation_run(run_id):
    run=row("SELECT * FROM asset_depreciation_runs WHERE id=:id",{"id":run_id})
    if not run:
        raise ValueError("دفعة الإهلاك غير موجودة.")
    if run.get("journal_id"):
        return run["journal_id"]
    lines=rows("""SELECT dl.*,fa.name,ac.depreciation_expense_account_id,
      ac.accumulated_depr_account_id
      FROM asset_depreciation_lines dl
      JOIN fixed_assets fa ON fa.id=dl.asset_id
      JOIN asset_categories ac ON ac.id=fa.category_id
      WHERE dl.run_id=:id""",{"id":run_id})
    journal_lines=[]
    for x in lines:
        if not x["depreciation_expense_account_id"] or not x["accumulated_depr_account_id"]:
            raise ValueError(f"فئة الأصل {x['name']} لا تحتوي على حسابات الإهلاك.")
        journal_lines.extend([
            {"account_id":x["depreciation_expense_account_id"],"debit":x["depreciation_amount"],
             "line_description":f"إهلاك الأصل {x['name']}"},
            {"account_id":x["accumulated_depr_account_id"],"credit":x["depreciation_amount"],
             "line_description":f"مجمع إهلاك الأصل {x['name']}"},
        ])
    jid=create_system_journal(run["period_date"],f"إهلاك الأصول {run['run_no']}",
                              run["run_no"],"ASSET_DEPRECIATION",run_id,journal_lines)
    execute("UPDATE asset_depreciation_runs SET journal_id=:jid,status='مرحّل' WHERE id=:id",
            {"jid":jid,"id":run_id})
    return jid



def next_payroll_run_number(period_end):
    year = period_end.year if hasattr(period_end,"year") else datetime.strptime(period_end,"%Y-%m-%d").year
    count = db.session.execute(text("""SELECT COUNT(*) FROM payroll_runs
        WHERE EXTRACT(YEAR FROM period_end)=:y"""),{"y":year}).scalar() or 0
    return f"PAY-{year}-{count+1:05d}"

def payroll_configuration():
    config=row("SELECT * FROM payroll_settings WHERE id=1")
    if not config:
        execute("""INSERT INTO payroll_settings(id,working_days_per_month,
                   working_hours_per_day,overtime_multiplier)
                   VALUES(1,30,8,1.5)""")
        config=row("SELECT * FROM payroll_settings WHERE id=1")
    return config

def calculate_employee_payroll(employee, period_start, period_end):
    cfg=payroll_configuration()
    working_days=float(cfg["working_days_per_month"] or 30)
    hours_per_day=float(cfg["working_hours_per_day"] or 8)
    overtime_multiplier=float(cfg["overtime_multiplier"] or 1.5)

    attendance=row("""SELECT
      COALESCE(SUM(overtime_hours),0) overtime_hours,
      COALESCE(SUM(CASE WHEN status='غائب' THEN 1 ELSE 0 END),0) absence_days,
      COALESCE(SUM(absence_hours),0) absence_hours
      FROM attendance_records
      WHERE employee_id=:employee AND attendance_date BETWEEN :start AND :end""",
      {"employee":employee["id"],"start":period_start,"end":period_end})

    basic=float(employee["basic_salary"] or 0)
    allowances=(float(employee.get("allowances") or 0)
                +float(employee.get("housing_allowance") or 0)
                +float(employee.get("transport_allowance") or 0)
                +float(employee.get("other_allowance") or 0))
    hourly_rate=basic/working_days/hours_per_day if working_days and hours_per_day else 0
    overtime_hours=float(attendance["overtime_hours"] or 0)
    overtime_amount=round(overtime_hours*hourly_rate*overtime_multiplier,2)
    absence_days=float(attendance["absence_days"] or 0)
    absence_hours=float(attendance["absence_hours"] or 0)
    absence_deduction=round((absence_days*basic/working_days)+(absence_hours*hourly_rate),2)

    adjustments=rows("""SELECT adjustment_type,amount FROM employee_salary_adjustments
      WHERE employee_id=:employee AND active=1 AND adjustment_date<=:end
      AND (recurring=1 OR adjustment_date BETWEEN :start AND :end)""",
      {"employee":employee["id"],"start":period_start,"end":period_end})
    extra_earnings=sum(float(x["amount"]) for x in adjustments if x["adjustment_type"]=="استحقاق")
    other_deductions=sum(float(x["amount"]) for x in adjustments if x["adjustment_type"]=="استقطاع")

    gross=round(basic+allowances+overtime_amount+extra_earnings,2)
    deductions=round(absence_deduction+other_deductions,2)
    net=round(max(gross-deductions,0),2)
    return {
        "basic_salary":basic,"allowances":round(allowances+extra_earnings,2),
        "overtime_hours":overtime_hours,"overtime_amount":overtime_amount,
        "absence_days":absence_days,"absence_deduction":absence_deduction,
        "other_deductions":round(other_deductions,2),"gross_salary":gross,
        "total_deductions":deductions,"net_salary":net
    }

def post_payroll_run(run_id):
    run=row("SELECT * FROM payroll_runs WHERE id=:id",{"id":run_id})
    if not run:
        raise ValueError("مسير الرواتب غير موجود.")
    if run.get("journal_id"):
        return run["journal_id"]
    cfg=payroll_configuration()
    required=["salary_expense_account_id","payroll_payable_account_id","deduction_account_id"]
    missing=[x for x in required if not cfg.get(x)]
    if missing:
        raise ValueError("أكمل ربط حسابات الرواتب من إعدادات الرواتب.")

    lines=[]
    for x in rows("""SELECT pl.*,e.name FROM payroll_lines pl
                     JOIN employees e ON e.id=pl.employee_id
                     WHERE pl.run_id=:id""",{"id":run_id}):
        common={"cost_center_id":x["cost_center_id"],
                "line_description":f"راتب الموظف {x['name']}"}
        lines.append({"account_id":cfg["salary_expense_account_id"],
                      "debit":x["gross_salary"],**common})
        lines.append({"account_id":cfg["payroll_payable_account_id"],
                      "credit":x["net_salary"],**common})
        if float(x["total_deductions"] or 0)>0:
            lines.append({"account_id":cfg["deduction_account_id"],
                          "credit":x["total_deductions"],**common})
    jid=create_system_journal(run["payment_date"] or run["period_end"],
                              f"مسير الرواتب {run['run_no']}",run["run_no"],
                              "PAYROLL",run_id,lines)
    execute("""UPDATE payroll_runs SET journal_id=:jid,status='مرحّل'
               WHERE id=:id""",{"jid":jid,"id":run_id})
    return jid



def next_project_number():
    count=db.session.execute(text("SELECT COUNT(*) FROM projects")).scalar() or 0
    return f"PRJ-{count+1:05d}"

def next_certificate_number(cert_date):
    year=cert_date.year if hasattr(cert_date,"year") else datetime.strptime(cert_date,"%Y-%m-%d").year
    count=db.session.execute(text("""SELECT COUNT(*) FROM progress_certificates
        WHERE EXTRACT(YEAR FROM certificate_date)=:y"""),{"y":year}).scalar() or 0
    return f"IPC-{year}-{count+1:05d}"

def project_financial_summary(project_id):
    project=row("SELECT * FROM projects WHERE id=:id",{"id":project_id})
    revenue=row("""SELECT COALESCE(SUM(total),0) value
                   FROM progress_certificates
                   WHERE project_id=:id AND status IN ('معتمد','مفوتر')""",{"id":project_id})["value"]
    cost=row("""SELECT COALESCE(SUM(amount),0) value
                FROM project_cost_entries WHERE project_id=:id""",{"id":project_id})["value"]
    certified=row("""SELECT COALESCE(SUM(gross_work_value),0) value
                     FROM progress_certificates
                     WHERE project_id=:id AND status IN ('معتمد','مفوتر')""",{"id":project_id})["value"]
    contract=float(project["contract_value"] or 0) if project else 0
    completion=round(float(certified or 0)/contract*100,2) if contract else 0
    return {
        "revenue":round(float(revenue or 0),2),
        "cost":round(float(cost or 0),2),
        "profit":round(float(revenue or 0)-float(cost or 0),2),
        "certified":round(float(certified or 0),2),
        "completion":min(completion,100)
    }



def next_crm_number(table_name, prefix):
    allowed={"crm_leads":"lead_no","crm_opportunities":"opportunity_no"}
    if table_name not in allowed:
        raise ValueError("Invalid CRM sequence")
    count=db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar() or 0
    return f"{prefix}-{count+1:06d}"

def crm_pipeline_summary():
    return rows("""SELECT stage,COUNT(*) count,
      COALESCE(SUM(estimated_value),0) value,
      COALESCE(SUM(weighted_value),0) weighted
      FROM crm_opportunities WHERE status='مفتوحة'
      GROUP BY stage ORDER BY
      CASE stage WHEN 'تأهيل' THEN 1 WHEN 'عرض سعر' THEN 2
      WHEN 'تفاوض' THEN 3 WHEN 'فوز' THEN 4 WHEN 'خسارة' THEN 5 ELSE 9 END""")



SECURITY_MODULES = {
    "dashboard":"لوحة التحكم","accounting":"المحاسبة","sales":"المبيعات",
    "purchases":"المشتريات","inventory":"المخزون","treasury":"الخزينة",
    "reports":"التقارير","assets":"الأصول الثابتة","payroll":"الموارد البشرية والرواتب",
    "projects":"المشاريع","crm":"إدارة علاقات العملاء","settings":"الإعدادات",
    "users":"المستخدمون والصلاحيات"
}
SECURITY_ACTIONS = {
    "view":"عرض","create":"إضافة","edit":"تعديل","delete":"حذف",
    "approve":"اعتماد","post":"ترحيل","print":"طباعة","export":"تصدير"
}

ENDPOINT_MODULE_MAP = {
    "dashboard":"dashboard",
    "journal_entries":"accounting","journal_view":"accounting","multi_journal":"accounting",
    "chart_of_accounts":"accounting","accounting_report":"accounting",
    "invoice_edit":"sales","invoice_submit_approval":"sales","invoice_return_to_draft":"sales","invoice_approve":"sales","invoice_copy":"sales","invoice_delete":"sales",
    "sales_center":"sales","sales_dashboard":"sales","sales_quotations":"sales",
    "sales_orders":"sales","sales_deliveries":"sales","invoices":"sales",
    "sales_returns":"sales","sales_tracking":"sales",
    "procurement_center":"purchases","purchase_requisitions":"purchases",
    "purchase_orders":"purchases","goods_receipts":"purchases","supplier_invoices":"purchases",
    "inventory":"inventory","inventory_item_view":"inventory","inventory_movements":"inventory",
    "inventory_transfers":"inventory","inventory_counts":"inventory",
    "treasury":"treasury","treasury_view":"treasury","receivables":"treasury",
    "bi_dashboard":"reports","bi_finance":"reports","bi_projects":"reports","bi_hr":"reports","bi_export":"reports",
    "reports":"reports","reports_center":"reports","financial_statements_center":"reports",
    "income_statement":"reports","balance_sheet":"reports","cash_flow_statement":"reports",
    "fixed_assets_center":"assets","fixed_assets":"assets","asset_categories":"assets",
    "asset_depreciation":"assets","fixed_assets_report":"assets",
    "hr_complete_center":"payroll",
    "hr_employee_profile":"payroll",
    "hr_contracts":"payroll",
    "hr_leave_types":"payroll",
    "hr_leaves":"payroll",
    "hr_leave_approve":"payroll",
    "hr_documents":"payroll",
    "hr_warnings":"payroll",
    "hr_end_of_service":"payroll",
    "hr_recruitment":"payroll",
    "hr_report":"payroll",
    "hr_report_export":"payroll",
    "hr_payroll_center":"payroll","departments":"payroll","attendance":"payroll",
    "salary_adjustments":"payroll","payroll_runs":"payroll","payroll_report":"payroll",
    "projects_center":"projects","project_new":"projects","project_view":"projects",
    "crm_center":"crm","crm_leads":"crm","crm_opportunities":"crm",
    "crm_activities":"crm","crm_pipeline":"crm","crm_report":"crm",
    "documents_center":"settings","document_categories":"settings","document_new":"settings","document_view":"settings","document_add_version":"settings","document_status_update":"settings","documents_report":"reports","documents_report_export":"reports",
    "data_import_center":"settings","data_import_template":"settings","data_import_upload":"settings","data_import_confirm":"settings","data_import_errors":"settings","data_import_history_detail":"settings",
    "global_search":"dashboard","smart_lookup":"dashboard","smart_entity_details":"dashboard","quick_create":"dashboard",
    "settings":"settings","branches":"settings","cost_centers":"settings",
    "users_admin":"users","roles_admin":"users","security_audit":"users",
}

def seed_security_data():
    for module_key,module_name in SECURITY_MODULES.items():
        for action_key,action_name in SECURITY_ACTIONS.items():
            permission_key=f"{module_key}.{action_key}"
            db.session.execute(text("""INSERT INTO system_permissions(
                permission_key,module_name,action_name,display_name)
                VALUES(:key,:module,:action,:display)
                ON CONFLICT(permission_key) DO NOTHING"""),
                {"key":permission_key,"module":module_key,"action":action_key,
                 "display":f"{action_name} - {module_name}"})
    db.session.execute(text("""INSERT INTO system_roles(code,name,description,is_system,active)
        VALUES('ADMIN','مدير النظام','صلاحيات كاملة',1,1)
        ON CONFLICT(code) DO NOTHING"""))
    admin_role=db.session.execute(text("SELECT id FROM system_roles WHERE code='ADMIN'")).scalar()
    db.session.execute(text("""INSERT INTO role_permissions(role_id,permission_id,allowed)
        SELECT :role,id,1 FROM system_permissions
        ON CONFLICT(role_id,permission_id) DO UPDATE SET allowed=1"""),{"role":admin_role})
    admin_user=db.session.execute(text("SELECT id FROM users WHERE role='admin' ORDER BY id LIMIT 1")).scalar()
    if admin_user:
        db.session.execute(text("""INSERT INTO user_role_assignments(user_id,role_id)
            VALUES(:user,:role) ON CONFLICT DO NOTHING"""),
            {"user":admin_user,"role":admin_role})
    db.session.commit()

def user_permission_keys(user_id):
    if not user_id:
        return set()
    role_name=db.session.execute(text("SELECT role FROM users WHERE id=:id"),{"id":user_id}).scalar()
    if role_name=="admin":
        return {"*"}
    result=db.session.execute(text("""SELECT DISTINCT p.permission_key
      FROM user_role_assignments ura
      JOIN system_roles r ON r.id=ura.role_id AND r.active=1
      JOIN role_permissions rp ON rp.role_id=r.id AND rp.allowed=1
      JOIN system_permissions p ON p.id=rp.permission_id
      WHERE ura.user_id=:user"""),{"user":user_id}).scalars().all()
    return set(result)

def has_permission(permission_key):
    permissions=session.get("permission_keys")
    if permissions is None:
        permissions=list(user_permission_keys(session.get("user_id")))
        session["permission_keys"]=permissions
    return "*" in permissions or permission_key in permissions

def infer_security_action():
    if request.method=="GET":
        endpoint=request.endpoint or ""
        if "export" in endpoint:
            return "export"
        if "print" in endpoint or "payslip" in endpoint:
            return "print"
        return "view"
    endpoint=request.endpoint or ""
    if any(x in endpoint for x in ("approve","post")):
        return "approve" if "approve" in endpoint else "post"
    if any(x in endpoint for x in ("delete","remove")):
        return "delete"
    if any(x in endpoint for x in ("edit","update")):
        return "edit"
    return "create"

def permission_required(permission_key):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            if not has_permission(permission_key):
                db.session.execute(text("""INSERT INTO access_denied_logs(
                    user_id,username,endpoint,permission_key,ip_address,created_at)
                    VALUES(:user,:username,:endpoint,:permission,:ip,:created)"""),
                    {"user":session.get("user_id"),"username":session.get("username"),
                     "endpoint":request.endpoint,"permission":permission_key,
                     "ip":request.headers.get("X-Forwarded-For",request.remote_addr),
                     "created":datetime.now()})
                db.session.commit()
                flash("ليس لديك صلاحية لتنفيذ هذه العملية","danger")
                return redirect(url_for("dashboard"))
            return fn(*args,**kwargs)
        return wrapper
    return decorator



DMS_ENTITY_TYPES = [
    "عام", "عميل", "مورد", "موظف", "مشروع", "عقد",
    "فاتورة مبيعات", "فاتورة مورد", "طلب شراء", "أمر شراء",
    "استلام بضاعة", "صنف مخزون", "أصل ثابت", "قيد محاسبي", "مركز تكلفة"
]

APPROVAL_DOCUMENT_TYPES = {
    "طلب شراء":"purchase_requisition",
    "أمر شراء":"purchase_order",
    "فاتورة مورد":"supplier_invoice",
    "مصروف":"expense",
    "سند صرف":"treasury_payment",
    "عرض سعر":"sales_quotation",
    "أمر بيع":"sales_order",
    "مستخلص مشروع":"progress_certificate",
    "أمر تغيير":"variation_order",
}

def next_approval_request_no():
    count=db.session.execute(text("SELECT COUNT(*) FROM approval_requests")).scalar() or 0
    return f"APR-{count+1:07d}"

def resolve_approval_workflow(document_type, amount):
    return row("""SELECT * FROM approval_workflows
                  WHERE document_type=:doc AND active=1
                  AND min_amount<=:amount
                  AND (max_amount IS NULL OR max_amount>=:amount)
                  ORDER BY min_amount DESC,id DESC LIMIT 1""",
               {"doc":document_type,"amount":float(amount or 0)})

def submit_for_approval(document_type, document_id, document_no, amount, notes=""):
    existing=row("""SELECT * FROM approval_requests
                    WHERE document_type=:doc AND document_id=:id""",
                 {"doc":document_type,"id":document_id})
    if existing:
        return existing["id"]
    workflow=resolve_approval_workflow(document_type,amount)
    if not workflow:
        raise ValueError(f"لا يوجد مسار اعتماد نشط للمستند: {document_type}")
    no=next_approval_request_no()
    execute("""INSERT INTO approval_requests(request_no,workflow_id,document_type,
      document_id,document_no,document_amount,requester_user_id,current_step_order,
      status,submitted_at,notes)
      VALUES(:no,:workflow,:doc,:doc_id,:doc_no,:amount,:user,1,'قيد الاعتماد',
      :submitted,:notes)""",
      {"no":no,"workflow":workflow["id"],"doc":document_type,
       "doc_id":document_id,"doc_no":document_no,"amount":float(amount or 0),
       "user":session.get("user_id"),"submitted":datetime.now(),"notes":notes})
    request_id=row("SELECT id FROM approval_requests WHERE request_no=:no",{"no":no})["id"]
    audit("SUBMIT_APPROVAL",document_type,f"إرسال {document_no} للاعتماد")
    return request_id

def approval_user_can_act(approval_request):
    step=row("""SELECT s.*,r.code role_code FROM approval_workflow_steps s
                LEFT JOIN system_roles r ON r.id=s.role_id
                WHERE s.workflow_id=:workflow AND s.step_order=:step""",
             {"workflow":approval_request["workflow_id"],
              "step":approval_request["current_step_order"]})
    if not step:
        return False,None
    username=session.get("username","")
    if step["approver_username"] and step["approver_username"]==username:
        return True,step
    if step["role_id"]:
        count=db.session.execute(text("""SELECT COUNT(*) FROM user_role_assignments
          WHERE user_id=:user AND role_id=:role"""),
          {"user":session.get("user_id"),"role":step["role_id"]}).scalar() or 0
        if count:
            return True,step
    if "*" in session.get("permission_keys",[]):
        return True,step
    return False,step

def apply_approval_action(request_id, action_type, notes=""):
    req=row("SELECT * FROM approval_requests WHERE id=:id",{"id":request_id})
    if not req:
        raise ValueError("طلب الاعتماد غير موجود.")
    if req["status"] not in ("قيد الاعتماد","معاد للتعديل"):
        raise ValueError("طلب الاعتماد مغلق.")
    allowed,step=approval_user_can_act(req)
    if not allowed:
        raise PermissionError("ليس لديك صلاحية تنفيذ خطوة الاعتماد الحالية.")
    execute("""INSERT INTO approval_actions(request_id,step_id,action_by_user_id,
      action_type,action_notes,action_at)
      VALUES(:request,:step,:user,:action,:notes,:dt)""",
      {"request":request_id,"step":step["id"] if step else None,
       "user":session.get("user_id"),"action":action_type,
       "notes":notes,"dt":datetime.now()})
    if action_type=="رفض":
        execute("""UPDATE approval_requests SET status='مرفوض',completed_at=:dt
                   WHERE id=:id""",{"dt":datetime.now(),"id":request_id})
    elif action_type=="إعادة":
        execute("""UPDATE approval_requests SET status='معاد للتعديل'
                   WHERE id=:id""",{"id":request_id})
    else:
        next_step=row("""SELECT * FROM approval_workflow_steps
                         WHERE workflow_id=:workflow AND step_order>:step
                         ORDER BY step_order LIMIT 1""",
                      {"workflow":req["workflow_id"],
                       "step":req["current_step_order"]})
        if next_step:
            execute("""UPDATE approval_requests SET current_step_order=:next,
                       status='قيد الاعتماد' WHERE id=:id""",
                    {"next":next_step["step_order"],"id":request_id})
        else:
            execute("""UPDATE approval_requests SET status='معتمد',completed_at=:dt
                       WHERE id=:id""",{"dt":datetime.now(),"id":request_id})
    audit("APPROVAL_ACTION",req["document_type"],
          f"{action_type} على {req['document_no']}")
    return req["document_type"],req["document_id"]

def approval_request_for(document_type,document_id):
    return row("""SELECT ar.*,aw.name workflow_name
                  FROM approval_requests ar
                  JOIN approval_workflows aw ON aw.id=ar.workflow_id
                  WHERE ar.document_type=:doc AND ar.document_id=:id""",
               {"doc":document_type,"id":document_id})



def next_budget_number(fiscal_year):
    count=db.session.execute(text("""SELECT COUNT(*) FROM budget_headers
        WHERE fiscal_year=:y"""),{"y":fiscal_year}).scalar() or 0
    return f"BUD-{fiscal_year}-{count+1:04d}"

def budget_actual_amount(account_id, fiscal_year, month_no, cost_center_id=None, branch_id=None):
    conditions=["l.account_id=:account","EXTRACT(YEAR FROM j.journal_date)=:year",
                "EXTRACT(MONTH FROM j.journal_date)=:month","j.status='مرحّل'"]
    params={"account":account_id,"year":fiscal_year,"month":month_no}
    if cost_center_id:
        conditions.append("l.cost_center_id=:cc")
        params["cc"]=cost_center_id
    data=row(f"""SELECT COALESCE(SUM(l.debit-l.credit),0) amount
                 FROM journal_entry_lines l
                 JOIN journal_entries j ON j.id=l.journal_id
                 WHERE {' AND '.join(conditions)}""",params)
    return round(float(data["amount"] or 0),2)

def budget_summary_rows(budget_id):
    budget=row("SELECT * FROM budget_headers WHERE id=:id",{"id":budget_id})
    if not budget:
        return []
    lines=rows("""SELECT bl.*,a.account_code,a.account_name_ar,a.account_type,
      cc.code cost_center_code,cc.name cost_center_name
      FROM budget_lines bl
      JOIN chart_of_accounts a ON a.id=bl.account_id
      LEFT JOIN cost_centers cc ON cc.id=bl.cost_center_id
      WHERE bl.budget_id=:id ORDER BY a.account_code,bl.month_no""",{"id":budget_id})
    output=[]
    for x in lines:
        item=dict(x)
        actual=budget_actual_amount(x["account_id"],budget["fiscal_year"],x["month_no"],
                                    x["cost_center_id"],budget["branch_id"])
        # For revenue accounts, actual normally carries a credit sign.
        if x["account_type"] in ("إيراد","التزام","حقوق ملكية"):
            actual=-actual
        item["actual_amount"]=actual
        item["variance_amount"]=round(float(x["budget_amount"] or 0)-actual,2)
        item["variance_percent"]=round(
            item["variance_amount"]/float(x["budget_amount"] or 1)*100,2
        ) if float(x["budget_amount"] or 0) else 0
        output.append(item)
    return output



def next_hr_number(table_name, prefix):
    allowed={
        "leave_requests":"request_no",
        "employee_contracts":"contract_no",
        "recruitment_candidates":"candidate_no"
    }
    if table_name not in allowed:
        raise ValueError("Invalid HR sequence")
    count=db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar() or 0
    return f"{prefix}-{count+1:06d}"

def calculate_leave_days(start_date, end_date):
    start=datetime.strptime(str(start_date),"%Y-%m-%d").date()
    end=datetime.strptime(str(end_date),"%Y-%m-%d").date()
    if end < start:
        raise ValueError("تاريخ نهاية الإجازة يجب أن يكون بعد تاريخ البداية.")
    return (end-start).days+1

def calculate_end_of_service(employee_id, service_end_date, other_dues=0, deductions=0):
    employee=row("SELECT * FROM employees WHERE id=:id",{"id":employee_id})
    if not employee:
        raise ValueError("الموظف غير موجود.")
    start=employee.get("hire_date")
    if not start:
        contract=row("""SELECT start_date FROM employee_contracts
                        WHERE employee_id=:id ORDER BY start_date LIMIT 1""",{"id":employee_id})
        start=contract["start_date"] if contract else None
    if not start:
        raise ValueError("تاريخ بداية خدمة الموظف غير مسجل.")
    end=datetime.strptime(str(service_end_date),"%Y-%m-%d").date()
    start_date=start if hasattr(start,"year") else datetime.strptime(str(start),"%Y-%m-%d").date()
    service_days=max((end-start_date).days,0)
    service_years=service_days/365.25
    basic=float(employee.get("basic_salary") or 0)
    if service_years <= 5:
        gratuity=(basic/2)*service_years
    else:
        gratuity=(basic/2)*5 + basic*(service_years-5)

    current_year=end.year
    leave_balance=row("""SELECT COALESCE(SUM(remaining),0) remaining
                         FROM employee_leave_balances
                         WHERE employee_id=:employee AND year=:year""",
                      {"employee":employee_id,"year":current_year})
    remaining=float(leave_balance["remaining"] or 0)
    leave_amount=(basic/30)*remaining
    net=gratuity+leave_amount+float(other_dues or 0)-float(deductions or 0)
    return {
        "service_start_date":start_date,
        "service_end_date":end,
        "service_years":round(service_years,4),
        "last_basic_salary":round(basic,2),
        "gratuity_amount":round(gratuity,2),
        "leave_balance_amount":round(leave_amount,2),
        "other_dues":round(float(other_dues or 0),2),
        "deductions":round(float(deductions or 0),2),
        "net_settlement":round(net,2),
    }



def next_contracting_number(table_name, prefix):
    allowed={
        "subcontractors":"code",
        "subcontract_contracts":"contract_no",
        "subcontract_certificates":"certificate_no",
        "variation_orders":"variation_no",
        "contract_extensions":"extension_no",
    }
    if table_name not in allowed:
        raise ValueError("Invalid contracting sequence")
    count=db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar() or 0
    return f"{prefix}-{count+1:06d}"

def subcontract_contract_summary(contract_id):
    contract=row("SELECT * FROM subcontract_contracts WHERE id=:id",{"id":contract_id})
    certified=row("""SELECT COALESCE(SUM(total),0) value
                     FROM subcontract_certificates
                     WHERE contract_id=:id AND status IN ('معتمد','مفوتر')""",
                  {"id":contract_id})["value"]
    retention=row("""SELECT COALESCE(SUM(retention_amount),0) value
                     FROM subcontract_certificates
                     WHERE contract_id=:id AND status IN ('معتمد','مفوتر')""",
                  {"id":contract_id})["value"]
    contract_value=float(contract["contract_value"] or 0) if contract else 0
    return {
        "certified":round(float(certified or 0),2),
        "retention":round(float(retention or 0),2),
        "remaining":round(contract_value-float(certified or 0),2),
        "completion":round(float(certified or 0)/contract_value*100,2) if contract_value else 0,
    }



def database_column_exists(table_name, column_name):
    """Return True when a PostgreSQL column exists in the current schema."""
    result = db.session.execute(text("""
        SELECT EXISTS(
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = current_schema()
              AND table_name = :table_name
              AND column_name = :column_name
        )
    """), {
        "table_name": table_name,
        "column_name": column_name
    }).scalar()
    return bool(result)


def database_table_exists(table_name):
    """Return True when a PostgreSQL table exists in the current schema."""
    result = db.session.execute(text("""
        SELECT EXISTS(
            SELECT 1
            FROM information_schema.tables
            WHERE table_schema = current_schema()
              AND table_name = :table_name
        )
    """), {"table_name": table_name}).scalar()
    return bool(result)


def safe_scalar(sql_text, params=None, default=0):
    """Run a scalar query without allowing a dashboard widget to crash the page."""
    try:
        result = db.session.execute(text(sql_text), params or {}).scalar()
        return default if result is None else result
    except Exception:
        db.session.rollback()
        return default


def ensure_hr_and_sales_schema():
    """Apply critical compatibility migrations for HR and sales dashboards."""
    statements = [
        "ALTER TABLE sales_returns ADD COLUMN IF NOT EXISTS cost_total NUMERIC(18,2) DEFAULT 0",
        "ALTER TABLE sales_returns ADD COLUMN IF NOT EXISTS status VARCHAR(40) DEFAULT 'معتمد'",
        "ALTER TABLE sales_return_items ADD COLUMN IF NOT EXISTS unit_cost NUMERIC(18,4) DEFAULT 0",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS name_en VARCHAR(255) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS department_id INTEGER",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS cost_center_id INTEGER",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS hire_date DATE",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS contract_end_date DATE",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS iqama_no VARCHAR(100) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS iqama_expiry DATE",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS passport_no VARCHAR(100) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS passport_expiry DATE",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS medical_insurance_expiry DATE",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS bank_iban VARCHAR(100) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS housing_allowance NUMERIC(18,2) DEFAULT 0",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS transport_allowance NUMERIC(18,2) DEFAULT 0",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS other_allowance NUMERIC(18,2) DEFAULT 0"
    ]
    for statement in statements:
        try:
            db.session.execute(text(statement))
            db.session.commit()
        except Exception:
            db.session.rollback()



def bi_safe_row(sql_text, params=None, defaults=None):
    """Return a row for BI widgets without allowing one widget to stop the dashboard."""
    try:
        result = row(sql_text, params or {})
        return result or (defaults or {})
    except Exception:
        db.session.rollback()
        return defaults or {}


def bi_safe_rows(sql_text, params=None):
    """Return rows for BI widgets and isolate database errors."""
    try:
        return rows(sql_text, params or {})
    except Exception:
        db.session.rollback()
        return []


def executive_dashboard_data():
    """Collect executive KPIs from all ERP modules."""
    warnings = []

    sales = bi_safe_row("""
        SELECT
          COALESCE(SUM(total),0) total_sales,
          COALESCE(SUM(vat),0) sales_vat,
          COUNT(*) invoice_count
        FROM invoices
        WHERE status='معتمدة'
          AND invoice_date >= DATE_TRUNC('year',CURRENT_DATE)
    """, defaults={"total_sales":0,"sales_vat":0,"invoice_count":0})

    purchases = bi_safe_row("""
        SELECT
          COALESCE(SUM(total),0) total_purchases,
          COALESCE(SUM(vat),0) purchase_vat,
          COUNT(*) invoice_count
        FROM supplier_invoices
        WHERE status IN ('معتمدة','مرحلة','مفوتر')
          AND invoice_date >= DATE_TRUNC('year',CURRENT_DATE)
    """, defaults={"total_purchases":0,"purchase_vat":0,"invoice_count":0})

    receivables = bi_safe_row("""
        SELECT COALESCE(SUM(
          i.total-COALESCE((SELECT SUM(a.allocated_amount)
          FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0)
        ),0) outstanding
        FROM invoices i WHERE i.status='معتمدة'
    """, defaults={"outstanding":0})

    payables = bi_safe_row("""
        SELECT COALESCE(SUM(
          si.total-COALESCE((SELECT SUM(a.allocated_amount)
          FROM supplier_payment_allocations a WHERE a.supplier_invoice_id=si.id),0)
        ),0) outstanding
        FROM supplier_invoices si
        WHERE si.status IN ('معتمدة','مرحلة','مفوتر')
    """, defaults={"outstanding":0})

    cash = bi_safe_row("""
        SELECT COALESCE(SUM(
          CASE WHEN transaction_type IN ('قبض','إيداع') THEN amount ELSE -amount END
        ),0) cash_balance
        FROM treasury_transactions
        WHERE status IN ('معتمد','مرحّل')
    """, defaults={"cash_balance":0})

    projects = bi_safe_row("""
        SELECT
          COUNT(*) FILTER (WHERE status='نشط') active_projects,
          COALESCE(SUM(contract_value),0) contract_value
        FROM projects
    """, defaults={"active_projects":0,"contract_value":0})

    hr = bi_safe_row("""
        SELECT
          COUNT(*) FILTER (WHERE active=1) active_employees,
          COALESCE(SUM(basic_salary),0) monthly_basic_payroll
        FROM employees
    """, defaults={"active_employees":0,"monthly_basic_payroll":0})

    inventory = bi_safe_row("""
        SELECT
          COUNT(*) item_count,
          COUNT(*) FILTER (WHERE active=1 AND quantity<=reorder_level) low_stock_count,
          COALESCE(SUM(quantity*COALESCE(unit_cost,0)),0) inventory_value
        FROM inventory
    """, defaults={"item_count":0,"low_stock_count":0,"inventory_value":0})

    gross_profit = round(
        float(sales.get("total_sales") or 0) - float(purchases.get("total_purchases") or 0), 2
    )

    monthly_sales = bi_safe_rows("""
        SELECT TO_CHAR(month_start,'YYYY-MM') month,
          COALESCE(SUM(i.total),0) sales
        FROM GENERATE_SERIES(
          DATE_TRUNC('month',CURRENT_DATE)-INTERVAL '11 months',
          DATE_TRUNC('month',CURRENT_DATE),
          INTERVAL '1 month'
        ) month_start
        LEFT JOIN invoices i
          ON DATE_TRUNC('month',i.invoice_date)=month_start
         AND i.status='معتمدة'
        GROUP BY month_start ORDER BY month_start
    """)

    monthly_purchases = bi_safe_rows("""
        SELECT TO_CHAR(month_start,'YYYY-MM') month,
          COALESCE(SUM(si.total),0) purchases
        FROM GENERATE_SERIES(
          DATE_TRUNC('month',CURRENT_DATE)-INTERVAL '11 months',
          DATE_TRUNC('month',CURRENT_DATE),
          INTERVAL '1 month'
        ) month_start
        LEFT JOIN supplier_invoices si
          ON DATE_TRUNC('month',si.invoice_date)=month_start
         AND si.status IN ('معتمدة','مرحلة','مفوتر')
        GROUP BY month_start ORDER BY month_start
    """)

    project_profitability = bi_safe_rows("""
        SELECT p.id,p.project_no,p.name,
          COALESCE((SELECT SUM(pc.total) FROM progress_certificates pc
                    WHERE pc.project_id=p.id AND pc.status IN ('معتمد','مفوتر')),0) revenue,
          COALESCE((SELECT SUM(pe.amount) FROM project_cost_entries pe
                    WHERE pe.project_id=p.id),0) cost
        FROM projects p
        ORDER BY revenue DESC LIMIT 8
    """)

    sales_by_customer = bi_safe_rows("""
        SELECT c.name,COALESCE(SUM(i.total),0) value
        FROM invoices i JOIN customers c ON c.id=i.customer_id
        WHERE i.status='معتمدة'
        GROUP BY c.id,c.name ORDER BY value DESC LIMIT 8
    """)

    overdue = bi_safe_row("""
        SELECT COUNT(*) overdue_count,
          COALESCE(SUM(i.total-COALESCE((SELECT SUM(a.allocated_amount)
            FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0)),0) overdue_value
        FROM invoices i
        WHERE i.status='معتمدة'
          AND COALESCE(i.due_date,i.invoice_date)<CURRENT_DATE
          AND i.total-COALESCE((SELECT SUM(a.allocated_amount)
            FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0)>0.005
    """, defaults={"overdue_count":0,"overdue_value":0})

    pending_approvals = bi_safe_row("""
        SELECT COUNT(*) count FROM approval_requests
        WHERE status IN ('قيد الاعتماد','معاد للتعديل')
    """, defaults={"count":0})

    expiring_hr = bi_safe_row("""
        SELECT COUNT(*) count FROM employees
        WHERE active=1 AND (
          iqama_expiry BETWEEN CURRENT_DATE AND CURRENT_DATE+INTERVAL '60 days'
          OR passport_expiry BETWEEN CURRENT_DATE AND CURRENT_DATE+INTERVAL '60 days'
          OR medical_insurance_expiry BETWEEN CURRENT_DATE AND CURRENT_DATE+INTERVAL '60 days'
          OR contract_end_date BETWEEN CURRENT_DATE AND CURRENT_DATE+INTERVAL '60 days'
        )
    """, defaults={"count":0})

    expiring_documents = bi_safe_row("""
        SELECT COUNT(*) count FROM documents_archive
        WHERE expiry_date BETWEEN CURRENT_DATE AND CURRENT_DATE+INTERVAL '60 days'
    """, defaults={"count":0})

    return {
        "sales": sales,
        "purchases": purchases,
        "receivables": receivables,
        "payables": payables,
        "cash": cash,
        "projects": projects,
        "hr": hr,
        "inventory": inventory,
        "gross_profit": gross_profit,
        "monthly_sales": monthly_sales,
        "monthly_purchases": monthly_purchases,
        "project_profitability": project_profitability,
        "sales_by_customer": sales_by_customer,
        "overdue": overdue,
        "pending_approvals": pending_approvals,
        "expiring_hr": expiring_hr,
        "expiring_documents": expiring_documents,
        "warnings": warnings,
    }



def prepare_invoice_items_from_form(form):
    item_names=form.getlist("item_name[]")
    descriptions=form.getlist("description[]")
    quantities=form.getlist("quantity[]")
    units=form.getlist("unit[]")
    unit_prices=form.getlist("unit_price[]")
    discounts=form.getlist("discount_rate[]")
    vat_rates=form.getlist("vat_rate[]")

    settings_row=row("SELECT vat_rate FROM settings WHERE id=1") or {}
    default_vat=float(settings_row.get("vat_rate") or 15)

    prepared=[]
    gross_subtotal=discount_total=vat_total=invoice_total=0.0
    for index,item_name in enumerate(item_names):
        item_name=(item_name or "").strip()
        if not item_name:
            continue
        quantity=float(quantities[index] or 0) if index<len(quantities) else 0
        unit_price=float(unit_prices[index] or 0) if index<len(unit_prices) else 0
        discount_rate=float(discounts[index] or 0) if index<len(discounts) else 0
        vat_rate=float(vat_rates[index] or default_vat) if index<len(vat_rates) else default_vat
        if quantity<=0:
            raise ValueError(f"الكمية في البند رقم {index+1} يجب أن تكون أكبر من صفر.")
        if unit_price<0 or discount_rate<0 or discount_rate>100 or vat_rate<0:
            raise ValueError(f"تحقق من السعر والخصم والضريبة في البند رقم {index+1}.")
        line_subtotal=round(quantity*unit_price,2)
        line_discount=round(line_subtotal*discount_rate/100,2)
        taxable=round(line_subtotal-line_discount,2)
        line_vat=round(taxable*vat_rate/100,2)
        line_total=round(taxable+line_vat,2)
        prepared.append({
            "item_name":item_name,
            "description":(descriptions[index] if index<len(descriptions) else "").strip(),
            "quantity":quantity,
            "unit":(units[index] if index<len(units) else "وحدة").strip() or "وحدة",
            "unit_price":unit_price,
            "discount_rate":discount_rate,
            "vat_rate":vat_rate,
            "line_subtotal":line_subtotal,
            "line_discount":line_discount,
            "line_vat":line_vat,
            "line_total":line_total,
        })
        gross_subtotal+=line_subtotal
        discount_total+=line_discount
        vat_total+=line_vat
        invoice_total+=line_total
    if not prepared:
        raise ValueError("يجب إضافة بند صحيح واحد على الأقل.")
    net_subtotal=round(gross_subtotal-discount_total,2)
    return prepared,round(gross_subtotal,2),round(discount_total,2),round(net_subtotal,2),round(vat_total,2),round(invoice_total,2)


def replace_invoice_items(invoice_id, prepared_items):
    execute("DELETE FROM invoice_items WHERE invoice_id=:id",{"id":invoice_id})
    for item in prepared_items:
        execute("""INSERT INTO invoice_items(
          invoice_id,item_name,description,quantity,unit,unit_price,
          discount_rate,vat_rate,line_subtotal,line_discount,line_vat,line_total)
          VALUES(:invoice_id,:item_name,:description,:quantity,:unit,:unit_price,
          :discount_rate,:vat_rate,:line_subtotal,:line_discount,:line_vat,:line_total)""",
          {"invoice_id":invoice_id,**item})


def invoice_is_editable(invoice):
    return bool(invoice and invoice.get("status") in ("مسودة","معاد للتعديل")
                and not invoice.get("journal_id"))



EXCEL_IMPORT_DEFINITIONS = {
    "customers": {
        "title": "العملاء",
        "headers": ["code","name","name_en","vat_no","phone","email","address","credit_limit"],
        "required": ["name"],
        "sample": ["CUST-001","شركة العميل","Customer Co.","300000000000003","0500000000","info@example.com","الدمام",50000],
    },
    "suppliers": {
        "title": "الموردون",
        "headers": ["code","name","name_en","vat_no","phone","email","address"],
        "required": ["name"],
        "sample": ["SUP-001","شركة المورد","Supplier Co.","300000000000003","0500000000","sales@example.com","الرياض"],
    },
    "inventory": {
        "title": "الأصناف",
        "headers": ["code","name","description","unit","quantity","unit_cost","reorder_level","active"],
        "required": ["name"],
        "sample": ["ITEM-001","كابل كهربائي","وصف الصنف","متر",100,25,20,1],
    },
    "employees": {
        "title": "الموظفون",
        "headers": ["employee_no","name","name_en","job_title","basic_salary","phone","email","hire_date","nationality"],
        "required": ["name"],
        "sample": ["EMP-001","اسم الموظف","Employee Name","محاسب",5000,"0500000000","employee@example.com","2026-01-01","سوداني"],
    },
    "accounts": {
        "title": "دليل الحسابات",
        "headers": ["account_code","account_name_ar","account_name_en","account_type","parent_code","accepts_entries","active"],
        "required": ["account_code","account_name_ar","account_type"],
        "sample": ["110101","النقدية","Cash","أصل","1101",1,1],
    },
    "cost_centers": {
        "title": "مراكز التكلفة",
        "headers": ["code","name","parent_code","active"],
        "required": ["code","name"],
        "sample": ["CC-001","المشروع الرئيسي","",1],
    },
}

IMPORT_HEADER_ALIASES = {
    "code": ["code", "customer code", "supplier code", "item code", "الكود", "كود", "رقم العميل", "رقم المورد", "رمز الصنف"],
    "name": ["name", "customer", "customer name", "client", "client name", "supplier", "supplier name", "item", "item name", "اسم", "اسم العميل", "العميل", "اسم المورد", "المورد", "اسم الصنف", "الصنف"],
    "name_en": ["name en", "english name", "name english", "الاسم الانجليزي", "الاسم بالانجليزية"],
    "vat_no": ["vat", "vat no", "vat number", "tax no", "tax number", "trn", "الرقم الضريبي", "رقم ضريبي"],
    "phone": ["phone", "mobile", "telephone", "tel", "cell", "contact no", "الجوال", "الهاتف", "رقم الجوال", "رقم الهاتف"],
    "email": ["email", "e-mail", "mail", "البريد", "البريد الالكتروني", "الايميل"],
    "address": ["address", "location", "العنوان", "الموقع"],
    "credit_limit": ["credit limit", "limit", "حد الائتمان", "الحد الائتماني"],
    "description": ["description", "details", "الوصف", "البيان"],
    "unit": ["unit", "uom", "الوحدة", "وحدة القياس"],
    "quantity": ["quantity", "qty", "stock", "الكمية", "الرصيد"],
    "unit_cost": ["unit cost", "cost", "average cost", "تكلفة الوحدة", "التكلفة"],
    "reorder_level": ["reorder level", "minimum stock", "حد اعادة الطلب", "الحد الادنى"],
    "active": ["active", "enabled", "نشط", "فعال"],
    "employee_no": ["employee no", "employee number", "emp no", "رقم الموظف", "كود الموظف"],
    "job_title": ["job title", "position", "المسمى الوظيفي", "الوظيفة"],
    "basic_salary": ["basic salary", "salary", "الراتب الاساسي", "الراتب"],
    "hire_date": ["hire date", "joining date", "date of joining", "تاريخ التعيين", "تاريخ المباشرة"],
    "nationality": ["nationality", "الجنسية"],
    "account_code": ["account code", "gl code", "كود الحساب", "رمز الحساب"],
    "account_name_ar": ["account name ar", "arabic account name", "اسم الحساب", "اسم الحساب عربي"],
    "account_name_en": ["account name en", "english account name", "اسم الحساب انجليزي"],
    "account_type": ["account type", "type", "نوع الحساب"],
    "parent_code": ["parent code", "parent account", "كود الحساب الاب", "الكود الاب"],
    "accepts_entries": ["accepts entries", "posting account", "يقبل القيود", "حساب حركة"],
}

def normalize_header_name(value):
    value = normalize_excel_value(value).lower().replace("_", " ").replace("-", " ")
    value = unicodedata.normalize("NFKC", value)
    value = re.sub(r"[\u064b-\u065f\u0670]", "", value)
    value = re.sub(r"[^\w\u0600-\u06ff ]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def clean_import_value(value, field=""):
    """Conservative cleaning: remove invisible characters and normalize common formats."""
    text_value=normalize_excel_value(value)
    text_value=unicodedata.normalize("NFKC",text_value)
    text_value=re.sub(r"[\u200b-\u200f\u202a-\u202e\ufeff]","",text_value).strip()
    text_value=re.sub(r"[ \t]+"," ",text_value)
    if field in {"phone","vat_no","employee_no"}:
        # Keep leading + for international phone numbers; VAT remains digits only.
        if field=="vat_no":
            return re.sub(r"\D","",text_value)
        return re.sub(r"[^0-9+]","",text_value)
    if field in {"credit_limit","quantity","unit_cost","reorder_level","basic_salary"}:
        return text_value.replace(",","")
    if field in {"active","accepts_entries"}:
        lowered=text_value.lower()
        if lowered in {"yes","true","active","enabled","نعم","نشط","فعال"}: return "1"
        if lowered in {"no","false","inactive","disabled","لا","غير نشط"}: return "0"
    return text_value


def _mapping_aliases(definition, module_name=None):
    aliases={}
    for field in definition["headers"]:
        aliases[normalize_header_name(field)]=(field,"اسم الحقل القياسي")
        for alias in IMPORT_HEADER_ALIASES.get(field,[]):
            aliases[normalize_header_name(alias)]=(field,"قاموس المرادفات")
    if module_name:
        try:
            saved=rows("""SELECT source_header,target_field FROM data_import_profile_aliases
                          WHERE module_name=:module""",{"module":module_name})
            for item in saved:
                aliases[normalize_header_name(item["source_header"])]=(item["target_field"],"مطابقة محفوظة")
        except Exception:
            db.session.rollback()
    return aliases


def smart_map_headers(raw_headers, definition, module_name=None, preferred_mapping=None):
    aliases=_mapping_aliases(definition,module_name)
    alias_names=list(aliases.keys())
    mapping,details,used={}, {}, set()
    preferred_mapping=preferred_mapping or {}
    for original in raw_headers:
        preferred=preferred_mapping.get(original)
        normalized=normalize_header_name(original)
        target=preferred if preferred in definition["headers"] else None
        confidence=100 if target else 0
        reason="قالب محفوظ" if target else ""
        if not target and normalized in aliases:
            target,reason=aliases[normalized]
            confidence=100
        if not target and normalized:
            candidates=difflib.get_close_matches(normalized,alias_names,n=1,cutoff=.72)
            if candidates:
                candidate=candidates[0]
                proposed,why=aliases[candidate]
                ratio=round(difflib.SequenceMatcher(None,normalized,candidate).ratio()*100)
                if proposed not in used:
                    target,confidence,reason=proposed,ratio,"مطابقة تقريبية"
        if target and target not in used:
            mapping[original]=target
            used.add(target)
            details[original]={"target":target,"confidence":confidence,"reason":reason}
        else:
            mapping[original]=""
            details[original]={"target":"","confidence":0,"reason":"غير مستخدم"}
    return mapping,details


def read_import_file(uploaded):
    filename = secure_filename(uploaded.filename or "")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "xlsx":
        workbook = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        first = next(iterator, None)
        if not first:
            raise ValueError("الملف فارغ.")
        return filename, [normalize_excel_value(v) for v in first], [list(v) for v in iterator]
    if extension == "csv":
        raw = uploaded.read()
        decoded = None
        for encoding in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                pass
        if decoded is None:
            raise ValueError("تعذر التعرف على ترميز ملف CSV.")
        sample=decoded[:4096]
        try:
            dialect=csv.Sniffer().sniff(sample,delimiters=",;\t|")
        except csv.Error:
            dialect=csv.excel
        parsed = list(csv.reader(io.StringIO(decoded),dialect))
        if not parsed:
            raise ValueError("الملف فارغ.")
        return filename, [normalize_excel_value(v) for v in parsed[0]], parsed[1:]
    raise ValueError("الأنواع المدعومة هي XLSX وCSV فقط.")


def import_quality_breakdown(preview, definition):
    if not preview:
        return {"overall":0,"completeness":0,"validity":0,"uniqueness":0,"formatting":0}
    required=definition.get("required",[])
    required_cells=max(len(preview)*max(len(required),1),1)
    completed=sum(1 for item in preview for f in required if str(item.get("data",{}).get(f,"")).strip())
    completeness=round(completed/required_cells*100)
    validity=round(sum(1 for item in preview if not item.get("errors"))/len(preview)*100)
    duplicates=sum(1 for item in preview if item.get("duplicate_status") in {"داخل الملف","موجود بالنظام"})
    uniqueness=round(max(0,1-duplicates/len(preview))*100)
    warning_rows=sum(1 for item in preview if item.get("warnings"))
    formatting=round(max(0,1-warning_rows/len(preview))*100)
    overall=round(completeness*.30+validity*.40+uniqueness*.20+formatting*.10)
    return {"overall":overall,"completeness":completeness,"validity":validity,"uniqueness":uniqueness,"formatting":formatting}


def import_quality_score(preview, definition=None):
    definition=definition or {"required":[]}
    return import_quality_breakdown(preview,definition)["overall"]


def find_existing_import_record(module_name,data):
    checks={
      "customers":("customers",[("code","code"),("vat_number","vat_no"),("email","email"),("name","name")]),
      "suppliers":("suppliers",[("code","code"),("vat_number","vat_no"),("email","email"),("name","name")]),
      "inventory":("inventory",[("code","code"),("name","name")]),
      "employees":("employees",[("employee_no","employee_no"),("email","email"),("name","name")]),
      "accounts":("chart_of_accounts",[("account_code","account_code")]),
      "cost_centers":("cost_centers",[("code","code"),("name","name")]),
    }
    if module_name not in checks: return None
    table,fields=checks[module_name]
    clauses=[]; params={}
    for db_field,input_field in fields:
        value=str(data.get(input_field) or "").strip()
        if value:
            key=f"v{len(params)}"; params[key]=value
            clauses.append(f"LOWER(CAST({db_field} AS TEXT))=LOWER(:{key})")
    if not clauses: return None
    try:
        return row(f"SELECT id FROM {table} WHERE {' OR '.join(clauses)} ORDER BY id LIMIT 1",params)
    except Exception:
        db.session.rollback(); return None


def preview_cache_path(token):
    return os.path.join(tempfile.gettempdir(),f"west_erp_import_{token}.json")


def save_import_preview(payload):
    token=uuid.uuid4().hex
    with open(preview_cache_path(token),"w",encoding="utf-8") as handle:
        json.dump(payload,handle,ensure_ascii=False,default=str)
    session["import_preview_token"]=token
    return token


def load_import_preview():
    token=session.get("import_preview_token")
    if not token: return {}
    try:
        with open(preview_cache_path(token),encoding="utf-8") as handle:
            return json.load(handle)
    except (OSError,ValueError):
        return {}


def clear_import_preview():
    token=session.pop("import_preview_token",None)
    if token:
        try: os.remove(preview_cache_path(token))
        except OSError: pass

def next_import_number():
    count=db.session.execute(text("SELECT COUNT(*) FROM data_import_jobs")).scalar() or 0
    return f"IMP-{count+1:07d}"

def normalize_excel_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value).strip()

def validate_import_row(module_name, row_data, row_no):
    definition=EXCEL_IMPORT_DEFINITIONS[module_name]
    errors=[]
    warnings=[]
    for field in definition["required"]:
        if not str(row_data.get(field,"")).strip():
            errors.append(f"الحقل {field} مطلوب")
    email=str(row_data.get("email") or "").strip()
    if email and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        errors.append("البريد الإلكتروني غير صحيح")
    vat=re.sub(r"\D", "", str(row_data.get("vat_no") or ""))
    if vat and len(vat)!=15:
        errors.append("الرقم الضريبي يجب أن يكون 15 رقمًا")
    phone=re.sub(r"\D", "", str(row_data.get("phone") or ""))
    if phone and len(phone)<8:
        warnings.append("رقم الهاتف قصير ويحتاج مراجعة")
    numeric_fields={"customers":["credit_limit"],"inventory":["quantity","unit_cost","reorder_level"],"employees":["basic_salary"]}.get(module_name,[])
    for field in numeric_fields:
        value=row_data.get(field)
        if value not in (None, ""):
            try:
                if float(value)<0:
                    errors.append(f"الحقل {field} لا يقبل قيمة سالبة")
            except (TypeError,ValueError):
                errors.append(f"الحقل {field} يجب أن يكون رقمًا")
    return errors,warnings

def import_excel_row(module_name, data, import_mode):
    updated=False
    if module_name=="customers":
        code=data.get("code") or None
        name_en=(data.get("name_en") or "").strip() or transliterate_arabic_name(data["name"])
        existing=row("""SELECT id FROM customers WHERE code=:code OR name=:name
                        ORDER BY id LIMIT 1""",{"code":code,"name":data["name"]}) if code else row(
                     "SELECT id FROM customers WHERE name=:name ORDER BY id LIMIT 1",{"name":data["name"]})
        if existing:
            if import_mode=="إضافة وتحديث":
                execute("""UPDATE customers SET name=:name,name_en=:name_en,vat_number=:vat,
                  phone=:phone,email=:email,address=:address,credit_limit=:limit
                  WHERE id=:id""",
                  {"name":data["name"],"name_en":name_en,
                   "vat":data.get("vat_no",""),"phone":data.get("phone",""),
                   "email":data.get("email",""),"address":data.get("address",""),
                   "limit":float(data.get("credit_limit") or 0),"id":existing["id"]})
                updated=True
            else:
                raise ValueError("العميل موجود مسبقًا.")
        else:
            execute("""INSERT INTO customers(code,name,name_en,vat_number,phone,email,
              address,credit_limit)
              VALUES(:code,:name,:name_en,:vat,:phone,:email,:address,:limit)""",
              {"code":code,"name":data["name"],"name_en":name_en,
               "vat":data.get("vat_no",""),"phone":data.get("phone",""),
               "email":data.get("email",""),"address":data.get("address",""),
               "limit":float(data.get("credit_limit") or 0)})

    elif module_name=="suppliers":
        code=data.get("code") or None
        name_en=(data.get("name_en") or "").strip() or transliterate_arabic_name(data["name"])
        existing=row("""SELECT id FROM suppliers WHERE code=:code OR name=:name
                        ORDER BY id LIMIT 1""",{"code":code,"name":data["name"]}) if code else row(
                     "SELECT id FROM suppliers WHERE name=:name ORDER BY id LIMIT 1",{"name":data["name"]})
        if existing:
            if import_mode=="إضافة وتحديث":
                execute("""UPDATE suppliers SET name=:name,name_en=:name_en,vat_number=:vat,
                  phone=:phone,email=:email,address=:address WHERE id=:id""",
                  {"name":data["name"],"name_en":name_en,
                   "vat":data.get("vat_no",""),"phone":data.get("phone",""),
                   "email":data.get("email",""),"address":data.get("address",""),
                   "id":existing["id"]})
                updated=True
            else:
                raise ValueError("المورد موجود مسبقًا.")
        else:
            execute("""INSERT INTO suppliers(code,name,name_en,vat_number,phone,email,
              address)
              VALUES(:code,:name,:name_en,:vat,:phone,:email,:address)""",
              {"code":code,"name":data["name"],"name_en":name_en,
               "vat":data.get("vat_no",""),"phone":data.get("phone",""),
               "email":data.get("email",""),"address":data.get("address","")})

    elif module_name=="inventory":
        code=(data.get("code") or "").strip() or next_inventory_sku()
        existing=row("""SELECT id FROM inventory WHERE code=:code OR name=:name
                        ORDER BY id LIMIT 1""",{"code":code,"name":data["name"]}) if code else row(
                     "SELECT id FROM inventory WHERE name=:name ORDER BY id LIMIT 1",{"name":data["name"]})
        payload={"code":code,"name":data["name"],"description":data.get("description",""),
                 "unit":normalize_item_unit(data.get("unit")),"quantity":float(data.get("quantity") or 0),
                 "unit_cost":float(data.get("unit_cost") or 0),
                 "reorder":float(data.get("reorder_level") or 0),
                 "active":int(float(data.get("active") or 1))}
        if existing:
            if import_mode=="إضافة وتحديث":
                execute("""UPDATE inventory SET name=:name,description=:description,
                  unit=:unit,quantity=:quantity,unit_cost=:unit_cost,
                  reorder_level=:reorder,active=:active WHERE id=:id""",
                  {**payload,"id":existing["id"]})
                updated=True
            else:
                raise ValueError("الصنف موجود مسبقًا.")
        else:
            execute("""INSERT INTO inventory(code,sku,name,description,unit,quantity,
              unit_cost,cost,reorder_level,active)
              VALUES(:code,:code,:name,:description,:unit,:quantity,:unit_cost,:unit_cost,:reorder,:active)""",
              payload)

    elif module_name=="employees":
        emp_no=data.get("employee_no") or None
        existing=row("""SELECT id FROM employees WHERE employee_no=:no OR name=:name
                        ORDER BY id LIMIT 1""",{"no":emp_no,"name":data["name"]}) if emp_no else row(
                     "SELECT id FROM employees WHERE name=:name ORDER BY id LIMIT 1",{"name":data["name"]})
        payload={"no":emp_no,"name":data["name"],
                 "name_en":(data.get("name_en") or "").strip() or transliterate_arabic_name(data["name"]),
                 "job":data.get("job_title",""),"salary":float(data.get("basic_salary") or 0),
                 "phone":data.get("phone",""),"email":data.get("email",""),
                 "hire":data.get("hire_date") or None,"nationality":data.get("nationality","")}
        if existing:
            if import_mode=="إضافة وتحديث":
                execute("""UPDATE employees SET name=:name,name_en=:name_en,
                  job_title=:job,basic_salary=:salary,phone=:phone,email=:email,
                  hire_date=:hire,nationality=:nationality WHERE id=:id""",
                  {**payload,"id":existing["id"]})
                updated=True
            else:
                raise ValueError("الموظف موجود مسبقًا.")
        else:
            execute("""INSERT INTO employees(employee_no,name,name_en,job_title,
              basic_salary,phone,email,hire_date,nationality,active)
              VALUES(:no,:name,:name_en,:job,:salary,:phone,:email,:hire,:nationality,1)""",
              payload)

    elif module_name=="accounts":
        existing=row("SELECT id FROM chart_of_accounts WHERE account_code=:code",
                     {"code":data["account_code"]})
        parent_id=None
        if data.get("parent_code"):
            parent=row("SELECT id FROM chart_of_accounts WHERE account_code=:code",
                       {"code":data["parent_code"]})
            if not parent:
                raise ValueError(f"الحساب الأب {data['parent_code']} غير موجود.")
            parent_id=parent["id"]
        payload={"code":data["account_code"],"ar":data["account_name_ar"],
                 "en":(data.get("account_name_en") or "").strip() or transliterate_arabic_name(data["account_name_ar"]),"type":data["account_type"],
                 "parent":parent_id,"accepts":int(float(data.get("accepts_entries") or 1)),
                 "active":int(float(data.get("active") or 1))}
        if existing:
            if import_mode=="إضافة وتحديث":
                execute("""UPDATE chart_of_accounts SET account_name_ar=:ar,
                  account_name_en=:en,account_type=:type,parent_id=:parent,
                  accepts_entries=:accepts,active=:active WHERE id=:id""",
                  {**payload,"id":existing["id"]})
                updated=True
            else:
                raise ValueError("الحساب موجود مسبقًا.")
        else:
            execute("""INSERT INTO chart_of_accounts(account_code,account_name_ar,
              account_name_en,account_type,parent_id,accepts_entries,active)
              VALUES(:code,:ar,:en,:type,:parent,:accepts,:active)""",payload)

    elif module_name=="cost_centers":
        existing=row("SELECT id FROM cost_centers WHERE code=:code",
                     {"code":data["code"]})
        parent_id=None
        if data.get("parent_code"):
            parent=row("SELECT id FROM cost_centers WHERE code=:code",
                       {"code":data["parent_code"]})
            if not parent:
                raise ValueError(f"مركز التكلفة الأب {data['parent_code']} غير موجود.")
            parent_id=parent["id"]
        payload={"code":data["code"],"name":data["name"],"parent":parent_id,
                 "active":int(float(data.get("active") or 1))}
        if existing:
            if import_mode=="إضافة وتحديث":
                execute("""UPDATE cost_centers SET name=:name,parent_id=:parent,
                           active=:active WHERE id=:id""",
                        {**payload,"id":existing["id"]})
                updated=True
            else:
                raise ValueError("مركز التكلفة موجود مسبقًا.")
        else:
            execute("""INSERT INTO cost_centers(code,name,parent_id,active)
                       VALUES(:code,:name,:parent,:active)""",payload)
    return updated



SMART_ENTITY_CONFIG = {
    "customers": {
        "table": "customers", "label": "العملاء",
        "display": "name", "code": "code", "tax": "vat_number",
        "account": "receivable_account_id"
    },
    "suppliers": {
        "table": "suppliers", "label": "الموردون",
        "display": "name", "code": "code", "tax": "vat_number",
        "account": "payable_account_id"
    },
    "items": {
        "table": "inventory", "label": "المواد",
        "display": "name", "code": "code", "tax": None,
        "account": None
    },
    "accounts": {
        "table": "chart_of_accounts", "label": "الحسابات",
        "display": "account_name_ar", "code": "account_code", "tax": None,
        "account": None
    },
    "cost_centers": {
        "table": "cost_centers", "label": "مراكز التكلفة",
        "display": "name", "code": "code", "tax": None,
        "account": None
    },
    "projects": {
        "table": "projects", "label": "المشاريع",
        "display": "name", "code": "project_no", "tax": None,
        "account": None
    },
}

def next_entity_code(table_name, prefix):
    count=db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar() or 0
    return f"{prefix}-{count+1:06d}"

def next_inventory_sku():
    """Generate one shared sequential item code for every item-creation screen."""
    if db.engine.dialect.name == "postgresql":
        db.session.execute(text("SELECT pg_advisory_xact_lock(2080001)"))
    # SKU is part of the original inventory table, so generation remains safe
    # even before optional compatibility columns are upgraded.
    values=rows("""SELECT COALESCE(sku,'') value FROM inventory
                   WHERE COALESCE(sku,'') ~ '^ITM-[0-9]+$'""") if db.engine.dialect.name=="postgresql" else rows(
                  "SELECT COALESCE(sku,'') value FROM inventory")
    highest=0
    for value in values:
        match=re.fullmatch(r"ITM-(\d+)",value["value"] or "")
        if match:
            highest=max(highest,int(match.group(1)))
    return f"ITM-{highest+1:06d}"

def normalize_item_unit(value):
    unit=(value or "وحدة").strip()
    if unit not in ITEM_UNITS:
        raise ValueError("نوع الوحدة المحدد غير صحيح.")
    return unit

def smart_entity_row(entity, entity_id):
    cfg=SMART_ENTITY_CONFIG.get(entity)
    if not cfg:
        return None
    fields=["id", f"{cfg['display']} AS name"]
    if cfg.get("code"):
        fields.append(f"{cfg['code']} AS code")
    if cfg.get("tax"):
        fields.append(f"{cfg['tax']} AS tax_number")
    if cfg.get("account"):
        fields.append(f"{cfg['account']} AS account_id")
    if entity=="items":
        fields.extend(["unit","unit_cost","quantity"])
    return row(f"SELECT {','.join(fields)} FROM {cfg['table']} WHERE id=:id",{"id":entity_id})

def create_quick_entity(entity, payload):
    name=(payload.get("name") or "").strip()
    if not name:
        raise ValueError("الاسم مطلوب.")
    if entity=="customer":
        code=(payload.get("code") or next_entity_code("customers","CUS")).strip()
        existing=row("SELECT id FROM customers WHERE code=:code OR name=:name",
                     {"code":code,"name":name})
        if existing:
            raise ValueError("العميل موجود مسبقًا.")
        execute("""INSERT INTO customers(code,name,name_en,vat_number,phone,email,address,
          credit_limit,receivable_account_id)
          VALUES(:code,:name,:name_en,:vat,:phone,:email,:address,:limit,:account)""",
          {"code":code,"name":name,"name_en":payload.get("name_en",""),
           "vat":payload.get("vat_number",""),"phone":payload.get("phone",""),
           "email":payload.get("email",""),"address":payload.get("address",""),
           "limit":float(payload.get("credit_limit") or 0),
           "account":payload.get("account_id") or None})
        new_id=row("SELECT id FROM customers WHERE code=:code",{"code":code})["id"]
        return "customers",smart_entity_row("customers",new_id)
    if entity=="supplier":
        code=(payload.get("code") or next_entity_code("suppliers","SUP")).strip()
        existing=row("SELECT id FROM suppliers WHERE code=:code OR name=:name",
                     {"code":code,"name":name})
        if existing:
            raise ValueError("المورد موجود مسبقًا.")
        execute("""INSERT INTO suppliers(code,name,name_en,vat_number,phone,email,address,
          payable_account_id)
          VALUES(:code,:name,:name_en,:vat,:phone,:email,:address,:account)""",
          {"code":code,"name":name,"name_en":payload.get("name_en",""),
           "vat":payload.get("vat_number",""),"phone":payload.get("phone",""),
           "email":payload.get("email",""),"address":payload.get("address",""),
           "account":payload.get("account_id") or None})
        new_id=row("SELECT id FROM suppliers WHERE code=:code",{"code":code})["id"]
        return "suppliers",smart_entity_row("suppliers",new_id)
    if entity=="item":
        code=next_inventory_sku()
        existing=row("SELECT id FROM inventory WHERE code=:code OR name=:name",
                     {"code":code,"name":name})
        if existing:
            raise ValueError("المادة موجودة مسبقًا.")
        execute("""INSERT INTO inventory(code,sku,name,description,unit,quantity,unit_cost,cost,
          reorder_level,active)
          VALUES(:code,:code,:name,:description,:unit,:quantity,:cost,:cost,:reorder,1)""",
          {"code":code,"name":name,"description":payload.get("description",""),
           "unit":normalize_item_unit(payload.get("unit")),
           "quantity":float(payload.get("quantity") or 0),
           "cost":float(payload.get("unit_cost") or 0),
           "reorder":float(payload.get("reorder_level") or 0)})
        new_id=row("SELECT id FROM inventory WHERE code=:code",{"code":code})["id"]
        return "items",smart_entity_row("items",new_id)
    raise ValueError("نوع السجل غير مدعوم.")


def init_db():
    statements = [
        """CREATE TABLE IF NOT EXISTS users(
            id SERIAL PRIMARY KEY,
            username VARCHAR(100) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            role VARCHAR(50) NOT NULL DEFAULT 'admin'
        )""",
        """CREATE TABLE IF NOT EXISTS settings(
            id INTEGER PRIMARY KEY,
            company_name_ar VARCHAR(255) DEFAULT 'اسم الشركة',
            company_name_en VARCHAR(255) DEFAULT 'Company Name',
            vat_number VARCHAR(50) DEFAULT '',
            cr_number VARCHAR(50) DEFAULT '',
            currency VARCHAR(10) DEFAULT 'SAR',
            address TEXT DEFAULT '',
            phone VARCHAR(50) DEFAULT '',
            email VARCHAR(255) DEFAULT '',
            logo_url TEXT DEFAULT '',
            logo_data TEXT DEFAULT '',
            logo_mime VARCHAR(50) DEFAULT '',
            vat_rate NUMERIC(5,2) DEFAULT 15.00
        )""",
        """CREATE TABLE IF NOT EXISTS branches(
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            city VARCHAR(150),
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS customers(
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            name_en VARCHAR(255) DEFAULT '',
            vat_number VARCHAR(50),
            phone VARCHAR(50),
            email VARCHAR(255),
            balance NUMERIC(18,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS suppliers(
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            name_en VARCHAR(255) DEFAULT '',
            vat_number VARCHAR(50),
            phone VARCHAR(50),
            email VARCHAR(255),
            balance NUMERIC(18,2) DEFAULT 0
        )""",

        """CREATE TABLE IF NOT EXISTS invoice_sequences(
            sequence_year INTEGER PRIMARY KEY,
            last_number INTEGER NOT NULL DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS invoices(
            id SERIAL PRIMARY KEY,
            invoice_no VARCHAR(100) UNIQUE NOT NULL,
            invoice_uuid VARCHAR(64),
            customer_id INTEGER NOT NULL REFERENCES customers(id),
            invoice_date DATE NOT NULL,
            subtotal NUMERIC(18,2) NOT NULL,
            vat NUMERIC(18,2) NOT NULL,
            total NUMERIC(18,2) NOT NULL,
            status VARCHAR(50) NOT NULL DEFAULT 'مسودة',
            branch_id INTEGER REFERENCES branches(id),
            notes TEXT,
            created_at TIMESTAMP NOT NULL
        )""",

        """CREATE TABLE IF NOT EXISTS invoice_items(
            id SERIAL PRIMARY KEY,
            invoice_id INTEGER NOT NULL REFERENCES invoices(id) ON DELETE CASCADE,
            item_name VARCHAR(255) NOT NULL,
            description TEXT DEFAULT '',
            quantity NUMERIC(18,3) NOT NULL DEFAULT 1,
            unit VARCHAR(50) DEFAULT 'وحدة',
            unit_price NUMERIC(18,2) NOT NULL DEFAULT 0,
            vat_rate NUMERIC(5,2) NOT NULL DEFAULT 15,
            line_subtotal NUMERIC(18,2) NOT NULL DEFAULT 0,
            line_vat NUMERIC(18,2) NOT NULL DEFAULT 0,
            line_total NUMERIC(18,2) NOT NULL DEFAULT 0
        )""",


        """CREATE TABLE IF NOT EXISTS chart_of_accounts(
            id SERIAL PRIMARY KEY,
            account_code VARCHAR(50) UNIQUE NOT NULL,
            account_name_ar VARCHAR(255) NOT NULL,
            account_name_en VARCHAR(255) DEFAULT '',
            account_type VARCHAR(50) NOT NULL,
            parent_id INTEGER REFERENCES chart_of_accounts(id),
            level INTEGER NOT NULL DEFAULT 1,
            accepts_entries INTEGER NOT NULL DEFAULT 1,
            normal_balance VARCHAR(20) DEFAULT 'مدين',
            statement_type VARCHAR(50) DEFAULT 'الميزانية العمومية',
            active INTEGER NOT NULL DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS cost_centers(
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            active INTEGER NOT NULL DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS journal_sequences(
            sequence_year INTEGER PRIMARY KEY,
            last_number INTEGER NOT NULL DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS journal_entries(
            id SERIAL PRIMARY KEY,
            journal_no VARCHAR(100) UNIQUE NOT NULL,
            journal_date DATE NOT NULL,
            reference VARCHAR(255) DEFAULT '',
            description TEXT DEFAULT '',
            status VARCHAR(50) NOT NULL DEFAULT 'مسودة',
            total_debit NUMERIC(18,2) NOT NULL DEFAULT 0,
            total_credit NUMERIC(18,2) NOT NULL DEFAULT 0,
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS journal_entry_lines(
            id SERIAL PRIMARY KEY,
            journal_id INTEGER NOT NULL REFERENCES journal_entries(id) ON DELETE CASCADE,
            account_id INTEGER NOT NULL REFERENCES chart_of_accounts(id),
            debit NUMERIC(18,2) NOT NULL DEFAULT 0,
            credit NUMERIC(18,2) NOT NULL DEFAULT 0,
            taxable INTEGER NOT NULL DEFAULT 0,
            tax_direction VARCHAR(50) DEFAULT 'غير مطبق',
            supplier_id INTEGER REFERENCES suppliers(id),
            customer_id INTEGER REFERENCES customers(id),
            party_type VARCHAR(20) DEFAULT '',
            tax_number VARCHAR(50) DEFAULT '',
            invoice_number VARCHAR(100) DEFAULT '',
            line_description TEXT DEFAULT '',
            cost_center_id INTEGER REFERENCES cost_centers(id)
        )""",

        """CREATE TABLE IF NOT EXISTS journal_batches(
            id SERIAL PRIMARY KEY,
            batch_no VARCHAR(100) UNIQUE NOT NULL,
            batch_date DATE NOT NULL,
            description TEXT DEFAULT '',
            status VARCHAR(50) NOT NULL DEFAULT 'مسودة',
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS journal_batch_groups(
            id SERIAL PRIMARY KEY,
            batch_id INTEGER NOT NULL REFERENCES journal_batches(id) ON DELETE CASCADE,
            group_no INTEGER NOT NULL,
            journal_no VARCHAR(100),
            total_debit NUMERIC(18,2) NOT NULL DEFAULT 0,
            total_credit NUMERIC(18,2) NOT NULL DEFAULT 0,
            status VARCHAR(50) NOT NULL DEFAULT 'مسودة'
        )""",
        """CREATE TABLE IF NOT EXISTS journal_batch_lines(
            id SERIAL PRIMARY KEY,
            group_id INTEGER NOT NULL REFERENCES journal_batch_groups(id) ON DELETE CASCADE,
            account_id INTEGER NOT NULL REFERENCES chart_of_accounts(id),
            debit NUMERIC(18,2) NOT NULL DEFAULT 0,
            credit NUMERIC(18,2) NOT NULL DEFAULT 0,
            taxable INTEGER NOT NULL DEFAULT 0,
            tax_direction VARCHAR(50) DEFAULT 'غير مطبق',
            party_type VARCHAR(20) DEFAULT '',
            supplier_id INTEGER REFERENCES suppliers(id),
            customer_id INTEGER REFERENCES customers(id),
            tax_number VARCHAR(50) DEFAULT '',
            invoice_number VARCHAR(100) DEFAULT '',
            invoice_date DATE,
            line_description TEXT DEFAULT '',
            cost_center_id INTEGER REFERENCES cost_centers(id)
        )""",
        """CREATE TABLE IF NOT EXISTS fiscal_years(
            id SERIAL PRIMARY KEY,name VARCHAR(100) NOT NULL,start_date DATE NOT NULL,
            end_date DATE NOT NULL,status VARCHAR(30) NOT NULL DEFAULT 'مفتوحة',
            UNIQUE(start_date,end_date)
        )""",
        """CREATE TABLE IF NOT EXISTS fiscal_periods(
            id SERIAL PRIMARY KEY,fiscal_year_id INTEGER NOT NULL REFERENCES fiscal_years(id) ON DELETE CASCADE,
            name VARCHAR(100) NOT NULL,start_date DATE NOT NULL,end_date DATE NOT NULL,
            status VARCHAR(30) NOT NULL DEFAULT 'مفتوحة'
        )""",
        """CREATE TABLE IF NOT EXISTS treasury_vouchers(
            id SERIAL PRIMARY KEY,
            voucher_no VARCHAR(100) UNIQUE NOT NULL,
            voucher_type VARCHAR(30) NOT NULL,
            voucher_date DATE NOT NULL,
            party_type VARCHAR(20) DEFAULT '',
            customer_id INTEGER REFERENCES customers(id),
            supplier_id INTEGER REFERENCES suppliers(id),
            cash_bank_account_id INTEGER NOT NULL REFERENCES chart_of_accounts(id),
            counter_account_id INTEGER NOT NULL REFERENCES chart_of_accounts(id),
            amount NUMERIC(18,2) NOT NULL,
            payment_method VARCHAR(30) DEFAULT 'نقدي',
            reference VARCHAR(100) DEFAULT '',
            description TEXT DEFAULT '',
            cost_center_id INTEGER REFERENCES cost_centers(id),
            status VARCHAR(30) DEFAULT 'مسودة',
            posting_status VARCHAR(30) DEFAULT 'غير مرحّل',
            journal_id INTEGER,
            reversed_by INTEGER,
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS purchase_requisitions(
            id SERIAL PRIMARY KEY,
            requisition_no VARCHAR(100) UNIQUE NOT NULL,
            requisition_date DATE NOT NULL,
            branch_id INTEGER REFERENCES branches(id),
            cost_center_id INTEGER REFERENCES cost_centers(id),
            requested_by VARCHAR(255) DEFAULT '',
            department VARCHAR(255) DEFAULT '',
            priority VARCHAR(30) DEFAULT 'عادية',
            reason TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'مسودة',
            total_estimated NUMERIC(18,2) DEFAULT 0,
            approved_by INTEGER,
            approved_at TIMESTAMP,
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS purchase_requisition_items(
            id SERIAL PRIMARY KEY,
            requisition_id INTEGER NOT NULL REFERENCES purchase_requisitions(id) ON DELETE CASCADE,
            item_code VARCHAR(100) DEFAULT '',
            item_name VARCHAR(255) NOT NULL,
            description TEXT DEFAULT '',
            quantity NUMERIC(18,3) NOT NULL,
            unit VARCHAR(50) DEFAULT '',
            estimated_price NUMERIC(18,2) DEFAULT 0,
            required_date DATE,
            suggested_supplier_id INTEGER REFERENCES suppliers(id)
        )""",
        """CREATE TABLE IF NOT EXISTS purchase_orders(
            id SERIAL PRIMARY KEY,
            po_no VARCHAR(100) UNIQUE NOT NULL,
            po_date DATE NOT NULL,
            requisition_id INTEGER REFERENCES purchase_requisitions(id),
            supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
            branch_id INTEGER REFERENCES branches(id),
            cost_center_id INTEGER REFERENCES cost_centers(id),
            payment_terms VARCHAR(255) DEFAULT '',
            delivery_terms VARCHAR(255) DEFAULT '',
            warehouse VARCHAR(255) DEFAULT '',
            notes TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'مسودة',
            subtotal NUMERIC(18,2) DEFAULT 0,
            vat NUMERIC(18,2) DEFAULT 0,
            total NUMERIC(18,2) DEFAULT 0,
            approved_by INTEGER,
            approved_at TIMESTAMP,
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS purchase_order_items(
            id SERIAL PRIMARY KEY,
            po_id INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
            item_code VARCHAR(100) DEFAULT '',
            item_name VARCHAR(255) NOT NULL,
            description TEXT DEFAULT '',
            quantity NUMERIC(18,3) NOT NULL,
            received_qty NUMERIC(18,3) DEFAULT 0,
            unit VARCHAR(50) DEFAULT '',
            unit_price NUMERIC(18,2) NOT NULL,
            vat_rate NUMERIC(5,2) DEFAULT 15
        )""",
        """CREATE TABLE IF NOT EXISTS goods_receipts(
            id SERIAL PRIMARY KEY,
            grn_no VARCHAR(100) UNIQUE NOT NULL,
            grn_date DATE NOT NULL,
            po_id INTEGER NOT NULL REFERENCES purchase_orders(id),
            supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
            warehouse VARCHAR(255) DEFAULT '',
            notes TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'معتمد',
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS goods_receipt_items(
            id SERIAL PRIMARY KEY,
            grn_id INTEGER NOT NULL REFERENCES goods_receipts(id) ON DELETE CASCADE,
            po_item_id INTEGER NOT NULL REFERENCES purchase_order_items(id),
            received_qty NUMERIC(18,3) NOT NULL,
            accepted_qty NUMERIC(18,3) NOT NULL,
            rejected_qty NUMERIC(18,3) DEFAULT 0,
            notes TEXT DEFAULT ''
        )""",
        """CREATE TABLE IF NOT EXISTS supplier_invoices(
            id SERIAL PRIMARY KEY,
            supplier_invoice_no VARCHAR(100) NOT NULL,
            internal_no VARCHAR(100) UNIQUE NOT NULL,
            invoice_date DATE NOT NULL,
            supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
            po_id INTEGER REFERENCES purchase_orders(id),
            grn_id INTEGER REFERENCES goods_receipts(id),
            cost_center_id INTEGER REFERENCES cost_centers(id),
            expense_or_inventory_account_id INTEGER NOT NULL REFERENCES chart_of_accounts(id),
            subtotal NUMERIC(18,2) NOT NULL,
            vat NUMERIC(18,2) DEFAULT 0,
            total NUMERIC(18,2) NOT NULL,
            status VARCHAR(50) DEFAULT 'مسودة',
            posting_status VARCHAR(30) DEFAULT 'غير مرحّل',
            journal_id INTEGER,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(supplier_id,supplier_invoice_no)
        )""",
        """CREATE TABLE IF NOT EXISTS approval_rules(
            id SERIAL PRIMARY KEY,
            document_type VARCHAR(50) NOT NULL,
            min_amount NUMERIC(18,2) DEFAULT 0,
            max_amount NUMERIC(18,2),
            approver_role VARCHAR(100) NOT NULL,
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS warehouses(
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            branch_id INTEGER REFERENCES branches(id),
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS inventory_categories(
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) UNIQUE,
            name VARCHAR(255) NOT NULL,
            parent_id INTEGER REFERENCES inventory_categories(id),
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS inventory_movements(
            id SERIAL PRIMARY KEY,
            movement_no VARCHAR(100) UNIQUE NOT NULL,
            movement_date DATE NOT NULL,
            movement_type VARCHAR(50) NOT NULL,
            item_id INTEGER NOT NULL REFERENCES inventory(id),
            warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
            destination_warehouse_id INTEGER REFERENCES warehouses(id),
            quantity NUMERIC(18,3) NOT NULL,
            unit_cost NUMERIC(18,4) DEFAULT 0,
            total_cost NUMERIC(18,2) DEFAULT 0,
            reference_type VARCHAR(50) DEFAULT '',
            reference_id INTEGER,
            reference_no VARCHAR(100) DEFAULT '',
            batch_no VARCHAR(100) DEFAULT '',
            serial_no VARCHAR(100) DEFAULT '',
            production_date DATE,
            expiry_date DATE,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS inventory_counts(
            id SERIAL PRIMARY KEY,
            count_no VARCHAR(100) UNIQUE NOT NULL,
            count_date DATE NOT NULL,
            warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
            status VARCHAR(30) DEFAULT 'مسودة',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS inventory_count_items(
            id SERIAL PRIMARY KEY,
            count_id INTEGER NOT NULL REFERENCES inventory_counts(id) ON DELETE CASCADE,
            item_id INTEGER NOT NULL REFERENCES inventory(id),
            system_qty NUMERIC(18,3) NOT NULL DEFAULT 0,
            counted_qty NUMERIC(18,3) NOT NULL DEFAULT 0,
            variance_qty NUMERIC(18,3) NOT NULL DEFAULT 0,
            unit_cost NUMERIC(18,4) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS sales_quotations(
            id SERIAL PRIMARY KEY, quotation_no VARCHAR(100) UNIQUE NOT NULL,
            quotation_date DATE NOT NULL, valid_until DATE,
            customer_id INTEGER NOT NULL REFERENCES customers(id),
            branch_id INTEGER REFERENCES branches(id), cost_center_id INTEGER REFERENCES cost_centers(id),
            status VARCHAR(40) DEFAULT 'مسودة', subtotal NUMERIC(18,2) DEFAULT 0,
            discount NUMERIC(18,2) DEFAULT 0, vat NUMERIC(18,2) DEFAULT 0,
            total NUMERIC(18,2) DEFAULT 0, notes TEXT DEFAULT '',
            converted_order_id INTEGER, created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS sales_quotation_items(
            id SERIAL PRIMARY KEY, quotation_id INTEGER NOT NULL REFERENCES sales_quotations(id) ON DELETE CASCADE,
            item_id INTEGER REFERENCES inventory(id), item_name VARCHAR(255) NOT NULL,
            quantity NUMERIC(18,3) NOT NULL, unit VARCHAR(50) DEFAULT 'وحدة',
            unit_price NUMERIC(18,2) NOT NULL, discount_rate NUMERIC(5,2) DEFAULT 0,
            vat_rate NUMERIC(5,2) DEFAULT 15, line_subtotal NUMERIC(18,2) DEFAULT 0,
            line_discount NUMERIC(18,2) DEFAULT 0, line_vat NUMERIC(18,2) DEFAULT 0,
            line_total NUMERIC(18,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS sales_orders(
            id SERIAL PRIMARY KEY, order_no VARCHAR(100) UNIQUE NOT NULL,
            order_date DATE NOT NULL, quotation_id INTEGER REFERENCES sales_quotations(id),
            customer_id INTEGER NOT NULL REFERENCES customers(id),
            branch_id INTEGER REFERENCES branches(id), cost_center_id INTEGER REFERENCES cost_centers(id),
            warehouse_id INTEGER REFERENCES warehouses(id), status VARCHAR(40) DEFAULT 'مسودة',
            subtotal NUMERIC(18,2) DEFAULT 0, discount NUMERIC(18,2) DEFAULT 0,
            vat NUMERIC(18,2) DEFAULT 0, total NUMERIC(18,2) DEFAULT 0,
            notes TEXT DEFAULT '', created_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS sales_order_items(
            id SERIAL PRIMARY KEY, order_id INTEGER NOT NULL REFERENCES sales_orders(id) ON DELETE CASCADE,
            item_id INTEGER REFERENCES inventory(id), item_name VARCHAR(255) NOT NULL,
            quantity NUMERIC(18,3) NOT NULL, delivered_qty NUMERIC(18,3) DEFAULT 0,
            unit VARCHAR(50) DEFAULT 'وحدة', unit_price NUMERIC(18,2) NOT NULL,
            discount_rate NUMERIC(5,2) DEFAULT 0, vat_rate NUMERIC(5,2) DEFAULT 15,
            line_subtotal NUMERIC(18,2) DEFAULT 0, line_discount NUMERIC(18,2) DEFAULT 0,
            line_vat NUMERIC(18,2) DEFAULT 0, line_total NUMERIC(18,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS sales_deliveries(
            id SERIAL PRIMARY KEY, delivery_no VARCHAR(100) UNIQUE NOT NULL,
            delivery_date DATE NOT NULL, order_id INTEGER NOT NULL REFERENCES sales_orders(id),
            customer_id INTEGER NOT NULL REFERENCES customers(id),
            warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
            status VARCHAR(40) DEFAULT 'معتمد', notes TEXT DEFAULT '',
            invoice_id INTEGER, created_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS sales_delivery_items(
            id SERIAL PRIMARY KEY, delivery_id INTEGER NOT NULL REFERENCES sales_deliveries(id) ON DELETE CASCADE,
            order_item_id INTEGER NOT NULL REFERENCES sales_order_items(id),
            item_id INTEGER REFERENCES inventory(id), quantity NUMERIC(18,3) NOT NULL,
            unit_cost NUMERIC(18,4) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS sales_returns(
            id SERIAL PRIMARY KEY,
            return_no VARCHAR(100) UNIQUE NOT NULL,
            return_date DATE NOT NULL,
            invoice_id INTEGER NOT NULL REFERENCES invoices(id),
            customer_id INTEGER NOT NULL REFERENCES customers(id),
            warehouse_id INTEGER NOT NULL REFERENCES warehouses(id),
            subtotal NUMERIC(18,2) DEFAULT 0,
            vat NUMERIC(18,2) DEFAULT 0,
            total NUMERIC(18,2) DEFAULT 0,
            cost_total NUMERIC(18,2) DEFAULT 0,
            status VARCHAR(40) DEFAULT 'معتمد',
            journal_id INTEGER,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS sales_return_items(
            id SERIAL PRIMARY KEY,
            return_id INTEGER NOT NULL REFERENCES sales_returns(id) ON DELETE CASCADE,
            invoice_item_id INTEGER REFERENCES invoice_items(id),
            item_id INTEGER REFERENCES inventory(id),
            item_name VARCHAR(255) NOT NULL,
            quantity NUMERIC(18,3) NOT NULL,
            unit VARCHAR(50) DEFAULT 'وحدة',
            unit_price NUMERIC(18,2) NOT NULL,
            vat_rate NUMERIC(5,2) DEFAULT 15,
            unit_cost NUMERIC(18,4) DEFAULT 0,
            line_subtotal NUMERIC(18,2) DEFAULT 0,
            line_vat NUMERIC(18,2) DEFAULT 0,
            line_total NUMERIC(18,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS invoice_payment_allocations(
            id SERIAL PRIMARY KEY,
            invoice_id INTEGER NOT NULL REFERENCES invoices(id),
            voucher_id INTEGER NOT NULL REFERENCES treasury_vouchers(id),
            allocated_amount NUMERIC(18,2) NOT NULL,
            allocation_date DATE NOT NULL,
            created_by INTEGER,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(invoice_id,voucher_id)
        )""",
        """CREATE TABLE IF NOT EXISTS asset_categories(
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            useful_life_months INTEGER NOT NULL DEFAULT 60,
            depreciation_method VARCHAR(50) DEFAULT 'القسط الثابت',
            asset_account_id INTEGER REFERENCES chart_of_accounts(id),
            accumulated_depr_account_id INTEGER REFERENCES chart_of_accounts(id),
            depreciation_expense_account_id INTEGER REFERENCES chart_of_accounts(id),
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS fixed_assets(
            id SERIAL PRIMARY KEY,
            asset_no VARCHAR(100) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            name_en VARCHAR(255) DEFAULT '',
            category_id INTEGER REFERENCES asset_categories(id),
            branch_id INTEGER REFERENCES branches(id),
            cost_center_id INTEGER REFERENCES cost_centers(id),
            supplier_id INTEGER REFERENCES suppliers(id),
            purchase_date DATE NOT NULL,
            capitalization_date DATE NOT NULL,
            purchase_cost NUMERIC(18,2) NOT NULL DEFAULT 0,
            residual_value NUMERIC(18,2) DEFAULT 0,
            useful_life_months INTEGER NOT NULL DEFAULT 60,
            depreciation_method VARCHAR(50) DEFAULT 'القسط الثابت',
            accumulated_depreciation NUMERIC(18,2) DEFAULT 0,
            net_book_value NUMERIC(18,2) DEFAULT 0,
            serial_number VARCHAR(100) DEFAULT '',
            location VARCHAR(255) DEFAULT '',
            status VARCHAR(50) DEFAULT 'نشط',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS asset_depreciation_runs(
            id SERIAL PRIMARY KEY,
            run_no VARCHAR(100) UNIQUE NOT NULL,
            period_date DATE NOT NULL,
            status VARCHAR(40) DEFAULT 'مسودة',
            total_amount NUMERIC(18,2) DEFAULT 0,
            journal_id INTEGER,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS asset_depreciation_lines(
            id SERIAL PRIMARY KEY,
            run_id INTEGER NOT NULL REFERENCES asset_depreciation_runs(id) ON DELETE CASCADE,
            asset_id INTEGER NOT NULL REFERENCES fixed_assets(id),
            depreciation_amount NUMERIC(18,2) NOT NULL DEFAULT 0,
            accumulated_before NUMERIC(18,2) DEFAULT 0,
            accumulated_after NUMERIC(18,2) DEFAULT 0,
            net_book_value_after NUMERIC(18,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS asset_disposals(
            id SERIAL PRIMARY KEY,
            disposal_no VARCHAR(100) UNIQUE NOT NULL,
            asset_id INTEGER NOT NULL REFERENCES fixed_assets(id),
            disposal_date DATE NOT NULL,
            disposal_type VARCHAR(50) NOT NULL,
            proceeds NUMERIC(18,2) DEFAULT 0,
            gain_loss NUMERIC(18,2) DEFAULT 0,
            journal_id INTEGER,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS departments(
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            manager_employee_id INTEGER,
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS payroll_settings(
            id INTEGER PRIMARY KEY,
            salary_expense_account_id INTEGER REFERENCES chart_of_accounts(id),
            payroll_payable_account_id INTEGER REFERENCES chart_of_accounts(id),
            deduction_account_id INTEGER REFERENCES chart_of_accounts(id),
            working_days_per_month NUMERIC(8,2) DEFAULT 30,
            working_hours_per_day NUMERIC(8,2) DEFAULT 8,
            overtime_multiplier NUMERIC(8,2) DEFAULT 1.5
        )""",
        """CREATE TABLE IF NOT EXISTS attendance_records(
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id),
            attendance_date DATE NOT NULL,
            status VARCHAR(30) DEFAULT 'حاضر',
            check_in TIME,
            check_out TIME,
            overtime_hours NUMERIC(8,2) DEFAULT 0,
            absence_hours NUMERIC(8,2) DEFAULT 0,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(employee_id,attendance_date)
        )""",
        """CREATE TABLE IF NOT EXISTS payroll_runs(
            id SERIAL PRIMARY KEY,
            run_no VARCHAR(100) UNIQUE NOT NULL,
            period_start DATE NOT NULL,
            period_end DATE NOT NULL,
            payment_date DATE,
            status VARCHAR(40) DEFAULT 'مسودة',
            total_gross NUMERIC(18,2) DEFAULT 0,
            total_deductions NUMERIC(18,2) DEFAULT 0,
            total_net NUMERIC(18,2) DEFAULT 0,
            journal_id INTEGER,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS payroll_lines(
            id SERIAL PRIMARY KEY,
            run_id INTEGER NOT NULL REFERENCES payroll_runs(id) ON DELETE CASCADE,
            employee_id INTEGER NOT NULL REFERENCES employees(id),
            basic_salary NUMERIC(18,2) DEFAULT 0,
            allowances NUMERIC(18,2) DEFAULT 0,
            overtime_hours NUMERIC(8,2) DEFAULT 0,
            overtime_amount NUMERIC(18,2) DEFAULT 0,
            absence_days NUMERIC(8,2) DEFAULT 0,
            absence_deduction NUMERIC(18,2) DEFAULT 0,
            other_deductions NUMERIC(18,2) DEFAULT 0,
            gross_salary NUMERIC(18,2) DEFAULT 0,
            total_deductions NUMERIC(18,2) DEFAULT 0,
            net_salary NUMERIC(18,2) DEFAULT 0,
            cost_center_id INTEGER REFERENCES cost_centers(id),
            notes TEXT DEFAULT '',
            UNIQUE(run_id,employee_id)
        )""",
        """CREATE TABLE IF NOT EXISTS employee_salary_adjustments(
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id),
            adjustment_date DATE NOT NULL,
            adjustment_type VARCHAR(30) NOT NULL,
            amount NUMERIC(18,2) NOT NULL,
            recurring INTEGER DEFAULT 0,
            description TEXT DEFAULT '',
            active INTEGER DEFAULT 1,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS projects(
            id SERIAL PRIMARY KEY,
            project_no VARCHAR(100) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            name_en VARCHAR(255) DEFAULT '',
            customer_id INTEGER REFERENCES customers(id),
            branch_id INTEGER REFERENCES branches(id),
            cost_center_id INTEGER REFERENCES cost_centers(id),
            project_manager VARCHAR(255) DEFAULT '',
            start_date DATE,
            end_date DATE,
            contract_value NUMERIC(18,2) DEFAULT 0,
            retention_rate NUMERIC(8,2) DEFAULT 0,
            advance_rate NUMERIC(8,2) DEFAULT 0,
            status VARCHAR(50) DEFAULT 'نشط',
            location VARCHAR(255) DEFAULT '',
            description TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS project_contracts(
            id SERIAL PRIMARY KEY,
            contract_no VARCHAR(100) UNIQUE NOT NULL,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            contract_date DATE NOT NULL,
            contract_type VARCHAR(50) DEFAULT 'مقطوعية',
            contract_value NUMERIC(18,2) DEFAULT 0,
            retention_rate NUMERIC(8,2) DEFAULT 0,
            advance_amount NUMERIC(18,2) DEFAULT 0,
            tax_rate NUMERIC(8,2) DEFAULT 15,
            status VARCHAR(50) DEFAULT 'ساري',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS project_boq_items(
            id SERIAL PRIMARY KEY,
            project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
            item_code VARCHAR(100) NOT NULL,
            description VARCHAR(500) NOT NULL,
            unit VARCHAR(50) DEFAULT 'وحدة',
            quantity NUMERIC(18,3) DEFAULT 0,
            unit_rate NUMERIC(18,2) DEFAULT 0,
            total_value NUMERIC(18,2) DEFAULT 0,
            previous_qty NUMERIC(18,3) DEFAULT 0,
            current_qty NUMERIC(18,3) DEFAULT 0,
            cumulative_qty NUMERIC(18,3) DEFAULT 0,
            completion_percent NUMERIC(8,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS progress_certificates(
            id SERIAL PRIMARY KEY,
            certificate_no VARCHAR(100) UNIQUE NOT NULL,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            contract_id INTEGER REFERENCES project_contracts(id),
            certificate_date DATE NOT NULL,
            period_from DATE,
            period_to DATE,
            gross_work_value NUMERIC(18,2) DEFAULT 0,
            retention_amount NUMERIC(18,2) DEFAULT 0,
            advance_recovery NUMERIC(18,2) DEFAULT 0,
            other_deductions NUMERIC(18,2) DEFAULT 0,
            subtotal NUMERIC(18,2) DEFAULT 0,
            vat NUMERIC(18,2) DEFAULT 0,
            total NUMERIC(18,2) DEFAULT 0,
            status VARCHAR(50) DEFAULT 'مسودة',
            invoice_id INTEGER,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS progress_certificate_items(
            id SERIAL PRIMARY KEY,
            certificate_id INTEGER NOT NULL REFERENCES progress_certificates(id) ON DELETE CASCADE,
            boq_item_id INTEGER NOT NULL REFERENCES project_boq_items(id),
            previous_qty NUMERIC(18,3) DEFAULT 0,
            current_qty NUMERIC(18,3) DEFAULT 0,
            cumulative_qty NUMERIC(18,3) DEFAULT 0,
            unit_rate NUMERIC(18,2) DEFAULT 0,
            current_value NUMERIC(18,2) DEFAULT 0,
            cumulative_value NUMERIC(18,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS project_cost_entries(
            id SERIAL PRIMARY KEY,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            cost_date DATE NOT NULL,
            cost_type VARCHAR(50) NOT NULL,
            reference VARCHAR(100) DEFAULT '',
            description TEXT NOT NULL,
            amount NUMERIC(18,2) NOT NULL DEFAULT 0,
            supplier_id INTEGER REFERENCES suppliers(id),
            employee_id INTEGER REFERENCES employees(id),
            journal_id INTEGER REFERENCES journal_entries(id),
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS crm_leads(
            id SERIAL PRIMARY KEY,
            lead_no VARCHAR(100) UNIQUE NOT NULL,
            company_name VARCHAR(255) NOT NULL,
            contact_name VARCHAR(255) DEFAULT '',
            phone VARCHAR(100) DEFAULT '',
            email VARCHAR(255) DEFAULT '',
            source VARCHAR(100) DEFAULT '',
            industry VARCHAR(100) DEFAULT '',
            city VARCHAR(100) DEFAULT '',
            assigned_to VARCHAR(255) DEFAULT '',
            status VARCHAR(50) DEFAULT 'جديد',
            priority VARCHAR(30) DEFAULT 'متوسطة',
            estimated_value NUMERIC(18,2) DEFAULT 0,
            probability NUMERIC(8,2) DEFAULT 0,
            expected_close_date DATE,
            notes TEXT DEFAULT '',
            converted_customer_id INTEGER REFERENCES customers(id),
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS crm_opportunities(
            id SERIAL PRIMARY KEY,
            opportunity_no VARCHAR(100) UNIQUE NOT NULL,
            lead_id INTEGER REFERENCES crm_leads(id),
            customer_id INTEGER REFERENCES customers(id),
            title VARCHAR(255) NOT NULL,
            stage VARCHAR(50) DEFAULT 'تأهيل',
            estimated_value NUMERIC(18,2) DEFAULT 0,
            probability NUMERIC(8,2) DEFAULT 0,
            weighted_value NUMERIC(18,2) DEFAULT 0,
            expected_close_date DATE,
            sales_person VARCHAR(255) DEFAULT '',
            status VARCHAR(40) DEFAULT 'مفتوحة',
            lost_reason TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            quotation_id INTEGER REFERENCES sales_quotations(id),
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS crm_activities(
            id SERIAL PRIMARY KEY,
            lead_id INTEGER REFERENCES crm_leads(id),
            opportunity_id INTEGER REFERENCES crm_opportunities(id),
            customer_id INTEGER REFERENCES customers(id),
            activity_type VARCHAR(50) NOT NULL,
            subject VARCHAR(255) NOT NULL,
            activity_date TIMESTAMP NOT NULL,
            due_date TIMESTAMP,
            assigned_to VARCHAR(255) DEFAULT '',
            status VARCHAR(40) DEFAULT 'مفتوحة',
            outcome TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS system_roles(
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            description TEXT DEFAULT '',
            is_system INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS system_permissions(
            id SERIAL PRIMARY KEY,
            permission_key VARCHAR(150) UNIQUE NOT NULL,
            module_name VARCHAR(100) NOT NULL,
            action_name VARCHAR(50) NOT NULL,
            display_name VARCHAR(255) NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS role_permissions(
            role_id INTEGER NOT NULL REFERENCES system_roles(id) ON DELETE CASCADE,
            permission_id INTEGER NOT NULL REFERENCES system_permissions(id) ON DELETE CASCADE,
            allowed INTEGER DEFAULT 1,
            PRIMARY KEY(role_id,permission_id)
        )""",
        """CREATE TABLE IF NOT EXISTS user_role_assignments(
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            role_id INTEGER NOT NULL REFERENCES system_roles(id) ON DELETE CASCADE,
            PRIMARY KEY(user_id,role_id)
        )""",
        """CREATE TABLE IF NOT EXISTS access_denied_logs(
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            username VARCHAR(100),
            endpoint VARCHAR(255),
            permission_key VARCHAR(150),
            ip_address VARCHAR(100),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS approval_workflows(
            id SERIAL PRIMARY KEY,
            code VARCHAR(100) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            document_type VARCHAR(100) NOT NULL,
            min_amount NUMERIC(18,2) DEFAULT 0,
            max_amount NUMERIC(18,2),
            active INTEGER DEFAULT 1,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS approval_workflow_steps(
            id SERIAL PRIMARY KEY,
            workflow_id INTEGER NOT NULL REFERENCES approval_workflows(id) ON DELETE CASCADE,
            step_order INTEGER NOT NULL,
            role_id INTEGER REFERENCES system_roles(id),
            approver_username VARCHAR(100) DEFAULT '',
            action_name VARCHAR(50) DEFAULT 'اعتماد',
            is_final INTEGER DEFAULT 0,
            UNIQUE(workflow_id,step_order)
        )""",
        """CREATE TABLE IF NOT EXISTS approval_requests(
            id SERIAL PRIMARY KEY,
            request_no VARCHAR(100) UNIQUE NOT NULL,
            workflow_id INTEGER NOT NULL REFERENCES approval_workflows(id),
            document_type VARCHAR(100) NOT NULL,
            document_id INTEGER NOT NULL,
            document_no VARCHAR(100) DEFAULT '',
            document_amount NUMERIC(18,2) DEFAULT 0,
            requester_user_id INTEGER REFERENCES users(id),
            current_step_order INTEGER DEFAULT 1,
            status VARCHAR(50) DEFAULT 'قيد الاعتماد',
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            notes TEXT DEFAULT '',
            UNIQUE(document_type,document_id)
        )""",
        """CREATE TABLE IF NOT EXISTS approval_actions(
            id SERIAL PRIMARY KEY,
            request_id INTEGER NOT NULL REFERENCES approval_requests(id) ON DELETE CASCADE,
            step_id INTEGER REFERENCES approval_workflow_steps(id),
            action_by_user_id INTEGER REFERENCES users(id),
            action_type VARCHAR(50) NOT NULL,
            action_notes TEXT DEFAULT '',
            action_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS budget_headers(
            id SERIAL PRIMARY KEY,
            budget_no VARCHAR(100) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            fiscal_year INTEGER NOT NULL,
            branch_id INTEGER REFERENCES branches(id),
            status VARCHAR(50) DEFAULT 'مسودة',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS budget_lines(
            id SERIAL PRIMARY KEY,
            budget_id INTEGER NOT NULL REFERENCES budget_headers(id) ON DELETE CASCADE,
            account_id INTEGER NOT NULL REFERENCES chart_of_accounts(id),
            cost_center_id INTEGER REFERENCES cost_centers(id),
            month_no INTEGER NOT NULL,
            budget_amount NUMERIC(18,2) DEFAULT 0,
            forecast_amount NUMERIC(18,2) DEFAULT 0,
            notes TEXT DEFAULT '',
            UNIQUE(budget_id,account_id,cost_center_id,month_no)
        )""",
        """CREATE TABLE IF NOT EXISTS budget_revisions(
            id SERIAL PRIMARY KEY,
            budget_id INTEGER NOT NULL REFERENCES budget_headers(id) ON DELETE CASCADE,
            revision_no INTEGER NOT NULL,
            revision_date DATE NOT NULL,
            reason TEXT DEFAULT '',
            approved_by INTEGER REFERENCES users(id),
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS employee_contracts(
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
            contract_no VARCHAR(100) UNIQUE NOT NULL,
            contract_type VARCHAR(50) DEFAULT 'محدد المدة',
            start_date DATE NOT NULL,
            end_date DATE,
            probation_end_date DATE,
            basic_salary NUMERIC(18,2) DEFAULT 0,
            housing_allowance NUMERIC(18,2) DEFAULT 0,
            transport_allowance NUMERIC(18,2) DEFAULT 0,
            other_allowance NUMERIC(18,2) DEFAULT 0,
            working_hours NUMERIC(8,2) DEFAULT 8,
            annual_leave_days INTEGER DEFAULT 21,
            status VARCHAR(50) DEFAULT 'ساري',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS leave_types(
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            paid INTEGER DEFAULT 1,
            annual_entitlement NUMERIC(8,2) DEFAULT 0,
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS employee_leave_balances(
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
            leave_type_id INTEGER NOT NULL REFERENCES leave_types(id),
            year INTEGER NOT NULL,
            opening_balance NUMERIC(8,2) DEFAULT 0,
            accrued NUMERIC(8,2) DEFAULT 0,
            used NUMERIC(8,2) DEFAULT 0,
            remaining NUMERIC(8,2) DEFAULT 0,
            UNIQUE(employee_id,leave_type_id,year)
        )""",
        """CREATE TABLE IF NOT EXISTS leave_requests(
            id SERIAL PRIMARY KEY,
            request_no VARCHAR(100) UNIQUE NOT NULL,
            employee_id INTEGER NOT NULL REFERENCES employees(id),
            leave_type_id INTEGER NOT NULL REFERENCES leave_types(id),
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            days NUMERIC(8,2) NOT NULL DEFAULT 0,
            reason TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'مسودة',
            approved_by INTEGER REFERENCES users(id),
            approved_at TIMESTAMP,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS employee_documents(
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
            document_type VARCHAR(100) NOT NULL,
            document_no VARCHAR(150) DEFAULT '',
            issue_date DATE,
            expiry_date DATE,
            file_name VARCHAR(255) DEFAULT '',
            file_url TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'ساري',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS employee_warnings(
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id),
            warning_date DATE NOT NULL,
            warning_type VARCHAR(100) DEFAULT 'إنذار كتابي',
            subject VARCHAR(255) NOT NULL,
            details TEXT NOT NULL,
            action_required TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'مفتوح',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS employee_end_of_service(
            id SERIAL PRIMARY KEY,
            employee_id INTEGER NOT NULL REFERENCES employees(id),
            calculation_date DATE NOT NULL,
            service_start_date DATE NOT NULL,
            service_end_date DATE NOT NULL,
            service_years NUMERIC(10,4) DEFAULT 0,
            last_basic_salary NUMERIC(18,2) DEFAULT 0,
            gratuity_amount NUMERIC(18,2) DEFAULT 0,
            leave_balance_amount NUMERIC(18,2) DEFAULT 0,
            other_dues NUMERIC(18,2) DEFAULT 0,
            deductions NUMERIC(18,2) DEFAULT 0,
            net_settlement NUMERIC(18,2) DEFAULT 0,
            status VARCHAR(50) DEFAULT 'مسودة',
            journal_id INTEGER REFERENCES journal_entries(id),
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS recruitment_candidates(
            id SERIAL PRIMARY KEY,
            candidate_no VARCHAR(100) UNIQUE NOT NULL,
            full_name VARCHAR(255) NOT NULL,
            nationality VARCHAR(100) DEFAULT '',
            phone VARCHAR(100) DEFAULT '',
            email VARCHAR(255) DEFAULT '',
            job_title VARCHAR(255) DEFAULT '',
            department_id INTEGER REFERENCES departments(id),
            expected_salary NUMERIC(18,2) DEFAULT 0,
            source VARCHAR(100) DEFAULT '',
            stage VARCHAR(50) DEFAULT 'جديد',
            interview_date TIMESTAMP,
            interview_result TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS subcontractors(
            id SERIAL PRIMARY KEY,
            code VARCHAR(100) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            name_en VARCHAR(255) DEFAULT '',
            tax_no VARCHAR(100) DEFAULT '',
            commercial_registration VARCHAR(100) DEFAULT '',
            phone VARCHAR(100) DEFAULT '',
            email VARCHAR(255) DEFAULT '',
            address TEXT DEFAULT '',
            specialty VARCHAR(255) DEFAULT '',
            active INTEGER DEFAULT 1,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS subcontract_contracts(
            id SERIAL PRIMARY KEY,
            contract_no VARCHAR(100) UNIQUE NOT NULL,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            subcontractor_id INTEGER NOT NULL REFERENCES subcontractors(id),
            contract_date DATE NOT NULL,
            start_date DATE,
            end_date DATE,
            contract_value NUMERIC(18,2) DEFAULT 0,
            retention_rate NUMERIC(8,2) DEFAULT 0,
            advance_amount NUMERIC(18,2) DEFAULT 0,
            tax_rate NUMERIC(8,2) DEFAULT 15,
            status VARCHAR(50) DEFAULT 'ساري',
            scope_of_work TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS subcontract_boq_items(
            id SERIAL PRIMARY KEY,
            contract_id INTEGER NOT NULL REFERENCES subcontract_contracts(id) ON DELETE CASCADE,
            item_code VARCHAR(100) NOT NULL,
            description VARCHAR(500) NOT NULL,
            unit VARCHAR(50) DEFAULT 'وحدة',
            quantity NUMERIC(18,3) DEFAULT 0,
            unit_rate NUMERIC(18,2) DEFAULT 0,
            total_value NUMERIC(18,2) DEFAULT 0,
            cumulative_qty NUMERIC(18,3) DEFAULT 0,
            UNIQUE(contract_id,item_code)
        )""",
        """CREATE TABLE IF NOT EXISTS subcontract_certificates(
            id SERIAL PRIMARY KEY,
            certificate_no VARCHAR(100) UNIQUE NOT NULL,
            contract_id INTEGER NOT NULL REFERENCES subcontract_contracts(id),
            certificate_date DATE NOT NULL,
            period_from DATE,
            period_to DATE,
            gross_value NUMERIC(18,2) DEFAULT 0,
            retention_amount NUMERIC(18,2) DEFAULT 0,
            advance_recovery NUMERIC(18,2) DEFAULT 0,
            other_deductions NUMERIC(18,2) DEFAULT 0,
            subtotal NUMERIC(18,2) DEFAULT 0,
            vat NUMERIC(18,2) DEFAULT 0,
            total NUMERIC(18,2) DEFAULT 0,
            status VARCHAR(50) DEFAULT 'مسودة',
            supplier_invoice_id INTEGER,
            notes TEXT DEFAULT '',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS subcontract_certificate_items(
            id SERIAL PRIMARY KEY,
            certificate_id INTEGER NOT NULL REFERENCES subcontract_certificates(id) ON DELETE CASCADE,
            boq_item_id INTEGER NOT NULL REFERENCES subcontract_boq_items(id),
            previous_qty NUMERIC(18,3) DEFAULT 0,
            current_qty NUMERIC(18,3) DEFAULT 0,
            cumulative_qty NUMERIC(18,3) DEFAULT 0,
            unit_rate NUMERIC(18,2) DEFAULT 0,
            current_value NUMERIC(18,2) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS variation_orders(
            id SERIAL PRIMARY KEY,
            variation_no VARCHAR(100) UNIQUE NOT NULL,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            contract_id INTEGER REFERENCES project_contracts(id),
            variation_date DATE NOT NULL,
            title VARCHAR(255) NOT NULL,
            description TEXT DEFAULT '',
            reason TEXT DEFAULT '',
            value NUMERIC(18,2) DEFAULT 0,
            time_extension_days INTEGER DEFAULT 0,
            status VARCHAR(50) DEFAULT 'مسودة',
            approval_request_id INTEGER REFERENCES approval_requests(id),
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS contract_extensions(
            id SERIAL PRIMARY KEY,
            project_id INTEGER NOT NULL REFERENCES projects(id),
            contract_id INTEGER REFERENCES project_contracts(id),
            extension_no VARCHAR(100) UNIQUE NOT NULL,
            request_date DATE NOT NULL,
            approved_date DATE,
            old_end_date DATE,
            new_end_date DATE NOT NULL,
            extension_days INTEGER DEFAULT 0,
            reason TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'مسودة',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS document_categories(
            id SERIAL PRIMARY KEY,
            code VARCHAR(100) UNIQUE NOT NULL,
            name VARCHAR(255) NOT NULL,
            retention_years INTEGER DEFAULT 5,
            requires_expiry INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS documents_archive(
            id SERIAL PRIMARY KEY,
            document_no VARCHAR(100) UNIQUE NOT NULL,
            title VARCHAR(255) NOT NULL,
            category_id INTEGER REFERENCES document_categories(id),
            entity_type VARCHAR(100) DEFAULT '',
            entity_id INTEGER,
            entity_name VARCHAR(255) DEFAULT '',
            issue_date DATE,
            expiry_date DATE,
            version_no VARCHAR(50) DEFAULT '1.0',
            status VARCHAR(50) DEFAULT 'ساري',
            confidentiality VARCHAR(50) DEFAULT 'داخلي',
            file_name VARCHAR(255) DEFAULT '',
            file_url TEXT DEFAULT '',
            keywords TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            uploaded_by INTEGER REFERENCES users(id),
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS document_versions(
            id SERIAL PRIMARY KEY,
            document_id INTEGER NOT NULL REFERENCES documents_archive(id) ON DELETE CASCADE,
            version_no VARCHAR(50) NOT NULL,
            file_name VARCHAR(255) DEFAULT '',
            file_url TEXT DEFAULT '',
            change_notes TEXT DEFAULT '',
            uploaded_by INTEGER REFERENCES users(id),
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS document_access_logs(
            id SERIAL PRIMARY KEY,
            document_id INTEGER REFERENCES documents_archive(id) ON DELETE CASCADE,
            user_id INTEGER REFERENCES users(id),
            action_type VARCHAR(50) NOT NULL,
            ip_address VARCHAR(100) DEFAULT '',
            action_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS data_import_jobs(
            id SERIAL PRIMARY KEY,
            import_no VARCHAR(100) UNIQUE NOT NULL,
            module_name VARCHAR(100) NOT NULL,
            file_name VARCHAR(255) DEFAULT '',
            import_mode VARCHAR(50) DEFAULT 'إضافة فقط',
            total_rows INTEGER DEFAULT 0,
            success_rows INTEGER DEFAULT 0,
            updated_rows INTEGER DEFAULT 0,
            failed_rows INTEGER DEFAULT 0,
            status VARCHAR(50) DEFAULT 'مكتمل',
            error_details TEXT DEFAULT '',
            imported_by INTEGER REFERENCES users(id),
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS data_import_profiles(
            id SERIAL PRIMARY KEY,
            profile_name VARCHAR(150) NOT NULL,
            module_name VARCHAR(100) NOT NULL,
            source_system VARCHAR(100) DEFAULT '',
            mapping_json TEXT NOT NULL,
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS data_import_profile_aliases(
            id SERIAL PRIMARY KEY,
            module_name VARCHAR(100) NOT NULL,
            source_header VARCHAR(255) NOT NULL,
            target_field VARCHAR(100) NOT NULL,
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(module_name,source_header)
        )""",
        """CREATE TABLE IF NOT EXISTS data_import_job_rows(
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL REFERENCES data_import_jobs(id) ON DELETE CASCADE,
            row_no INTEGER NOT NULL,
            row_data TEXT NOT NULL,
            row_status VARCHAR(30) DEFAULT 'قيد الانتظار',
            action_type VARCHAR(30) DEFAULT '',
            error_message TEXT DEFAULT '',
            processed_at TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS data_import_job_events(
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL REFERENCES data_import_jobs(id) ON DELETE CASCADE,
            event_type VARCHAR(50) NOT NULL,
            message TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE INDEX IF NOT EXISTS idx_import_job_rows_status
            ON data_import_job_rows(job_id,row_status,id)""",
        """CREATE TABLE IF NOT EXISTS audit_logs(
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            username VARCHAR(100),
            action VARCHAR(100) NOT NULL,
            entity VARCHAR(100) NOT NULL,
            details TEXT DEFAULT '',
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )""",
        """CREATE TABLE IF NOT EXISTS expenses(
            id SERIAL PRIMARY KEY,
            expense_date DATE NOT NULL,
            category VARCHAR(255) NOT NULL,
            description TEXT,
            amount NUMERIC(18,2) NOT NULL,
            vat NUMERIC(18,2) DEFAULT 0,
            total NUMERIC(18,2) NOT NULL,
            branch_id INTEGER REFERENCES branches(id)
        )""",
        """CREATE TABLE IF NOT EXISTS inventory(
            id SERIAL PRIMARY KEY,
            sku VARCHAR(100) UNIQUE,
            name VARCHAR(255) NOT NULL,
            quantity NUMERIC(18,3) DEFAULT 0,
            unit VARCHAR(50) DEFAULT 'وحدة',
            cost NUMERIC(18,2) DEFAULT 0,
            sale_price NUMERIC(18,2) DEFAULT 0,
            reorder_level NUMERIC(18,3) DEFAULT 0
        )""",
        """CREATE TABLE IF NOT EXISTS employees(
            id SERIAL PRIMARY KEY,
            employee_no VARCHAR(100) UNIQUE,
            name VARCHAR(255) NOT NULL,
            job_title VARCHAR(255),
            branch_id INTEGER REFERENCES branches(id),
            basic_salary NUMERIC(18,2) DEFAULT 0,
            allowances NUMERIC(18,2) DEFAULT 0,
            active INTEGER DEFAULT 1
        )"""
    ]
    for statement in statements:
        db.session.execute(text(statement))

    # Safe schema upgrades for existing Neon databases.
    migrations = [
        "ALTER TABLE settings ADD COLUMN IF NOT EXISTS address TEXT DEFAULT ''",
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS name_en VARCHAR(255) DEFAULT ''",
        "ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS name_en VARCHAR(255) DEFAULT ''",
        "ALTER TABLE purchase_requisition_items ADD COLUMN IF NOT EXISTS ordered_qty NUMERIC(18,3) DEFAULT 0",
        "ALTER TABLE purchase_requisition_items ADD COLUMN IF NOT EXISTS inventory_item_id INTEGER REFERENCES inventory(id)",
        "ALTER TABLE purchase_order_items ADD COLUMN IF NOT EXISTS requisition_item_id INTEGER REFERENCES purchase_requisition_items(id)",
        "ALTER TABLE purchase_orders ADD COLUMN IF NOT EXISTS warehouse_id INTEGER REFERENCES warehouses(id)",
        "ALTER TABLE purchase_order_items ADD COLUMN IF NOT EXISTS inventory_item_id INTEGER REFERENCES inventory(id)",
        "ALTER TABLE goods_receipts ADD COLUMN IF NOT EXISTS warehouse_id INTEGER REFERENCES warehouses(id)",
        "ALTER TABLE settings ADD COLUMN IF NOT EXISTS phone VARCHAR(50) DEFAULT ''",
        "ALTER TABLE settings ADD COLUMN IF NOT EXISTS email VARCHAR(255) DEFAULT ''",
        "ALTER TABLE settings ADD COLUMN IF NOT EXISTS logo_url TEXT DEFAULT ''",
        "ALTER TABLE settings ADD COLUMN IF NOT EXISTS logo_data TEXT DEFAULT ''",
        "ALTER TABLE settings ADD COLUMN IF NOT EXISTS logo_mime VARCHAR(50) DEFAULT ''",
        "ALTER TABLE settings ADD COLUMN IF NOT EXISTS vat_rate NUMERIC(5,2) DEFAULT 15.00",
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS invoice_uuid VARCHAR(64)",
        "ALTER TABLE journal_entry_lines ADD COLUMN IF NOT EXISTS invoice_date DATE",
        "ALTER TABLE journal_entry_lines ADD COLUMN IF NOT EXISTS customer_id INTEGER REFERENCES customers(id)",
        "ALTER TABLE journal_entry_lines ADD COLUMN IF NOT EXISTS party_type VARCHAR(20) DEFAULT ''",
        "ALTER TABLE journal_entry_lines ADD COLUMN IF NOT EXISTS entity_id INTEGER",
        "ALTER TABLE journal_batch_lines ADD COLUMN IF NOT EXISTS entity_id INTEGER",
        "ALTER TABLE journal_entry_lines ADD COLUMN IF NOT EXISTS tax_number VARCHAR(50) DEFAULT ''",
        "ALTER TABLE chart_of_accounts ADD COLUMN IF NOT EXISTS normal_balance VARCHAR(20) DEFAULT 'مدين'",
        "ALTER TABLE chart_of_accounts ADD COLUMN IF NOT EXISTS statement_type VARCHAR(50) DEFAULT 'الميزانية العمومية'",
        'ALTER TABLE invoice_items ADD COLUMN IF NOT EXISTS item_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS sales_order_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS delivery_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS cost_center_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS warehouse_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS draft_number VARCHAR(100)',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS official_invoice_no VARCHAR(100)',
        'CREATE UNIQUE INDEX IF NOT EXISTS uq_invoices_draft_number ON invoices(draft_number) WHERE draft_number IS NOT NULL',
        'CREATE UNIQUE INDEX IF NOT EXISTS uq_invoices_official_no ON invoices(official_invoice_no) WHERE official_invoice_no IS NOT NULL',
        'ALTER TABLE customers ADD COLUMN IF NOT EXISTS code VARCHAR(100)',
        "ALTER TABLE customers ADD COLUMN IF NOT EXISTS address TEXT DEFAULT ''",
        'ALTER TABLE customers ADD COLUMN IF NOT EXISTS credit_limit NUMERIC(18,2) DEFAULT 0',
        'ALTER TABLE customers ADD COLUMN IF NOT EXISTS receivable_account_id INTEGER',
        'ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS code VARCHAR(100)',
        "ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS address TEXT DEFAULT ''",
        'ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS payable_account_id INTEGER',
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS code VARCHAR(100)',
        "ALTER TABLE inventory ADD COLUMN IF NOT EXISTS description TEXT DEFAULT ''",
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS unit_cost NUMERIC(18,2) DEFAULT 0',
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS reorder_level NUMERIC(18,3) DEFAULT 0',
        'ALTER TABLE journal_entry_lines ADD COLUMN IF NOT EXISTS party_account_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS project_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS approved_by INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS copied_from_id INTEGER',
        'ALTER TABLE invoice_items ADD COLUMN IF NOT EXISTS discount_rate NUMERIC(5,2) DEFAULT 0',
        'ALTER TABLE invoice_items ADD COLUMN IF NOT EXISTS line_discount NUMERIC(18,2) DEFAULT 0',
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS nationality VARCHAR(100) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS passport_no VARCHAR(100) DEFAULT ''",
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS passport_expiry DATE',
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS iqama_expiry DATE',
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS medical_insurance_expiry DATE',
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS social_insurance_no VARCHAR(100) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS emergency_contact VARCHAR(255) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS emergency_phone VARCHAR(100) DEFAULT ''",
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS termination_date DATE',
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS termination_reason TEXT DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS full_name VARCHAR(255) DEFAULT ''",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS email VARCHAR(255) DEFAULT ''",
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS active INTEGER DEFAULT 1',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS last_login TIMESTAMP',
        'ALTER TABLE users ADD COLUMN IF NOT EXISTS must_change_password INTEGER DEFAULT 0',
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS name_en VARCHAR(255) DEFAULT ''",
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS department_id INTEGER',
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS cost_center_id INTEGER',
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS hire_date DATE',
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS contract_end_date DATE',
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS iqama_no VARCHAR(100) DEFAULT ''",
        "ALTER TABLE employees ADD COLUMN IF NOT EXISTS bank_iban VARCHAR(100) DEFAULT ''",
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS housing_allowance NUMERIC(18,2) DEFAULT 0',
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS transport_allowance NUMERIC(18,2) DEFAULT 0',
        'ALTER TABLE employees ADD COLUMN IF NOT EXISTS other_allowance NUMERIC(18,2) DEFAULT 0',
        'ALTER TABLE sales_returns ADD COLUMN IF NOT EXISTS cost_total NUMERIC(18,2) DEFAULT 0',
        'ALTER TABLE sales_returns ADD COLUMN IF NOT EXISTS journal_id INTEGER',
        "ALTER TABLE sales_returns ADD COLUMN IF NOT EXISTS status VARCHAR(40) DEFAULT 'معتمد'",
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS payment_status VARCHAR(30) DEFAULT \'غير مسددة\'',
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS payment_terms VARCHAR(255) DEFAULT ''",
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS payment_method VARCHAR(50) DEFAULT ''",
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS sales_person VARCHAR(255) DEFAULT ''",
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS customer_reference VARCHAR(100) DEFAULT ''",
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS discount NUMERIC(18,2) DEFAULT 0",
        "ALTER TABLE sales_quotations ADD COLUMN IF NOT EXISTS sales_person VARCHAR(255) DEFAULT ''",
        "ALTER TABLE sales_quotations ADD COLUMN IF NOT EXISTS payment_terms VARCHAR(255) DEFAULT ''",
        "ALTER TABLE sales_orders ADD COLUMN IF NOT EXISTS sales_person VARCHAR(255) DEFAULT ''",
        "ALTER TABLE sales_orders ADD COLUMN IF NOT EXISTS payment_terms VARCHAR(255) DEFAULT ''",
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS due_date DATE',
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS payment_status VARCHAR(30) DEFAULT 'غير مسددة'",
        "ALTER TABLE inventory ADD COLUMN IF NOT EXISTS name_en VARCHAR(255) DEFAULT ''",
        "ALTER TABLE inventory ADD COLUMN IF NOT EXISTS barcode VARCHAR(100) DEFAULT ''",
        "ALTER TABLE inventory ADD COLUMN IF NOT EXISTS qr_code VARCHAR(255) DEFAULT ''",
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS category_id INTEGER',
        "ALTER TABLE inventory ADD COLUMN IF NOT EXISTS item_type VARCHAR(30) DEFAULT 'مخزني'",
        "ALTER TABLE inventory ADD COLUMN IF NOT EXISTS valuation_method VARCHAR(30) DEFAULT 'متوسط متحرك'",
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS min_level NUMERIC(18,3) DEFAULT 0',
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS max_level NUMERIC(18,3) DEFAULT 0',
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS track_batch INTEGER DEFAULT 0',
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS track_serial INTEGER DEFAULT 0',
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS track_expiry INTEGER DEFAULT 0',
        'ALTER TABLE inventory ADD COLUMN IF NOT EXISTS active INTEGER DEFAULT 1',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS customer_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS supplier_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS sales_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS purchases_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS vat_input_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS vat_output_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS cash_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS bank_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS inventory_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS cost_of_sales_account_id INTEGER',
        'ALTER TABLE settings ADD COLUMN IF NOT EXISTS retained_earnings_account_id INTEGER',
        'ALTER TABLE invoices ADD COLUMN IF NOT EXISTS journal_id INTEGER',
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS posting_status VARCHAR(30) DEFAULT 'غير مرحّل'",
        "ALTER TABLE journal_entries ADD COLUMN IF NOT EXISTS source_type VARCHAR(50) DEFAULT 'MANUAL'",
        'ALTER TABLE journal_entries ADD COLUMN IF NOT EXISTS source_id INTEGER',
        'ALTER TABLE journal_entries ADD COLUMN IF NOT EXISTS posted_at TIMESTAMP',
        'ALTER TABLE journal_entries ADD COLUMN IF NOT EXISTS reversal_of INTEGER',
        'ALTER TABLE journal_entries ADD COLUMN IF NOT EXISTS reversed_by INTEGER',
        'ALTER TABLE expenses ADD COLUMN IF NOT EXISTS supplier_id INTEGER',
        'ALTER TABLE expenses ADD COLUMN IF NOT EXISTS expense_account_id INTEGER',
        'ALTER TABLE expenses ADD COLUMN IF NOT EXISTS payment_account_id INTEGER',
        'ALTER TABLE expenses ADD COLUMN IF NOT EXISTS cost_center_id INTEGER',
        'ALTER TABLE expenses ADD COLUMN IF NOT EXISTS taxable INTEGER DEFAULT 0',
        "ALTER TABLE expenses ADD COLUMN IF NOT EXISTS invoice_number VARCHAR(100) DEFAULT ''",
        'ALTER TABLE expenses ADD COLUMN IF NOT EXISTS invoice_date DATE',
        "ALTER TABLE expenses ADD COLUMN IF NOT EXISTS status VARCHAR(30) DEFAULT 'مسودة'",
        'ALTER TABLE expenses ADD COLUMN IF NOT EXISTS journal_id INTEGER',
        "ALTER TABLE expenses ADD COLUMN IF NOT EXISTS posting_status VARCHAR(30) DEFAULT 'غير مرحّل'"
    ]
    for migration in migrations:
        db.session.execute(text(migration))

    default_password_hash = generate_password_hash("admin123")
    db.session.execute(
        text("""
            INSERT INTO users(id, username, password, role)
            VALUES(1,'admin',:password,'admin')
            ON CONFLICT (id) DO NOTHING
        """),
        {"password": default_password_hash}
    )
    db.session.execute(text("""
        INSERT INTO settings(id)
        VALUES(1)
        ON CONFLICT (id) DO NOTHING
    """))
    db.session.execute(text("""
        INSERT INTO branches(id,name,city,active)
        VALUES(1,'الفرع الرئيسي','الدمام',1)
        ON CONFLICT (id) DO NOTHING
    """))

    defaults = [
        ("1000","الأصول","Assets","أصل",None,1,0),
        ("1100","الأصول المتداولة","Current Assets","أصل","1000",2,0),
        ("1110","الصندوق","Cash","أصل","1100",3,1),
        ("1120","البنوك","Banks","أصل","1100",3,1),
        ("1130","العملاء","Accounts Receivable","أصل","1100",3,1),
        ("1140","ضريبة القيمة المضافة - مدخلات","VAT Input","أصل","1100",3,1),
        ("1200","المخزون","Inventory","أصل","1000",2,1),
        ("2000","الخصوم","Liabilities","خصم",None,1,0),
        ("2100","الموردون","Accounts Payable","خصم","2000",2,1),
        ("2200","ضريبة القيمة المضافة - مخرجات","VAT Output","خصم","2000",2,1),
        ("3000","حقوق الملكية","Equity","حقوق ملكية",None,1,0),
        ("3100","رأس المال","Capital","حقوق ملكية","3000",2,1),
        ("4000","الإيرادات","Revenue","إيراد",None,1,0),
        ("4100","إيرادات المبيعات","Sales Revenue","إيراد","4000",2,1),
        ("5000","تكلفة المبيعات والمشاريع","Cost of Sales","مصروف",None,1,0),
        ("5100","مواد المشاريع","Project Materials","مصروف","5000",2,1),
        ("6000","المصروفات","Expenses","مصروف",None,1,0),
        ("6100","مصروفات إدارية","Administrative Expenses","مصروف","6000",2,1),
    ]
    ids = {}
    for code, ar, en, typ, parent_code, level, accepts in defaults:
        parent_id = ids.get(parent_code)
        existing = db.session.execute(text("SELECT id FROM chart_of_accounts WHERE account_code=:c"), {"c": code}).scalar()
        if existing:
            ids[code] = existing
        else:
            new_id = db.session.execute(
                text("""INSERT INTO chart_of_accounts(
                    account_code,account_name_ar,account_name_en,account_type,
                    parent_id,level,accepts_entries,active)
                    VALUES(:c,:ar,:en,:t,:p,:l,:a,1) RETURNING id"""),
                {"c":code,"ar":ar,"en":en,"t":typ,"p":parent_id,"l":level,"a":accepts}
            ).scalar_one()
            ids[code] = new_id
    db.session.execute(text("""
        INSERT INTO cost_centers(code,name,active)
        VALUES('CC-001','الإدارة العامة',1)
        ON CONFLICT (code) DO NOTHING
    """))
    db.session.execute(text("""
        INSERT INTO warehouses(code,name,active)
        VALUES('WH-001','المستودع الرئيسي',1)
        ON CONFLICT (code) DO NOTHING
    """))

    for min_amount,max_amount,role in [
        (0,5000,"مدير القسم"),
        (5000.01,20000,"المدير المالي"),
        (20000.01,None,"المدير العام")
    ]:
        db.session.execute(text("""INSERT INTO approval_rules(
            document_type,min_amount,max_amount,approver_role,active)
            SELECT 'طلب شراء',:mn,:mx,:role,1
            WHERE NOT EXISTS(
              SELECT 1 FROM approval_rules WHERE document_type='طلب شراء'
              AND min_amount=:mn AND max_amount IS NOT DISTINCT FROM :mx
            )"""), {"mn":min_amount,"mx":max_amount,"role":role})

    current_year = datetime.now().year
    fy_id = db.session.execute(text("""INSERT INTO fiscal_years(name,start_date,end_date,status)
        VALUES(:n,:s,:e,'مفتوحة') ON CONFLICT(start_date,end_date)
        DO UPDATE SET name=EXCLUDED.name RETURNING id"""),
        {"n":f"السنة المالية {current_year}","s":f"{current_year}-01-01","e":f"{current_year}-12-31"}).scalar_one()
    import calendar
    for m in range(1,13):
        last=calendar.monthrange(current_year,m)[1]
        db.session.execute(text("""INSERT INTO fiscal_periods(fiscal_year_id,name,start_date,end_date,status)
            SELECT :fy,:n,:s,:e,'مفتوحة' WHERE NOT EXISTS(
            SELECT 1 FROM fiscal_periods WHERE start_date=:s AND end_date=:e)"""),
            {"fy":fy_id,"n":f"{current_year}-{m:02d}","s":f"{current_year}-{m:02d}-01",
             "e":f"{current_year}-{m:02d}-{last:02d}"})
    db.session.execute(text("""UPDATE settings SET
      customer_account_id=COALESCE(customer_account_id,:c),
      supplier_account_id=COALESCE(supplier_account_id,:s),
      sales_account_id=COALESCE(sales_account_id,:sales),
      purchases_account_id=COALESCE(purchases_account_id,:p),
      vat_input_account_id=COALESCE(vat_input_account_id,:vi),
      vat_output_account_id=COALESCE(vat_output_account_id,:vo),
      cash_account_id=COALESCE(cash_account_id,:cash),
      bank_account_id=COALESCE(bank_account_id,:bank),
      inventory_account_id=COALESCE(inventory_account_id,:inv),
      cost_of_sales_account_id=COALESCE(cost_of_sales_account_id,:cos),
      retained_earnings_account_id=COALESCE(retained_earnings_account_id,:re)
      WHERE id=1"""),{"c":ids.get("1130"),"s":ids.get("2100"),"sales":ids.get("4100"),
      "p":ids.get("5100"),"vi":ids.get("1140"),"vo":ids.get("2200"),"cash":ids.get("1110"),
      "bank":ids.get("1120"),"inv":ids.get("1200"),"cos":ids.get("5000"),"re":ids.get("3100")})

    db.session.commit()

@app.before_request
def ensure_database():
    if not getattr(app, "_db_initialized", False):
        init_db()
        seed_security_data()
        app._db_initialized = True

    if session.get("user_id") and request.endpoint not in (
        "login","logout","static","health","change_password"
    ):
        module_key=ENDPOINT_MODULE_MAP.get(request.endpoint)
        if module_key:
            permission_key=f"{module_key}.{infer_security_action()}"
            if not has_permission(permission_key):
                try:
                    db.session.execute(text("""INSERT INTO access_denied_logs(
                        user_id,username,endpoint,permission_key,ip_address,created_at)
                        VALUES(:user,:username,:endpoint,:permission,:ip,:created)"""),
                        {"user":session.get("user_id"),"username":session.get("username"),
                         "endpoint":request.endpoint,"permission":permission_key,
                         "ip":request.headers.get("X-Forwarded-For",request.remote_addr),
                         "created":datetime.now()})
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                flash("ليس لديك صلاحية للوصول إلى هذه الشاشة","danger")
                return redirect(url_for("dashboard"))

@app.context_processor
def inject_settings():
    settings = row("SELECT * FROM settings WHERE id=1")
    return {"app_settings": settings, "app_version": APP_VERSION, "item_units": ITEM_UNITS,
            "journal_account_types": JOURNAL_ACCOUNT_TYPES,
            "can": has_permission, "current_username": session.get("username")}

@app.route("/health")
def health():
    db.session.execute(text("SELECT 1"))
    return {"status": "ok"}, 200

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        user = row("SELECT * FROM users WHERE username=:u", {"u": username})

        password_ok = False
        if user:
            stored_password = user["password"] or ""
            if stored_password.startswith(("pbkdf2:", "scrypt:")):
                password_ok = check_password_hash(stored_password, password)
            else:
                # Automatic migration for the old plaintext admin password.
                password_ok = stored_password == password
                if password_ok:
                    execute(
                        "UPDATE users SET password=:password WHERE id=:id",
                        {"password": generate_password_hash(password), "id": user["id"]},
                    )

        if user and password_ok:
            if not user.get("active",1):
                flash("هذا المستخدم موقوف","danger")
                return render_template("login.html")
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["permission_keys"] = list(user_permission_keys(user["id"]))
            execute("UPDATE users SET last_login=:dt WHERE id=:id",
                    {"dt":datetime.now(),"id":user["id"]})
            audit("LOGIN", "USER", f"تسجيل دخول المستخدم {username}")
            return redirect(url_for("dashboard"))

        flash("بيانات الدخول غير صحيحة", "danger")

    return render_template("login.html")

@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_password = request.form["current_password"]
        new_password = request.form["new_password"]
        confirm_password = request.form["confirm_password"]

        if len(new_password) < 8:
            flash("كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل", "danger")
            return redirect(url_for("change_password"))

        if new_password != confirm_password:
            flash("تأكيد كلمة المرور غير مطابق", "danger")
            return redirect(url_for("change_password"))

        user = row("SELECT * FROM users WHERE id=:id", {"id": session["user_id"]})
        stored_password = user["password"] or ""
        valid_current = (
            check_password_hash(stored_password, current_password)
            if stored_password.startswith(("pbkdf2:", "scrypt:"))
            else stored_password == current_password
        )

        if not valid_current:
            flash("كلمة المرور الحالية غير صحيحة", "danger")
            return redirect(url_for("change_password"))

        execute(
            "UPDATE users SET password=:password WHERE id=:id",
            {"password": generate_password_hash(new_password), "id": session["user_id"]},
        )
        audit("UPDATE", "USER", "تم تغيير كلمة المرور")
        flash("تم تغيير كلمة المرور بنجاح", "success")
        return redirect(url_for("dashboard"))

    return render_template("change_password.html")


@app.route("/audit-logs")
@login_required
def audit_logs():
    logs = rows("""
        SELECT * FROM audit_logs
        ORDER BY id DESC
        LIMIT 300
    """)
    return render_template("audit_logs.html", logs=logs)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def dashboard():
    sales = row("SELECT COALESCE(SUM(total),0) s FROM invoices WHERE status='معتمدة'")["s"]
    expenses_total = row("SELECT COALESCE(SUM(total),0) s FROM expenses")["s"]
    customers_count = row("SELECT COUNT(*) c FROM customers")["c"]
    employees_count = row("SELECT COUNT(*) c FROM employees WHERE active=1")["c"]
    recent = rows("""
        SELECT i.*, c.name customer_name FROM invoices i
        JOIN customers c ON c.id=i.customer_id
        ORDER BY i.id DESC LIMIT 5
    """)
    return render_template("dashboard.html", sales=float(sales), expenses=float(expenses_total),
                           customers=customers_count, employees=employees_count, recent=recent)

@app.route("/settings", methods=["GET","POST"])
@login_required
def settings():
    current = row("SELECT * FROM settings WHERE id=1")

    if request.method == "POST":
        logo_data = current["logo_data"] or ""
        logo_mime = current["logo_mime"] or ""

        uploaded_logo = request.files.get("logo_file")
        remove_logo = request.form.get("remove_logo") == "1"

        if remove_logo:
            logo_data = ""
            logo_mime = ""
        elif uploaded_logo and uploaded_logo.filename:
            allowed_mimes = {"image/png", "image/jpeg", "image/webp"}
            if uploaded_logo.mimetype not in allowed_mimes:
                flash("صيغة الشعار غير مدعومة. استخدم PNG أو JPG أو WEBP.", "danger")
                return redirect(url_for("settings"))

            raw = uploaded_logo.read()
            if len(raw) > 2 * 1024 * 1024:
                flash("حجم الشعار يجب ألا يتجاوز 2 ميجابايت.", "danger")
                return redirect(url_for("settings"))

            logo_data = base64.b64encode(raw).decode("ascii")
            logo_mime = uploaded_logo.mimetype

        execute("""UPDATE settings
                   SET company_name_ar=:ar, company_name_en=:en, vat_number=:vat,
                       cr_number=:cr, currency=:cur, address=:address, phone=:phone,
                       email=:email, logo_url=:logo_url, logo_data=:logo_data,
                       logo_mime=:logo_mime, vat_rate=:vat_rate,
                       customer_account_id=:customer_account_id,supplier_account_id=:supplier_account_id,
                       sales_account_id=:sales_account_id,purchases_account_id=:purchases_account_id,
                       vat_input_account_id=:vat_input_account_id,vat_output_account_id=:vat_output_account_id,
                       cash_account_id=:cash_account_id,bank_account_id=:bank_account_id,
                       inventory_account_id=:inventory_account_id,cost_of_sales_account_id=:cost_of_sales_account_id,
                       retained_earnings_account_id=:retained_earnings_account_id
                   WHERE id=1""",
                {"ar": request.form["company_name_ar"],
                 "en": request.form["company_name_en"],
                 "vat": request.form["vat_number"],
                 "cr": request.form["cr_number"],
                 "cur": request.form["currency"],
                 "address": request.form.get("address", ""),
                 "phone": request.form.get("phone", ""),
                 "email": request.form.get("email", ""),
                 "logo_url": request.form.get("logo_url", ""),
                 "logo_data": logo_data,
                 "logo_mime": logo_mime,
                 "vat_rate": float(request.form.get("vat_rate", 15) or 15),
                 "customer_account_id":request.form.get("customer_account_id") or None,
                 "supplier_account_id":request.form.get("supplier_account_id") or None,
                 "sales_account_id":request.form.get("sales_account_id") or None,
                 "purchases_account_id":request.form.get("purchases_account_id") or None,
                 "vat_input_account_id":request.form.get("vat_input_account_id") or None,
                 "vat_output_account_id":request.form.get("vat_output_account_id") or None,
                 "cash_account_id":request.form.get("cash_account_id") or None,
                 "bank_account_id":request.form.get("bank_account_id") or None,
                 "inventory_account_id":request.form.get("inventory_account_id") or None,
                 "cost_of_sales_account_id":request.form.get("cost_of_sales_account_id") or None,
                 "retained_earnings_account_id":request.form.get("retained_earnings_account_id") or None})

        audit("UPDATE", "SETTINGS", "تم تحديث إعدادات الشركة والشعار")
        flash("تم حفظ إعدادات الشركة والشعار", "success")
        return redirect(url_for("settings"))

    return render_template("settings.html",row=current,
        accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                         WHERE active=1 AND accepts_entries=1 ORDER BY account_code"""))

@app.route("/branches", methods=["GET","POST"])
@login_required
def branches():
    if request.method == "POST":
        execute("INSERT INTO branches(name,city) VALUES(:n,:c)",
                {"n": request.form["name"], "c": request.form["city"]})
        audit("CREATE", "BRANCH", f"إضافة فرع: {request.form['name']}")
        flash("تمت إضافة الفرع", "success")
    return render_template("branches.html", rows=rows("SELECT * FROM branches ORDER BY id DESC"))

@app.route("/customers", methods=["GET","POST"])
@login_required
def customers():
    if request.method == "POST":
        name_ar = request.form["name"].strip()
        name_en = request.form.get("name_en","").strip() or transliterate_arabic_name(name_ar)
        party_account=row("""SELECT id,account_name_ar FROM chart_of_accounts WHERE id=:id AND active=1
          AND accepts_entries=1 AND account_type='أصل'""",{"id":request.form.get("receivable_account_id")})
        if not party_account:
            flash("اختر حساب عميل صحيحًا من دليل الحسابات","danger")
            return redirect(url_for("customers"))
        if row("SELECT id FROM customers WHERE receivable_account_id=:id",{"id":party_account["id"]}):
            flash("هذا الحساب مرتبط بعميل آخر مسبقًا","danger")
            return redirect(url_for("customers"))
        name_ar=party_account["account_name_ar"]
        name_en=request.form.get("name_en","").strip() or transliterate_arabic_name(name_ar)
        execute("""INSERT INTO customers(name,name_en,vat_number,phone,email,receivable_account_id)
                   VALUES(:n,:ne,:v,:p,:e,:account)""",
                {"n":name_ar,"ne":name_en,"v":request.form["vat_number"],
                 "p":request.form["phone"],"e":request.form["email"],
                 "account":request.form.get("receivable_account_id") or None})
        audit("CREATE","CUSTOMER",f"إضافة عميل: {name_ar} / {name_en}")
        flash("تمت إضافة العميل", "success")
    return render_template("customers.html", rows=rows("SELECT * FROM customers ORDER BY id DESC"),
      party_accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
        WHERE active=1 AND accepts_entries=1 AND account_type='أصل'
        AND NOT EXISTS (SELECT 1 FROM customers c WHERE c.receivable_account_id=chart_of_accounts.id)
        ORDER BY account_code"""))

@app.route("/suppliers", methods=["GET","POST"])
@login_required
def suppliers():
    if request.method == "POST":
        name_ar = request.form["name"].strip()
        name_en = request.form.get("name_en","").strip() or transliterate_arabic_name(name_ar)
        party_account=row("""SELECT id,account_name_ar FROM chart_of_accounts WHERE id=:id AND active=1
          AND accepts_entries=1 AND account_type IN ('خصم','التزام')""",{"id":request.form.get("payable_account_id")})
        if not party_account:
            flash("اختر حساب مورد صحيحًا من دليل الحسابات","danger")
            return redirect(url_for("suppliers"))
        if row("SELECT id FROM suppliers WHERE payable_account_id=:id",{"id":party_account["id"]}):
            flash("هذا الحساب مرتبط بمورد آخر مسبقًا","danger")
            return redirect(url_for("suppliers"))
        name_ar=party_account["account_name_ar"]
        name_en=request.form.get("name_en","").strip() or transliterate_arabic_name(name_ar)
        execute("""INSERT INTO suppliers(name,name_en,vat_number,phone,email,payable_account_id)
                   VALUES(:n,:ne,:v,:p,:e,:account)""",
                {"n":name_ar,"ne":name_en,"v":request.form["vat_number"],
                 "p":request.form["phone"],"e":request.form["email"],
                 "account":request.form.get("payable_account_id") or None})
        audit("CREATE","SUPPLIER",f"إضافة مورد: {name_ar} / {name_en}")
        flash("تمت إضافة المورد", "success")
    return render_template("suppliers.html", rows=rows("SELECT * FROM suppliers ORDER BY id DESC"),
      party_accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
        WHERE active=1 AND accepts_entries=1 AND account_type IN ('خصم','التزام')
        AND NOT EXISTS (SELECT 1 FROM suppliers s WHERE s.payable_account_id=chart_of_accounts.id)
        ORDER BY account_code"""))

def party_crud_config(party_type):
    if party_type=="customer":
        return {"table":"customers","label":"العميل","list_endpoint":"customers","statement_type":"customer","account_field":"receivable_account_id","account_types":("أصل",)}
    if party_type=="supplier":
        return {"table":"suppliers","label":"المورد","list_endpoint":"suppliers","statement_type":"supplier","account_field":"payable_account_id","account_types":("خصم","التزام")}
    return None

@app.route("/parties/<party_type>/<int:party_id>/edit",methods=["GET","POST"])
@login_required
def party_edit(party_type,party_id):
    cfg=party_crud_config(party_type)
    if not cfg:
        return "نوع السجل غير صحيح",404
    party=row(f"SELECT id,name,name_en,vat_number,phone,email,{cfg['account_field']} AS party_account_id FROM {cfg['table']} WHERE id=:id",{"id":party_id})
    if not party:
        return f"{cfg['label']} غير موجود",404
    if request.method=="POST":
        name=(request.form.get("name") or "").strip()
        if not name:
            flash("الاسم مطلوب","danger")
            return redirect(request.url)
        duplicate=row(f"SELECT id FROM {cfg['table']} WHERE LOWER(name)=LOWER(:name) AND id<>:id",
                      {"name":name,"id":party_id})
        if duplicate:
            flash(f"يوجد {cfg['label']} آخر بالاسم نفسه","danger")
            return redirect(request.url)
        requested_account=request.form.get("party_account_id")
        account_types=cfg["account_types"]
        placeholders=",".join(f":type{i}" for i in range(len(account_types)))
        account_params={f"type{i}":value for i,value in enumerate(account_types)}
        valid_account=row(f"""SELECT id,account_name_ar FROM chart_of_accounts WHERE id=:account_id AND active=1
          AND accepts_entries=1 AND account_type IN ({placeholders})""",
          {**account_params,"account_id":requested_account})
        if not valid_account:
            flash(f"اختر حساب {cfg['label']} صحيحًا من دليل الحسابات","danger")
            return redirect(request.url)
        linked_elsewhere=row(f"SELECT id FROM {cfg['table']} WHERE {cfg['account_field']}=:account AND id<>:id",{"account":requested_account,"id":party_id})
        if linked_elsewhere:
            flash("هذا الحساب مرتبط بجهة أخرى مسبقًا","danger")
            return redirect(request.url)
        name=valid_account["account_name_ar"]
        duplicate=row(f"SELECT id FROM {cfg['table']} WHERE LOWER(name)=LOWER(:name) AND id<>:id",{"name":name,"id":party_id})
        if duplicate:
            flash(f"يوجد {cfg['label']} آخر بالاسم نفسه","danger")
            return redirect(request.url)
        name_en=(request.form.get("name_en") or "").strip() or transliterate_arabic_name(name)
        execute(f"""UPDATE {cfg['table']} SET name=:name,name_en=:name_en,
          vat_number=:vat,phone=:phone,email=:email,{cfg['account_field']}=:party_account_id WHERE id=:id""",
          {"name":name,"name_en":name_en,"vat":(request.form.get("vat_number") or "").strip(),
           "phone":(request.form.get("phone") or "").strip(),"email":(request.form.get("email") or "").strip(),
           "party_account_id":requested_account,
           "id":party_id})
        audit("UPDATE",party_type.upper(),f"تعديل {cfg['label']}: {name}")
        flash(f"تم تعديل بيانات {cfg['label']}","success")
        return redirect(url_for(cfg["list_endpoint"]))
    account_types=cfg["account_types"]
    placeholders=",".join(f":type{i}" for i in range(len(account_types)))
    params={f"type{i}":value for i,value in enumerate(account_types)}
    party_accounts=rows(f"""SELECT id,account_code,account_name_ar FROM chart_of_accounts
      WHERE active=1 AND accepts_entries=1 AND account_type IN ({placeholders}) ORDER BY account_code""",params)
    return render_template("party_edit.html",party=party,party_type=party_type,cfg=cfg,party_accounts=party_accounts)

@app.route("/parties/<party_type>/<int:party_id>/delete",methods=["POST"])
@login_required
def party_delete(party_type,party_id):
    cfg=party_crud_config(party_type)
    if not cfg:
        return "نوع السجل غير صحيح",404
    party=row(f"SELECT id,name FROM {cfg['table']} WHERE id=:id",{"id":party_id})
    if not party:
        flash(f"{cfg['label']} غير موجود","danger")
        return redirect(url_for(cfg["list_endpoint"]))
    try:
        db.session.execute(text(f"DELETE FROM {cfg['table']} WHERE id=:id"),{"id":party_id})
        db.session.commit()
        audit("DELETE",party_type.upper(),f"حذف {cfg['label']}: {party['name']}")
        flash(f"تم حذف {cfg['label']} بنجاح","success")
    except Exception:
        db.session.rollback()
        flash(f"لا يمكن حذف {cfg['label']} لوجود فواتير أو قيود أو حركات مرتبطة به. يمكن تعديل بياناته بدلاً من الحذف.","danger")
    return redirect(url_for(cfg["list_endpoint"]))

@app.route("/api/suppliers/quick-create", methods=["POST"])
@login_required
def supplier_quick_create():
    data=request.get_json(silent=True) or {}
    name=(data.get("name") or "").strip()
    if not name:
        return {"error":"اسم المورد مطلوب"},400
    existing=row("SELECT id,name,name_en,vat_number FROM suppliers WHERE LOWER(name)=LOWER(:name) LIMIT 1",{"name":name})
    if existing:
        return {"supplier":dict(existing),"created":False}
    name_en=(data.get("name_en") or "").strip() or transliterate_arabic_name(name)
    execute("""INSERT INTO suppliers(name,name_en,vat_number,phone,email)
               VALUES(:name,:name_en,:vat,:phone,:email)""",
            {"name":name,"name_en":name_en,"vat":(data.get("vat_number") or "").strip(),
             "phone":(data.get("phone") or "").strip(),"email":(data.get("email") or "").strip()})
    supplier=row("SELECT id,name,name_en,vat_number FROM suppliers WHERE LOWER(name)=LOWER(:name) ORDER BY id DESC LIMIT 1",{"name":name})
    audit("CREATE","SUPPLIER",f"إضافة مورد سريعة من أمر الشراء: {name}")
    return {"supplier":dict(supplier),"created":True},201

@app.route("/invoices", methods=["GET","POST"])
@login_required
def invoices():
    if request.method=="POST":
        try:
            prepared_items,gross_subtotal,discount_total,net_subtotal,invoice_vat,invoice_total = prepare_invoice_items_from_form(request.form)

            customer_id=request.form.get("customer_id")
            invoice_date=request.form.get("invoice_date")
            if not customer_id:
                raise ValueError("اختر العميل قبل حفظ الفاتورة.")
            if not invoice_date:
                raise ValueError("أدخل تاريخ الفاتورة.")

            customer=row("SELECT id FROM customers WHERE id=:id",{"id":customer_id})
            if not customer:
                raise ValueError("العميل المحدد غير موجود.")

            requested_status=request.form.get("status","مسودة")
            if requested_status not in ("مسودة","معتمدة"):
                requested_status="مسودة"

            invoice_no=next_invoice_draft_number()
            invoice_uuid=str(uuid.uuid4())

            # One database transaction: invoice + items.
            db.session.execute(text("""INSERT INTO invoices(
              invoice_no,draft_number,invoice_uuid,customer_id,invoice_date,due_date,
              subtotal,discount,vat,total,status,branch_id,cost_center_id,project_id,
              payment_terms,payment_method,sales_person,customer_reference,
              notes,created_at,updated_at,posting_status)
              VALUES(:no,:draft_number,:uuid,:customer,:invoice_date,:due_date,:subtotal,:discount,
              :vat,:total,'مسودة',:branch,:cost_center,:project,:payment_terms,
              :payment_method,:sales_person,:customer_reference,:notes,:created,:updated,
              'غير مرحّل')"""),
              {"no":invoice_no,"draft_number":invoice_no,"uuid":invoice_uuid,"customer":customer_id,
               "invoice_date":invoice_date,
               "due_date":request.form.get("due_date") or None,
               "subtotal":net_subtotal,"discount":discount_total,
               "vat":invoice_vat,"total":invoice_total,
               "branch":request.form.get("branch_id") or None,
               "cost_center":request.form.get("cost_center_id") or None,
               "project":request.form.get("project_id") or None,
               "payment_terms":request.form.get("payment_terms",""),
               "payment_method":request.form.get("payment_method",""),
               "sales_person":request.form.get("sales_person",""),
               "customer_reference":request.form.get("customer_reference",""),
               "notes":request.form.get("notes",""),
               "created":datetime.now(),"updated":datetime.now()})

            invoice_id=db.session.execute(
                text("SELECT id FROM invoices WHERE invoice_no=:no"),
                {"no":invoice_no}
            ).scalar()

            for item in prepared_items:
                db.session.execute(text("""INSERT INTO invoice_items(
                  invoice_id,item_name,description,quantity,unit,unit_price,
                  discount_rate,vat_rate,line_subtotal,line_discount,line_vat,line_total)
                  VALUES(:invoice_id,:item_name,:description,:quantity,:unit,:unit_price,
                  :discount_rate,:vat_rate,:line_subtotal,:line_discount,:line_vat,:line_total)"""),
                  {"invoice_id":invoice_id,**item})

            db.session.commit()
            audit("CREATE","INVOICE",f"إنشاء فاتورة {invoice_no}")

            if requested_status=="معتمدة":
                try:
                    draft_number,official_number=assign_official_invoice_number(invoice_id)
                    post_invoice_to_ledger(invoice_id)
                    execute("""UPDATE invoices
                               SET status='معتمدة',approved_by=:user,
                                   approved_at=:dt,updated_at=:dt
                               WHERE id=:id""",
                            {"user":session.get("user_id"),
                             "dt":datetime.now(),"id":invoice_id})
                    audit("APPROVE","INVOICE",
                          f"اعتماد الفاتورة وتغيير الرقم من {draft_number} إلى {official_number}")
                    flash(f"تم الحفظ والاعتماد. الرقم النهائي: {official_number}.","success")
                except Exception as exc:
                    db.session.rollback()
                    try:
                        execute("""UPDATE invoices
                                   SET invoice_no=COALESCE(draft_number,:draft_number),
                                       official_invoice_no=NULL,
                                       status='مسودة',posting_status='خطأ'
                                   WHERE id=:id""",
                                {"draft_number":invoice_no,"id":invoice_id})
                    except Exception:
                        db.session.rollback()
                    flash(f"تم حفظ الفاتورة برقمها الأولي، لكن تعذر اعتمادها: {exc}","warning")
            else:
                flash(f"تم حفظ الفاتورة برقم {invoice_no}.","success")

            return redirect(url_for(
                "invoice_view",
                invoice_id=invoice_id,
                print=1 if request.form.get("print_after_save")=="1" else None
            ))

        except Exception as exc:
            db.session.rollback()
            app.logger.exception("Invoice save failed")
            flash(f"تعذر حفظ الفاتورة: {exc}","danger")
            return redirect(url_for("invoices"))

    invoice_rows=rows("""SELECT i.*,c.name customer_name,b.name branch_name
      FROM invoices i JOIN customers c ON c.id=i.customer_id
      LEFT JOIN branches b ON b.id=i.branch_id ORDER BY i.id DESC""")
    return render_template("invoices.html",rows=invoice_rows,
      customers=rows("SELECT * FROM customers ORDER BY name"),
      branches=rows("SELECT * FROM branches WHERE active=1 ORDER BY name"),
      inventory_items=rows("SELECT * FROM inventory ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"),
      projects=rows("SELECT id,project_no,name FROM projects ORDER BY project_no"))


@app.route("/expenses", methods=["GET","POST"])
@login_required
def expenses():
    if request.method=="POST":
        amount=round(float(request.form["amount"] or 0),2)
        taxable=1 if request.form.get("taxable")=="1" else 0
        rate=float(row("SELECT vat_rate FROM settings WHERE id=1")["vat_rate"] or 15)
        vat=round(amount*rate/100,2) if taxable else 0
        total=round(amount+vat,2)
        execute("""INSERT INTO expenses(expense_date,category,description,amount,vat,total,
          branch_id,supplier_id,expense_account_id,payment_account_id,cost_center_id,taxable,
          invoice_number,invoice_date,status,posting_status)
          VALUES(:dt,:cat,:des,:amt,:vat,:tot,:bid,:sid,:ea,:pa,:cc,:tax,:inv,:idate,:status,'غير مرحّل')""",
          {"dt":request.form["expense_date"],"cat":request.form["category"],
           "des":request.form.get("description",""),"amt":amount,"vat":vat,"tot":total,
           "bid":request.form.get("branch_id") or None,"sid":request.form.get("supplier_id") or None,
           "ea":request.form.get("expense_account_id") or None,"pa":request.form.get("payment_account_id") or None,
           "cc":request.form.get("cost_center_id") or None,"tax":taxable,
           "inv":request.form.get("invoice_number",""),"idate":request.form.get("invoice_date") or None,
           "status":request.form.get("status","مسودة")})
        expense_id=row("SELECT MAX(id) id FROM expenses")["id"]
        if request.form.get("status")=="معتمدة":
            try:
                post_expense_to_ledger(expense_id)
                flash("تم تسجيل المصروف وترحيله محاسبيًا","success")
            except Exception as exc:
                db.session.rollback()
                execute("UPDATE expenses SET status='مسودة',posting_status='خطأ' WHERE id=:id",{"id":expense_id})
                flash(f"تم حفظ المصروف كمسودة ولم يُرحّل: {exc}","danger")
        else: flash("تم حفظ المصروف كمسودة","success")
        audit("CREATE","EXPENSE",f"إضافة مصروف: {request.form['category']}")
        return redirect(url_for("expenses"))
    expense_rows=rows("""SELECT e.*,b.name branch_name,s.name supplier_name,
      a.account_name_ar expense_account_name,p.account_name_ar payment_account_name,
      cc.name cost_center_name,j.journal_no FROM expenses e
      LEFT JOIN branches b ON b.id=e.branch_id LEFT JOIN suppliers s ON s.id=e.supplier_id
      LEFT JOIN chart_of_accounts a ON a.id=e.expense_account_id
      LEFT JOIN chart_of_accounts p ON p.id=e.payment_account_id
      LEFT JOIN cost_centers cc ON cc.id=e.cost_center_id
      LEFT JOIN journal_entries j ON j.id=e.journal_id ORDER BY e.id DESC""")
    accts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                  WHERE active=1 AND accepts_entries=1 ORDER BY account_code""")
    return render_template("expenses.html",rows=expense_rows,
      branches=rows("SELECT * FROM branches WHERE active=1 ORDER BY name"),
      suppliers=rows("SELECT * FROM suppliers ORDER BY name"),accounts=accts,
      centers=rows("SELECT * FROM cost_centers WHERE active=1 ORDER BY code"))


@app.route("/inventory", methods=["GET","POST"])
@login_required
def inventory():
    if request.method=="POST":
        sku=next_inventory_sku()
        name=request.form["name"].strip()
        execute("""INSERT INTO inventory(
          sku,code,name,name_en,barcode,category_id,item_type,unit,cost,unit_cost,sale_price,reorder_level,
          min_level,max_level,valuation_method,track_batch,track_serial,track_expiry,active,quantity)
          VALUES(:sku,:sku,:name,:name_en,:barcode,:category,:type,:unit,:cost,:cost,:sale,:reorder,
          :min,:max,:valuation,:batch,:serial,:expiry,1,0)""",
          {"sku":sku,"name":name,"name_en":request.form.get("name_en",""),
           "barcode":request.form.get("barcode",""),"category":request.form.get("category_id") or None,
           "type":request.form.get("item_type","مخزني"),"unit":normalize_item_unit(request.form.get("unit")),
           "cost":float(request.form.get("cost") or 0),"sale":float(request.form.get("sale_price") or 0),
           "reorder":float(request.form.get("reorder_level") or 0),
           "min":float(request.form.get("min_level") or 0),"max":float(request.form.get("max_level") or 0),
           "valuation":request.form.get("valuation_method","متوسط متحرك"),
           "batch":1 if request.form.get("track_batch") else 0,
           "serial":1 if request.form.get("track_serial") else 0,
           "expiry":1 if request.form.get("track_expiry") else 0})
        item_id=row("SELECT id FROM inventory WHERE sku IS NOT DISTINCT FROM :sku AND name=:name ORDER BY id DESC LIMIT 1",
                    {"sku":sku,"name":name})["id"]
        opening=float(request.form.get("opening_quantity") or 0)
        if opening>0:
            warehouse_id=request.form.get("warehouse_id",type=int)
            if not warehouse_id:
                flash("اختر المستودع للرصيد الافتتاحي","danger")
                return redirect(url_for("inventory"))
            record_inventory_movement(
                request.form.get("opening_date") or datetime.now().date().isoformat(),
                "رصيد افتتاحي",item_id,warehouse_id,opening,
                float(request.form.get("cost") or 0),notes="رصيد افتتاحي"
            )
        audit("CREATE","INVENTORY",f"إضافة صنف: {name}")
        flash("تم إنشاء بطاقة الصنف","success")
        return redirect(url_for("inventory_item_view",item_id=item_id))
    item_rows=rows("""SELECT i.*,c.name category_name,
      COALESCE((SELECT SUM(CASE
        WHEN m.movement_type IN ('رصيد افتتاحي','استلام','تسوية زيادة','تحويل وارد','مرتجع مبيعات') THEN m.quantity
        WHEN m.movement_type IN ('صرف','تسوية نقص','تحويل صادر','بيع','مرتجع مشتريات') THEN -m.quantity ELSE 0 END)
        FROM inventory_movements m WHERE m.item_id=i.id),0) stock_qty
      FROM inventory i LEFT JOIN inventory_categories c ON c.id=i.category_id
      ORDER BY i.id DESC""")
    return render_template("inventory.html",rows=item_rows,
      warehouses=rows("SELECT * FROM warehouses WHERE active=1 ORDER BY code"),
      categories=rows("SELECT * FROM inventory_categories WHERE active=1 ORDER BY name"))

@app.route("/inventory/items/<int:item_id>")
@login_required
def inventory_item_view(item_id):
    item=row("""SELECT i.*,c.name category_name FROM inventory i
                LEFT JOIN inventory_categories c ON c.id=i.category_id WHERE i.id=:id""",{"id":item_id})
    if not item:return "الصنف غير موجود",404
    balances=rows("""SELECT w.id,w.code,w.name,
      COALESCE(SUM(CASE
        WHEN m.movement_type IN ('رصيد افتتاحي','استلام','تسوية زيادة','تحويل وارد','مرتجع مبيعات') THEN m.quantity
        WHEN m.movement_type IN ('صرف','تسوية نقص','تحويل صادر','بيع','مرتجع مشتريات') THEN -m.quantity ELSE 0 END),0) quantity
      FROM warehouses w LEFT JOIN inventory_movements m ON m.warehouse_id=w.id AND m.item_id=:item
      GROUP BY w.id,w.code,w.name ORDER BY w.code""",{"item":item_id})
    movements=rows("""SELECT m.*,w.code warehouse_code,w.name warehouse_name,
                      d.code destination_code,d.name destination_name
      FROM inventory_movements m JOIN warehouses w ON w.id=m.warehouse_id
      LEFT JOIN warehouses d ON d.id=m.destination_warehouse_id
      WHERE m.item_id=:item ORDER BY m.movement_date DESC,m.id DESC""",{"item":item_id})
    return render_template("inventory_item_view.html",item=item,balances=balances,movements=movements)

@app.route("/inventory/movements",methods=["GET","POST"])
@login_required
def inventory_movements():
    if request.method=="POST":
        try:
            record_inventory_movement(
              request.form["movement_date"],request.form["movement_type"],
              int(request.form["item_id"]),int(request.form["warehouse_id"]),
              request.form["quantity"],request.form.get("unit_cost") or 0,
              reference_no=request.form.get("reference_no",""),
              batch_no=request.form.get("batch_no",""),serial_no=request.form.get("serial_no",""),
              production_date=request.form.get("production_date") or None,
              expiry_date=request.form.get("expiry_date") or None,
              notes=request.form.get("notes",""))
            flash("تم تسجيل حركة المخزون","success")
        except Exception as exc:
            flash(str(exc),"danger")
        return redirect(url_for("inventory_movements"))
    data=rows("""SELECT m.*,i.sku,i.name item_name,w.code warehouse_code,w.name warehouse_name
                 FROM inventory_movements m JOIN inventory i ON i.id=m.item_id
                 JOIN warehouses w ON w.id=m.warehouse_id
                 ORDER BY m.movement_date DESC,m.id DESC LIMIT 500""")
    return render_template("inventory_movements.html",movements=data,
      items=rows("SELECT id,sku,name,unit,cost FROM inventory WHERE active=1 ORDER BY name"),
      warehouses=rows("SELECT id,code,name FROM warehouses WHERE active=1 ORDER BY code"))

@app.route("/inventory/transfers",methods=["GET","POST"])
@login_required
def inventory_transfers():
    if request.method=="POST":
        item_id=int(request.form["item_id"]); source=int(request.form["source_warehouse_id"])
        destination=int(request.form["destination_warehouse_id"])
        if source==destination:
            flash("يجب اختيار مستودعين مختلفين","danger")
            return redirect(url_for("inventory_transfers"))
        qty=float(request.form["quantity"]); dt=request.form["transfer_date"]
        item=row("SELECT cost FROM inventory WHERE id=:id",{"id":item_id})
        try:
            ref=f"TR-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            record_inventory_movement(dt,"تحويل صادر",item_id,source,qty,item["cost"],
                                      destination_warehouse_id=destination,reference_type="TRANSFER",
                                      reference_no=ref,notes=request.form.get("notes",""))
            record_inventory_movement(dt,"تحويل وارد",item_id,destination,qty,item["cost"],
                                      destination_warehouse_id=source,reference_type="TRANSFER",
                                      reference_no=ref,notes=request.form.get("notes",""))
            flash("تم تحويل المخزون بين المستودعات","success")
        except Exception as exc:
            flash(str(exc),"danger")
        return redirect(url_for("inventory_transfers"))
    return render_template("inventory_transfers.html",
      items=rows("SELECT id,sku,name,unit FROM inventory WHERE active=1 ORDER BY name"),
      warehouses=rows("SELECT id,code,name FROM warehouses WHERE active=1 ORDER BY code"),
      transfers=rows("""SELECT m.*,i.name item_name,w.name source_name,d.name destination_name
        FROM inventory_movements m JOIN inventory i ON i.id=m.item_id
        JOIN warehouses w ON w.id=m.warehouse_id
        LEFT JOIN warehouses d ON d.id=m.destination_warehouse_id
        WHERE m.movement_type='تحويل صادر' ORDER BY m.id DESC LIMIT 200"""))

@app.route("/inventory/counts",methods=["GET","POST"])
@login_required
def inventory_counts():
    if request.method=="POST":
        count_date=request.form["count_date"]; warehouse_id=int(request.form["warehouse_id"])
        no=next_inventory_count_number(count_date)
        execute("""INSERT INTO inventory_counts(count_no,count_date,warehouse_id,status,notes,created_by,created_at)
          VALUES(:no,:dt,:warehouse,'معتمد',:notes,:uid,:created)""",
          {"no":no,"dt":count_date,"warehouse":warehouse_id,"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        count_id=row("SELECT id FROM inventory_counts WHERE count_no=:no",{"no":no})["id"]
        item_ids=request.form.getlist("item_id[]"); counted=request.form.getlist("counted_qty[]")
        for idx,item_id in enumerate(item_ids):
            item=row("SELECT cost FROM inventory WHERE id=:id",{"id":item_id})
            system=warehouse_stock(int(item_id),warehouse_id)
            actual=float(counted[idx] or 0); variance=round(actual-system,3)
            execute("""INSERT INTO inventory_count_items(count_id,item_id,system_qty,counted_qty,variance_qty,unit_cost)
                       VALUES(:count,:item,:system,:actual,:variance,:cost)""",
                    {"count":count_id,"item":item_id,"system":system,"actual":actual,
                     "variance":variance,"cost":item["cost"]})
            if variance>0:
                record_inventory_movement(count_date,"تسوية زيادة",int(item_id),warehouse_id,variance,item["cost"],
                                          reference_type="STOCK_COUNT",reference_id=count_id,reference_no=no)
            elif variance<0:
                record_inventory_movement(count_date,"تسوية نقص",int(item_id),warehouse_id,abs(variance),item["cost"],
                                          reference_type="STOCK_COUNT",reference_id=count_id,reference_no=no)
        flash(f"تم اعتماد الجرد {no} وترحيل الفروقات","success")
        return redirect(url_for("inventory_counts"))
    items=rows("SELECT id,sku,name,unit,cost FROM inventory WHERE active=1 ORDER BY name")
    return render_template("inventory_counts.html",
      items=items,warehouses=rows("SELECT id,code,name FROM warehouses WHERE active=1 ORDER BY code"),
      counts=rows("""SELECT c.*,w.code warehouse_code,w.name warehouse_name
                     FROM inventory_counts c JOIN warehouses w ON w.id=c.warehouse_id
                     ORDER BY c.count_date DESC,c.id DESC"""))

@app.route("/inventory/warehouses",methods=["GET","POST"])
@login_required
def warehouses():
    if request.method=="POST":
        execute("""INSERT INTO warehouses(code,name,branch_id,active)
                   VALUES(:code,:name,:branch,1) ON CONFLICT(code) DO NOTHING""",
                {"code":request.form["code"],"name":request.form["name"],
                 "branch":request.form.get("branch_id") or None})
        flash("تم حفظ المستودع","success")
        return redirect(url_for("warehouses"))
    return render_template("warehouses.html",
      warehouses=rows("""SELECT w.*,b.name branch_name FROM warehouses w
                         LEFT JOIN branches b ON b.id=w.branch_id ORDER BY w.code"""),
      branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"))

@app.route("/api/warehouses/quick-create", methods=["POST"])
@login_required
def warehouse_quick_create():
    data=request.get_json(silent=True) or {}
    code=(data.get("code") or "").strip()
    name=(data.get("name") or "").strip()
    if not code or not name:
        return {"error":"كود واسم المستودع مطلوبان"},400
    existing=row("SELECT id,code,name FROM warehouses WHERE LOWER(code)=LOWER(:code) LIMIT 1",{"code":code})
    if existing:
        return {"warehouse":dict(existing),"created":False}
    execute("INSERT INTO warehouses(code,name,branch_id,active) VALUES(:code,:name,:branch,1)",
            {"code":code,"name":name,"branch":data.get("branch_id") or None})
    warehouse=row("SELECT id,code,name FROM warehouses WHERE LOWER(code)=LOWER(:code) LIMIT 1",{"code":code})
    audit("CREATE","WAREHOUSE",f"إضافة مستودع سريعة من أمر الشراء: {code} - {name}")
    return {"warehouse":dict(warehouse),"created":True},201

@app.route("/api/inventory/quick-create", methods=["POST"])
@login_required
def inventory_quick_create():
    try:
        data=request.get_json(silent=True) or {}
        name=(data.get("name") or "").strip()
        if not name:
            return {"ok":False,"error":"اسم الصنف مطلوب"},400
        unit=normalize_item_unit(data.get("unit"))
        cost=float(data.get("cost") or 0)
        existing=row("""SELECT id,sku,name,unit,cost FROM inventory
                        WHERE LOWER(name)=LOWER(:name) LIMIT 1""",{"name":name})
        if existing:
            return {"ok":True,"item":dict(existing),"created":False}
        sku=next_inventory_sku()
        result=db.session.execute(text("""INSERT INTO inventory(
          sku,code,name,quantity,unit,cost,unit_cost,sale_price,reorder_level,active)
          VALUES(:sku,:sku,:name,0,:unit,:cost,:cost,0,0,1)
          RETURNING id,sku,name,unit,cost"""),
          {"sku":sku,"name":name,"unit":unit,"cost":cost}).mappings().first()
        db.session.commit()
        audit("CREATE","INVENTORY_ITEM",f"إضافة صنف سريعة: {sku} - {name}")
        return {"ok":True,"item":dict(result),"created":True},201
    except Exception as exc:
        db.session.rollback()
        app.logger.exception("Quick inventory item creation failed")
        return {"ok":False,"error":f"تعذر حفظ الصنف: {exc}"},400

@app.route("/inventory/export.xlsx")
@login_required
def inventory_export():
    data=rows("""SELECT sku,name,name_en,barcode,unit,quantity,cost,sale_price,reorder_level,
                 min_level,max_level,valuation_method,active FROM inventory ORDER BY sku,name""")
    headers=["الكود","اسم الصنف","English Name","الباركود","الوحدة","الرصيد","متوسط التكلفة",
             "سعر البيع","إعادة الطلب","الحد الأدنى","الحد الأعلى","طريقة التقييم","الحالة"]
    keys=["sku","name","name_en","barcode","unit","quantity","cost","sale_price","reorder_level",
          "min_level","max_level","valuation_method","active"]
    return xlsx_response("inventory.xlsx","المخزون",headers,[[r.get(k) for k in keys] for r in data])


@app.route("/employees", methods=["GET","POST"])
@login_required
def employees():
    if request.method == "POST":
        execute("""INSERT INTO employees(employee_no,name,job_title,branch_id,basic_salary,allowances)
                   VALUES(:eno,:n,:job,:bid,:sal,:allow)""",
                {"eno": request.form["employee_no"] or None, "n": request.form["name"],
                 "job": request.form["job_title"], "bid": request.form["branch_id"],
                 "sal": float(request.form["basic_salary"] or 0),
                 "allow": float(request.form["allowances"] or 0)})
        audit("CREATE", "EMPLOYEE", f"إضافة موظف: {request.form['name']}")
        flash("تمت إضافة الموظف", "success")
    employee_rows = rows("""SELECT e.*, b.name branch_name FROM employees e
                            LEFT JOIN branches b ON b.id=e.branch_id ORDER BY e.id DESC""")
    return render_template("employees.html", rows=employee_rows,
                           branches=rows("SELECT * FROM branches WHERE active=1 ORDER BY name"))



def zatca_tlv_base64(seller_name, vat_number, invoice_datetime, total, vat_total):
    """
    Create the QR payload using only the five standard TLV fields:
    1 seller name
    2 VAT number
    3 invoice date/time
    4 invoice total including VAT
    5 VAT total

    The invoice number and UUID are intentionally NOT included in the QR payload.
    """
    values = [
        (1, str(seller_name or "")),
        (2, str(vat_number or "")),
        (3, str(invoice_datetime or "")),
        (4, f"{Decimal(str(total)):.2f}"),
        (5, f"{Decimal(str(vat_total)):.2f}"),
    ]

    payload = bytearray()
    for tag, value in values:
        encoded = value.encode("utf-8")
        if len(encoded) > 255:
            raise ValueError("QR field is too long")
        payload.extend(bytes([tag, len(encoded)]))
        payload.extend(encoded)

    return base64.b64encode(payload).decode("ascii")


def qr_data_uri(data):
    qr = qrcode.QRCode(version=None, box_size=7, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    image = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")


@app.route("/invoice/<int:invoice_id>")
@login_required
def invoice_view(invoice_id):
    invoice = row("""
        SELECT i.*, c.name customer_name, c.name_en customer_name_en, c.vat_number customer_vat,
               c.phone customer_phone, c.email customer_email,
               b.name branch_name
        FROM invoices i
        JOIN customers c ON c.id=i.customer_id
        LEFT JOIN branches b ON b.id=i.branch_id
        WHERE i.id=:id
    """, {"id": invoice_id})

    if not invoice:
        return "الفاتورة غير موجودة", 404

    items = rows("""
        SELECT *
        FROM invoice_items
        WHERE invoice_id=:invoice_id
        ORDER BY id
    """, {"invoice_id": invoice_id})

    company = row("SELECT * FROM settings WHERE id=1")
    issue_time = invoice["created_at"].isoformat() if invoice["created_at"] else str(invoice["invoice_date"])

    tlv = zatca_tlv_base64(
        company["company_name_ar"],
        company["vat_number"],
        issue_time,
        invoice["total"],
        invoice["vat"]
    )
    qr_image = qr_data_uri(tlv)

    return render_template(
        "invoice_print.html",
        invoice=invoice,
        items=items,
        company=company,
        qr_image=qr_image,
        qr_payload=tlv,
        print_lang=request.args.get("lang","ar") if request.args.get("lang","ar") in ("ar","en","bi") else "ar"
    )



@app.route("/chart-of-accounts", methods=["GET","POST"])
@login_required
def chart_of_accounts():
    if request.method == "POST":
        linked_party=(request.form.get("linked_party") or "").strip()
        linked_type=None; linked_id=None; linked_record=None
        if linked_party:
            try:
                linked_type,raw_id=linked_party.split(":",1); linked_id=int(raw_id)
            except (ValueError,TypeError):
                flash("الجهة المرتبطة غير صحيحة","danger")
                return redirect(url_for("chart_of_accounts"))
            if linked_type=="customer":
                linked_record=row("SELECT id,name,name_en,receivable_account_id linked_account_id FROM customers WHERE id=:id",{"id":linked_id})
                linked_account_type="أصل"
            elif linked_type=="supplier":
                linked_record=row("SELECT id,name,name_en,payable_account_id linked_account_id FROM suppliers WHERE id=:id",{"id":linked_id})
                linked_account_type="خصم"
            else:
                linked_record=None
            if not linked_record or linked_record["linked_account_id"]:
                flash("العميل أو المورد غير موجود أو مرتبط بحساب مسبقًا","danger")
                return redirect(url_for("chart_of_accounts"))
        parent_id = request.form.get("parent_id") or None
        parent_level = 0
        if parent_id:
            parent = row("SELECT level,account_type FROM chart_of_accounts WHERE id=:id", {"id": parent_id})
            parent_level = parent["level"] if parent else 0

        account_type = linked_account_type if linked_record else (parent["account_type"] if parent_id and parent else request.form["account_type"])
        if linked_record and parent_id and parent and parent["account_type"] != account_type:
            flash("نوع الحساب الأب لا يتوافق مع نوع الجهة المختارة","danger")
            return redirect(url_for("chart_of_accounts"))
        normal_balance = "مدين" if account_type in ("أصل", "مصروف") else "دائن"
        statement_type = "قائمة الدخل" if account_type in ("إيراد", "مصروف") else "الميزانية العمومية"
        account_name_ar=linked_record["name"] if linked_record else request.form["account_name_ar"].strip()
        account_name_en=(linked_record["name_en"] or "").strip() if linked_record else request.form.get("account_name_en","").strip()
        if not account_name_en:
            account_name_en=transliterate_arabic_name(account_name_ar)

        execute("""INSERT INTO chart_of_accounts(
            account_code,account_name_ar,account_name_en,account_type,
            parent_id,level,accepts_entries,normal_balance,statement_type,active)
            VALUES(:code,:ar,:en,:type,:parent,:level,:accepts,:normal_balance,:statement_type,:active)""",
            {"code":request.form["account_code"].strip(),
             "ar":account_name_ar,
             "en":account_name_en,
             "type":account_type,
             "parent":parent_id,
             "level":parent_level+1,
             "accepts":1 if request.form.get("accepts_entries")=="1" else 0,
             "normal_balance":normal_balance,
             "statement_type":statement_type,
             "active":1 if request.form.get("active")=="1" else 0})
        if linked_record:
            created_account=row("SELECT id FROM chart_of_accounts WHERE account_code=:code",{"code":request.form["account_code"].strip()})
            if linked_type=="customer":
                execute("UPDATE customers SET receivable_account_id=:account WHERE id=:id",{"account":created_account["id"],"id":linked_id})
            else:
                execute("UPDATE suppliers SET payable_account_id=:account WHERE id=:id",{"account":created_account["id"],"id":linked_id})
        audit("CREATE", "ACCOUNT", f"إضافة حساب {request.form['account_code']}")
        flash("تمت إضافة الحساب", "success")
        return redirect(url_for("chart_of_accounts"))

    q = request.args.get("q", "").strip()
    account_type = request.args.get("account_type", "").strip()
    active = request.args.get("active", "").strip()

    conditions = []
    params = {}
    if q:
        conditions.append("(a.account_code ILIKE :q OR a.account_name_ar ILIKE :q OR a.account_name_en ILIKE :q)")
        params["q"] = f"%{q}%"
    if account_type:
        conditions.append("a.account_type=:account_type")
        params["account_type"] = account_type
    if active in ("0","1"):
        conditions.append("a.active=:active")
        params["active"] = int(active)

    where = " WHERE " + " AND ".join(conditions) if conditions else ""

    accounts = rows(f"""
        SELECT a.*, p.account_name_ar parent_name,
               COALESCE(ch.children_count,0) children_count,
               COALESCE(mv.entries_count,0) entries_count,
               COALESCE(mv.total_debit,0) total_debit,
               COALESCE(mv.total_credit,0) total_credit,
               COALESCE(mv.total_debit,0)-COALESCE(mv.total_credit,0) balance
        FROM chart_of_accounts a
        LEFT JOIN chart_of_accounts p ON p.id=a.parent_id
        LEFT JOIN (
            SELECT parent_id, COUNT(*) children_count
            FROM chart_of_accounts
            WHERE parent_id IS NOT NULL
            GROUP BY parent_id
        ) ch ON ch.parent_id=a.id
        LEFT JOIN (
            SELECT account_id, COUNT(*) entries_count,
                   SUM(debit) total_debit, SUM(credit) total_credit
            FROM journal_entry_lines
            GROUP BY account_id
        ) mv ON mv.account_id=a.id
        {where}
        ORDER BY a.account_code
    """, params)

    parents = rows("""
        SELECT id,account_code,account_name_ar,level,account_type
        FROM chart_of_accounts
        ORDER BY account_code
    """)

    return render_template(
        "chart_of_accounts.html",
        accounts=accounts,
        parents=parents,
        q=q,
        selected_type=account_type,
        selected_active=active,
        customers_for_accounts=rows("SELECT id,name,name_en,receivable_account_id FROM customers ORDER BY name"),
        suppliers_for_accounts=rows("SELECT id,name,name_en,payable_account_id FROM suppliers ORDER BY name")
    )



@app.route("/chart-of-accounts/next-code")
@login_required
def chart_next_code():
    parent_id = request.args.get("parent_id", type=int)
    if not parent_id:
        return {"code": "", "message": "اختر الحساب الرئيسي أولاً"}

    parent = row("""
        SELECT id, account_code, level
        FROM chart_of_accounts
        WHERE id=:id
    """, {"id": parent_id})

    if not parent:
        return {"code": "", "message": "الحساب الرئيسي غير موجود"}, 404

    children = rows("""
        SELECT account_code
        FROM chart_of_accounts
        WHERE parent_id=:id
        ORDER BY account_code
    """, {"id": parent_id})

    parent_code = str(parent["account_code"]).strip()
    numeric_codes = []

    for child in children:
        code = str(child["account_code"]).strip()
        if code.isdigit():
            numeric_codes.append(int(code))

    if parent_code.isdigit():
        parent_number = int(parent_code)
        parent_level = int(parent["level"] or 1)

        # 1000 -> 1100, 1200 ...
        # 1100 -> 1110, 1120 ...
        # 1110 -> 1111, 1112 ...
        step = 10 ** max(0, 3 - parent_level)

        if numeric_codes:
            candidate = max(numeric_codes) + step
        else:
            candidate = parent_number + step

        width = max(len(parent_code), len(str(candidate)))
        suggested_code = str(candidate).zfill(width)
    else:
        # Alphanumeric fallback: ABC -> ABC-01, ABC-02 ...
        sequence = 1
        existing = {str(c["account_code"]).strip() for c in children}
        while f"{parent_code}-{sequence:02d}" in existing:
            sequence += 1
        suggested_code = f"{parent_code}-{sequence:02d}"

    return {
        "code": suggested_code,
        "parent_code": parent_code,
        "message": "تم اقتراح الكود تلقائيًا"
    }

@app.route("/chart-of-accounts/fill-english-names",methods=["POST"])
@login_required
def chart_fill_english_names():
    missing=rows("""SELECT id,account_name_ar FROM chart_of_accounts
      WHERE account_name_en IS NULL OR TRIM(account_name_en)='' ORDER BY id""")
    updated=0
    try:
        for account in missing:
            english_name=transliterate_arabic_name(account["account_name_ar"])
            if english_name:
                db.session.execute(text("UPDATE chart_of_accounts SET account_name_en=:name WHERE id=:id"),
                  {"name":english_name,"id":account["id"]})
                updated+=1
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        flash(f"تعذر استكمال الأسماء الإنجليزية: {exc}","danger")
        return redirect(url_for("chart_of_accounts"))
    audit("UPDATE","ACCOUNT",f"توليد الأسماء الإنجليزية لـ {updated} حساب")
    flash(f"تم توليد الاسم الإنجليزي لـ {updated} حساب، دون تعديل الأسماء الموجودة","success")
    return redirect(url_for("chart_of_accounts"))


@app.route("/chart-of-accounts/<int:account_id>")
@login_required
def account_view(account_id):
    account = row("""
        SELECT a.*,p.account_name_ar parent_name,
               COALESCE(mv.entries_count,0) entries_count,
               COALESCE(mv.total_debit,0) total_debit,
               COALESCE(mv.total_credit,0) total_credit,
               COALESCE(mv.total_debit,0)-COALESCE(mv.total_credit,0) balance
        FROM chart_of_accounts a
        LEFT JOIN chart_of_accounts p ON p.id=a.parent_id
        LEFT JOIN (
            SELECT account_id,COUNT(*) entries_count,
                   SUM(debit) total_debit,SUM(credit) total_credit
            FROM journal_entry_lines GROUP BY account_id
        ) mv ON mv.account_id=a.id
        WHERE a.id=:id
    """, {"id": account_id})
    if not account:
        return "الحساب غير موجود", 404

    children = rows("""
        SELECT id,account_code,account_name_ar,accepts_entries,active
        FROM chart_of_accounts WHERE parent_id=:id ORDER BY account_code
    """, {"id": account_id})

    movements = rows("""
        SELECT j.journal_no,j.journal_date,j.status,l.debit,l.credit,
               l.line_description,l.invoice_number,l.invoice_date
        FROM journal_entry_lines l
        JOIN journal_entries j ON j.id=l.journal_id
        WHERE l.account_id=:id
        ORDER BY j.journal_date DESC,j.id DESC,l.id DESC
        LIMIT 200
    """, {"id": account_id})

    company = row("SELECT * FROM settings WHERE id=1")
    return render_template("account_view.html", account=account, children=children,
                           movements=movements, company=company)

@app.route("/chart-of-accounts/<int:account_id>/edit", methods=["GET","POST"])
@login_required
def account_edit(account_id):
    account = row("SELECT * FROM chart_of_accounts WHERE id=:id", {"id": account_id})
    if not account:
        return "الحساب غير موجود", 404

    if request.method == "POST":
        parent_id = request.form.get("parent_id") or None
        if parent_id and int(parent_id) == account_id:
            flash("لا يمكن جعل الحساب أبًا لنفسه", "danger")
            return redirect(url_for("account_edit", account_id=account_id))

        parent_level = 0
        if parent_id:
            parent = row("SELECT level FROM chart_of_accounts WHERE id=:id", {"id": parent_id})
            parent_level = parent["level"] if parent else 0

        account_name_ar=request.form["account_name_ar"].strip()
        account_name_en=request.form.get("account_name_en","").strip() or transliterate_arabic_name(account_name_ar)
        execute("""UPDATE chart_of_accounts SET
            account_code=:code,account_name_ar=:ar,account_name_en=:en,
            account_type=:type,parent_id=:parent,level=:level,
            accepts_entries=:accepts,normal_balance=:normal_balance,
            statement_type=:statement_type,active=:active
            WHERE id=:id""",
            {"code":request.form["account_code"].strip(),
             "ar":account_name_ar,
             "en":account_name_en,
             "type":request.form["account_type"],
             "parent":parent_id,
             "level":parent_level+1,
             "accepts":1 if request.form.get("accepts_entries")=="1" else 0,
             "normal_balance":request.form["normal_balance"],
             "statement_type":request.form["statement_type"],
             "active":1 if request.form.get("active")=="1" else 0,
             "id":account_id})
        audit("UPDATE", "ACCOUNT", f"تعديل الحساب {request.form['account_code']}")
        flash("تم تعديل الحساب", "success")
        return redirect(url_for("account_view", account_id=account_id))

    parents = rows("""
        SELECT id,account_code,account_name_ar,level
        FROM chart_of_accounts WHERE id<>:id ORDER BY account_code
    """, {"id": account_id})
    return render_template("account_edit.html", account=account, parents=parents)

@app.route("/chart-of-accounts/<int:account_id>/delete", methods=["POST"])
@login_required
def account_delete(account_id):
    account = row("SELECT * FROM chart_of_accounts WHERE id=:id", {"id": account_id})
    if not account:
        return "الحساب غير موجود", 404

    children_count = row("SELECT COUNT(*) count FROM chart_of_accounts WHERE parent_id=:id", {"id": account_id})["count"]
    entries_count = row("SELECT COUNT(*) count FROM journal_entry_lines WHERE account_id=:id", {"id": account_id})["count"]

    if children_count > 0:
        flash("لا يمكن حذف الحساب لأنه يحتوي على حسابات فرعية", "danger")
    elif entries_count > 0:
        flash("لا يمكن حذف الحساب لأنه مستخدم في قيود يومية. يمكنك تعطيله بدلًا من حذفه.", "danger")
    else:
        execute("DELETE FROM chart_of_accounts WHERE id=:id", {"id": account_id})
        audit("DELETE", "ACCOUNT", f"حذف الحساب {account['account_code']}")
        flash("تم حذف الحساب", "success")
    return redirect(url_for("chart_of_accounts"))

@app.route("/chart-of-accounts/export.xlsx")
@login_required
def chart_export():
    data = rows("""
        SELECT a.account_code,a.account_name_ar,a.account_name_en,a.account_type,
               p.account_code parent_code,p.account_name_ar parent_name,
               a.level,a.accepts_entries,a.normal_balance,a.statement_type,a.active,
               COALESCE(mv.entries_count,0) entries_count,
               COALESCE(mv.total_debit,0) total_debit,
               COALESCE(mv.total_credit,0) total_credit,
               COALESCE(mv.total_debit,0)-COALESCE(mv.total_credit,0) balance
        FROM chart_of_accounts a
        LEFT JOIN chart_of_accounts p ON p.id=a.parent_id
        LEFT JOIN (
            SELECT account_id,COUNT(*) entries_count,
                   SUM(debit) total_debit,SUM(credit) total_credit
            FROM journal_entry_lines GROUP BY account_id
        ) mv ON mv.account_id=a.id
        ORDER BY a.account_code
    """)
    headers = ["الكود","اسم الحساب","الاسم الإنجليزي","النوع","كود الحساب الأب",
               "الحساب الأب","المستوى","يقبل حركة","طبيعة الرصيد","القائمة المالية",
               "الحالة","عدد القيود","إجمالي المدين","إجمالي الدائن","الرصيد"]
    keys = ["account_code","account_name_ar","account_name_en","account_type",
            "parent_code","parent_name","level","accepts_entries","normal_balance",
            "statement_type","active","entries_count","total_debit","total_credit","balance"]
    records = [[r.get(k) for k in keys] for r in data]
    return xlsx_response("chart_of_accounts.xlsx","دليل الحسابات",headers,records)



@app.route("/cost-centers", methods=["GET","POST"])
@login_required
def cost_centers():
    if request.method=="POST":
        execute("INSERT INTO cost_centers(code,name,active) VALUES(:c,:n,1)",
                {"c":request.form["code"].strip(),"n":request.form["name"].strip()})
        flash("تمت إضافة مركز التكلفة","success")
        return redirect(url_for("cost_centers"))
    return render_template("cost_centers.html",
        centers=rows("SELECT * FROM cost_centers ORDER BY code"))

@app.route("/journal-entries", methods=["GET","POST"])
@login_required
def journal_entries():
    if request.method=="POST":
        account_ids=request.form.getlist("account_id[]")
        debits=request.form.getlist("debit[]")
        credits=request.form.getlist("credit[]")
        taxable_values=request.form.getlist("taxable[]")
        tax_directions=request.form.getlist("tax_direction[]")
        supplier_ids=request.form.getlist("supplier_id[]")
        customer_ids=request.form.getlist("customer_id[]")
        party_types=request.form.getlist("party_type[]")
        entity_ids=request.form.getlist("entity_id[]")
        tax_numbers=request.form.getlist("tax_number[]")
        invoice_numbers=request.form.getlist("invoice_number[]")
        invoice_dates=request.form.getlist("invoice_date[]")
        party_account_ids=request.form.getlist("party_account_id[]")
        descriptions=request.form.getlist("line_description[]")
        cost_center_ids=request.form.getlist("cost_center_id[]")
        lines=[]; total_debit=0.0; total_credit=0.0
        for i,account_id in enumerate(account_ids):
            if not account_id: continue
            debit=float(debits[i] or 0); credit=float(credits[i] or 0)
            if debit>0 and credit>0:
                flash(f"السطر {i+1}: لا يمكن إدخال مدين ودائن معًا","danger")
                return redirect(url_for("journal_entries"))
            if debit<=0 and credit<=0:
                flash(f"السطر {i+1}: أدخل مبلغ مدين أو دائن","danger")
                return redirect(url_for("journal_entries"))
            account=row("SELECT accepts_entries FROM chart_of_accounts WHERE id=:id",{"id":account_id})
            if not account or account["accepts_entries"]!=1:
                flash(f"السطر {i+1}: الحساب لا يقبل حركة","danger")
                return redirect(url_for("journal_entries"))
            taxable=1 if i<len(taxable_values) and taxable_values[i]=="1" else 0
            direction=tax_directions[i] if i<len(tax_directions) else "غير مطبق"
            party_type = party_types[i] if i < len(party_types) else ""
            supplier_id = (supplier_ids[i] or None) if i < len(supplier_ids) else None
            customer_id = (customer_ids[i] or None) if i < len(customer_ids) else None
            tax_number = tax_numbers[i].strip() if i < len(tax_numbers) else ""
            invoice_number = invoice_numbers[i].strip() if i < len(invoice_numbers) else ""
            invoice_date = invoice_dates[i] if i < len(invoice_dates) and invoice_dates[i] else None
            party_account_id = (party_account_ids[i] or None) if i < len(party_account_ids) else None
            entity_id = (entity_ids[i] or None) if i < len(entity_ids) else None

            if party_type not in JOURNAL_ACCOUNT_TYPES:
                flash(f"السطر {i+1}: نوع الحساب غير صحيح","danger")
                return redirect(url_for("journal_entries"))

            if party_type == "مورد":
                customer_id = None
                party = row("SELECT vat_number,payable_account_id FROM suppliers WHERE id=:id", {"id": supplier_id}) if supplier_id else None
                tax_number = (party["vat_number"] or "") if party else ""
                party_account_id = party.get("payable_account_id") if party else party_account_id
                entity_id = supplier_id
            elif party_type == "عميل":
                supplier_id = None
                party = row("SELECT vat_number,receivable_account_id FROM customers WHERE id=:id", {"id": customer_id}) if customer_id else None
                tax_number = (party["vat_number"] or "") if party else ""
                party_account_id = party.get("receivable_account_id") if party else party_account_id
                entity_id = customer_id
            elif party_type in ("موظف","مندوب مبيعات"):
                supplier_id = None; customer_id = None; tax_number = ""
                employee=row("SELECT id FROM employees WHERE id=:id AND active=1",{"id":entity_id}) if entity_id else None
                if not employee:
                    flash(f"السطر {i+1}: اختر {party_type}","danger")
                    return redirect(url_for("journal_entries"))
            else:
                supplier_id = None
                customer_id = None
                tax_number = ""
                entity_id = None

            if taxable:
                if direction == "غير مطبق":
                    flash(f"السطر {i+1}: حدد مدخلات أو مخرجات","danger")
                    return redirect(url_for("journal_entries"))
                if party_type not in ("مورد", "عميل"):
                    flash(f"السطر {i+1}: حدد نوع الطرف مورد أو عميل","danger")
                    return redirect(url_for("journal_entries"))
                if party_type == "مورد" and not supplier_id:
                    flash(f"السطر {i+1}: اختر المورد","danger")
                    return redirect(url_for("journal_entries"))
                if party_type == "عميل" and not customer_id:
                    flash(f"السطر {i+1}: اختر العميل","danger")
                    return redirect(url_for("journal_entries"))
                if not tax_number:
                    flash(f"السطر {i+1}: الرقم الضريبي غير موجود في بيانات الطرف","danger")
                    return redirect(url_for("journal_entries"))
                if not invoice_number:
                    flash(f"السطر {i+1}: رقم الفاتورة مطلوب للعملية الخاضعة للضريبة","danger")
                    return redirect(url_for("journal_entries"))
                if not invoice_date:
                    flash(f"السطر {i+1}: تاريخ الفاتورة مطلوب للعملية الخاضعة للضريبة","danger")
                    return redirect(url_for("journal_entries"))

            lines.append({"account_id":account_id,"debit":round(debit,2),"credit":round(credit,2),
                          "taxable":taxable,"tax_direction":direction,
                          "supplier_id":supplier_id,
                          "customer_id":customer_id,
                          "party_type":party_type,
                          "tax_number":tax_number,
                          "invoice_number":invoice_number,
                          "invoice_date":invoice_date,
                          "party_account_id":party_account_id,
                          "entity_id":entity_id,
                          "line_description":descriptions[i] if i<len(descriptions) else "",
                          "cost_center_id":(cost_center_ids[i] or None) if i<len(cost_center_ids) else None})
            total_debit+=debit; total_credit+=credit
        if len(lines)<2:
            flash("يجب أن يحتوي القيد على سطرين على الأقل","danger")
            return redirect(url_for("journal_entries"))
        total_debit=round(total_debit,2); total_credit=round(total_credit,2)
        if total_debit!=total_credit:
            flash(f"القيد غير متوازن. الفرق {abs(total_debit-total_credit):.2f}","danger")
            return redirect(url_for("journal_entries"))
        try:
            ensure_open_period(request.form["journal_date"])
        except Exception as exc:
            flash(str(exc),"danger")
            return redirect(url_for("journal_entries"))
        journal_status=request.form.get("status","مسودة")
        if journal_status not in ("مسودة","مرحّل"):
            flash("حالة القيد غير صحيحة","danger")
            return redirect(url_for("journal_entries"))
        posted_at=datetime.now() if journal_status=="مرحّل" else None
        journal_no=next_journal_number(request.form["journal_date"])
        execute("""INSERT INTO journal_entries(
            journal_no,journal_date,reference,description,status,total_debit,total_credit,created_by,created_at,posted_at)
            VALUES(:no,:date,:ref,:desc,:status,:debit,:credit,:user,:created,:posted_at)""",
            {"no":journal_no,"date":request.form["journal_date"],
             "ref":request.form.get("reference",""),"desc":request.form.get("description",""),
             "status":journal_status,"debit":total_debit,"credit":total_credit,
             "user":session.get("user_id"),"created":datetime.now(),"posted_at":posted_at})
        journal_id=row("SELECT id FROM journal_entries WHERE journal_no=:no",{"no":journal_no})["id"]
        for line in lines:
            execute("""INSERT INTO journal_entry_lines(
                journal_id,account_id,debit,credit,taxable,tax_direction,supplier_id,customer_id,
                party_type,tax_number,invoice_number,invoice_date,party_account_id,entity_id,line_description,cost_center_id)
                VALUES(:journal_id,:account_id,:debit,:credit,:taxable,:tax_direction,
                :supplier_id,:customer_id,:party_type,:tax_number,:invoice_number,:invoice_date,
                :party_account_id,:entity_id,:line_description,:cost_center_id)""",
                {"journal_id":journal_id,**line})
        audit("POST" if journal_status=="مرحّل" else "CREATE","JOURNAL",f"{'ترحيل' if journal_status=='مرحّل' else 'حفظ'} القيد {journal_no}")
        flash(f"تم {'حفظ وترحيل' if journal_status=='مرحّل' else 'حفظ'} القيد {journal_no}","success")
        if request.form.get("print_after_save") == "1":
            return redirect(url_for("journal_view", journal_id=journal_id, print=1))
        return redirect(url_for("journal_entries"))
    q = request.args.get("q", "").strip()
    journal_params = {}
    journal_where = ""
    if q:
        journal_where = " WHERE journal_no ILIKE :q OR description ILIKE :q OR reference ILIKE :q "
        journal_params["q"] = f"%{q}%"
    return render_template("journal_entries.html",
        journals=rows(f"SELECT * FROM journal_entries {journal_where} ORDER BY journal_date DESC,id DESC", journal_params),
        q=q,
        accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                         WHERE active=1 AND accepts_entries=1 ORDER BY account_code"""),
        suppliers=rows("SELECT id,name,vat_number FROM suppliers ORDER BY name"),
        customers=rows("SELECT id,name,vat_number FROM customers ORDER BY name"),
        employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"),
        centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"))

@app.route("/journal-entries/<int:journal_id>/post",methods=["POST"])
@login_required
def journal_post(journal_id):
    journal=row("SELECT * FROM journal_entries WHERE id=:id",{"id":journal_id})
    if not journal:
        return "القيد غير موجود",404
    if journal["status"]=="مرحّل":
        flash("القيد مرحّل مسبقًا","info")
        return redirect(url_for("journal_entries"))
    totals=row("""SELECT COALESCE(SUM(debit),0) debit,COALESCE(SUM(credit),0) credit,
      COUNT(*) line_count FROM journal_entry_lines WHERE journal_id=:id""",{"id":journal_id})
    debit=round(float(totals["debit"] or 0),2);credit=round(float(totals["credit"] or 0),2)
    if totals["line_count"]<2 or debit<=0 or debit!=credit:
        flash("لا يمكن ترحيل القيد: يجب أن يحتوي على سطرين متوازنين على الأقل","danger")
        return redirect(url_for("journal_entries"))
    try:
        ensure_open_period(journal["journal_date"])
    except Exception as exc:
        flash(str(exc),"danger")
        return redirect(url_for("journal_entries"))
    execute("""UPDATE journal_entries SET status='مرحّل',posted_at=:posted,
      total_debit=:debit,total_credit=:credit WHERE id=:id AND status<>'مرحّل'""",
      {"posted":datetime.now(),"debit":debit,"credit":credit,"id":journal_id})
    audit("POST","JOURNAL",f"ترحيل القيد {journal['journal_no']}")
    flash(f"تم ترحيل القيد {journal['journal_no']} وظهر في الأستاذ والتقارير","success")
    return redirect(url_for("journal_entries"))



# ==================== Complete CRUD / View / Print / Export ====================

MASTER_MODULES = {
    "customers": {
        "table": "customers", "title": "العملاء", "entity": "CUSTOMER",
        "fields": ["name", "vat_number", "phone", "email"],
        "labels": ["الاسم", "الرقم الضريبي", "الهاتف", "البريد الإلكتروني"],
    },
    "suppliers": {
        "table": "suppliers", "title": "الموردون", "entity": "SUPPLIER",
        "fields": ["name", "vat_number", "phone", "email"],
        "labels": ["الاسم", "الرقم الضريبي", "الهاتف", "البريد الإلكتروني"],
    },
    "inventory": {
        "table": "inventory", "title": "المخزون", "entity": "INVENTORY",
        "fields": ["sku", "name", "quantity", "unit", "cost", "sale_price", "reorder_level"],
        "labels": ["الكود", "اسم الصنف", "الكمية", "الوحدة", "التكلفة", "سعر البيع", "حد إعادة الطلب"],
    },
    "employees": {
        "table": "employees", "title": "الموظفون", "entity": "EMPLOYEE",
        "fields": ["employee_no", "name", "job_title", "basic_salary", "allowances", "active"],
        "labels": ["رقم الموظف", "الاسم", "المسمى الوظيفي", "الراتب الأساسي", "البدلات", "الحالة"],
    },
    "branches": {
        "table": "branches", "title": "الفروع", "entity": "BRANCH",
        "fields": ["name", "city", "active"],
        "labels": ["اسم الفرع", "المدينة", "الحالة"],
    },
    "cost-centers": {
        "table": "cost_centers", "title": "مراكز التكلفة", "entity": "COST_CENTER",
        "fields": ["code", "name", "active"],
        "labels": ["الكود", "اسم مركز التكلفة", "الحالة"],
    },
}

@app.route("/manage/<module_name>")
@login_required
def manage_module(module_name):
    cfg = MASTER_MODULES.get(module_name)
    if not cfg:
        return "الوحدة غير موجودة", 404
    q = request.args.get("q", "").strip()
    where = ""
    params = {}
    if q:
        searchable = [f"CAST({f} AS TEXT) ILIKE :q" for f in cfg["fields"]]
        where = " WHERE " + " OR ".join(searchable)
        params["q"] = f"%{q}%"
    data = rows(f"SELECT * FROM {cfg['table']}{where} ORDER BY id DESC", params)
    return render_template("module_manage.html", module_name=module_name, cfg=cfg, data=data, q=q)

@app.route("/manage/<module_name>/<int:record_id>")
@login_required
def module_view(module_name, record_id):
    cfg = MASTER_MODULES.get(module_name)
    if not cfg:
        return "الوحدة غير موجودة", 404
    record = row(f"SELECT * FROM {cfg['table']} WHERE id=:id", {"id": record_id})
    if not record:
        return "السجل غير موجود", 404
    return render_template("module_view.html", module_name=module_name, cfg=cfg, record=record)

@app.route("/manage/<module_name>/<int:record_id>/edit", methods=["GET", "POST"])
@login_required
def module_edit(module_name, record_id):
    cfg = MASTER_MODULES.get(module_name)
    if not cfg:
        return "الوحدة غير موجودة", 404
    record = row(f"SELECT * FROM {cfg['table']} WHERE id=:id", {"id": record_id})
    if not record:
        return "السجل غير موجود", 404
    if request.method == "POST":
        assignments = ", ".join([f"{f}=:{f}" for f in cfg["fields"]])
        params = {"id": record_id}
        for field in cfg["fields"]:
            value = request.form.get(field, "")
            if field in ("quantity", "cost", "sale_price", "reorder_level", "basic_salary", "allowances"):
                value = float(value or 0)
            elif field == "active":
                value = 1 if value == "1" else 0
            params[field] = value
        execute(f"UPDATE {cfg['table']} SET {assignments} WHERE id=:id", params)
        audit("UPDATE", cfg["entity"], f"تعديل السجل رقم {record_id}")
        flash("تم حفظ التعديل", "success")
        return redirect(url_for("module_view", module_name=module_name, record_id=record_id))
    return render_template("module_edit.html", module_name=module_name, cfg=cfg, record=record)

@app.route("/manage/<module_name>/<int:record_id>/delete", methods=["POST"])
@login_required
def module_delete(module_name, record_id):
    cfg = MASTER_MODULES.get(module_name)
    if not cfg:
        return "الوحدة غير موجودة", 404
    try:
        execute(f"DELETE FROM {cfg['table']} WHERE id=:id", {"id": record_id})
        audit("DELETE", cfg["entity"], f"حذف السجل رقم {record_id}")
        flash("تم حذف السجل", "success")
    except Exception:
        db.session.rollback()
        flash("تعذر الحذف لأن السجل مرتبط بعمليات أخرى. يمكن تعطيله بدلًا من حذفه.", "danger")
    return redirect(url_for("manage_module", module_name=module_name))

@app.route("/manage/<module_name>/export.xlsx")
@login_required
def module_export(module_name):
    cfg = MASTER_MODULES.get(module_name)
    if not cfg:
        return "الوحدة غير موجودة", 404
    data = rows(f"SELECT * FROM {cfg['table']} ORDER BY id")
    records = [[r.get(f) for f in cfg["fields"]] for r in data]
    return xlsx_response(f"{module_name}.xlsx", cfg["title"], cfg["labels"], records)

@app.route("/journal-entries/<int:journal_id>")
@login_required
def journal_view(journal_id):
    journal = row("SELECT * FROM journal_entries WHERE id=:id", {"id": journal_id})
    if not journal:
        return "القيد غير موجود", 404
    lines = rows("""
        SELECT l.*, a.account_code, a.account_name_ar,
               s.name supplier_name, cu.name customer_name,
               c.code cost_center_code, c.name cost_center_name
        FROM journal_entry_lines l
        JOIN chart_of_accounts a ON a.id=l.account_id
        LEFT JOIN suppliers s ON s.id=l.supplier_id
        LEFT JOIN customers cu ON cu.id=l.customer_id
        LEFT JOIN cost_centers c ON c.id=l.cost_center_id
        WHERE l.journal_id=:id ORDER BY l.id
    """, {"id": journal_id})
    company = row("SELECT * FROM settings WHERE id=1")
    return render_template("journal_view.html", journal=journal, lines=lines, company=company)

@app.route("/journal-entries/<int:journal_id>/delete", methods=["POST"])
@login_required
def journal_delete(journal_id):
    journal = row("SELECT * FROM journal_entries WHERE id=:id", {"id": journal_id})
    if not journal:
        return "القيد غير موجود", 404
    if journal["status"] == "مرحّل":
        flash("لا يمكن حذف قيد مرحّل. يجب إنشاء قيد عكسي.", "danger")
    else:
        execute("DELETE FROM journal_entries WHERE id=:id", {"id": journal_id})
        audit("DELETE", "JOURNAL", f"حذف القيد {journal['journal_no']}")
        flash("تم حذف القيد المسودة", "success")
    return redirect(url_for("journal_entries"))

@app.route("/journal-entries/export.xlsx")
@login_required
def journal_export():
    q = request.args.get("q", "").strip()
    params = {}
    where = ""
    if q:
        where = """ WHERE j.journal_no ILIKE :q OR j.description ILIKE :q
                    OR a.account_code ILIKE :q OR a.account_name_ar ILIKE :q
                    OR COALESCE(l.invoice_number,'') ILIKE :q """
        params["q"] = f"%{q}%"
    data = rows(f"""
        SELECT j.journal_no,j.journal_date,j.status,j.reference,j.description,
               a.account_code,a.account_name_ar,l.debit,l.credit,l.taxable,
               l.tax_direction,l.party_type,s.name supplier_name,cu.name customer_name,
               l.tax_number,l.invoice_number,l.invoice_date,
               l.line_description,c.code cost_center_code,c.name cost_center_name
        FROM journal_entries j
        JOIN journal_entry_lines l ON l.journal_id=j.id
        JOIN chart_of_accounts a ON a.id=l.account_id
        LEFT JOIN suppliers s ON s.id=l.supplier_id
        LEFT JOIN customers cu ON cu.id=l.customer_id
        LEFT JOIN cost_centers c ON c.id=l.cost_center_id
        {where}
        ORDER BY j.journal_date DESC,j.id DESC,l.id
    """, params)
    headers = ["رقم القيد","تاريخ القيد","الحالة","المرجع","البيان العام","الكود","اسم الحساب",
               "مدين","دائن","خاضع للضريبة","مدخلات/مخرجات","نوع الطرف","المورد","العميل","الرقم الضريبي","رقم الفاتورة",
               "تاريخ الفاتورة","البيان","كود مركز التكلفة","مركز التكلفة"]
    records = [[r.get(k) for k in [
        "journal_no","journal_date","status","reference","description","account_code",
        "account_name_ar","debit","credit","taxable","tax_direction","party_type","supplier_name",
        "customer_name","tax_number","invoice_number","invoice_date","line_description",
        "cost_center_code","cost_center_name"
    ]] for r in data]
    return xlsx_response("journal_entries.xlsx", "قيود اليومية", headers, records)


@app.route("/invoices/<int:invoice_id>/edit",methods=["GET","POST"])
@login_required
def invoice_edit(invoice_id):
    invoice=row("""SELECT i.*,c.name customer_name FROM invoices i
                   JOIN customers c ON c.id=i.customer_id WHERE i.id=:id""",{"id":invoice_id})
    if not invoice:
        return "الفاتورة غير موجودة",404
    if not invoice_is_editable(invoice):
        flash("لا يمكن تعديل الفاتورة بعد الاعتماد أو الترحيل. استخدم المرتجع أو الإشعار الدائن للتصحيح.","danger")
        return redirect(url_for("invoice_view",invoice_id=invoice_id))

    if request.method=="POST":
        try:
            prepared,gross,discount,net_subtotal,vat,total=prepare_invoice_items_from_form(request.form)
            execute("""UPDATE invoices SET customer_id=:customer,invoice_date=:invoice_date,
              due_date=:due_date,subtotal=:subtotal,discount=:discount,vat=:vat,total=:total,
              branch_id=:branch,cost_center_id=:cost_center,project_id=:project,
              payment_terms=:payment_terms,payment_method=:payment_method,
              sales_person=:sales_person,customer_reference=:customer_reference,
              notes=:notes,updated_at=:updated WHERE id=:id""",
              {"customer":request.form["customer_id"],
               "invoice_date":request.form["invoice_date"],
               "due_date":request.form.get("due_date") or None,
               "subtotal":net_subtotal,"discount":discount,"vat":vat,"total":total,
               "branch":request.form.get("branch_id") or None,
               "cost_center":request.form.get("cost_center_id") or None,
               "project":request.form.get("project_id") or None,
               "payment_terms":request.form.get("payment_terms",""),
               "payment_method":request.form.get("payment_method",""),
               "sales_person":request.form.get("sales_person",""),
               "customer_reference":request.form.get("customer_reference",""),
               "notes":request.form.get("notes",""),"updated":datetime.now(),"id":invoice_id})
            replace_invoice_items(invoice_id,prepared)
            audit("UPDATE","INVOICE",f"تعديل مسودة الفاتورة {invoice['invoice_no']}")
            flash("تم حفظ تعديلات مسودة الفاتورة.","success")
            return redirect(url_for("invoice_view",invoice_id=invoice_id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc),"danger")

    return render_template("invoice_edit.html",invoice=invoice,
      items=rows("SELECT * FROM invoice_items WHERE invoice_id=:id ORDER BY id",{"id":invoice_id}),
      customers=rows("SELECT * FROM customers ORDER BY name"),
      branches=rows("SELECT * FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"),
      projects=rows("SELECT id,project_no,name FROM projects ORDER BY project_no"),
      inventory_items=rows("SELECT * FROM inventory ORDER BY name"))


@app.route("/invoices/<int:invoice_id>/submit-approval",methods=["POST"])
@login_required
def invoice_submit_approval(invoice_id):
    invoice=row("SELECT * FROM invoices WHERE id=:id",{"id":invoice_id})
    if not invoice:
        return "الفاتورة غير موجودة",404
    if not invoice_is_editable(invoice):
        flash("هذه الفاتورة ليست مسودة قابلة للإرسال.","danger")
    else:
        execute("""UPDATE invoices SET status='بانتظار الاعتماد',updated_at=:dt
                   WHERE id=:id""",{"dt":datetime.now(),"id":invoice_id})
        audit("SUBMIT_APPROVAL","INVOICE",f"إرسال الفاتورة {invoice['invoice_no']} للاعتماد")
        flash("تم إرسال الفاتورة للاعتماد، وتم قفل تعديلها مؤقتًا.","success")
    return redirect(url_for("invoice_view",invoice_id=invoice_id))


@app.route("/invoices/<int:invoice_id>/return-to-draft",methods=["POST"])
@login_required
def invoice_return_to_draft(invoice_id):
    invoice=row("SELECT * FROM invoices WHERE id=:id",{"id":invoice_id})
    if not invoice:
        return "الفاتورة غير موجودة",404
    if invoice.get("journal_id") or invoice["status"]=="معتمدة":
        flash("لا يمكن إعادة فاتورة مرحلة إلى المسودة.","danger")
    else:
        execute("""UPDATE invoices SET status='معاد للتعديل',updated_at=:dt
                   WHERE id=:id""",{"dt":datetime.now(),"id":invoice_id})
        audit("RETURN_DRAFT","INVOICE",f"إعادة الفاتورة {invoice['invoice_no']} للتعديل")
        flash("تمت إعادة الفاتورة للتعديل.","success")
    return redirect(url_for("invoice_view",invoice_id=invoice_id))


@app.route("/invoices/<int:invoice_id>/approve",methods=["POST"])
@login_required
def invoice_approve(invoice_id):
    invoice=row("SELECT * FROM invoices WHERE id=:id",{"id":invoice_id})
    if not invoice:
        return "الفاتورة غير موجودة",404
    if invoice.get("journal_id") or invoice["status"]=="معتمدة":
        flash("الفاتورة معتمدة ومرحلة بالفعل.","info")
        return redirect(url_for("invoice_view",invoice_id=invoice_id))
    if invoice["status"] not in ("مسودة","معاد للتعديل","بانتظار الاعتماد"):
        flash("حالة الفاتورة لا تسمح بالاعتماد.","danger")
        return redirect(url_for("invoice_view",invoice_id=invoice_id))

    old_number = invoice["invoice_no"]
    official_number = None

    try:
        draft_number,official_number=assign_official_invoice_number(invoice_id)
        post_invoice_to_ledger(invoice_id)

        execute("""UPDATE invoices
                   SET status='معتمدة',
                       approved_by=:user,
                       approved_at=:dt,
                       updated_at=:dt
                   WHERE id=:id""",
                {
                    "user":session.get("user_id"),
                    "dt":datetime.now(),
                    "id":invoice_id,
                })

        audit(
            "APPROVE",
            "INVOICE",
            f"اعتماد الفاتورة وتغيير الرقم من {draft_number} إلى {official_number}"
        )
        flash(
            f"تم اعتماد الفاتورة. تغير الرقم فقط من {draft_number} إلى {official_number}.",
            "success"
        )

    except Exception as exc:
        db.session.rollback()

        # Restore the date/time number if posting did not complete.
        current=row("SELECT journal_id FROM invoices WHERE id=:id",{"id":invoice_id})
        if not current or not current.get("journal_id"):
            try:
                execute("""UPDATE invoices
                           SET invoice_no=COALESCE(draft_number,:old_number),
                               official_invoice_no=NULL,
                               posting_status='خطأ',
                               updated_at=:dt
                           WHERE id=:id""",
                        {
                            "old_number":old_number,
                            "dt":datetime.now(),
                            "id":invoice_id,
                        })
            except Exception:
                db.session.rollback()

        app.logger.exception("Final invoice approval failed")
        flash(f"تعذر اعتماد وترحيل الفاتورة: {exc}","danger")

    return redirect(url_for("invoice_view",invoice_id=invoice_id))



@app.route("/invoices/<int:invoice_id>/copy",methods=["POST"])
@login_required
def invoice_copy(invoice_id):
    source=row("SELECT * FROM invoices WHERE id=:id",{"id":invoice_id})
    if not source:
        return "الفاتورة غير موجودة",404
    new_date=request.form.get("invoice_date") or datetime.now().date()
    new_no=next_invoice_draft_number()
    execute("""INSERT INTO invoices(invoice_no,draft_number,invoice_uuid,customer_id,invoice_date,
      due_date,subtotal,discount,vat,total,status,branch_id,cost_center_id,project_id,
      payment_terms,payment_method,sales_person,customer_reference,notes,created_at,
      updated_at,copied_from_id)
      VALUES(:no,:draft_number,:uuid,:customer,:invoice_date,:due_date,:subtotal,:discount,:vat,
      :total,'مسودة',:branch,:cost_center,:project,:payment_terms,:payment_method,
      :sales_person,:customer_reference,:notes,:created,:updated,:source)""",
      {"no":new_no,"draft_number":new_no,"uuid":str(uuid.uuid4()),"customer":source["customer_id"],
       "invoice_date":new_date,"due_date":source.get("due_date"),
       "subtotal":source["subtotal"],"discount":source.get("discount") or 0,
       "vat":source["vat"],"total":source["total"],"branch":source.get("branch_id"),
       "cost_center":source.get("cost_center_id"),"project":source.get("project_id"),
       "payment_terms":source.get("payment_terms") or "",
       "payment_method":source.get("payment_method") or "",
       "sales_person":source.get("sales_person") or "",
       "customer_reference":source.get("customer_reference") or "",
       "notes":f"نسخة من {source['invoice_no']} - {source.get('notes') or ''}",
       "created":datetime.now(),"updated":datetime.now(),"source":invoice_id})
    new_id=row("SELECT id FROM invoices WHERE invoice_no=:no",{"no":new_no})["id"]
    execute("""INSERT INTO invoice_items(invoice_id,item_name,description,quantity,unit,
      unit_price,discount_rate,vat_rate,line_subtotal,line_discount,line_vat,line_total)
      SELECT :new_id,item_name,description,quantity,unit,unit_price,
      COALESCE(discount_rate,0),vat_rate,line_subtotal,COALESCE(line_discount,0),
      line_vat,line_total FROM invoice_items WHERE invoice_id=:source""",
      {"new_id":new_id,"source":invoice_id})
    audit("COPY","INVOICE",f"نسخ الفاتورة {source['invoice_no']} إلى {new_no}")
    flash(f"تم إنشاء مسودة جديدة رقم {new_no} من الفاتورة الأصلية.","success")
    return redirect(url_for("invoice_edit",invoice_id=new_id))



@app.route("/invoices/<int:invoice_id>/delete", methods=["POST"])
@login_required
def invoice_delete(invoice_id):
    invoice = row("SELECT * FROM invoices WHERE id=:id", {"id": invoice_id})
    if not invoice:
        return "الفاتورة غير موجودة", 404
    if not invoice_is_editable(invoice):
        flash("يمكن حذف المسودة فقط. الفواتير المعتمدة أو المرسلة للاعتماد لا تُحذف.", "danger")
    else:
        execute("DELETE FROM invoices WHERE id=:id", {"id": invoice_id})
        audit("DELETE", "INVOICE", f"حذف الفاتورة {invoice['invoice_no']}")
        flash("تم حذف الفاتورة المسودة", "success")
    return redirect(url_for("invoices"))




@app.route("/multi-journal", methods=["GET", "POST"])
@login_required
def multi_journal():
    if request.method == "POST":
        group_numbers = request.form.getlist("group_no[]")
        account_ids = request.form.getlist("account_id[]")
        debits = request.form.getlist("debit[]")
        credits = request.form.getlist("credit[]")
        taxable_values = request.form.getlist("taxable[]")
        tax_directions = request.form.getlist("tax_direction[]")
        party_types = request.form.getlist("party_type[]")
        entity_ids = request.form.getlist("entity_id[]")
        supplier_ids = request.form.getlist("supplier_id[]")
        customer_ids = request.form.getlist("customer_id[]")
        tax_numbers = request.form.getlist("tax_number[]")
        invoice_numbers = request.form.getlist("invoice_number[]")
        invoice_dates = request.form.getlist("invoice_date[]")
        descriptions = request.form.getlist("line_description[]")
        cost_center_ids = request.form.getlist("cost_center_id[]")

        groups = {}
        for i, account_id in enumerate(account_ids):
            if not account_id:
                continue
            group_no = int(group_numbers[i] or 1)
            debit = float(debits[i] or 0)
            credit = float(credits[i] or 0)
            if debit > 0 and credit > 0:
                flash(f"المجموعة {group_no}، السطر {i+1}: لا يمكن إدخال مدين ودائن معًا", "danger")
                return redirect(url_for("multi_journal"))
            if debit <= 0 and credit <= 0:
                continue

            taxable = 1 if taxable_values[i] == "1" else 0
            party_type = party_types[i] if i < len(party_types) else ""
            entity_id = (entity_ids[i] or None) if i < len(entity_ids) else None
            supplier_id = (supplier_ids[i] or None) if i < len(supplier_ids) else None
            customer_id = (customer_ids[i] or None) if i < len(customer_ids) else None
            tax_number = tax_numbers[i].strip() if i < len(tax_numbers) else ""
            direction = tax_directions[i] if i < len(tax_directions) else "غير مطبق"
            if party_type not in JOURNAL_ACCOUNT_TYPES:
                flash(f"المجموعة {group_no}: نوع الحساب غير صحيح","danger")
                return redirect(url_for("multi_journal"))
            if party_type in ("موظف","مندوب مبيعات"):
                employee=row("SELECT id FROM employees WHERE id=:id AND active=1",{"id":entity_id}) if entity_id else None
                if not employee:
                    flash(f"المجموعة {group_no}: اختر {party_type}","danger")
                    return redirect(url_for("multi_journal"))
            elif party_type=="عميل":
                entity_id=customer_id
            elif party_type=="مورد":
                entity_id=supplier_id
            else:
                entity_id=None

            if taxable:
                if direction == "غير مطبق":
                    flash(f"المجموعة {group_no}: حدد مدخلات أو مخرجات", "danger")
                    return redirect(url_for("multi_journal"))
                if party_type not in ("مورد", "عميل"):
                    flash(f"المجموعة {group_no}: حدد موردًا أو عميلًا", "danger")
                    return redirect(url_for("multi_journal"))
                if not tax_number:
                    flash(f"المجموعة {group_no}: الرقم الضريبي مطلوب", "danger")
                    return redirect(url_for("multi_journal"))
                if not invoice_numbers[i] or not invoice_dates[i]:
                    flash(f"المجموعة {group_no}: رقم وتاريخ الفاتورة مطلوبان", "danger")
                    return redirect(url_for("multi_journal"))

            groups.setdefault(group_no, []).append({
                "account_id": account_id,
                "debit": round(debit, 2),
                "credit": round(credit, 2),
                "taxable": taxable,
                "tax_direction": direction,
                "party_type": party_type,
                "supplier_id": supplier_id if party_type == "مورد" else None,
                "customer_id": customer_id if party_type == "عميل" else None,
                "entity_id": entity_id,
                "tax_number": tax_number,
                "invoice_number": invoice_numbers[i] if i < len(invoice_numbers) else "",
                "invoice_date": invoice_dates[i] if i < len(invoice_dates) and invoice_dates[i] else None,
                "line_description": descriptions[i] if i < len(descriptions) else "",
                "cost_center_id": (cost_center_ids[i] or None) if i < len(cost_center_ids) else None,
            })

        if not groups:
            flash("أدخل مجموعة واحدة على الأقل", "danger")
            return redirect(url_for("multi_journal"))

        for group_no, lines in groups.items():
            td = round(sum(x["debit"] for x in lines), 2)
            tc = round(sum(x["credit"] for x in lines), 2)
            if td != tc:
                flash(f"المجموعة {group_no} غير متوازنة، الفرق {abs(td-tc):.2f}", "danger")
                return redirect(url_for("multi_journal"))

        batch_no = next_batch_number(request.form["batch_date"])
        execute("""INSERT INTO journal_batches(batch_no,batch_date,description,status,created_by,created_at)
                   VALUES(:no,:date,:description,:status,:user,:created)""",
                {"no": batch_no, "date": request.form["batch_date"],
                 "description": request.form.get("description",""),
                 "status": request.form["status"],
                 "user": session.get("user_id"), "created": datetime.now()})
        batch_id = row("SELECT id FROM journal_batches WHERE batch_no=:no", {"no": batch_no})["id"]

        for group_no, lines in sorted(groups.items()):
            td = round(sum(x["debit"] for x in lines), 2)
            tc = round(sum(x["credit"] for x in lines), 2)
            journal_no = next_journal_number(request.form["batch_date"])
            execute("""INSERT INTO journal_batch_groups(batch_id,group_no,journal_no,total_debit,total_credit,status)
                       VALUES(:batch,:group_no,:journal_no,:td,:tc,:status)""",
                    {"batch": batch_id, "group_no": group_no, "journal_no": journal_no,
                     "td": td, "tc": tc, "status": request.form["status"]})
            group_id = row("""SELECT id FROM journal_batch_groups
                              WHERE batch_id=:batch AND group_no=:group_no""",
                           {"batch": batch_id, "group_no": group_no})["id"]
            for line in lines:
                execute("""INSERT INTO journal_batch_lines(
                    group_id,account_id,debit,credit,taxable,tax_direction,party_type,
                    supplier_id,customer_id,entity_id,tax_number,invoice_number,invoice_date,
                    line_description,cost_center_id)
                    VALUES(:group_id,:account_id,:debit,:credit,:taxable,:tax_direction,:party_type,
                    :supplier_id,:customer_id,:entity_id,:tax_number,:invoice_number,:invoice_date,
                    :line_description,:cost_center_id)""", {"group_id": group_id, **line})

        audit("CREATE", "JOURNAL_BATCH", f"إنشاء دفعة قيود {batch_no}")
        flash(f"تم حفظ دفعة القيود {batch_no}", "success")
        if request.form.get("print_after_save") == "1":
            return redirect(url_for("multi_journal_view", batch_id=batch_id, print=1))
        return redirect(url_for("multi_journal_view", batch_id=batch_id))

    batches = rows("SELECT * FROM journal_batches ORDER BY batch_date DESC,id DESC")
    return render_template(
        "multi_journal.html",
        batches=batches,
        accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                         WHERE active=1 AND accepts_entries=1 ORDER BY account_code"""),
        suppliers=rows("SELECT id,name,vat_number FROM suppliers ORDER BY name"),
        customers=rows("SELECT id,name,vat_number FROM customers ORDER BY name"),
        employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"),
        centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code")
    )

@app.route("/multi-journal/<int:batch_id>")
@login_required
def multi_journal_view(batch_id):
    batch = row("SELECT * FROM journal_batches WHERE id=:id", {"id": batch_id})
    if not batch:
        return "الدفعة غير موجودة", 404
    groups = rows("SELECT * FROM journal_batch_groups WHERE batch_id=:id ORDER BY group_no", {"id": batch_id})
    lines = rows("""
        SELECT l.*,g.group_no,g.journal_no,a.account_code,a.account_name_ar,
               s.name supplier_name,cus.name customer_name,
               cc.code cost_center_code,cc.name cost_center_name
        FROM journal_batch_lines l
        JOIN journal_batch_groups g ON g.id=l.group_id
        JOIN chart_of_accounts a ON a.id=l.account_id
        LEFT JOIN suppliers s ON s.id=l.supplier_id
        LEFT JOIN customers cus ON cus.id=l.customer_id
        LEFT JOIN cost_centers cc ON cc.id=l.cost_center_id
        WHERE g.batch_id=:id ORDER BY g.group_no,l.id
    """, {"id": batch_id})
    company = row("SELECT * FROM settings WHERE id=1")
    return render_template("multi_journal_view.html", batch=batch, groups=groups, lines=lines, company=company, auto_print=request.args.get("print")=="1")


@app.route("/multi-journal/<int:batch_id>/group/<int:group_no>/print")
@login_required
def multi_journal_group_print(batch_id, group_no):
    batch = row("SELECT * FROM journal_batches WHERE id=:id", {"id": batch_id})
    if not batch:
        return "الدفعة غير موجودة", 404

    group = row("""
        SELECT *
        FROM journal_batch_groups
        WHERE batch_id=:batch_id AND group_no=:group_no
    """, {"batch_id": batch_id, "group_no": group_no})
    if not group:
        return "القيد غير موجود داخل الدفعة", 404

    lines = rows("""
        SELECT l.*,a.account_code,a.account_name_ar,
               s.name supplier_name,cus.name customer_name,
               cc.code cost_center_code,cc.name cost_center_name
        FROM journal_batch_lines l
        JOIN chart_of_accounts a ON a.id=l.account_id
        LEFT JOIN suppliers s ON s.id=l.supplier_id
        LEFT JOIN customers cus ON cus.id=l.customer_id
        LEFT JOIN cost_centers cc ON cc.id=l.cost_center_id
        WHERE l.group_id=:group_id
        ORDER BY l.id
    """, {"group_id": group["id"]})

    company = row("SELECT * FROM settings WHERE id=1")
    return render_template(
        "multi_journal_group_print.html",
        batch=batch,
        group=group,
        lines=lines,
        company=company
    )


@app.route("/multi-journal/<int:batch_id>/delete", methods=["POST"])
@login_required
def multi_journal_delete(batch_id):
    batch = row("SELECT * FROM journal_batches WHERE id=:id", {"id": batch_id})
    if not batch:
        return "الدفعة غير موجودة", 404
    if batch["status"] == "مرحّل":
        flash("لا يمكن حذف دفعة مرحّلة", "danger")
    else:
        execute("DELETE FROM journal_batches WHERE id=:id", {"id": batch_id})
        audit("DELETE", "JOURNAL_BATCH", f"حذف دفعة {batch['batch_no']}")
        flash("تم حذف الدفعة", "success")
    return redirect(url_for("multi_journal"))

@app.route("/multi-journal/<int:batch_id>/export.xlsx")
@login_required
def multi_journal_export(batch_id):
    data = rows("""
        SELECT b.batch_no,b.batch_date,g.group_no,g.journal_no,a.account_code,a.account_name_ar,
               l.debit,l.credit,l.taxable,l.tax_direction,l.party_type,
               COALESCE(s.name,c.name) party_name,l.tax_number,l.invoice_number,l.invoice_date,
               l.line_description,cc.code cost_center_code,cc.name cost_center_name
        FROM journal_batches b
        JOIN journal_batch_groups g ON g.batch_id=b.id
        JOIN journal_batch_lines l ON l.group_id=g.id
        JOIN chart_of_accounts a ON a.id=l.account_id
        LEFT JOIN suppliers s ON s.id=l.supplier_id
        LEFT JOIN customers c ON c.id=l.customer_id
        LEFT JOIN cost_centers cc ON cc.id=l.cost_center_id
        WHERE b.id=:id ORDER BY g.group_no,l.id
    """, {"id": batch_id})
    headers = ["رقم الدفعة","تاريخ الدفعة","المجموعة","رقم القيد","الكود","اسم الحساب",
               "مدين","دائن","خاضع للضريبة","مدخلات/مخرجات","نوع الطرف","اسم الطرف",
               "الرقم الضريبي","رقم الفاتورة","تاريخ الفاتورة","البيان",
               "كود مركز التكلفة","مركز التكلفة"]
    keys = ["batch_no","batch_date","group_no","journal_no","account_code","account_name_ar",
            "debit","credit","taxable","tax_direction","party_type","party_name","tax_number",
            "invoice_number","invoice_date","line_description","cost_center_code","cost_center_name"]
    records = [[r.get(k) for k in keys] for r in data]
    return xlsx_response("multi_journal.xlsx", "قيود متعددة", headers, records)




ACCOUNTING_REPORTS = {
    "ledger-summary": {"no": 1, "title": "حساب الأستاذ - إجمالي", "group": "الحسابات"},
    "ledger-detail": {"no": 2, "title": "حساب الأستاذ - تفصيلي", "group": "الحسابات"},
    "cost-center-movements-center": {"no": 3, "title": "حركات مراكز التكلفة - حسب المركز", "group": "مراكز التكلفة"},
    "cost-center-movements-account": {"no": 4, "title": "حركات مراكز التكلفة - حسب الحساب", "group": "مراكز التكلفة"},
    "journal-report": {"no": 5, "title": "القيود المحاسبية", "group": "الحسابات"},
    "coa-report": {"no": 6, "title": "الدليل الحسابي", "group": "الحسابات"},
    "account-balances": {"no": 7, "title": "أرصدة الحسابات", "group": "الحسابات"},
    "trial-balance": {"no": 8, "title": "ميزان المراجعة", "group": "الحسابات"},
    "cc-ledger-summary": {"no": 9, "title": "حساب الأستاذ لمراكز التكلفة - إجمالي", "group": "مراكز التكلفة"},
    "cc-ledger-detail": {"no": 10, "title": "حساب الأستاذ لمراكز التكلفة - تفصيلي", "group": "مراكز التكلفة"},
    "cc-trial-balance": {"no": 11, "title": "ميزان المراجعة - مراكز التكلفة", "group": "مراكز التكلفة"},
    "cc-balances": {"no": 12, "title": "أرصدة مراكز التكلفة", "group": "مراكز التكلفة"},
    "accounts-with-centers": {"no": 13, "title": "إجمالي الحسابات مع مراكز التكلفة", "group": "تحليل مشترك"},
    "centers-with-accounts": {"no": 14, "title": "إجمالي مراكز التكلفة مع الحسابات", "group": "تحليل مشترك"},
    "vat-report": {"no": 15, "title": "ضريبة القيمة المضافة", "group": "الضرائب"},
}

def report_filters():
    return {
        "date_from": request.args.get("date_from", ""),
        "date_to": request.args.get("date_to", ""),
        "account_id": request.args.get("account_id", type=int),
        "cost_center_id": request.args.get("cost_center_id", type=int),
        "status": request.args.get("status", "مرحّل"),
        "tax_direction": request.args.get("tax_direction", ""),
        "q": request.args.get("q", "").strip(),
    }

def journal_filter_sql(filters, line_alias="l", journal_alias="j"):
    conditions = []
    params = {}
    if filters["date_from"]:
        conditions.append(f"{journal_alias}.journal_date >= :date_from")
        params["date_from"] = filters["date_from"]
    if filters["date_to"]:
        conditions.append(f"{journal_alias}.journal_date <= :date_to")
        params["date_to"] = filters["date_to"]
    if filters["status"] and filters["status"] != "الكل":
        conditions.append(f"{journal_alias}.status = :status")
        params["status"] = filters["status"]
    if filters["account_id"]:
        conditions.append(f"{line_alias}.account_id = :account_id")
        params["account_id"] = filters["account_id"]
    if filters["cost_center_id"]:
        conditions.append(f"{line_alias}.cost_center_id = :cost_center_id")
        params["cost_center_id"] = filters["cost_center_id"]
    if filters["tax_direction"]:
        conditions.append(f"{line_alias}.tax_direction = :tax_direction")
        params["tax_direction"] = filters["tax_direction"]
    if filters["q"]:
        conditions.append(f"""(
            {journal_alias}.journal_no ILIKE :q OR
            COALESCE({journal_alias}.reference,'') ILIKE :q OR
            COALESCE({journal_alias}.description,'') ILIKE :q OR
            COALESCE({line_alias}.invoice_number,'') ILIKE :q OR
            COALESCE({line_alias}.line_description,'') ILIKE :q
        )""")
        params["q"] = f"%{filters['q']}%"
    return (" WHERE " + " AND ".join(conditions)) if conditions else "", params

def report_context(slug, filters):
    definition = ACCOUNTING_REPORTS[slug]
    where, params = journal_filter_sql(filters)
    headers, data, summary = [], [], {}

    if slug == "coa-report":
        headers = ["الكود","اسم الحساب","الاسم الإنجليزي","النوع","الحساب الأب","المستوى","طبيعة الرصيد","القائمة المالية","يقبل حركة","الحالة"]
        data = rows("""
            SELECT a.id,a.account_code,a.account_name_ar,a.account_name_en,a.account_type,
                   p.account_name_ar parent_name,a.level,a.normal_balance,a.statement_type,
                   a.accepts_entries,a.active
            FROM chart_of_accounts a
            LEFT JOIN chart_of_accounts p ON p.id=a.parent_id
            ORDER BY a.account_code
        """)

    elif slug in ("ledger-summary","account-balances","trial-balance"):
        headers = ["الكود","اسم الحساب","إجمالي المدين","إجمالي الدائن","الرصيد المدين","الرصيد الدائن","عدد الحركات"]
        data = rows(f"""
            SELECT a.id,a.account_code,a.account_name_ar,
                   COALESCE(SUM(l.debit),0) total_debit,
                   COALESCE(SUM(l.credit),0) total_credit,
                   GREATEST(COALESCE(SUM(l.debit-l.credit),0),0) debit_balance,
                   GREATEST(COALESCE(SUM(l.credit-l.debit),0),0) credit_balance,
                   COUNT(l.id) movement_count
            FROM chart_of_accounts a
            LEFT JOIN journal_entry_lines l ON l.account_id=a.id
            LEFT JOIN journal_entries j ON j.id=l.journal_id
            {where}
            GROUP BY a.id,a.account_code,a.account_name_ar
            ORDER BY a.account_code
        """, params)
        summary = {
            "total_debit": sum(float(x["total_debit"]) for x in data),
            "total_credit": sum(float(x["total_credit"]) for x in data),
            "debit_balance": sum(float(x["debit_balance"]) for x in data),
            "credit_balance": sum(float(x["credit_balance"]) for x in data),
        }

    elif slug == "ledger-detail":
        headers = ["التاريخ","رقم القيد","المرجع","البيان","رقم الفاتورة","تاريخ الفاتورة","مدين","دائن","الرصيد المتحرك","الحالة"]
        raw = rows(f"""
            SELECT j.id journal_id,j.journal_date,j.journal_no,j.reference,j.description,
                   l.invoice_number,l.invoice_date,l.line_description,l.debit,l.credit,j.status
            FROM journal_entry_lines l
            JOIN journal_entries j ON j.id=l.journal_id
            {where}
            ORDER BY j.journal_date,j.id,l.id
        """, params)
        running = 0.0
        data = []
        for item in raw:
            running += float(item["debit"]) - float(item["credit"])
            d = dict(item)
            d["running_balance"] = running
            data.append(d)
        summary = {
            "total_debit": sum(float(x["debit"]) for x in data),
            "total_credit": sum(float(x["credit"]) for x in data),
            "closing_balance": running,
        }

    elif slug == "journal-report":
        headers = ["تاريخ القيد","رقم القيد","الحالة","المرجع","البيان العام","الكود","اسم الحساب","مدين","دائن","اسم الطرف","الرقم الضريبي","رقم الفاتورة","تاريخ الفاتورة","مركز التكلفة"]
        data = rows(f"""
            SELECT j.id journal_id,j.journal_date,j.journal_no,j.status,j.reference,j.description,
                   a.account_code,a.account_name_ar,l.debit,l.credit,
                   COALESCE(s.name,cus.name,'') party_name,l.tax_number,l.invoice_number,l.invoice_date,
                   cc.code cost_center_code,cc.name cost_center_name
            FROM journal_entry_lines l
            JOIN journal_entries j ON j.id=l.journal_id
            JOIN chart_of_accounts a ON a.id=l.account_id
            LEFT JOIN suppliers s ON s.id=l.supplier_id
            LEFT JOIN customers cus ON cus.id=l.customer_id
            LEFT JOIN cost_centers cc ON cc.id=l.cost_center_id
            {where}
            ORDER BY j.journal_date DESC,j.id DESC,l.id
        """, params)
        summary = {
            "total_debit": sum(float(x["debit"]) for x in data),
            "total_credit": sum(float(x["credit"]) for x in data),
            "count": len(data),
        }

    elif slug in ("cost-center-movements-center","cc-ledger-detail"):
        headers = ["مركز التكلفة","التاريخ","رقم القيد","الكود","اسم الحساب","البيان","مدين","دائن","الرصيد","الحالة"]
        raw = rows(f"""
            SELECT j.id journal_id,cc.code cost_center_code,cc.name cost_center_name,
                   j.journal_date,j.journal_no,j.status,a.account_code,a.account_name_ar,
                   COALESCE(l.line_description,j.description) line_description,l.debit,l.credit
            FROM journal_entry_lines l
            JOIN journal_entries j ON j.id=l.journal_id
            JOIN chart_of_accounts a ON a.id=l.account_id
            JOIN cost_centers cc ON cc.id=l.cost_center_id
            {where}
            ORDER BY cc.code,j.journal_date,j.id,l.id
        """, params)
        balances = {}
        data = []
        for item in raw:
            key = item["cost_center_code"]
            balances[key] = balances.get(key, 0.0) + float(item["debit"]) - float(item["credit"])
            d = dict(item); d["running_balance"] = balances[key]; data.append(d)
        summary = {"total_debit": sum(float(x["debit"]) for x in data), "total_credit": sum(float(x["credit"]) for x in data)}

    elif slug == "cost-center-movements-account":
        headers = ["الكود","اسم الحساب","مركز التكلفة","تاريخ القيد","رقم القيد","البيان","مدين","دائن","الحالة"]
        data = rows(f"""
            SELECT j.id journal_id,a.account_code,a.account_name_ar,
                   cc.code cost_center_code,cc.name cost_center_name,
                   j.journal_date,j.journal_no,j.status,
                   COALESCE(l.line_description,j.description) line_description,l.debit,l.credit
            FROM journal_entry_lines l
            JOIN journal_entries j ON j.id=l.journal_id
            JOIN chart_of_accounts a ON a.id=l.account_id
            JOIN cost_centers cc ON cc.id=l.cost_center_id
            {where}
            ORDER BY a.account_code,cc.code,j.journal_date,j.id
        """, params)

    elif slug in ("cc-ledger-summary","cc-trial-balance","cc-balances"):
        headers = ["كود المركز","مركز التكلفة","إجمالي المدين","إجمالي الدائن","الرصيد المدين","الرصيد الدائن","عدد الحركات"]
        data = rows(f"""
            SELECT cc.id,cc.code cost_center_code,cc.name cost_center_name,
                   COALESCE(SUM(l.debit),0) total_debit,
                   COALESCE(SUM(l.credit),0) total_credit,
                   GREATEST(COALESCE(SUM(l.debit-l.credit),0),0) debit_balance,
                   GREATEST(COALESCE(SUM(l.credit-l.debit),0),0) credit_balance,
                   COUNT(l.id) movement_count
            FROM cost_centers cc
            LEFT JOIN journal_entry_lines l ON l.cost_center_id=cc.id
            LEFT JOIN journal_entries j ON j.id=l.journal_id
            {where}
            GROUP BY cc.id,cc.code,cc.name
            ORDER BY cc.code
        """, params)
        summary = {
            "total_debit": sum(float(x["total_debit"]) for x in data),
            "total_credit": sum(float(x["total_credit"]) for x in data),
        }

    elif slug == "accounts-with-centers":
        headers = ["كود الحساب","اسم الحساب","كود المركز","مركز التكلفة","إجمالي المدين","إجمالي الدائن","الرصيد"]
        data = rows(f"""
            SELECT a.account_code,a.account_name_ar,cc.code cost_center_code,cc.name cost_center_name,
                   SUM(l.debit) total_debit,SUM(l.credit) total_credit,SUM(l.debit-l.credit) balance
            FROM journal_entry_lines l
            JOIN journal_entries j ON j.id=l.journal_id
            JOIN chart_of_accounts a ON a.id=l.account_id
            LEFT JOIN cost_centers cc ON cc.id=l.cost_center_id
            {where}
            GROUP BY a.account_code,a.account_name_ar,cc.code,cc.name
            ORDER BY a.account_code,cc.code
        """, params)

    elif slug == "centers-with-accounts":
        headers = ["كود المركز","مركز التكلفة","كود الحساب","اسم الحساب","إجمالي المدين","إجمالي الدائن","الرصيد"]
        data = rows(f"""
            SELECT cc.code cost_center_code,cc.name cost_center_name,a.account_code,a.account_name_ar,
                   SUM(l.debit) total_debit,SUM(l.credit) total_credit,SUM(l.debit-l.credit) balance
            FROM journal_entry_lines l
            JOIN journal_entries j ON j.id=l.journal_id
            JOIN chart_of_accounts a ON a.id=l.account_id
            LEFT JOIN cost_centers cc ON cc.id=l.cost_center_id
            {where}
            GROUP BY cc.code,cc.name,a.account_code,a.account_name_ar
            ORDER BY cc.code,a.account_code
        """, params)

    elif slug == "vat-report":
        vat_where, vat_params = journal_filter_sql(filters)
        vat_condition = "l.taxable=1"
        vat_where = vat_where + (" AND " if vat_where else " WHERE ") + vat_condition
        headers = ["تاريخ القيد","رقم القيد","تاريخ الفاتورة","رقم الفاتورة","اسم المورد / العميل","الرقم الضريبي","مدخلات / مخرجات","المبلغ قبل الضريبة","ضريبة القيمة المضافة","الإجمالي","حالة البيانات"]
        raw = rows(f"""
            SELECT j.id journal_id,j.journal_date,j.journal_no,l.invoice_date,l.invoice_number,
                   COALESCE(s.name,cus.name,'') party_name,l.tax_number,l.tax_direction,
                   GREATEST(l.debit,l.credit) amount_before_tax
            FROM journal_entry_lines l
            JOIN journal_entries j ON j.id=l.journal_id
            LEFT JOIN suppliers s ON s.id=l.supplier_id
            LEFT JOIN customers cus ON cus.id=l.customer_id
            {vat_where}
            ORDER BY COALESCE(l.invoice_date,j.journal_date),j.id,l.id
        """, vat_params)
        vat_rate = float(row("SELECT vat_rate FROM settings WHERE id=1")["vat_rate"] or 15)
        data = []
        for item in raw:
            d = dict(item)
            base_amount = float(item["amount_before_tax"] or 0)
            d["vat_amount"] = round(base_amount * vat_rate / 100, 2)
            d["total_amount"] = round(base_amount + d["vat_amount"], 2)
            complete = all([item["invoice_date"], item["invoice_number"], item["party_name"], item["tax_number"], item["tax_direction"] in ("مدخلات","مخرجات")])
            d["compliance_status"] = "مكتمل" if complete else "بيانات ناقصة"
            data.append(d)
        output_rows = [x for x in data if x["tax_direction"] == "مخرجات"]
        input_rows = [x for x in data if x["tax_direction"] == "مدخلات"]
        summary = {
            "output_base": sum(x["amount_before_tax"] for x in output_rows),
            "output_vat": sum(x["vat_amount"] for x in output_rows),
            "input_base": sum(x["amount_before_tax"] for x in input_rows),
            "input_vat": sum(x["vat_amount"] for x in input_rows),
            "net_vat": sum(x["vat_amount"] for x in output_rows) - sum(x["vat_amount"] for x in input_rows),
        }

    return {
        "definition": definition,
        "headers": headers,
        "data": data,
        "summary": summary,
    }





@app.route("/sales")
@login_required
def sales_center():
    return render_template("sales_center.html",
      quotations=row("SELECT COUNT(*) c FROM sales_quotations")["c"],
      orders=row("SELECT COUNT(*) c FROM sales_orders")["c"],
      deliveries=row("SELECT COUNT(*) c FROM sales_deliveries")["c"],
      invoices=row("SELECT COUNT(*) c FROM invoices")["c"])

@app.route("/sales/quotations",methods=["GET","POST"])
@login_required
def sales_quotations():
    if request.method=="POST":
        try: lines,sub,disc,vat,total=parse_sales_lines(request.form)
        except Exception as exc:
            flash(str(exc),"danger"); return redirect(url_for("sales_quotations"))
        no=next_sales_no("sales_quotations","quotation_date","QT",request.form["quotation_date"])
        execute("""INSERT INTO sales_quotations(quotation_no,quotation_date,valid_until,customer_id,
          branch_id,cost_center_id,status,subtotal,discount,vat,total,notes,created_by,created_at)
          VALUES(:no,:dt,:valid,:customer,:branch,:cc,:status,:sub,:disc,:vat,:total,:notes,:uid,:created)""",
          {"no":no,"dt":request.form["quotation_date"],"valid":request.form.get("valid_until") or None,
           "customer":request.form["customer_id"],"branch":request.form.get("branch_id") or None,
           "cc":request.form.get("cost_center_id") or None,"status":request.form.get("status","مسودة"),
           "sub":sub,"disc":disc,"vat":vat,"total":total,"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        qid=row("SELECT id FROM sales_quotations WHERE quotation_no=:n",{"n":no})["id"]
        for x in lines:
            execute("""INSERT INTO sales_quotation_items(quotation_id,item_id,item_name,quantity,unit,
              unit_price,discount_rate,vat_rate,line_subtotal,line_discount,line_vat,line_total)
              VALUES(:qid,:item_id,:item_name,:quantity,:unit,:unit_price,:discount_rate,:vat_rate,
              :line_subtotal,:line_discount,:line_vat,:line_total)""",{"qid":qid,**x})
        flash(f"تم إنشاء عرض السعر {no}","success")
        return redirect(url_for("sales_quotation_view",quotation_id=qid))
    return render_template("sales_quotations.html",
      docs=rows("""SELECT q.*,c.name customer_name FROM sales_quotations q JOIN customers c ON c.id=q.customer_id ORDER BY q.id DESC"""),
      customers=rows("SELECT id,name,name_en FROM customers ORDER BY name"),
      items=rows("SELECT id,sku,name,unit,sale_price,quantity FROM inventory WHERE active=1 ORDER BY name"),
      branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"))

@app.route("/sales/quotations/<int:quotation_id>")
@login_required
def sales_quotation_view(quotation_id):
    doc=row("""SELECT q.*,c.name customer_name,c.name_en customer_name_en FROM sales_quotations q
               JOIN customers c ON c.id=q.customer_id WHERE q.id=:id""",{"id":quotation_id})
    items=rows("SELECT * FROM sales_quotation_items WHERE quotation_id=:id ORDER BY id",{"id":quotation_id})
    return render_template("sales_doc_view.html",doc=doc,items=items,title="عرض سعر / Quotation",
                           number=doc["quotation_no"],date=doc["quotation_date"],kind="quotation")

@app.route("/sales/quotations/<int:quotation_id>/convert",methods=["POST"])
@login_required
def sales_quotation_convert(quotation_id):
    q=row("SELECT * FROM sales_quotations WHERE id=:id",{"id":quotation_id})
    if q.get("converted_order_id"): return redirect(url_for("sales_order_view",order_id=q["converted_order_id"]))
    wh=row("SELECT id FROM warehouses WHERE active=1 ORDER BY id LIMIT 1")
    no=next_sales_no("sales_orders","order_date","SO",datetime.now().date().isoformat())
    execute("""INSERT INTO sales_orders(order_no,order_date,quotation_id,customer_id,branch_id,
      cost_center_id,warehouse_id,status,subtotal,discount,vat,total,notes,created_by,created_at)
      VALUES(:no,:dt,:qid,:customer,:branch,:cc,:wh,'مسودة',:sub,:disc,:vat,:total,:notes,:uid,:created)""",
      {"no":no,"dt":datetime.now().date(),"qid":quotation_id,"customer":q["customer_id"],
       "branch":q["branch_id"],"cc":q["cost_center_id"],"wh":wh["id"] if wh else None,
       "sub":q["subtotal"],"disc":q["discount"],"vat":q["vat"],"total":q["total"],
       "notes":f"من عرض السعر {q['quotation_no']}","uid":session.get("user_id"),"created":datetime.now()})
    oid=row("SELECT id FROM sales_orders WHERE order_no=:n",{"n":no})["id"]
    execute("""INSERT INTO sales_order_items(order_id,item_id,item_name,quantity,unit,unit_price,
      discount_rate,vat_rate,line_subtotal,line_discount,line_vat,line_total)
      SELECT :oid,item_id,item_name,quantity,unit,unit_price,discount_rate,vat_rate,
      line_subtotal,line_discount,line_vat,line_total FROM sales_quotation_items WHERE quotation_id=:qid""",
      {"oid":oid,"qid":quotation_id})
    execute("UPDATE sales_quotations SET status='محوّل',converted_order_id=:oid WHERE id=:id",
            {"oid":oid,"id":quotation_id})
    return redirect(url_for("sales_order_view",order_id=oid))

@app.route("/sales/orders",methods=["GET","POST"])
@login_required
def sales_orders():
    if request.method=="POST":
        try: lines,sub,disc,vat,total=parse_sales_lines(request.form)
        except Exception as exc:
            flash(str(exc),"danger"); return redirect(url_for("sales_orders"))
        no=next_sales_no("sales_orders","order_date","SO",request.form["order_date"])
        execute("""INSERT INTO sales_orders(order_no,order_date,customer_id,branch_id,cost_center_id,
          warehouse_id,status,subtotal,discount,vat,total,notes,created_by,created_at)
          VALUES(:no,:dt,:customer,:branch,:cc,:wh,:status,:sub,:disc,:vat,:total,:notes,:uid,:created)""",
          {"no":no,"dt":request.form["order_date"],"customer":request.form["customer_id"],
           "branch":request.form.get("branch_id") or None,"cc":request.form.get("cost_center_id") or None,
           "wh":request.form["warehouse_id"],"status":request.form.get("status","مسودة"),
           "sub":sub,"disc":disc,"vat":vat,"total":total,"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        oid=row("SELECT id FROM sales_orders WHERE order_no=:n",{"n":no})["id"]
        for x in lines:
            execute("""INSERT INTO sales_order_items(order_id,item_id,item_name,quantity,unit,unit_price,
              discount_rate,vat_rate,line_subtotal,line_discount,line_vat,line_total)
              VALUES(:oid,:item_id,:item_name,:quantity,:unit,:unit_price,:discount_rate,:vat_rate,
              :line_subtotal,:line_discount,:line_vat,:line_total)""",{"oid":oid,**x})
        return redirect(url_for("sales_order_view",order_id=oid))
    return render_template("sales_orders.html",
      docs=rows("""SELECT o.*,c.name customer_name,w.name warehouse_name FROM sales_orders o
                   JOIN customers c ON c.id=o.customer_id LEFT JOIN warehouses w ON w.id=o.warehouse_id ORDER BY o.id DESC"""),
      customers=rows("SELECT id,name,name_en FROM customers ORDER BY name"),
      items=rows("SELECT id,sku,name,unit,sale_price,quantity FROM inventory WHERE active=1 ORDER BY name"),
      branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"),
      warehouses=rows("SELECT id,code,name FROM warehouses WHERE active=1 ORDER BY code"))

@app.route("/sales/orders/<int:order_id>")
@login_required
def sales_order_view(order_id):
    doc=row("""SELECT o.*,c.name customer_name,c.name_en customer_name_en,w.name warehouse_name
               FROM sales_orders o JOIN customers c ON c.id=o.customer_id
               LEFT JOIN warehouses w ON w.id=o.warehouse_id WHERE o.id=:id""",{"id":order_id})
    items=rows("SELECT * FROM sales_order_items WHERE order_id=:id ORDER BY id",{"id":order_id})
    return render_template("sales_doc_view.html",doc=doc,items=items,title="أمر بيع / Sales Order",
                           number=doc["order_no"],date=doc["order_date"],kind="order")

@app.route("/sales/orders/<int:order_id>/approve",methods=["POST"])
@login_required
def sales_order_approve(order_id):
    o=row("SELECT * FROM sales_orders WHERE id=:id",{"id":order_id})
    for i in rows("SELECT * FROM sales_order_items WHERE order_id=:id",{"id":order_id}):
        if warehouse_stock(i["item_id"],o["warehouse_id"]) < float(i["quantity"])-float(i["delivered_qty"]):
            flash(f"الرصيد غير كافٍ للصنف {i['item_name']}","danger")
            return redirect(url_for("sales_order_view",order_id=order_id))
    execute("UPDATE sales_orders SET status='معتمد' WHERE id=:id",{"id":order_id})
    return redirect(url_for("sales_order_view",order_id=order_id))

@app.route("/sales/deliveries",methods=["GET","POST"])
@login_required
def sales_deliveries():
    if request.method=="POST":
        oid=int(request.form["order_id"]); order=row("SELECT * FROM sales_orders WHERE id=:id",{"id":oid})
        item_ids=request.form.getlist("order_item_id[]"); qtys=request.form.getlist("delivery_qty[]")
        lines=[]
        for i,item_id in enumerate(item_ids):
            oi=row("SELECT * FROM sales_order_items WHERE id=:id",{"id":item_id})
            qty=float(qtys[i] or 0); remain=float(oi["quantity"])-float(oi["delivered_qty"])
            if qty>remain+0.0001: flash("كمية التسليم تتجاوز المتبقي","danger"); return redirect(url_for("sales_deliveries"))
            if qty>0:
                if warehouse_stock(oi["item_id"],order["warehouse_id"])<qty:
                    flash(f"الرصيد غير كافٍ للصنف {oi['item_name']}","danger"); return redirect(url_for("sales_deliveries"))
                lines.append((oi,qty))
        if not lines: flash("أدخل كمية تسليم","danger"); return redirect(url_for("sales_deliveries"))
        no=next_sales_no("sales_deliveries","delivery_date","DN",request.form["delivery_date"])
        execute("""INSERT INTO sales_deliveries(delivery_no,delivery_date,order_id,customer_id,
          warehouse_id,status,notes,created_by,created_at)
          VALUES(:no,:dt,:oid,:customer,:wh,'معتمد',:notes,:uid,:created)""",
          {"no":no,"dt":request.form["delivery_date"],"oid":oid,"customer":order["customer_id"],
           "wh":order["warehouse_id"],"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        did=row("SELECT id FROM sales_deliveries WHERE delivery_no=:n",{"n":no})["id"]
        for oi,qty in lines:
            item=row("SELECT cost FROM inventory WHERE id=:id",{"id":oi["item_id"]})
            execute("""INSERT INTO sales_delivery_items(delivery_id,order_item_id,item_id,quantity,unit_cost)
                       VALUES(:did,:oi,:item,:qty,:cost)""",
                    {"did":did,"oi":oi["id"],"item":oi["item_id"],"qty":qty,"cost":item["cost"]})
            execute("UPDATE sales_order_items SET delivered_qty=delivered_qty+:q WHERE id=:id",
                    {"q":qty,"id":oi["id"]})
            record_inventory_movement(request.form["delivery_date"],"بيع",oi["item_id"],order["warehouse_id"],
                                      qty,item["cost"],reference_type="DELIVERY",reference_id=did,reference_no=no)
        remaining=row("SELECT COUNT(*) c FROM sales_order_items WHERE order_id=:id AND delivered_qty<quantity",
                      {"id":oid})["c"]
        execute("UPDATE sales_orders SET status=:s WHERE id=:id",
                {"s":"مكتمل التسليم" if remaining==0 else "تسليم جزئي","id":oid})
        return redirect(url_for("sales_delivery_view",delivery_id=did))
    return render_template("sales_deliveries.html",
      orders=rows("""SELECT o.id,o.order_no,c.name customer_name FROM sales_orders o
                     JOIN customers c ON c.id=o.customer_id WHERE o.status IN ('معتمد','تسليم جزئي') ORDER BY o.id DESC"""),
      docs=rows("""SELECT d.*,o.order_no,c.name customer_name FROM sales_deliveries d
                   JOIN sales_orders o ON o.id=d.order_id JOIN customers c ON c.id=d.customer_id ORDER BY d.id DESC"""))

@app.route("/sales/orders/<int:order_id>/items")
@login_required
def sales_order_items_api(order_id):
    return {"items":[dict(x) for x in rows("""SELECT id,item_name,quantity,delivered_qty
      FROM sales_order_items WHERE order_id=:id ORDER BY id""",{"id":order_id})]}

@app.route("/sales/deliveries/<int:delivery_id>")
@login_required
def sales_delivery_view(delivery_id):
    doc=row("""SELECT d.*,o.order_no,c.name customer_name,c.name_en customer_name_en,w.name warehouse_name
               FROM sales_deliveries d JOIN sales_orders o ON o.id=d.order_id
               JOIN customers c ON c.id=d.customer_id JOIN warehouses w ON w.id=d.warehouse_id
               WHERE d.id=:id""",{"id":delivery_id})
    items=rows("""SELECT di.*,oi.item_name,oi.unit FROM sales_delivery_items di
                  JOIN sales_order_items oi ON oi.id=di.order_item_id WHERE di.delivery_id=:id""",{"id":delivery_id})
    return render_template("sales_delivery_view.html",doc=doc,items=items)


@app.route("/sales/deliveries/<int:delivery_id>/create-invoice",methods=["POST"])
@login_required
def sales_delivery_create_invoice(delivery_id):
    try:
        invoice_id = create_invoice_from_sales_delivery(delivery_id)
        post_invoice_to_ledger(invoice_id)
        execute("UPDATE invoices SET status='معتمدة' WHERE id=:id",{"id":invoice_id})
        flash("تم إنشاء الفاتورة وترحيل قيد المبيعات وتكلفة البضاعة المباعة","success")
        return redirect(url_for("invoice_view",invoice_id=invoice_id))
    except Exception as exc:
        db.session.rollback()
        flash(str(exc),"danger")
        return redirect(url_for("sales_delivery_view",delivery_id=delivery_id))

@app.route("/sales/returns",methods=["GET","POST"])
@login_required
def sales_returns():
    if request.method=="POST":
        invoice_id=int(request.form["invoice_id"])
        invoice=row("SELECT * FROM invoices WHERE id=:id",{"id":invoice_id})
        if not invoice:
            flash("الفاتورة غير موجودة","danger")
            return redirect(url_for("sales_returns"))
        warehouse_id=int(request.form["warehouse_id"])
        ids=request.form.getlist("invoice_item_id[]")
        quantities=request.form.getlist("return_qty[]")
        prepared=[]; subtotal=vat_total=grand_total=cost_total=0.0
        for index,item_id in enumerate(ids):
            item=row("""SELECT ii.*,COALESCE(inv.cost,0) current_cost
                        FROM invoice_items ii LEFT JOIN inventory inv ON inv.id=ii.item_id
                        WHERE ii.id=:id AND ii.invoice_id=:invoice""",
                     {"id":item_id,"invoice":invoice_id})
            if not item:
                continue
            qty=float(quantities[index] or 0)
            previous=db.session.execute(text("""SELECT COALESCE(SUM(sri.quantity),0)
                FROM sales_return_items sri JOIN sales_returns sr ON sr.id=sri.return_id
                WHERE sri.invoice_item_id=:item AND sr.invoice_id=:invoice"""),
                {"item":item_id,"invoice":invoice_id}).scalar() or 0
            available=float(item["quantity"])-float(previous)
            if qty < 0 or qty > available + 0.0001:
                flash(f"كمية المرتجع للصنف {item['item_name']} تتجاوز المتاح {available:.3f}","danger")
                return redirect(url_for("sales_returns"))
            if qty == 0:
                continue
            base=round(qty*float(item["unit_price"]),2)
            tax=round(base*float(item["vat_rate"] or 0)/100,2)
            cost=round(qty*float(item["current_cost"] or 0),2)
            prepared.append((item,qty,base,tax,cost))
            subtotal+=base;vat_total+=tax;grand_total+=base+tax;cost_total+=cost
        if not prepared:
            flash("أدخل كمية مرتجع واحدة على الأقل","danger")
            return redirect(url_for("sales_returns"))

        no=next_sales_no("sales_returns","return_date","SR",request.form["return_date"]) \
           if "sales_returns" in {"sales_returns"} else ""
        # next_sales_no does not originally include returns; generate safely.
        year=datetime.strptime(request.form["return_date"],"%Y-%m-%d").year
        seq=db.session.execute(text("SELECT COUNT(*) FROM sales_returns WHERE EXTRACT(YEAR FROM return_date)=:y"),
                               {"y":year}).scalar() or 0
        no=f"SR-{year}-{seq+1:06d}"

        execute("""INSERT INTO sales_returns(
          return_no,return_date,invoice_id,customer_id,warehouse_id,subtotal,vat,total,
          cost_total,status,notes,created_by,created_at)
          VALUES(:no,:dt,:invoice,:customer,:warehouse,:sub,:vat,:total,:cost,
          'معتمد',:notes,:uid,:created)""",
          {"no":no,"dt":request.form["return_date"],"invoice":invoice_id,
           "customer":invoice["customer_id"],"warehouse":warehouse_id,
           "sub":round(subtotal,2),"vat":round(vat_total,2),"total":round(grand_total,2),
           "cost":round(cost_total,2),"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        return_id=row("SELECT id FROM sales_returns WHERE return_no=:no",{"no":no})["id"]

        for item,qty,base,tax,cost in prepared:
            unit_cost=float(item["current_cost"] or 0)
            execute("""INSERT INTO sales_return_items(
              return_id,invoice_item_id,item_id,item_name,quantity,unit,unit_price,vat_rate,
              unit_cost,line_subtotal,line_vat,line_total)
              VALUES(:return_id,:invoice_item,:item_id,:name,:qty,:unit,:price,:rate,
              :cost,:sub,:vat,:total)""",
              {"return_id":return_id,"invoice_item":item["id"],"item_id":item["item_id"],
               "name":item["item_name"],"qty":qty,"unit":item["unit"],
               "price":item["unit_price"],"rate":item["vat_rate"],"cost":unit_cost,
               "sub":base,"vat":tax,"total":round(base+tax,2)})
            if item["item_id"]:
                record_inventory_movement(
                    request.form["return_date"],"مرتجع مبيعات",item["item_id"],warehouse_id,
                    qty,unit_cost,reference_type="SALES_RETURN",reference_id=return_id,
                    reference_no=no,notes=f"مرتجع للفاتورة {invoice['invoice_no']}"
                )
        post_sales_return_to_ledger(return_id)
        audit("CREATE","SALES_RETURN",f"إنشاء مرتجع مبيعات {no}")
        flash(f"تم إنشاء المرتجع {no} وإعادة المخزون وترحيل القيد","success")
        return redirect(url_for("sales_return_view",return_id=return_id))

    return render_template("sales_returns.html",
      invoices=rows("""SELECT i.id,i.invoice_no,i.invoice_date,c.name customer_name
                       FROM invoices i JOIN customers c ON c.id=i.customer_id
                       WHERE i.status='معتمدة' ORDER BY i.id DESC"""),
      warehouses=rows("SELECT id,code,name FROM warehouses WHERE active=1 ORDER BY code"),
      docs=rows("""SELECT r.*,i.invoice_no,c.name customer_name,j.journal_no
                   FROM sales_returns r JOIN invoices i ON i.id=r.invoice_id
                   JOIN customers c ON c.id=r.customer_id
                   LEFT JOIN journal_entries j ON j.id=r.journal_id ORDER BY r.id DESC"""))

@app.route("/sales/invoices/<int:invoice_id>/return-items")
@login_required
def sales_invoice_return_items(invoice_id):
    data=rows("""SELECT ii.id,ii.item_id,ii.item_name,ii.quantity,ii.unit,ii.unit_price,ii.vat_rate,
      ii.quantity-COALESCE((SELECT SUM(sri.quantity) FROM sales_return_items sri
      JOIN sales_returns sr ON sr.id=sri.return_id
      WHERE sri.invoice_item_id=ii.id AND sr.invoice_id=:invoice),0) returnable_qty
      FROM invoice_items ii WHERE ii.invoice_id=:invoice ORDER BY ii.id""",{"invoice":invoice_id})
    return {"items":[dict(x) for x in data]}

@app.route("/sales/returns/<int:return_id>")
@login_required
def sales_return_view(return_id):
    doc=row("""SELECT r.*,i.invoice_no,c.name customer_name,c.name_en customer_name_en,
                      w.name warehouse_name,j.journal_no
               FROM sales_returns r JOIN invoices i ON i.id=r.invoice_id
               JOIN customers c ON c.id=r.customer_id
               JOIN warehouses w ON w.id=r.warehouse_id
               LEFT JOIN journal_entries j ON j.id=r.journal_id
               WHERE r.id=:id""",{"id":return_id})
    if not doc:
        return "مرتجع المبيعات غير موجود",404
    items=rows("SELECT * FROM sales_return_items WHERE return_id=:id ORDER BY id",{"id":return_id})
    return render_template("sales_return_view.html",doc=doc,items=items)

@app.route("/sales/dashboard")
@login_required
def sales_dashboard():
    """Sales dashboard that remains available during schema upgrades."""
    summary = {
        "total_sales": 0,
        "total_vat": 0,
        "invoice_count": 0,
    }
    return_summary = {
        "total_returns": 0,
        "return_count": 0,
    }
    profit = {
        "net_sales": 0,
        "net_cost": 0,
        "gross_profit": 0,
    }
    monthly = []
    customers = []
    items = []
    dashboard_warnings = []

    try:
        summary_row = row("""
            SELECT
                COALESCE(SUM(total), 0) AS total_sales,
                COALESCE(SUM(vat), 0) AS total_vat,
                COUNT(*) AS invoice_count
            FROM invoices
            WHERE status = 'معتمدة'
        """)
        if summary_row:
            summary = summary_row
    except Exception as exc:
        db.session.rollback()
        dashboard_warnings.append(f"تعذر قراءة ملخص الفواتير: {exc}")

    try:
        if database_table_exists("sales_returns"):
            return_row = row("""
                SELECT
                    COALESCE(SUM(total), 0) AS total_returns,
                    COUNT(*) AS return_count
                FROM sales_returns
            """)
            if return_row:
                return_summary = return_row
    except Exception as exc:
        db.session.rollback()
        dashboard_warnings.append(f"تعذر قراءة مرتجعات المبيعات: {exc}")

    total_sales = float(summary.get("total_sales") or 0)
    total_returns = float(return_summary.get("total_returns") or 0)
    net_sales = round(total_sales - total_returns, 2)

    sales_cost = 0.0
    return_cost = 0.0

    try:
        if database_table_exists("sales_delivery_items"):
            sales_cost = float(safe_scalar("""
                SELECT COALESCE(SUM(
                    COALESCE(quantity, 0) * COALESCE(unit_cost, 0)
                ), 0)
                FROM sales_delivery_items
            """, default=0) or 0)
    except Exception:
        db.session.rollback()

    try:
        if (
            database_table_exists("sales_returns")
            and database_column_exists("sales_returns", "cost_total")
        ):
            return_cost = float(safe_scalar("""
                SELECT COALESCE(SUM(cost_total), 0)
                FROM sales_returns
            """, default=0) or 0)
        elif (
            database_table_exists("sales_return_items")
            and database_column_exists("sales_return_items", "unit_cost")
        ):
            return_cost = float(safe_scalar("""
                SELECT COALESCE(SUM(
                    COALESCE(quantity, 0) * COALESCE(unit_cost, 0)
                ), 0)
                FROM sales_return_items
            """, default=0) or 0)
    except Exception:
        db.session.rollback()

    net_cost = round(sales_cost - return_cost, 2)
    profit = {
        "net_sales": net_sales,
        "net_cost": net_cost,
        "gross_profit": round(net_sales - net_cost, 2),
    }

    try:
        monthly = rows("""
            SELECT
                TO_CHAR(invoice_date, 'YYYY-MM') AS month,
                COALESCE(SUM(total), 0) AS sales,
                COUNT(*) AS invoice_count
            FROM invoices
            WHERE status = 'معتمدة'
            GROUP BY TO_CHAR(invoice_date, 'YYYY-MM')
            ORDER BY month DESC
            LIMIT 12
        """)
    except Exception as exc:
        db.session.rollback()
        dashboard_warnings.append(f"تعذر قراءة المبيعات الشهرية: {exc}")

    try:
        customers = rows("""
            SELECT c.name, COALESCE(SUM(i.total), 0) AS total
            FROM invoices i
            JOIN customers c ON c.id = i.customer_id
            WHERE i.status = 'معتمدة'
            GROUP BY c.id, c.name
            ORDER BY total DESC
            LIMIT 10
        """)
    except Exception as exc:
        db.session.rollback()
        dashboard_warnings.append(f"تعذر قراءة أفضل العملاء: {exc}")

    try:
        items = rows("""
            SELECT
                ii.item_name,
                COALESCE(SUM(ii.quantity), 0) AS quantity,
                COALESCE(SUM(ii.line_total), 0) AS sales
            FROM invoice_items ii
            JOIN invoices i ON i.id = ii.invoice_id
            WHERE i.status = 'معتمدة'
            GROUP BY ii.item_name
            ORDER BY sales DESC
            LIMIT 10
        """)
    except Exception as exc:
        db.session.rollback()
        dashboard_warnings.append(f"تعذر قراءة أفضل الأصناف: {exc}")

    return render_template(
        "sales_dashboard.html",
        summary=summary,
        return_summary=return_summary,
        profit=profit,
        monthly=monthly,
        top_customers=customers,
        top_items=items,
        dashboard_warnings=dashboard_warnings,
    )

@app.route("/sales/reports")
@login_required
def sales_reports():
    date_from=request.args.get("date_from","")
    date_to=request.args.get("date_to","")
    customer_id=request.args.get("customer_id","")
    conditions=["i.status='معتمدة'"];params={}
    if date_from:
        conditions.append("i.invoice_date>=:date_from");params["date_from"]=date_from
    if date_to:
        conditions.append("i.invoice_date<=:date_to");params["date_to"]=date_to
    if customer_id:
        conditions.append("i.customer_id=:customer");params["customer"]=customer_id
    where=" AND ".join(conditions)
    data=rows(f"""SELECT i.*,c.name customer_name,
      COALESCE((SELECT SUM(di.quantity*di.unit_cost) FROM sales_delivery_items di
      JOIN sales_deliveries d ON d.id=di.delivery_id WHERE d.id=i.delivery_id),0) cost,
      i.subtotal-COALESCE((SELECT SUM(di.quantity*di.unit_cost) FROM sales_delivery_items di
      JOIN sales_deliveries d ON d.id=di.delivery_id WHERE d.id=i.delivery_id),0) gross_profit
      FROM invoices i JOIN customers c ON c.id=i.customer_id
      WHERE {where} ORDER BY i.invoice_date DESC,i.id DESC""",params)
    return render_template("sales_reports.html",rows=data,
      customers=rows("SELECT id,name FROM customers ORDER BY name"),
      date_from=date_from,date_to=date_to,customer_id=customer_id)

@app.route("/sales/reports.xlsx")
@login_required
def sales_reports_export():
    data=rows("""SELECT i.invoice_no,i.invoice_date,c.name customer_name,i.subtotal,i.vat,i.total,
      COALESCE((SELECT SUM(di.quantity*di.unit_cost) FROM sales_delivery_items di
      JOIN sales_deliveries d ON d.id=di.delivery_id WHERE d.id=i.delivery_id),0) cost
      FROM invoices i JOIN customers c ON c.id=i.customer_id
      WHERE i.status='معتمدة' ORDER BY i.invoice_date DESC,i.id DESC""")
    return xlsx_response("sales_profit_report.xlsx","المبيعات والأرباح",
      ["رقم الفاتورة","التاريخ","العميل","قبل الضريبة","الضريبة","الإجمالي","التكلفة","مجمل الربح"],
      [[r["invoice_no"],r["invoice_date"],r["customer_name"],r["subtotal"],r["vat"],
        r["total"],r["cost"],float(r["subtotal"])-float(r["cost"])] for r in data])


@app.route("/procurement")
@login_required
def procurement_center():
    stats = {
        "requisitions": row("SELECT COUNT(*) count FROM purchase_requisitions")["count"],
        "orders": row("SELECT COUNT(*) count FROM purchase_orders")["count"],
        "receipts": row("SELECT COUNT(*) count FROM goods_receipts")["count"],
        "supplier_invoices": row("SELECT COUNT(*) count FROM supplier_invoices")["count"],
    }
    return render_template("procurement_center.html",stats=stats)

@app.route("/purchase-requisitions",methods=["GET","POST"])
@login_required
def purchase_requisitions():
    if request.method=="POST":
        inventory_item_ids=request.form.getlist("inventory_item_id[]")
        item_names=request.form.getlist("item_name[]")
        quantities=request.form.getlist("quantity[]")
        units=request.form.getlist("unit[]")
        prices=request.form.getlist("estimated_price[]")
        descriptions=request.form.getlist("description[]")
        required_dates=request.form.getlist("required_date[]")
        supplier_ids=request.form.getlist("suggested_supplier_id[]")
        item_codes=request.form.getlist("item_code[]")
        items=[]; total=0
        for i,name in enumerate(item_names):
            if not name.strip(): continue
            inventory_item_id=((inventory_item_ids[i] if i<len(inventory_item_ids) else "") or None)
            inventory_item=row("SELECT id,sku,name,unit FROM inventory WHERE id=:id AND active=1",{"id":inventory_item_id}) if inventory_item_id else None
            if not inventory_item:
                flash(f"اختر صنف مخزون صحيحًا للسطر: {name}","danger")
                return redirect(request.url)
            qty=float(quantities[i] or 0); price=float(prices[i] or 0)
            if qty<=0: continue
            total += qty*price
            items.append({"name":inventory_item["name"],"qty":qty,"unit":normalize_item_unit(units[i] or inventory_item["unit"]),
                          "price":price,"description":descriptions[i],
                          "required_date":required_dates[i] or None,
                          "supplier_id":supplier_ids[i] or None,
                          "item_code":item_codes[i] or inventory_item["sku"],
                          "inventory_item_id":inventory_item["id"]})
        if not items:
            flash("أدخل صنفًا واحدًا على الأقل","danger")
            return redirect(url_for("purchase_requisitions"))
        no=next_document_number("purchase_requisitions","requisition_date","PR",request.form["requisition_date"])
        execute("""INSERT INTO purchase_requisitions(
          requisition_no,requisition_date,branch_id,cost_center_id,requested_by,
          department,priority,reason,notes,status,total_estimated,created_by,created_at)
          VALUES(:no,:dt,:branch,:cc,:requested,:department,:priority,:reason,
          :notes,:status,:total,:uid,:created)""",
          {"no":no,"dt":request.form["requisition_date"],
           "branch":request.form.get("branch_id") or None,
           "cc":request.form.get("cost_center_id") or None,
           "requested":request.form.get("requested_by",""),
           "department":request.form.get("department",""),
           "priority":request.form.get("priority","عادية"),
           "reason":request.form.get("reason",""),
           "notes":request.form.get("notes",""),
           "status":request.form.get("status","مسودة"),
           "total":round(total,2),"uid":session.get("user_id"),"created":datetime.now()})
        req_id=row("SELECT id FROM purchase_requisitions WHERE requisition_no=:n",{"n":no})["id"]
        for x in items:
            execute("""INSERT INTO purchase_requisition_items(
              requisition_id,inventory_item_id,item_code,item_name,description,quantity,unit,
              estimated_price,required_date,suggested_supplier_id)
              VALUES(:rid,:inventory_item,:code,:name,:des,:qty,:unit,:price,:rd,:sid)""",
              {"rid":req_id,"code":x["item_code"],"name":x["name"],"des":x["description"],
               "qty":x["qty"],"unit":x["unit"],"price":x["price"],
               "rd":x["required_date"],"sid":x["supplier_id"],"inventory_item":x["inventory_item_id"]})
        audit("CREATE","PURCHASE_REQUISITION",f"إنشاء طلب شراء {no}")
        flash(f"تم إنشاء طلب الشراء {no}","success")
        return redirect(url_for("purchase_requisition_view",req_id=req_id))
    reqs=rows("""SELECT pr.*,b.name branch_name,cc.name cost_center_name
                 FROM purchase_requisitions pr
                 LEFT JOIN branches b ON b.id=pr.branch_id
                 LEFT JOIN cost_centers cc ON cc.id=pr.cost_center_id
                 ORDER BY pr.requisition_date DESC,pr.id DESC""")
    return render_template("purchase_requisitions.html",reqs=reqs,
      branches=rows("SELECT * FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT * FROM cost_centers WHERE active=1 ORDER BY code"),
      suppliers=[dict(x) for x in rows("SELECT id,name,name_en,vat_number FROM suppliers ORDER BY name")],
      inventory_items=[dict(x) for x in rows("SELECT id,sku,name,unit,cost FROM inventory WHERE active=1 ORDER BY name")])

@app.route("/purchase-requisitions/<int:req_id>")
@login_required
def purchase_requisition_view(req_id):
    req=row("""SELECT pr.*,b.name branch_name,cc.name cost_center_name
               FROM purchase_requisitions pr
               LEFT JOIN branches b ON b.id=pr.branch_id
               LEFT JOIN cost_centers cc ON cc.id=pr.cost_center_id
               WHERE pr.id=:id""",{"id":req_id})
    if not req:return "طلب الشراء غير موجود",404
    items=rows("""SELECT i.*,s.name supplier_name FROM purchase_requisition_items i
                  LEFT JOIN suppliers s ON s.id=i.suggested_supplier_id
                  WHERE i.requisition_id=:id ORDER BY i.id""",{"id":req_id})
    return render_template("purchase_requisition_view.html",req=req,items=items,
                           approver=get_required_approver("طلب شراء",float(req["total_estimated"])))

@app.route("/purchase-requisitions/<int:req_id>/approve",methods=["POST"])
@login_required
def purchase_requisition_approve(req_id):
    execute("""UPDATE purchase_requisitions SET status='معتمد',
               approved_by=:u,approved_at=:t WHERE id=:id""",
            {"u":session.get("user_id"),"t":datetime.now(),"id":req_id})
    audit("APPROVE","PURCHASE_REQUISITION",f"اعتماد طلب شراء {req_id}")
    flash("تم اعتماد طلب الشراء","success")
    return redirect(url_for("purchase_requisition_view",req_id=req_id))


@app.route("/purchase-requisitions/<int:req_id>/open-items")
@login_required
def purchase_requisition_open_items(req_id):
    req=row("""SELECT pr.*,b.name branch_name,cc.name cost_center_name
               FROM purchase_requisitions pr
               LEFT JOIN branches b ON b.id=pr.branch_id
               LEFT JOIN cost_centers cc ON cc.id=pr.cost_center_id
               WHERE pr.id=:id""",{"id":req_id})
    if not req:
        return {"error":"طلب الشراء غير موجود"},404
    if req["status"] not in ("معتمد","تم إصدار أمر شراء جزئيًا"):
        return {"error":"طلب الشراء غير متاح للتحويل إلى أمر شراء"},400
    items=rows("""SELECT pri.id,pri.item_code,pri.item_name,pri.description,pri.quantity,COALESCE(pri.ordered_qty,0) ordered_qty,
                  quantity-COALESCE(ordered_qty,0) remaining_qty,unit,estimated_price,suggested_supplier_id
                  ,COALESCE(pri.inventory_item_id,(SELECT inv.id FROM inventory inv
                    WHERE (NULLIF(pri.item_code,'') IS NOT NULL AND inv.sku=pri.item_code)
                       OR LOWER(inv.name)=LOWER(pri.item_name)
                    ORDER BY CASE WHEN inv.sku=pri.item_code THEN 0 ELSE 1 END LIMIT 1)) inventory_item_id
                  FROM purchase_requisition_items pri
                  WHERE requisition_id=:id AND COALESCE(ordered_qty,0)<quantity ORDER BY pri.id""",{"id":req_id})
    return {"requisition":dict(req),"items":[dict(x) for x in items]}

@app.route("/purchase-orders",methods=["GET","POST"])
@login_required
def purchase_orders():
    if request.method=="POST":
        req_id=request.form.get("requisition_id") or None
        if req_id:
            req=row("SELECT status FROM purchase_requisitions WHERE id=:id",{"id":req_id})
            if not req or req["status"] not in ("معتمد","تم إصدار أمر شراء جزئيًا"):
                flash("لا يمكن إصدار أمر شراء من طلب غير معتمد","danger")
                return redirect(url_for("purchase_orders"))
        names=request.form.getlist("item_name[]")
        req_item_ids=request.form.getlist("requisition_item_id[]")
        qtys=request.form.getlist("quantity[]")
        prices=request.form.getlist("unit_price[]")
        vats=request.form.getlist("vat_rate[]")
        units=request.form.getlist("unit[]")
        codes=request.form.getlist("item_code[]")
        descriptions=request.form.getlist("description[]")
        inventory_item_ids=request.form.getlist("inventory_item_id[]")
        items=[]; subtotal=0; vat=0
        for i,n in enumerate(names):
            if not n.strip(): continue
            inventory_item_id=((inventory_item_ids[i] if i<len(inventory_item_ids) else "") or None)
            inventory_item=row("SELECT id,sku,name,unit FROM inventory WHERE id=:id AND active=1",{"id":inventory_item_id}) if inventory_item_id else None
            if not inventory_item:
                flash(f"اختر صنف مخزون صحيحًا للسطر: {n}","danger")
                return redirect(request.url)
            q=float(qtys[i] or 0); p=float(prices[i] or 0); vr=float(vats[i] or 0)
            if q<=0:continue
            base=q*p; tax=base*vr/100
            subtotal+=base; vat+=tax
            req_item_id = (req_item_ids[i] if i < len(req_item_ids) else "") or None
            if req_item_id:
                source_item=row("""SELECT id,quantity,COALESCE(ordered_qty,0) ordered_qty,requisition_id
                                   FROM purchase_requisition_items WHERE id=:id""",{"id":req_item_id})
                if not source_item or not req_id or int(source_item["requisition_id"]) != int(req_id):
                    flash("يوجد صنف غير مرتبط بطلب الشراء المحدد","danger")
                    return redirect(url_for("purchase_orders"))
                remaining=float(source_item["quantity"])-float(source_item["ordered_qty"] or 0)
                if q > remaining + 1e-9:
                    flash(f"كمية الصنف {n} تتجاوز الكمية المتبقية في طلب الشراء","danger")
                    return redirect(url_for("purchase_orders"))
            items.append({"name":inventory_item["name"],"qty":q,"price":p,"vat":vr,
                          "unit":normalize_item_unit(units[i] or inventory_item["unit"]),"code":codes[i] or inventory_item["sku"],"description":descriptions[i],
                          "requisition_item_id":req_item_id,
                          "inventory_item_id":inventory_item["id"]})
        if not items:
            flash("أدخل صنفًا واحدًا على الأقل","danger")
            return redirect(url_for("purchase_orders"))
        no=next_document_number("purchase_orders","po_date","PO",request.form["po_date"])
        execute("""INSERT INTO purchase_orders(
          po_no,po_date,requisition_id,supplier_id,branch_id,cost_center_id,
          payment_terms,delivery_terms,warehouse,warehouse_id,notes,status,subtotal,vat,total,
          created_by,created_at)
          VALUES(:no,:dt,:req,:supplier,:branch,:cc,:pay,:delivery,:warehouse,
          :warehouse_id,:notes,:status,:sub,:vat,:total,:uid,:created)""",
          {"no":no,"dt":request.form["po_date"],"req":req_id,
           "supplier":request.form["supplier_id"],"branch":request.form.get("branch_id") or None,
           "cc":request.form.get("cost_center_id") or None,
           "pay":request.form.get("payment_terms",""),
           "delivery":request.form.get("delivery_terms",""),
           "warehouse":request.form.get("warehouse_name",""),
           "warehouse_id":request.form.get("warehouse_id") or None,
           "notes":request.form.get("notes",""),
           "status":request.form.get("status","مسودة"),
           "sub":round(subtotal,2),"vat":round(vat,2),"total":round(subtotal+vat,2),
           "uid":session.get("user_id"),"created":datetime.now()})
        po_id=row("SELECT id FROM purchase_orders WHERE po_no=:n",{"n":no})["id"]
        for x in items:
            execute("""INSERT INTO purchase_order_items(
              po_id,requisition_item_id,inventory_item_id,item_code,item_name,description,quantity,unit,unit_price,vat_rate)
              VALUES(:po,:req_item,:inventory_item,:code,:name,:des,:qty,:unit,:price,:vat)""",
              {"po":po_id,"req_item":x["requisition_item_id"],"code":x["code"],"name":x["name"],"des":x["description"],
               "qty":x["qty"],"unit":x["unit"],"price":x["price"],"vat":x["vat"],
               "inventory_item":x["inventory_item_id"]})
            if x["requisition_item_id"]:
                execute("""UPDATE purchase_requisition_items
                           SET ordered_qty=COALESCE(ordered_qty,0)+:qty WHERE id=:id""",
                        {"qty":x["qty"],"id":x["requisition_item_id"]})
        if req_id:
            open_count=row("""SELECT COUNT(*) count FROM purchase_requisition_items
                              WHERE requisition_id=:id AND COALESCE(ordered_qty,0)<quantity""",{"id":req_id})["count"]
            execute("UPDATE purchase_requisitions SET status=:status WHERE id=:id",
                    {"status":"تم إصدار أمر شراء بالكامل" if open_count==0 else "تم إصدار أمر شراء جزئيًا","id":req_id})
        audit("CREATE","PURCHASE_ORDER",f"إنشاء أمر شراء {no}")
        flash(f"تم إنشاء أمر الشراء {no}","success")
        return redirect(url_for("purchase_order_view",po_id=po_id))
    return render_template("purchase_orders.html",
      orders=rows("""SELECT po.*,s.name supplier_name FROM purchase_orders po
                     JOIN suppliers s ON s.id=po.supplier_id
                     ORDER BY po.po_date DESC,po.id DESC"""),
      requisitions=rows("""SELECT id,requisition_no,total_estimated,status FROM purchase_requisitions
                           WHERE status IN ('معتمد','تم إصدار أمر شراء جزئيًا') ORDER BY requisition_date DESC"""),
      suppliers=rows("SELECT id,name,name_en,vat_number FROM suppliers ORDER BY name"),
      warehouses=rows("SELECT id,code,name FROM warehouses WHERE active=1 ORDER BY code,name"),
      inventory_items=[dict(x) for x in rows("SELECT id,sku,name,unit,cost FROM inventory WHERE active=1 ORDER BY name")],
      branches=rows("SELECT * FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT * FROM cost_centers WHERE active=1 ORDER BY code"))

@app.route("/purchase-orders/<int:po_id>")
@login_required
def purchase_order_view(po_id):
    po=row("""SELECT po.*,s.name supplier_name,s.name_en supplier_name_en,
              s.vat_number supplier_vat,b.name branch_name,cc.name cost_center_name
              ,pr.requisition_no source_requisition_no
              FROM purchase_orders po JOIN suppliers s ON s.id=po.supplier_id
              LEFT JOIN purchase_requisitions pr ON pr.id=po.requisition_id
              LEFT JOIN branches b ON b.id=po.branch_id
              LEFT JOIN cost_centers cc ON cc.id=po.cost_center_id WHERE po.id=:id""",{"id":po_id})
    if not po:return "أمر الشراء غير موجود",404
    items=rows("SELECT * FROM purchase_order_items WHERE po_id=:id ORDER BY id",{"id":po_id})
    company=row("SELECT * FROM settings WHERE id=1")
    return render_template("purchase_order_view.html",po=po,items=items,company=company)

@app.route("/purchase-orders/<int:po_id>/approve",methods=["POST"])
@login_required
def purchase_order_approve(po_id):
    execute("""UPDATE purchase_orders SET status='معتمد',
               approved_by=:u,approved_at=:t WHERE id=:id""",
            {"u":session.get("user_id"),"t":datetime.now(),"id":po_id})
    flash("تم اعتماد أمر الشراء","success")
    return redirect(url_for("purchase_order_view",po_id=po_id))

@app.route("/goods-receipts",methods=["GET","POST"])
@login_required
def goods_receipts():
    if request.method=="POST":
        po_id=int(request.form["po_id"])
        po=row("SELECT * FROM purchase_orders WHERE id=:id",{"id":po_id})
        if not po or po["status"] not in ("معتمد","مستلم جزئيًا"):
            flash("يجب اعتماد أمر الشراء أولًا","danger")
            return redirect(url_for("goods_receipts"))
        item_ids=request.form.getlist("po_item_id[]")
        qtys=request.form.getlist("received_qty[]")
        accepted=request.form.getlist("accepted_qty[]")
        rejected=request.form.getlist("rejected_qty[]")
        notes=request.form.getlist("line_notes[]")
        lines=[]
        for i,item_id in enumerate(item_ids):
            item=row("SELECT * FROM purchase_order_items WHERE id=:id",{"id":item_id})
            if not item:continue
            rq=float(qtys[i] or 0); aq=float(accepted[i] or 0); rej=float(rejected[i] or 0)
            remaining=float(item["quantity"])-float(item["received_qty"])
            if rq<0 or rq>remaining+1e-9:
                flash(f"الكمية المستلمة للصنف {item['item_name']} تتجاوز المتبقي","danger")
                return redirect(url_for("goods_receipts"))
            if abs((aq+rej)-rq)>0.001:
                flash("المقبول + المرفوض يجب أن يساوي المستلم","danger")
                return redirect(url_for("goods_receipts"))
            if rq>0:lines.append((item_id,rq,aq,rej,notes[i]))
        if not lines:
            flash("أدخل كمية استلام","danger")
            return redirect(url_for("goods_receipts"))
        no=next_document_number("goods_receipts","grn_date","GRN",request.form["grn_date"])
        warehouse_id=request.form.get("warehouse_id") or po.get("warehouse_id")
        warehouse_record=row("SELECT id,code,name FROM warehouses WHERE id=:id AND active=1",{"id":warehouse_id}) if warehouse_id else None
        if not warehouse_record:
            flash("اختر مستودعًا صحيحًا قبل الاستلام","danger")
            return redirect(url_for("goods_receipts",po_id=po_id))
        execute("""INSERT INTO goods_receipts(
          grn_no,grn_date,po_id,supplier_id,warehouse,warehouse_id,notes,status,created_by,created_at)
          VALUES(:no,:dt,:po,:supplier,:wh,:warehouse_id,:notes,'معتمد',:uid,:created)""",
          {"no":no,"dt":request.form["grn_date"],"po":po_id,"supplier":po["supplier_id"],
           "wh":warehouse_record["name"],"warehouse_id":warehouse_record["id"],"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        grn_id=row("SELECT id FROM goods_receipts WHERE grn_no=:n",{"n":no})["id"]
        for item_id,rq,aq,rej,note in lines:
            execute("""INSERT INTO goods_receipt_items(
              grn_id,po_item_id,received_qty,accepted_qty,rejected_qty,notes)
              VALUES(:grn,:item,:r,:a,:rej,:notes)""",
              {"grn":grn_id,"item":item_id,"r":rq,"a":aq,"rej":rej,"notes":note})
            execute("""UPDATE purchase_order_items SET received_qty=received_qty+:q
                       WHERE id=:id""",{"q":rq,"id":item_id})
            if aq>0:
                inv_item=row("""SELECT id,cost FROM inventory
                                WHERE (NULLIF(:sku,'') IS NOT NULL AND sku=:sku)
                                   OR LOWER(name)=LOWER(:name)
                                ORDER BY CASE WHEN sku=:sku THEN 0 ELSE 1 END LIMIT 1""",
                             {"sku":item["item_code"] or "","name":item["item_name"]})
                if not inv_item:
                    generated_sku=item["item_code"] or next_inventory_sku()
                    execute("""INSERT INTO inventory(sku,code,name,quantity,unit,cost,unit_cost,sale_price,reorder_level,active)
                               VALUES(:sku,:sku,:name,0,:unit,:cost,:cost,0,0,1)""",
                            {"sku":generated_sku,"name":item["item_name"],
                             "unit":normalize_item_unit(item["unit"]),"cost":item["unit_price"]})
                    inv_item=row("SELECT id,cost FROM inventory WHERE name=:name ORDER BY id DESC LIMIT 1",
                                 {"name":item["item_name"]})
                if warehouse_record:
                    record_inventory_movement(
                      request.form["grn_date"],"استلام",inv_item["id"],warehouse_record["id"],aq,
                      item["unit_price"],reference_type="GRN",reference_id=grn_id,
                      reference_no=no,notes=f"استلام أمر الشراء {po['po_no']}")

        remaining=row("""SELECT COUNT(*) count FROM purchase_order_items
                         WHERE po_id=:po AND received_qty<quantity""",{"po":po_id})["count"]
        execute("UPDATE purchase_orders SET status=:s WHERE id=:id",
                {"s":"مستلم بالكامل" if remaining==0 else "مستلم جزئيًا","id":po_id})
        audit("CREATE","GOODS_RECEIPT",f"إنشاء استلام {no}")
        flash(f"تم إنشاء استلام المواد {no}","success")
        return redirect(url_for("goods_receipt_view",grn_id=grn_id))
    return render_template("goods_receipts.html",
      pos=rows("""SELECT po.id,po.po_no,po.po_date,s.name supplier_name
                  FROM purchase_orders po JOIN suppliers s ON s.id=po.supplier_id
                  WHERE po.status IN ('معتمد','مستلم جزئيًا') ORDER BY po.po_date DESC"""),
      receipts=rows("""SELECT g.*,po.po_no,s.name supplier_name FROM goods_receipts g
                       JOIN purchase_orders po ON po.id=g.po_id
                       JOIN suppliers s ON s.id=g.supplier_id
                       ORDER BY g.grn_date DESC,g.id DESC"""))

@app.route("/purchase-orders/<int:po_id>/open-items")
@login_required
def purchase_order_open_items(po_id):
    po=row("SELECT * FROM purchase_orders WHERE id=:id",{"id":po_id})
    items=rows("""SELECT id,item_code,item_name,quantity,received_qty,unit
                  FROM purchase_order_items WHERE po_id=:id ORDER BY id""",{"id":po_id})
    return {"po":dict(po) if po else None,"items":[dict(x) for x in items]}

@app.route("/goods-receipts/<int:grn_id>")
@login_required
def goods_receipt_view(grn_id):
    grn=row("""SELECT g.*,po.po_no,s.name supplier_name FROM goods_receipts g
               JOIN purchase_orders po ON po.id=g.po_id
               JOIN suppliers s ON s.id=g.supplier_id WHERE g.id=:id""",{"id":grn_id})
    if not grn:return "الاستلام غير موجود",404
    items=rows("""SELECT gi.*,pi.item_code,pi.item_name,pi.unit
                  FROM goods_receipt_items gi JOIN purchase_order_items pi ON pi.id=gi.po_item_id
                  WHERE gi.grn_id=:id ORDER BY gi.id""",{"id":grn_id})
    return render_template("goods_receipt_view.html",grn=grn,items=items)

@app.route("/supplier-invoices",methods=["GET","POST"])
@login_required
def supplier_invoices():
    if request.method=="POST":
        subtotal=round(float(request.form["subtotal"] or 0),2)
        vat=round(float(request.form["vat"] or 0),2)
        total=round(subtotal+vat,2)
        internal=next_document_number("supplier_invoices","invoice_date","SI",request.form["invoice_date"])
        try:
            execute("""INSERT INTO supplier_invoices(
              supplier_invoice_no,internal_no,invoice_date,supplier_id,po_id,grn_id,
              cost_center_id,expense_or_inventory_account_id,subtotal,vat,total,status,
              posting_status,notes,created_by,created_at)
              VALUES(:supplier_no,:internal,:dt,:supplier,:po,:grn,:cc,:account,
              :subtotal,:vat,:total,:status,'غير مرحّل',:notes,:uid,:created)""",
              {"supplier_no":request.form["supplier_invoice_no"],"internal":internal,
               "dt":request.form["invoice_date"],"supplier":request.form["supplier_id"],
               "po":request.form.get("po_id") or None,"grn":request.form.get("grn_id") or None,
               "cc":request.form.get("cost_center_id") or None,
               "account":request.form["expense_or_inventory_account_id"],
               "subtotal":subtotal,"vat":vat,"total":total,
               "status":request.form.get("status","مسودة"),
               "notes":request.form.get("notes",""),
               "uid":session.get("user_id"),"created":datetime.now()})
        except Exception:
            db.session.rollback()
            flash("رقم فاتورة المورد مستخدم مسبقًا لهذا المورد","danger")
            return redirect(url_for("supplier_invoices"))
        inv_id=row("SELECT id FROM supplier_invoices WHERE internal_no=:n",{"n":internal})["id"]
        if request.form.get("status")=="معتمدة":
            try:
                post_supplier_invoice(inv_id)
                flash(f"تم حفظ وترحيل فاتورة المورد {internal}","success")
            except Exception as exc:
                db.session.rollback()
                execute("UPDATE supplier_invoices SET status='مسودة',posting_status='خطأ' WHERE id=:id",{"id":inv_id})
                flash(f"تم حفظها كمسودة ولم تُرحّل: {exc}","danger")
        else:flash(f"تم حفظ فاتورة المورد {internal} كمسودة","success")
        return redirect(url_for("supplier_invoice_view",invoice_id=inv_id))
    accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                     WHERE active=1 AND accepts_entries=1 ORDER BY account_code""")
    return render_template("supplier_invoices.html",
      invoices=rows("""SELECT si.*,s.name supplier_name,j.journal_no FROM supplier_invoices si
                       JOIN suppliers s ON s.id=si.supplier_id
                       LEFT JOIN journal_entries j ON j.id=si.journal_id
                       ORDER BY si.invoice_date DESC,si.id DESC"""),
      suppliers=rows("SELECT id,name,name_en FROM suppliers ORDER BY name"),
      pos=rows("SELECT id,po_no,supplier_id,total FROM purchase_orders ORDER BY po_date DESC"),
      grns=rows("SELECT id,grn_no,supplier_id,po_id FROM goods_receipts ORDER BY grn_date DESC"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"),
      accounts=accounts)

@app.route("/supplier-invoices/<int:invoice_id>")
@login_required
def supplier_invoice_view(invoice_id):
    inv=row("""SELECT si.*,s.name supplier_name,s.name_en supplier_name_en,
              s.vat_number supplier_vat,po.po_no,g.grn_no,a.account_code,a.account_name_ar,
              j.journal_no FROM supplier_invoices si
              JOIN suppliers s ON s.id=si.supplier_id
              LEFT JOIN purchase_orders po ON po.id=si.po_id
              LEFT JOIN goods_receipts g ON g.id=si.grn_id
              JOIN chart_of_accounts a ON a.id=si.expense_or_inventory_account_id
              LEFT JOIN journal_entries j ON j.id=si.journal_id WHERE si.id=:id""",{"id":invoice_id})
    if not inv:return "فاتورة المورد غير موجودة",404
    return render_template("supplier_invoice_view.html",inv=inv)

@app.route("/supplier-invoices/<int:invoice_id>/post",methods=["POST"])
@login_required
def supplier_invoice_post(invoice_id):
    try:
        post_supplier_invoice(invoice_id)
        flash("تم ترحيل فاتورة المورد","success")
    except Exception as exc:flash(str(exc),"danger")
    return redirect(url_for("supplier_invoice_view",invoice_id=invoice_id))




@app.route("/sales/tracking")
@login_required
def sales_tracking():
    query=request.args.get("q","").strip()
    conditions=[];params={}
    if query:
        conditions.append("""(q.quotation_no ILIKE :q OR o.order_no ILIKE :q
                           OR d.delivery_no ILIKE :q OR i.invoice_no ILIKE :q
                           OR c.name ILIKE :q)""")
        params["q"]=f"%{query}%"
    where=" WHERE "+" AND ".join(conditions) if conditions else ""
    docs=rows(f"""SELECT
      q.id quotation_id,q.quotation_no,q.status quotation_status,
      o.id order_id,o.order_no,o.status order_status,
      d.id delivery_id,d.delivery_no,d.status delivery_status,
      i.id invoice_id,i.invoice_no,i.status invoice_status,i.payment_status,
      c.name customer_name,
      COALESCE(i.total,o.total,q.total,0) total,
      COALESCE(i.invoice_date,d.delivery_date,o.order_date,q.quotation_date) document_date,
      COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                WHERE a.invoice_id=i.id),0) paid,
      COALESCE(i.total,0)-COALESCE((SELECT SUM(a.allocated_amount)
                FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0) outstanding
      FROM sales_quotations q
      LEFT JOIN sales_orders o ON o.quotation_id=q.id
      LEFT JOIN sales_deliveries d ON d.order_id=o.id
      LEFT JOIN invoices i ON i.delivery_id=d.id
      JOIN customers c ON c.id=q.customer_id
      {where}
      ORDER BY document_date DESC,q.id DESC""",params)

    direct_orders=rows(f"""SELECT
      NULL quotation_id,NULL quotation_no,NULL quotation_status,
      o.id order_id,o.order_no,o.status order_status,
      d.id delivery_id,d.delivery_no,d.status delivery_status,
      i.id invoice_id,i.invoice_no,i.status invoice_status,i.payment_status,
      c.name customer_name,
      COALESCE(i.total,o.total,0) total,
      COALESCE(i.invoice_date,d.delivery_date,o.order_date) document_date,
      COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                WHERE a.invoice_id=i.id),0) paid,
      COALESCE(i.total,0)-COALESCE((SELECT SUM(a.allocated_amount)
                FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0) outstanding
      FROM sales_orders o
      LEFT JOIN sales_deliveries d ON d.order_id=o.id
      LEFT JOIN invoices i ON i.delivery_id=d.id
      JOIN customers c ON c.id=o.customer_id
      WHERE o.quotation_id IS NULL
      ORDER BY document_date DESC,o.id DESC""")
    return render_template("sales_tracking.html",docs=list(docs)+list(direct_orders),q=query)

@app.route("/sales/timeline")
@login_required
def sales_timeline():
    invoice_id=request.args.get("invoice_id",type=int)
    delivery_id=request.args.get("delivery_id",type=int)
    order_id=request.args.get("order_id",type=int)
    quotation_id=request.args.get("quotation_id",type=int)
    timeline=sales_document_timeline(invoice_id,delivery_id,order_id,quotation_id)
    return render_template("sales_timeline.html",timeline=timeline)

@app.route("/invoices/<int:invoice_id>/commercial-data",methods=["GET","POST"])
@login_required
def invoice_commercial_data(invoice_id):
    inv=row("""SELECT i.*,c.name customer_name FROM invoices i
               JOIN customers c ON c.id=i.customer_id WHERE i.id=:id""",{"id":invoice_id})
    if not inv:
        return "الفاتورة غير موجودة",404
    if request.method=="POST":
        due_date=request.form.get("due_date") or None
        execute("""UPDATE invoices SET due_date=:due,payment_terms=:terms,
          payment_method=:method,sales_person=:sales_person,
          customer_reference=:customer_reference,notes=:notes
          WHERE id=:id""",
          {"due":due_date,"terms":request.form.get("payment_terms",""),
           "method":request.form.get("payment_method",""),
           "sales_person":request.form.get("sales_person",""),
           "customer_reference":request.form.get("customer_reference",""),
           "notes":request.form.get("notes",""),"id":invoice_id})
        audit("UPDATE","INVOICE",f"تحديث البيانات التجارية للفاتورة {inv['invoice_no']}")
        flash("تم تحديث بيانات الفاتورة","success")
        return redirect(url_for("invoice_view",invoice_id=invoice_id))
    return render_template("invoice_commercial_data.html",invoice=inv)


@app.route("/receivables",methods=["GET","POST"])
@login_required
def receivables():
    if request.method=="POST":
        customer_id=int(request.form["customer_id"])
        receipt_date=request.form["receipt_date"]
        cash_account_id=int(request.form["cash_bank_account_id"])
        invoice_ids=request.form.getlist("invoice_id[]")
        amounts=request.form.getlist("allocated_amount[]")
        allocations=[]
        for idx,invoice_id in enumerate(invoice_ids):
            amount=round(float(amounts[idx] or 0),2)
            if amount>0:
                allocations.append({"invoice_id":int(invoice_id),"amount":amount})
        total=round(sum(x["amount"] for x in allocations),2)
        if total<=0:
            flash("أدخل مبلغ سداد لفاتورة واحدة على الأقل","danger")
            return redirect(url_for("receivables"))

        settings_row=row("SELECT customer_account_id FROM settings WHERE id=1")
        if not settings_row or not settings_row["customer_account_id"]:
            flash("حدد حساب العملاء الافتراضي من الإعدادات أولًا","danger")
            return redirect(url_for("receivables"))

        try:
            ensure_open_period(receipt_date)
            voucher_no=next_treasury_number("قبض",receipt_date)
            execute("""INSERT INTO treasury_vouchers(
                voucher_no,voucher_type,voucher_date,party_type,customer_id,
                cash_bank_account_id,counter_account_id,amount,payment_method,
                reference,description,status,posting_status,created_by,created_at)
                VALUES(:no,'قبض',:dt,'عميل',:customer,:cash,:counter,:amount,:method,
                :reference,:description,'معتمد','غير مرحّل',:uid,:created)""",
                {"no":voucher_no,"dt":receipt_date,"customer":customer_id,
                 "cash":cash_account_id,"counter":settings_row["customer_account_id"],
                 "amount":total,"method":request.form.get("payment_method","تحويل بنكي"),
                 "reference":request.form.get("reference",""),
                 "description":request.form.get("description","تحصيل فواتير عميل"),
                 "uid":session.get("user_id"),"created":datetime.now()})
            voucher_id=row("SELECT id FROM treasury_vouchers WHERE voucher_no=:no",
                           {"no":voucher_no})["id"]
            post_treasury_voucher(voucher_id)
            allocate_customer_receipt(voucher_id,allocations)
            audit("CREATE","CUSTOMER_RECEIPT",f"تحصيل وربط الفواتير بسند {voucher_no}")
            flash(f"تم إنشاء وترحيل سند القبض {voucher_no} وربطه بالفواتير","success")
            return redirect(url_for("treasury_view",voucher_id=voucher_id))
        except Exception as exc:
            db.session.rollback()
            flash(str(exc),"danger")
            return redirect(url_for("receivables"))

    open_invoices=rows("""SELECT i.id,i.invoice_no,i.invoice_date,
      COALESCE(i.due_date,i.invoice_date) due_date,i.customer_id,c.name customer_name,
      i.total,COALESCE((SELECT SUM(a.allocated_amount)
        FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0) paid,
      i.total-COALESCE((SELECT SUM(a.allocated_amount)
        FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0) outstanding,
      i.payment_status
      FROM invoices i JOIN customers c ON c.id=i.customer_id
      WHERE i.status='معتمدة'
        AND i.total-COALESCE((SELECT SUM(a.allocated_amount)
          FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0)>0.005
      ORDER BY c.name,i.invoice_date,i.id""")
    accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                     WHERE active=1 AND accepts_entries=1
                     AND account_type='أصل' ORDER BY account_code""")
    return render_template("receivables.html",open_invoices=open_invoices,accounts=accounts,
      customers=rows("SELECT id,name,name_en FROM customers ORDER BY name"))

@app.route("/receivables/customer/<int:customer_id>/open-invoices")
@login_required
def customer_open_invoices(customer_id):
    data=rows("""SELECT i.id,i.invoice_no,i.invoice_date,
      COALESCE(i.due_date,i.invoice_date) due_date,i.total,
      COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                WHERE a.invoice_id=i.id),0) paid,
      i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                WHERE a.invoice_id=i.id),0) outstanding
      FROM invoices i WHERE i.customer_id=:customer AND i.status='معتمدة'
      AND i.total-COALESCE((SELECT SUM(a.allocated_amount)
                FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0)>0.005
      ORDER BY i.invoice_date,i.id""",{"customer":customer_id})
    return {"items":[dict(x) for x in data]}

@app.route("/receivables/aging")
@login_required
def receivables_aging():
    data=rows("""SELECT c.id,c.name,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date)<=30 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) bucket_0_30,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date) BETWEEN 31 AND 60 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) bucket_31_60,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date) BETWEEN 61 AND 90 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) bucket_61_90,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date)>90 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) bucket_over_90,
      SUM(i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0)) total_outstanding
      FROM invoices i JOIN customers c ON c.id=i.customer_id
      WHERE i.status='معتمدة'
      AND i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                            WHERE a.invoice_id=i.id),0)>0.005
      GROUP BY c.id,c.name ORDER BY total_outstanding DESC""")
    return render_template("receivables_aging.html",rows=data)

@app.route("/receivables/aging.xlsx")
@login_required
def receivables_aging_export():
    data=rows("""SELECT c.name,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date)<=30 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) b1,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date) BETWEEN 31 AND 60 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) b2,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date) BETWEEN 61 AND 90 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) b3,
      SUM(CASE WHEN CURRENT_DATE-COALESCE(i.due_date,i.invoice_date)>90 THEN
        i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                          WHERE a.invoice_id=i.id),0) ELSE 0 END) b4,
      SUM(i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                           WHERE a.invoice_id=i.id),0)) total
      FROM invoices i JOIN customers c ON c.id=i.customer_id
      WHERE i.status='معتمدة'
      AND i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                            WHERE a.invoice_id=i.id),0)>0.005
      GROUP BY c.id,c.name ORDER BY total DESC""")
    return xlsx_response("receivables_aging.xlsx","أعمار ديون العملاء",
      ["العميل","0-30 يوم","31-60 يوم","61-90 يوم","أكثر من 90 يوم","الإجمالي"],
      [[r["name"],r["b1"],r["b2"],r["b3"],r["b4"],r["total"]] for r in data])


@app.route("/treasury", methods=["GET","POST"])
@login_required
def treasury():
    if request.method == "POST":
        voucher_type = request.form["voucher_type"]
        voucher_date = request.form["voucher_date"]
        amount = round(float(request.form["amount"] or 0),2)
        if amount <= 0:
            flash("المبلغ يجب أن يكون أكبر من صفر","danger")
            return redirect(url_for("treasury"))

        try:
            ensure_open_period(voucher_date)
        except Exception as exc:
            flash(str(exc),"danger")
            return redirect(url_for("treasury"))

        voucher_no = next_treasury_number(voucher_type, voucher_date)
        execute("""INSERT INTO treasury_vouchers(
            voucher_no,voucher_type,voucher_date,party_type,customer_id,supplier_id,
            cash_bank_account_id,counter_account_id,amount,payment_method,reference,
            description,cost_center_id,status,posting_status,created_by,created_at)
            VALUES(:no,:type,:dt,:ptype,:cid,:sid,:cash,:counter,:amount,:method,
                   :ref,:des,:cc,:status,'غير مرحّل',:uid,:created)""",
            {"no":voucher_no,"type":voucher_type,"dt":voucher_date,
             "ptype":request.form.get("party_type",""),
             "cid":request.form.get("customer_id") or None,
             "sid":request.form.get("supplier_id") or None,
             "cash":request.form["cash_bank_account_id"],
             "counter":request.form["counter_account_id"],
             "amount":amount,"method":request.form.get("payment_method","نقدي"),
             "ref":request.form.get("reference",""),
             "des":request.form.get("description",""),
             "cc":request.form.get("cost_center_id") or None,
             "status":request.form.get("status","مسودة"),
             "uid":session.get("user_id"),"created":datetime.now()})
        voucher_id = row("SELECT id FROM treasury_vouchers WHERE voucher_no=:n",{"n":voucher_no})["id"]

        if request.form.get("status") == "معتمد":
            try:
                post_treasury_voucher(voucher_id)
                flash(f"تم حفظ وترحيل السند {voucher_no}","success")
            except Exception as exc:
                db.session.rollback()
                execute("UPDATE treasury_vouchers SET status='مسودة',posting_status='خطأ' WHERE id=:id",
                        {"id":voucher_id})
                flash(f"تم حفظ السند كمسودة ولم يُرحّل: {exc}","danger")
        else:
            flash(f"تم حفظ السند {voucher_no} كمسودة","success")

        if request.form.get("print_after_save") == "1":
            return redirect(url_for("treasury_view",voucher_id=voucher_id,print=1))
        return redirect(url_for("treasury_view",voucher_id=voucher_id))

    q = request.args.get("q","").strip()
    conditions=[]; params={}
    if q:
        conditions.append("""(tv.voucher_no ILIKE :q OR COALESCE(tv.reference,'') ILIKE :q
                           OR COALESCE(tv.description,'') ILIKE :q
                           OR COALESCE(c.name,'') ILIKE :q OR COALESCE(s.name,'') ILIKE :q)""")
        params["q"]=f"%{q}%"
    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    vouchers = rows(f"""SELECT tv.*,c.name customer_name,s.name supplier_name,
                        a.account_name_ar cash_account_name,b.account_name_ar counter_account_name,
                        j.journal_no
                        FROM treasury_vouchers tv
                        LEFT JOIN customers c ON c.id=tv.customer_id
                        LEFT JOIN suppliers s ON s.id=tv.supplier_id
                        JOIN chart_of_accounts a ON a.id=tv.cash_bank_account_id
                        JOIN chart_of_accounts b ON b.id=tv.counter_account_id
                        LEFT JOIN journal_entries j ON j.id=tv.journal_id
                        {where}
                        ORDER BY tv.voucher_date DESC,tv.id DESC""",params)
    accounts = rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                       WHERE active=1 AND accepts_entries=1 ORDER BY account_code""")
    return render_template("treasury.html",vouchers=vouchers,accounts=accounts,
        customers=rows("SELECT id,name,name_en,vat_number FROM customers ORDER BY name"),
        suppliers=rows("SELECT id,name,name_en,vat_number FROM suppliers ORDER BY name"),
        centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"),
        q=q)

@app.route("/treasury/<int:voucher_id>")
@login_required
def treasury_view(voucher_id):
    voucher = row("""SELECT tv.*,c.name customer_name,c.name_en customer_name_en,
                     s.name supplier_name,s.name_en supplier_name_en,
                     a.account_code cash_account_code,a.account_name_ar cash_account_name,
                     b.account_code counter_account_code,b.account_name_ar counter_account_name,
                     cc.code cost_center_code,cc.name cost_center_name,j.journal_no
                     FROM treasury_vouchers tv
                     LEFT JOIN customers c ON c.id=tv.customer_id
                     LEFT JOIN suppliers s ON s.id=tv.supplier_id
                     JOIN chart_of_accounts a ON a.id=tv.cash_bank_account_id
                     JOIN chart_of_accounts b ON b.id=tv.counter_account_id
                     LEFT JOIN cost_centers cc ON cc.id=tv.cost_center_id
                     LEFT JOIN journal_entries j ON j.id=tv.journal_id
                     WHERE tv.id=:id""",{"id":voucher_id})
    if not voucher:
        return "السند غير موجود",404
    company=row("SELECT * FROM settings WHERE id=1")
    return render_template("treasury_view.html",voucher=voucher,company=company,
                           print_lang=request.args.get("lang","ar"))

@app.route("/treasury/<int:voucher_id>/post", methods=["POST"])
@login_required
def treasury_post(voucher_id):
    try:
        post_treasury_voucher(voucher_id)
        flash("تم ترحيل السند محاسبيًا","success")
    except Exception as exc:
        flash(str(exc),"danger")
    return redirect(url_for("treasury_view",voucher_id=voucher_id))

@app.route("/treasury/<int:voucher_id>/delete", methods=["POST"])
@login_required
def treasury_delete(voucher_id):
    v=row("SELECT * FROM treasury_vouchers WHERE id=:id",{"id":voucher_id})
    if not v:
        return "السند غير موجود",404
    if v["journal_id"]:
        flash("لا يمكن حذف سند مرحّل. يجب إنشاء قيد عكسي.","danger")
    else:
        execute("DELETE FROM treasury_vouchers WHERE id=:id",{"id":voucher_id})
        audit("DELETE","TREASURY",f"حذف السند {v['voucher_no']}")
        flash("تم حذف السند","success")
    return redirect(url_for("treasury"))

@app.route("/treasury/export.xlsx")
@login_required
def treasury_export():
    data=rows("""SELECT voucher_no,voucher_type,voucher_date,party_type,amount,
                 payment_method,reference,description,status,posting_status
                 FROM treasury_vouchers ORDER BY voucher_date DESC,id DESC""")
    headers=["رقم السند","النوع","التاريخ","نوع الطرف","المبلغ","طريقة الدفع",
             "المرجع","البيان","الحالة","الترحيل"]
    keys=["voucher_no","voucher_type","voucher_date","party_type","amount",
          "payment_method","reference","description","status","posting_status"]
    return xlsx_response("treasury_vouchers.xlsx","سندات الخزينة",headers,
                         [[r.get(k) for k in keys] for r in data])


@app.route("/fiscal-periods",methods=["GET","POST"])
@login_required
def fiscal_periods():
    if request.method=="POST":
        pid=request.form.get("period_id",type=int); action=request.form.get("action")
        if pid and action in ("open","close"):
            status="مفتوحة" if action=="open" else "مغلقة"
            execute("UPDATE fiscal_periods SET status=:s WHERE id=:id",{"s":status,"id":pid})
            audit("PERIOD","FISCAL_PERIOD",f"الفترة {pid}: {status}")
        return redirect(url_for("fiscal_periods"))
    return render_template("fiscal_periods.html",periods=rows("""SELECT p.*,y.name fiscal_year_name
      FROM fiscal_periods p JOIN fiscal_years y ON y.id=p.fiscal_year_id ORDER BY p.start_date DESC"""))

@app.route("/invoices/<int:invoice_id>/post",methods=["POST"])
@login_required
def invoice_post(invoice_id):
    try:
        post_invoice_to_ledger(invoice_id)
        execute("UPDATE invoices SET status='معتمدة' WHERE id=:id",{"id":invoice_id})
        flash("تم ترحيل الفاتورة محاسبيًا","success")
    except Exception as exc: flash(str(exc),"danger")
    return redirect(url_for("invoice_view",invoice_id=invoice_id))

@app.route("/expenses/<int:expense_id>/post",methods=["POST"])
@login_required
def expense_post(expense_id):
    try:
        post_expense_to_ledger(expense_id)
        execute("UPDATE expenses SET status='معتمدة' WHERE id=:id",{"id":expense_id})
        flash("تم ترحيل المصروف محاسبيًا","success")
    except Exception as exc: flash(str(exc),"danger")
    return redirect(url_for("expenses"))




@app.route("/notifications")
@login_required
def notifications_center():
    notifications=[]
    for r in rows("""SELECT id,sku,name,quantity,reorder_level FROM inventory
                     WHERE active=1 AND quantity<=reorder_level
                     ORDER BY quantity,name LIMIT 100"""):
        notifications.append({"level":"warning","title":"مخزون منخفض",
          "message":f"{r['sku'] or ''} - {r['name']}، الرصيد {r['quantity']} وحد إعادة الطلب {r['reorder_level']}",
          "url":url_for("inventory_item_view",item_id=r["id"])})

    for r in rows("""SELECT i.id,i.invoice_no,c.name customer_name,
      i.total-COALESCE((SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
                        WHERE a.invoice_id=i.id),0) outstanding
      FROM invoices i JOIN customers c ON c.id=i.customer_id
      WHERE i.status='معتمدة' AND COALESCE(i.due_date,i.invoice_date)<CURRENT_DATE
      AND i.total-COALESCE((SELECT SUM(a.allocated_amount)
                           FROM invoice_payment_allocations a WHERE a.invoice_id=i.id),0)>0.005
      ORDER BY COALESCE(i.due_date,i.invoice_date) LIMIT 100"""):
        notifications.append({"level":"danger","title":"فاتورة عميل متأخرة",
          "message":f"{r['invoice_no']} - {r['customer_name']}، المتبقي {float(r['outstanding']):.2f}",
          "url":url_for("invoice_view",invoice_id=r["id"])})

    for r in rows("""SELECT id,journal_no,description FROM journal_entries
                     WHERE status<>'مرحّل' ORDER BY journal_date,id LIMIT 100"""):
        notifications.append({"level":"info","title":"قيد غير مرحّل",
          "message":f"{r['journal_no']} - {r['description'] or ''}",
          "url":url_for("journal_view",journal_id=r["id"])})

    return render_template("notifications_center.html",notifications=notifications)

@app.route("/system-health")
@login_required
def system_health():
    checks=[]
    try:
        db.session.execute(text("SELECT 1"))
        checks.append({"name":"قاعدة البيانات","status":"سليم","ok":True})
    except Exception as exc:
        checks.append({"name":"قاعدة البيانات","status":str(exc),"ok":False})
    checks += [
        {"name":"إصدار النظام","status":APP_VERSION,"ok":True},
        {"name":"المستخدمون","status":row("SELECT COUNT(*) c FROM users")["c"],"ok":True},
        {"name":"القيود","status":row("SELECT COUNT(*) c FROM journal_entries")["c"],"ok":True},
        {"name":"الفواتير","status":row("SELECT COUNT(*) c FROM invoices")["c"],"ok":True},
        {"name":"الأصناف","status":row("SELECT COUNT(*) c FROM inventory")["c"],"ok":True},
    ]
    return render_template("system_health.html",checks=checks)













@app.route("/bi")
@login_required
def bi_dashboard():
    return render_template("bi_dashboard.html",data=executive_dashboard_data())

@app.route("/bi/finance")
@login_required
def bi_finance():
    from_date=request.args.get("from_date","")
    to_date=request.args.get("to_date","")
    conditions=["j.status='مرحّل'"]
    params={}
    if from_date:
        conditions.append("j.journal_date>=:from_date")
        params["from_date"]=from_date
    if to_date:
        conditions.append("j.journal_date<=:to_date")
        params["to_date"]=to_date
    financial=bi_safe_rows(f"""
      SELECT a.account_type,
        COALESCE(SUM(l.debit),0) debit,
        COALESCE(SUM(l.credit),0) credit,
        COALESCE(SUM(l.debit-l.credit),0) balance
      FROM journal_entry_lines l
      JOIN journal_entries j ON j.id=l.journal_id
      JOIN chart_of_accounts a ON a.id=l.account_id
      WHERE {' AND '.join(conditions)}
      GROUP BY a.account_type ORDER BY a.account_type
    """,params)
    return render_template("bi_finance.html",rows=financial,
                           from_date=from_date,to_date=to_date)

@app.route("/bi/projects")
@login_required
def bi_projects():
    data=bi_safe_rows("""
      SELECT p.id,p.project_no,p.name,p.status,p.contract_value,
        COALESCE((SELECT SUM(pc.gross_work_value) FROM progress_certificates pc
          WHERE pc.project_id=p.id AND pc.status IN ('معتمد','مفوتر')),0) certified,
        COALESCE((SELECT SUM(pc.total) FROM progress_certificates pc
          WHERE pc.project_id=p.id AND pc.status IN ('معتمد','مفوتر')),0) revenue,
        COALESCE((SELECT SUM(pe.amount) FROM project_cost_entries pe
          WHERE pe.project_id=p.id),0) cost
      FROM projects p ORDER BY p.project_no
    """)
    return render_template("bi_projects.html",rows=data)

@app.route("/bi/hr")
@login_required
def bi_hr():
    departments_data=bi_safe_rows("""
      SELECT COALESCE(d.name,'بدون قسم') department_name,
        COUNT(e.id) employee_count,
        COALESCE(SUM(e.basic_salary),0) basic_payroll
      FROM employees e
      LEFT JOIN departments d ON d.id=e.department_id
      WHERE e.active=1
      GROUP BY d.id,d.name ORDER BY employee_count DESC
    """)
    alerts=bi_safe_rows("""
      SELECT e.id,e.employee_no,e.name,
        LEAST(
          COALESCE(e.iqama_expiry,DATE '9999-12-31'),
          COALESCE(e.passport_expiry,DATE '9999-12-31'),
          COALESCE(e.medical_insurance_expiry,DATE '9999-12-31'),
          COALESCE(e.contract_end_date,DATE '9999-12-31')
        ) nearest_expiry
      FROM employees e WHERE e.active=1
      ORDER BY nearest_expiry LIMIT 20
    """)
    return render_template("bi_hr.html",departments=departments_data,alerts=alerts)

@app.route("/bi/export.xlsx")
@login_required
def bi_export():
    data=executive_dashboard_data()
    output=[
      ["إجمالي المبيعات",data["sales"].get("total_sales",0)],
      ["إجمالي المشتريات",data["purchases"].get("total_purchases",0)],
      ["الفرق الإجمالي",data["gross_profit"]],
      ["الذمم المدينة",data["receivables"].get("outstanding",0)],
      ["الذمم الدائنة",data["payables"].get("outstanding",0)],
      ["رصيد الخزينة",data["cash"].get("cash_balance",0)],
      ["قيمة المخزون",data["inventory"].get("inventory_value",0)],
      ["المشاريع النشطة",data["projects"].get("active_projects",0)],
      ["الموظفون النشطون",data["hr"].get("active_employees",0)],
      ["طلبات الاعتماد المعلقة",data["pending_approvals"].get("count",0)],
    ]
    return xlsx_response("executive_dashboard.xlsx","المؤشرات التنفيذية",
                         ["المؤشر","القيمة"],output)


@app.route("/documents")
@login_required
def documents_center():
    q=request.args.get("q","").strip()
    category=request.args.get("category_id","").strip()
    entity_type=request.args.get("entity_type","").strip()
    status=request.args.get("status","").strip()
    conditions=[];params={}
    if q:
        conditions.append("""(d.document_no ILIKE :q OR d.title ILIKE :q
                           OR d.entity_name ILIKE :q OR d.keywords ILIKE :q)""")
        params["q"]=f"%{q}%"
    if category:
        conditions.append("d.category_id=:category")
        params["category"]=int(category)
    if entity_type:
        conditions.append("d.entity_type=:entity_type")
        params["entity_type"]=entity_type
    if status:
        conditions.append("d.status=:status")
        params["status"]=status
    where=" WHERE "+" AND ".join(conditions) if conditions else ""
    docs=rows(f"""SELECT d.*,c.name category_name,u.username uploaded_by_name
                  FROM documents_archive d
                  LEFT JOIN document_categories c ON c.id=d.category_id
                  LEFT JOIN users u ON u.id=d.uploaded_by
                  {where}
                  ORDER BY d.uploaded_at DESC,d.id DESC""",params)
    stats={
      "documents":row("SELECT COUNT(*) c FROM documents_archive")["c"],
      "expiring":row("""SELECT COUNT(*) c FROM documents_archive
                        WHERE expiry_date BETWEEN CURRENT_DATE
                        AND CURRENT_DATE+INTERVAL '60 days'""")["c"],
      "expired":row("""SELECT COUNT(*) c FROM documents_archive
                       WHERE expiry_date<CURRENT_DATE""")["c"],
      "versions":row("SELECT COUNT(*) c FROM document_versions")["c"],
    }
    return render_template("documents_center.html",documents=docs,stats=stats,q=q,
      selected_category=category,selected_entity_type=entity_type,
      selected_status=status,
      categories=rows("SELECT id,name FROM document_categories WHERE active=1 ORDER BY name"),
      entity_types=DMS_ENTITY_TYPES)

@app.route("/documents/categories",methods=["GET","POST"])
@login_required
def document_categories():
    if request.method=="POST":
        execute("""INSERT INTO document_categories(code,name,retention_years,
          requires_expiry,active) VALUES(:code,:name,:years,:expiry,1)
          ON CONFLICT(code) DO UPDATE SET name=EXCLUDED.name,
          retention_years=EXCLUDED.retention_years,
          requires_expiry=EXCLUDED.requires_expiry""",
          {"code":request.form["code"].upper().strip(),
           "name":request.form["name"],
           "years":int(request.form.get("retention_years") or 5),
           "expiry":1 if request.form.get("requires_expiry") else 0})
        flash("تم حفظ تصنيف المستند","success")
        return redirect(url_for("document_categories"))
    return render_template("document_categories.html",
      rows=rows("SELECT * FROM document_categories ORDER BY code"))

@app.route("/documents/new",methods=["GET","POST"])
@login_required
def document_new():
    if request.method=="POST":
        no=next_document_number()
        execute("""INSERT INTO documents_archive(document_no,title,category_id,
          entity_type,entity_id,entity_name,issue_date,expiry_date,version_no,
          status,confidentiality,file_name,file_url,keywords,notes,uploaded_by,
          uploaded_at)
          VALUES(:no,:title,:category,:entity_type,:entity_id,:entity_name,
          :issue,:expiry,:version,:status,:confidentiality,:file_name,:file_url,
          :keywords,:notes,:user,:dt)""",
          {"no":no,"title":request.form["title"],
           "category":request.form.get("category_id") or None,
           "entity_type":request.form.get("entity_type","عام"),
           "entity_id":request.form.get("entity_id") or None,
           "entity_name":request.form.get("entity_name",""),
           "issue":request.form.get("issue_date") or None,
           "expiry":request.form.get("expiry_date") or None,
           "version":request.form.get("version_no","1.0"),
           "status":request.form.get("status","ساري"),
           "confidentiality":request.form.get("confidentiality","داخلي"),
           "file_name":request.form.get("file_name",""),
           "file_url":request.form.get("file_url",""),
           "keywords":request.form.get("keywords",""),
           "notes":request.form.get("notes",""),
           "user":session.get("user_id"),"dt":datetime.now()})
        document_id=row("SELECT id FROM documents_archive WHERE document_no=:no",{"no":no})["id"]
        execute("""INSERT INTO document_versions(document_id,version_no,file_name,
          file_url,change_notes,uploaded_by,uploaded_at)
          VALUES(:document,:version,:file_name,:file_url,'الإصدار الأول',:user,:dt)""",
          {"document":document_id,"version":request.form.get("version_no","1.0"),
           "file_name":request.form.get("file_name",""),
           "file_url":request.form.get("file_url",""),
           "user":session.get("user_id"),"dt":datetime.now()})
        audit("CREATE","DOCUMENT",f"إنشاء المستند {no}")
        flash(f"تم أرشفة المستند {no}","success")
        return redirect(url_for("document_view",document_id=document_id))
    return render_template("document_form.html",
      categories=rows("SELECT id,name FROM document_categories WHERE active=1 ORDER BY name"),
      entity_types=DMS_ENTITY_TYPES)

@app.route("/documents/<int:document_id>")
@login_required
def document_view(document_id):
    doc=row("""SELECT d.*,c.name category_name,u.username uploaded_by_name
               FROM documents_archive d
               LEFT JOIN document_categories c ON c.id=d.category_id
               LEFT JOIN users u ON u.id=d.uploaded_by
               WHERE d.id=:id""",{"id":document_id})
    if not doc:
        return "المستند غير موجود",404
    log_document_access(document_id,"عرض")
    return render_template("document_view.html",document=doc,
      versions=rows("""SELECT v.*,u.username uploaded_by_name
                       FROM document_versions v
                       LEFT JOIN users u ON u.id=v.uploaded_by
                       WHERE v.document_id=:id
                       ORDER BY v.uploaded_at DESC,v.id DESC""",{"id":document_id}),
      access_logs=rows("""SELECT l.*,u.username
                          FROM document_access_logs l
                          LEFT JOIN users u ON u.id=l.user_id
                          WHERE l.document_id=:id
                          ORDER BY l.action_at DESC LIMIT 100""",{"id":document_id}))

@app.route("/documents/<int:document_id>/version",methods=["POST"])
@login_required
def document_add_version(document_id):
    doc=row("SELECT * FROM documents_archive WHERE id=:id",{"id":document_id})
    if not doc:
        return "المستند غير موجود",404
    version=request.form["version_no"]
    execute("""INSERT INTO document_versions(document_id,version_no,file_name,
      file_url,change_notes,uploaded_by,uploaded_at)
      VALUES(:document,:version,:file_name,:file_url,:notes,:user,:dt)""",
      {"document":document_id,"version":version,
       "file_name":request.form.get("file_name",""),
       "file_url":request.form.get("file_url",""),
       "notes":request.form.get("change_notes",""),
       "user":session.get("user_id"),"dt":datetime.now()})
    execute("""UPDATE documents_archive SET version_no=:version,
      file_name=:file_name,file_url=:file_url WHERE id=:id""",
      {"version":version,"file_name":request.form.get("file_name",""),
       "file_url":request.form.get("file_url",""),"id":document_id})
    log_document_access(document_id,"إضافة إصدار")
    flash("تمت إضافة إصدار جديد","success")
    return redirect(url_for("document_view",document_id=document_id))

@app.route("/documents/<int:document_id>/status",methods=["POST"])
@login_required
def document_status_update(document_id):
    execute("""UPDATE documents_archive SET status=:status WHERE id=:id""",
            {"status":request.form["status"],"id":document_id})
    log_document_access(document_id,"تحديث الحالة")
    flash("تم تحديث حالة المستند","success")
    return redirect(url_for("document_view",document_id=document_id))

@app.route("/documents/report")
@login_required
def documents_report():
    data=rows("""SELECT d.document_no,d.title,c.name category_name,d.entity_type,
      d.entity_name,d.issue_date,d.expiry_date,d.version_no,d.status,
      d.confidentiality,d.file_url
      FROM documents_archive d
      LEFT JOIN document_categories c ON c.id=d.category_id
      ORDER BY d.document_no""")
    return render_template("documents_report.html",rows=data)

@app.route("/documents/report.xlsx")
@login_required
def documents_report_export():
    data=rows("""SELECT d.document_no,d.title,c.name category_name,d.entity_type,
      d.entity_name,d.issue_date,d.expiry_date,d.version_no,d.status,
      d.confidentiality,d.file_url
      FROM documents_archive d
      LEFT JOIN document_categories c ON c.id=d.category_id
      ORDER BY d.document_no""")
    return xlsx_response("documents_archive.xlsx","أرشيف المستندات",
      ["رقم المستند","العنوان","التصنيف","نوع الارتباط","الجهة",
       "تاريخ الإصدار","تاريخ الانتهاء","الإصدار","الحالة","السرية","الرابط"],
      [[r["document_no"],r["title"],r["category_name"],r["entity_type"],
        r["entity_name"],r["issue_date"],r["expiry_date"],r["version_no"],
        r["status"],r["confidentiality"],r["file_url"]] for r in data])


@app.route("/contracts")
@login_required
def contracts_center():
    stats={
      "subcontractors":row("SELECT COUNT(*) c FROM subcontractors WHERE active=1")["c"],
      "contracts":row("SELECT COUNT(*) c FROM subcontract_contracts WHERE status='ساري'")["c"],
      "variations":row("SELECT COUNT(*) c FROM variation_orders WHERE status IN ('مسودة','قيد الاعتماد')")["c"],
      "extensions":row("SELECT COUNT(*) c FROM contract_extensions WHERE status IN ('مسودة','قيد الاعتماد')")["c"],
    }
    return render_template("contracts_center.html",stats=stats,
      contracts=rows("""SELECT sc.*,p.project_no,p.name project_name,s.name subcontractor_name,
        COALESCE((SELECT SUM(c.total) FROM subcontract_certificates c
                  WHERE c.contract_id=sc.id AND c.status IN ('معتمد','مفوتر')),0) certified
        FROM subcontract_contracts sc
        JOIN projects p ON p.id=sc.project_id
        JOIN subcontractors s ON s.id=sc.subcontractor_id
        ORDER BY sc.id DESC LIMIT 20"""),
      variations=rows("""SELECT v.*,p.project_no,p.name project_name
                         FROM variation_orders v JOIN projects p ON p.id=v.project_id
                         ORDER BY v.id DESC LIMIT 20"""))

@app.route("/contracts/subcontractors",methods=["GET","POST"])
@login_required
def subcontractors():
    if request.method=="POST":
        code=request.form.get("code") or next_contracting_number("subcontractors","SUB")
        execute("""INSERT INTO subcontractors(code,name,name_en,tax_no,
          commercial_registration,phone,email,address,specialty,active,
          created_by,created_at)
          VALUES(:code,:name,:name_en,:tax,:cr,:phone,:email,:address,:specialty,
          1,:uid,:created)
          ON CONFLICT(code) DO UPDATE SET name=EXCLUDED.name,
          phone=EXCLUDED.phone,email=EXCLUDED.email,specialty=EXCLUDED.specialty""",
          {"code":code,"name":request.form["name"],
           "name_en":request.form.get("name_en",""),
           "tax":request.form.get("tax_no",""),
           "cr":request.form.get("commercial_registration",""),
           "phone":request.form.get("phone",""),
           "email":request.form.get("email",""),
           "address":request.form.get("address",""),
           "specialty":request.form.get("specialty",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash("تم حفظ المقاول الباطن","success")
        return redirect(url_for("subcontractors"))
    return render_template("subcontractors.html",
      rows=rows("SELECT * FROM subcontractors ORDER BY name"))

@app.route("/contracts/subcontracts",methods=["GET","POST"])
@login_required
def subcontract_contracts():
    if request.method=="POST":
        no=request.form.get("contract_no") or next_contracting_number("subcontract_contracts","SC")
        execute("""INSERT INTO subcontract_contracts(contract_no,project_id,
          subcontractor_id,contract_date,start_date,end_date,contract_value,
          retention_rate,advance_amount,tax_rate,status,scope_of_work,notes,
          created_by,created_at)
          VALUES(:no,:project,:subcontractor,:contract_date,:start,:end,:value,
          :retention,:advance,:tax,:status,:scope,:notes,:uid,:created)""",
          {"no":no,"project":request.form["project_id"],
           "subcontractor":request.form["subcontractor_id"],
           "contract_date":request.form["contract_date"],
           "start":request.form.get("start_date") or None,
           "end":request.form.get("end_date") or None,
           "value":float(request.form.get("contract_value") or 0),
           "retention":float(request.form.get("retention_rate") or 0),
           "advance":float(request.form.get("advance_amount") or 0),
           "tax":float(request.form.get("tax_rate") or 15),
           "status":request.form.get("status","ساري"),
           "scope":request.form.get("scope_of_work",""),
           "notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash("تم حفظ عقد المقاول الباطن","success")
        return redirect(url_for("subcontract_contracts"))
    return render_template("subcontract_contracts.html",
      contracts=rows("""SELECT sc.*,p.project_no,p.name project_name,
        s.name subcontractor_name FROM subcontract_contracts sc
        JOIN projects p ON p.id=sc.project_id
        JOIN subcontractors s ON s.id=sc.subcontractor_id
        ORDER BY sc.id DESC"""),
      projects=rows("SELECT id,project_no,name FROM projects ORDER BY project_no"),
      subcontractors=rows("SELECT id,code,name FROM subcontractors WHERE active=1 ORDER BY name"))

@app.route("/contracts/subcontracts/<int:contract_id>")
@login_required
def subcontract_contract_view(contract_id):
    contract=row("""SELECT sc.*,p.project_no,p.name project_name,
      s.name subcontractor_name,s.tax_no,s.phone
      FROM subcontract_contracts sc
      JOIN projects p ON p.id=sc.project_id
      JOIN subcontractors s ON s.id=sc.subcontractor_id
      WHERE sc.id=:id""",{"id":contract_id})
    if not contract:
        return "العقد غير موجود",404
    return render_template("subcontract_contract_view.html",
      contract=contract,summary=subcontract_contract_summary(contract_id),
      boq=rows("""SELECT * FROM subcontract_boq_items
                  WHERE contract_id=:id ORDER BY id""",{"id":contract_id}),
      certificates=rows("""SELECT * FROM subcontract_certificates
                           WHERE contract_id=:id ORDER BY id DESC""",{"id":contract_id}))

@app.route("/contracts/subcontracts/<int:contract_id>/boq",methods=["POST"])
@login_required
def subcontract_boq_add(contract_id):
    qty=float(request.form.get("quantity") or 0)
    rate=float(request.form.get("unit_rate") or 0)
    execute("""INSERT INTO subcontract_boq_items(contract_id,item_code,description,
      unit,quantity,unit_rate,total_value)
      VALUES(:contract,:code,:description,:unit,:qty,:rate,:total)
      ON CONFLICT(contract_id,item_code) DO UPDATE SET
      description=EXCLUDED.description,unit=EXCLUDED.unit,
      quantity=EXCLUDED.quantity,unit_rate=EXCLUDED.unit_rate,
      total_value=EXCLUDED.total_value""",
      {"contract":contract_id,"code":request.form["item_code"],
       "description":request.form["description"],
       "unit":request.form.get("unit","وحدة"),"qty":qty,"rate":rate,
       "total":round(qty*rate,2)})
    flash("تم حفظ بند العقد","success")
    return redirect(url_for("subcontract_contract_view",contract_id=contract_id))

@app.route("/contracts/subcontracts/<int:contract_id>/certificate",methods=["GET","POST"])
@login_required
def subcontract_certificate_new(contract_id):
    contract=row("""SELECT sc.*,p.project_no,p.name project_name,
      s.name subcontractor_name FROM subcontract_contracts sc
      JOIN projects p ON p.id=sc.project_id
      JOIN subcontractors s ON s.id=sc.subcontractor_id
      WHERE sc.id=:id""",{"id":contract_id})
    if request.method=="POST":
        no=next_contracting_number("subcontract_certificates","SPC")
        item_ids=request.form.getlist("boq_item_id[]")
        current_qtys=request.form.getlist("current_qty[]")
        prepared=[];gross=0
        for idx,item_id in enumerate(item_ids):
            item=row("""SELECT * FROM subcontract_boq_items
                        WHERE id=:id AND contract_id=:contract""",
                     {"id":item_id,"contract":contract_id})
            if not item:
                continue
            current=float(current_qtys[idx] or 0)
            previous=float(item["cumulative_qty"] or 0)
            cumulative=previous+current
            if cumulative > float(item["quantity"] or 0)+0.0001:
                flash(f"الكمية التراكمية للبند {item['item_code']} تجاوزت كمية العقد","danger")
                return redirect(url_for("subcontract_certificate_new",contract_id=contract_id))
            current_value=round(current*float(item["unit_rate"] or 0),2)
            prepared.append((item,previous,current,cumulative,current_value))
            gross+=current_value

        retention_rate=float(contract["retention_rate"] or 0)
        retention=round(gross*retention_rate/100,2)
        advance_recovery=float(request.form.get("advance_recovery") or 0)
        other=float(request.form.get("other_deductions") or 0)
        subtotal=round(gross-retention-advance_recovery-other,2)
        vat=round(subtotal*float(contract["tax_rate"] or 15)/100,2)
        total=round(subtotal+vat,2)

        execute("""INSERT INTO subcontract_certificates(certificate_no,contract_id,
          certificate_date,period_from,period_to,gross_value,retention_amount,
          advance_recovery,other_deductions,subtotal,vat,total,status,notes,
          created_by,created_at)
          VALUES(:no,:contract,:dt,:from,:to,:gross,:retention,:advance,:other,
          :subtotal,:vat,:total,'مسودة',:notes,:uid,:created)""",
          {"no":no,"contract":contract_id,"dt":request.form["certificate_date"],
           "from":request.form.get("period_from") or None,
           "to":request.form.get("period_to") or None,
           "gross":round(gross,2),"retention":retention,
           "advance":advance_recovery,"other":other,"subtotal":subtotal,
           "vat":vat,"total":total,"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        cert_id=row("SELECT id FROM subcontract_certificates WHERE certificate_no=:no",{"no":no})["id"]
        for item,previous,current,cumulative,current_value in prepared:
            execute("""INSERT INTO subcontract_certificate_items(certificate_id,
              boq_item_id,previous_qty,current_qty,cumulative_qty,unit_rate,current_value)
              VALUES(:cert,:item,:previous,:current,:cumulative,:rate,:value)""",
              {"cert":cert_id,"item":item["id"],"previous":previous,
               "current":current,"cumulative":cumulative,
               "rate":item["unit_rate"],"value":current_value})
            execute("""UPDATE subcontract_boq_items SET cumulative_qty=:qty
                       WHERE id=:id""",{"qty":cumulative,"id":item["id"]})
        flash(f"تم إنشاء مستخلص المقاول {no}","success")
        return redirect(url_for("subcontract_certificate_view",certificate_id=cert_id))
    return render_template("subcontract_certificate_form.html",
      contract=contract,
      boq=rows("""SELECT * FROM subcontract_boq_items
                  WHERE contract_id=:id ORDER BY id""",{"id":contract_id}))

@app.route("/contracts/subcontract-certificates/<int:certificate_id>")
@login_required
def subcontract_certificate_view(certificate_id):
    cert=row("""SELECT c.*,sc.contract_no,p.project_no,p.name project_name,
      s.name subcontractor_name
      FROM subcontract_certificates c
      JOIN subcontract_contracts sc ON sc.id=c.contract_id
      JOIN projects p ON p.id=sc.project_id
      JOIN subcontractors s ON s.id=sc.subcontractor_id
      WHERE c.id=:id""",{"id":certificate_id})
    if not cert:
        return "المستخلص غير موجود",404
    return render_template("subcontract_certificate_view.html",cert=cert,
      items=rows("""SELECT i.*,b.item_code,b.description,b.unit,b.quantity contract_qty
                    FROM subcontract_certificate_items i
                    JOIN subcontract_boq_items b ON b.id=i.boq_item_id
                    WHERE i.certificate_id=:id ORDER BY i.id""",{"id":certificate_id}))

@app.route("/contracts/subcontract-certificates/<int:certificate_id>/approve",methods=["POST"])
@login_required
def subcontract_certificate_approve(certificate_id):
    execute("""UPDATE subcontract_certificates SET status='معتمد'
               WHERE id=:id""",{"id":certificate_id})
    flash("تم اعتماد مستخلص المقاول","success")
    return redirect(url_for("subcontract_certificate_view",certificate_id=certificate_id))

@app.route("/contracts/variations",methods=["GET","POST"])
@login_required
def variation_orders():
    if request.method=="POST":
        no=request.form.get("variation_no") or next_contracting_number("variation_orders","VO")
        execute("""INSERT INTO variation_orders(variation_no,project_id,contract_id,
          variation_date,title,description,reason,value,time_extension_days,status,
          created_by,created_at)
          VALUES(:no,:project,:contract,:dt,:title,:description,:reason,:value,
          :days,'مسودة',:uid,:created)""",
          {"no":no,"project":request.form["project_id"],
           "contract":request.form.get("contract_id") or None,
           "dt":request.form["variation_date"],"title":request.form["title"],
           "description":request.form.get("description",""),
           "reason":request.form.get("reason",""),
           "value":float(request.form.get("value") or 0),
           "days":int(request.form.get("time_extension_days") or 0),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash(f"تم إنشاء أمر التغيير {no}","success")
        return redirect(url_for("variation_orders"))
    return render_template("variation_orders.html",
      rows=rows("""SELECT v.*,p.project_no,p.name project_name,pc.contract_no
                   FROM variation_orders v
                   JOIN projects p ON p.id=v.project_id
                   LEFT JOIN project_contracts pc ON pc.id=v.contract_id
                   ORDER BY v.id DESC"""),
      projects=rows("SELECT id,project_no,name FROM projects ORDER BY project_no"),
      contracts=rows("SELECT id,contract_no,project_id FROM project_contracts ORDER BY contract_no"))

@app.route("/contracts/variations/<int:variation_id>/submit",methods=["POST"])
@login_required
def variation_submit(variation_id):
    variation=row("""SELECT * FROM variation_orders WHERE id=:id""",{"id":variation_id})
    if not variation:
        return "أمر التغيير غير موجود",404
    try:
        approval_id=submit_for_approval("أمر تغيير",variation["id"],
                                        variation["variation_no"],variation["value"],
                                        variation["title"])
        execute("""UPDATE variation_orders SET status='قيد الاعتماد',
                   approval_request_id=:approval WHERE id=:id""",
                {"approval":approval_id,"id":variation_id})
        flash("تم إرسال أمر التغيير للاعتماد","success")
    except Exception as exc:
        db.session.rollback()
        flash(str(exc),"danger")
    return redirect(url_for("variation_orders"))

@app.route("/contracts/extensions",methods=["GET","POST"])
@login_required
def contract_extensions():
    if request.method=="POST":
        no=request.form.get("extension_no") or next_contracting_number("contract_extensions","EXT")
        old_end=request.form.get("old_end_date") or None
        new_end=request.form["new_end_date"]
        days=0
        if old_end:
            days=(datetime.strptime(new_end,"%Y-%m-%d").date()-
                  datetime.strptime(old_end,"%Y-%m-%d").date()).days
        execute("""INSERT INTO contract_extensions(project_id,contract_id,extension_no,
          request_date,old_end_date,new_end_date,extension_days,reason,status,
          created_by,created_at)
          VALUES(:project,:contract,:no,:request_date,:old_end,:new_end,:days,
          :reason,'مسودة',:uid,:created)""",
          {"project":request.form["project_id"],
           "contract":request.form.get("contract_id") or None,
           "no":no,"request_date":request.form["request_date"],
           "old_end":old_end,"new_end":new_end,"days":days,
           "reason":request.form.get("reason",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash("تم حفظ طلب تمديد العقد","success")
        return redirect(url_for("contract_extensions"))
    return render_template("contract_extensions.html",
      rows=rows("""SELECT e.*,p.project_no,p.name project_name,pc.contract_no
                   FROM contract_extensions e
                   JOIN projects p ON p.id=e.project_id
                   LEFT JOIN project_contracts pc ON pc.id=e.contract_id
                   ORDER BY e.id DESC"""),
      projects=rows("SELECT id,project_no,name FROM projects ORDER BY project_no"),
      contracts=rows("SELECT id,contract_no,project_id FROM project_contracts ORDER BY contract_no"))

@app.route("/contracts/report")
@login_required
def contracts_report():
    data=rows("""SELECT sc.contract_no,p.project_no,p.name project_name,
      s.name subcontractor_name,sc.contract_value,sc.retention_rate,sc.advance_amount,
      sc.start_date,sc.end_date,sc.status,
      COALESCE((SELECT SUM(c.total) FROM subcontract_certificates c
                WHERE c.contract_id=sc.id AND c.status IN ('معتمد','مفوتر')),0) certified
      FROM subcontract_contracts sc
      JOIN projects p ON p.id=sc.project_id
      JOIN subcontractors s ON s.id=sc.subcontractor_id
      ORDER BY sc.id DESC""")
    return render_template("contracts_report.html",rows=data)

@app.route("/contracts/report.xlsx")
@login_required
def contracts_report_export():
    data=rows("""SELECT sc.contract_no,p.project_no,p.name project_name,
      s.name subcontractor_name,sc.contract_value,sc.retention_rate,sc.advance_amount,
      sc.start_date,sc.end_date,sc.status,
      COALESCE((SELECT SUM(c.total) FROM subcontract_certificates c
                WHERE c.contract_id=sc.id AND c.status IN ('معتمد','مفوتر')),0) certified
      FROM subcontract_contracts sc
      JOIN projects p ON p.id=sc.project_id
      JOIN subcontractors s ON s.id=sc.subcontractor_id
      ORDER BY sc.id DESC""")
    return xlsx_response("subcontract_contracts.xlsx","عقود المقاولين",
      ["العقد","المشروع","اسم المشروع","المقاول","قيمة العقد","الاحتجاز%",
       "الدفعة المقدمة","البداية","النهاية","المعتمد","المتبقي","الحالة"],
      [[r["contract_no"],r["project_no"],r["project_name"],r["subcontractor_name"],
        r["contract_value"],r["retention_rate"],r["advance_amount"],r["start_date"],
        r["end_date"],r["certified"],float(r["contract_value"])-float(r["certified"]),
        r["status"]] for r in data])


@app.route("/hr")
@login_required
def hr_complete_center():
    stats={
      "employees":row("SELECT COUNT(*) c FROM employees WHERE active=1")["c"],
      "contracts_expiring":row("""SELECT COUNT(*) c FROM employee_contracts
        WHERE status='ساري' AND end_date IS NOT NULL
        AND end_date BETWEEN CURRENT_DATE AND CURRENT_DATE+INTERVAL '60 days'""")["c"],
      "documents_expiring":row("""SELECT COUNT(*) c FROM employee_documents
        WHERE expiry_date IS NOT NULL
        AND expiry_date BETWEEN CURRENT_DATE AND CURRENT_DATE+INTERVAL '60 days'""")["c"],
      "pending_leaves":row("SELECT COUNT(*) c FROM leave_requests WHERE status='قيد الاعتماد'")["c"],
    }
    alerts=rows("""SELECT e.id,e.name,'إقامة' alert_type,e.iqama_expiry expiry_date
                   FROM employees e
                   WHERE e.iqama_expiry IS NOT NULL
                   AND e.iqama_expiry<=CURRENT_DATE+INTERVAL '60 days'
                   UNION ALL
                   SELECT e.id,e.name,'جواز سفر',e.passport_expiry
                   FROM employees e
                   WHERE e.passport_expiry IS NOT NULL
                   AND e.passport_expiry<=CURRENT_DATE+INTERVAL '60 days'
                   UNION ALL
                   SELECT e.id,e.name,'تأمين طبي',e.medical_insurance_expiry
                   FROM employees e
                   WHERE e.medical_insurance_expiry IS NOT NULL
                   AND e.medical_insurance_expiry<=CURRENT_DATE+INTERVAL '60 days'
                   ORDER BY expiry_date""")
    return render_template("hr_complete_center.html",stats=stats,alerts=alerts)

@app.route("/hr/employees/<int:employee_id>")
@login_required
def hr_employee_profile(employee_id):
    employee=row("""SELECT e.*,d.name department_name,b.name branch_name,
      cc.name cost_center_name FROM employees e
      LEFT JOIN departments d ON d.id=e.department_id
      LEFT JOIN branches b ON b.id=e.branch_id
      LEFT JOIN cost_centers cc ON cc.id=e.cost_center_id
      WHERE e.id=:id""",{"id":employee_id})
    if not employee:
        return "الموظف غير موجود",404
    return render_template("hr_employee_profile.html",employee=employee,
      contracts=rows("""SELECT * FROM employee_contracts
                        WHERE employee_id=:id ORDER BY start_date DESC""",{"id":employee_id}),
      leaves=rows("""SELECT lr.*,lt.name leave_type_name FROM leave_requests lr
                    JOIN leave_types lt ON lt.id=lr.leave_type_id
                    WHERE lr.employee_id=:id ORDER BY lr.start_date DESC""",{"id":employee_id}),
      documents=rows("""SELECT * FROM employee_documents
                        WHERE employee_id=:id ORDER BY expiry_date NULLS LAST""",{"id":employee_id}),
      warnings=rows("""SELECT * FROM employee_warnings
                       WHERE employee_id=:id ORDER BY warning_date DESC""",{"id":employee_id}),
      settlements=rows("""SELECT * FROM employee_end_of_service
                          WHERE employee_id=:id ORDER BY calculation_date DESC""",{"id":employee_id}))

@app.route("/hr/contracts",methods=["GET","POST"])
@login_required
def hr_contracts():
    if request.method=="POST":
        no=request.form.get("contract_no") or next_hr_number("employee_contracts","CTR")
        execute("""INSERT INTO employee_contracts(employee_id,contract_no,contract_type,
          start_date,end_date,probation_end_date,basic_salary,housing_allowance,
          transport_allowance,other_allowance,working_hours,annual_leave_days,
          status,notes,created_by,created_at)
          VALUES(:employee,:no,:type,:start,:end,:probation,:basic,:housing,
          :transport,:other,:hours,:leave_days,:status,:notes,:uid,:created)""",
          {"employee":request.form["employee_id"],"no":no,
           "type":request.form.get("contract_type","محدد المدة"),
           "start":request.form["start_date"],
           "end":request.form.get("end_date") or None,
           "probation":request.form.get("probation_end_date") or None,
           "basic":float(request.form.get("basic_salary") or 0),
           "housing":float(request.form.get("housing_allowance") or 0),
           "transport":float(request.form.get("transport_allowance") or 0),
           "other":float(request.form.get("other_allowance") or 0),
           "hours":float(request.form.get("working_hours") or 8),
           "leave_days":int(request.form.get("annual_leave_days") or 21),
           "status":request.form.get("status","ساري"),
           "notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        execute("""UPDATE employees SET basic_salary=:basic,
          housing_allowance=:housing,transport_allowance=:transport,
          other_allowance=:other,hire_date=COALESCE(hire_date,:start),
          contract_end_date=:end WHERE id=:employee""",
          {"basic":float(request.form.get("basic_salary") or 0),
           "housing":float(request.form.get("housing_allowance") or 0),
           "transport":float(request.form.get("transport_allowance") or 0),
           "other":float(request.form.get("other_allowance") or 0),
           "start":request.form["start_date"],
           "end":request.form.get("end_date") or None,
           "employee":request.form["employee_id"]})
        flash("تم حفظ عقد الموظف","success")
        return redirect(url_for("hr_contracts"))
    return render_template("hr_contracts.html",
      contracts=rows("""SELECT c.*,e.employee_no,e.name FROM employee_contracts c
                        JOIN employees e ON e.id=c.employee_id
                        ORDER BY c.start_date DESC,c.id DESC"""),
      employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"))

@app.route("/hr/leave-types",methods=["GET","POST"])
@login_required
def hr_leave_types():
    if request.method=="POST":
        execute("""INSERT INTO leave_types(code,name,paid,annual_entitlement,active)
                   VALUES(:code,:name,:paid,:entitlement,1)
                   ON CONFLICT(code) DO UPDATE SET name=EXCLUDED.name,
                   paid=EXCLUDED.paid,annual_entitlement=EXCLUDED.annual_entitlement""",
                {"code":request.form["code"],"name":request.form["name"],
                 "paid":1 if request.form.get("paid") else 0,
                 "entitlement":float(request.form.get("annual_entitlement") or 0)})
        flash("تم حفظ نوع الإجازة","success")
        return redirect(url_for("hr_leave_types"))
    return render_template("hr_leave_types.html",
      rows=rows("SELECT * FROM leave_types ORDER BY code"))

@app.route("/hr/leaves",methods=["GET","POST"])
@login_required
def hr_leaves():
    if request.method=="POST":
        days=calculate_leave_days(request.form["start_date"],request.form["end_date"])
        no=next_hr_number("leave_requests","LEV")
        execute("""INSERT INTO leave_requests(request_no,employee_id,leave_type_id,
          start_date,end_date,days,reason,status,created_by,created_at)
          VALUES(:no,:employee,:type,:start,:end,:days,:reason,'قيد الاعتماد',
          :uid,:created)""",
          {"no":no,"employee":request.form["employee_id"],
           "type":request.form["leave_type_id"],
           "start":request.form["start_date"],"end":request.form["end_date"],
           "days":days,"reason":request.form.get("reason",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash(f"تم إنشاء طلب الإجازة {no}","success")
        return redirect(url_for("hr_leaves"))
    return render_template("hr_leaves.html",
      requests=rows("""SELECT lr.*,e.employee_no,e.name,lt.name leave_type_name
                       FROM leave_requests lr
                       JOIN employees e ON e.id=lr.employee_id
                       JOIN leave_types lt ON lt.id=lr.leave_type_id
                       ORDER BY lr.id DESC"""),
      employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"),
      leave_types=rows("SELECT id,name FROM leave_types WHERE active=1 ORDER BY name"))

@app.route("/hr/leaves/<int:request_id>/approve",methods=["POST"])
@login_required
def hr_leave_approve(request_id):
    leave=row("SELECT * FROM leave_requests WHERE id=:id",{"id":request_id})
    if not leave:
        return "طلب الإجازة غير موجود",404
    year=leave["start_date"].year
    balance=row("""SELECT * FROM employee_leave_balances
                   WHERE employee_id=:employee AND leave_type_id=:type AND year=:year""",
                {"employee":leave["employee_id"],"type":leave["leave_type_id"],"year":year})
    if not balance:
        entitlement=row("SELECT annual_entitlement FROM leave_types WHERE id=:id",
                        {"id":leave["leave_type_id"]})
        opening=float(entitlement["annual_entitlement"] or 0)
        execute("""INSERT INTO employee_leave_balances(employee_id,leave_type_id,year,
          opening_balance,accrued,used,remaining)
          VALUES(:employee,:type,:year,:opening,0,0,:opening)""",
          {"employee":leave["employee_id"],"type":leave["leave_type_id"],
           "year":year,"opening":opening})
        balance=row("""SELECT * FROM employee_leave_balances
                       WHERE employee_id=:employee AND leave_type_id=:type AND year=:year""",
                    {"employee":leave["employee_id"],"type":leave["leave_type_id"],"year":year})
    if float(balance["remaining"] or 0) < float(leave["days"] or 0):
        flash("رصيد الإجازة غير كافٍ","danger")
        return redirect(url_for("hr_leaves"))
    execute("""UPDATE employee_leave_balances SET used=used+:days,
      remaining=remaining-:days WHERE id=:id""",
      {"days":leave["days"],"id":balance["id"]})
    execute("""UPDATE leave_requests SET status='معتمد',approved_by=:user,
      approved_at=:dt WHERE id=:id""",
      {"user":session.get("user_id"),"dt":datetime.now(),"id":request_id})
    flash("تم اعتماد الإجازة","success")
    return redirect(url_for("hr_leaves"))

@app.route("/hr/documents",methods=["GET","POST"])
@login_required
def hr_documents():
    if request.method=="POST":
        execute("""INSERT INTO employee_documents(employee_id,document_type,document_no,
          issue_date,expiry_date,file_name,file_url,status,notes,created_by,created_at)
          VALUES(:employee,:type,:no,:issue,:expiry,:file_name,:file_url,'ساري',
          :notes,:uid,:created)""",
          {"employee":request.form["employee_id"],
           "type":request.form["document_type"],
           "no":request.form.get("document_no",""),
           "issue":request.form.get("issue_date") or None,
           "expiry":request.form.get("expiry_date") or None,
           "file_name":request.form.get("file_name",""),
           "file_url":request.form.get("file_url",""),
           "notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash("تم حفظ مستند الموظف","success")
        return redirect(url_for("hr_documents"))
    return render_template("hr_documents.html",
      documents=rows("""SELECT d.*,e.employee_no,e.name FROM employee_documents d
                        JOIN employees e ON e.id=d.employee_id
                        ORDER BY d.expiry_date NULLS LAST,d.id DESC"""),
      employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"))

@app.route("/hr/warnings",methods=["GET","POST"])
@login_required
def hr_warnings():
    if request.method=="POST":
        execute("""INSERT INTO employee_warnings(employee_id,warning_date,warning_type,
          subject,details,action_required,status,created_by,created_at)
          VALUES(:employee,:dt,:type,:subject,:details,:action,'مفتوح',:uid,:created)""",
          {"employee":request.form["employee_id"],
           "dt":request.form["warning_date"],
           "type":request.form.get("warning_type","إنذار كتابي"),
           "subject":request.form["subject"],
           "details":request.form["details"],
           "action":request.form.get("action_required",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash("تم حفظ الإنذار","success")
        return redirect(url_for("hr_warnings"))
    return render_template("hr_warnings.html",
      warnings=rows("""SELECT w.*,e.employee_no,e.name FROM employee_warnings w
                       JOIN employees e ON e.id=w.employee_id
                       ORDER BY w.warning_date DESC,w.id DESC"""),
      employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"))

@app.route("/hr/end-of-service",methods=["GET","POST"])
@login_required
def hr_end_of_service():
    if request.method=="POST":
        calc=calculate_end_of_service(
            int(request.form["employee_id"]),
            request.form["service_end_date"],
            request.form.get("other_dues") or 0,
            request.form.get("deductions") or 0
        )
        execute("""INSERT INTO employee_end_of_service(employee_id,calculation_date,
          service_start_date,service_end_date,service_years,last_basic_salary,
          gratuity_amount,leave_balance_amount,other_dues,deductions,net_settlement,
          status,notes,created_by,created_at)
          VALUES(:employee,:calc_date,:start,:end,:years,:basic,:gratuity,:leave,
          :other,:deductions,:net,'مسودة',:notes,:uid,:created)""",
          {"employee":request.form["employee_id"],
           "calc_date":request.form.get("calculation_date") or datetime.now().date(),
           "start":calc["service_start_date"],"end":calc["service_end_date"],
           "years":calc["service_years"],"basic":calc["last_basic_salary"],
           "gratuity":calc["gratuity_amount"],"leave":calc["leave_balance_amount"],
           "other":calc["other_dues"],"deductions":calc["deductions"],
           "net":calc["net_settlement"],"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash("تم احتساب نهاية الخدمة","success")
        return redirect(url_for("hr_end_of_service"))
    return render_template("hr_end_of_service.html",
      settlements=rows("""SELECT s.*,e.employee_no,e.name FROM employee_end_of_service s
                          JOIN employees e ON e.id=s.employee_id
                          ORDER BY s.calculation_date DESC,s.id DESC"""),
      employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"))

@app.route("/hr/recruitment",methods=["GET","POST"])
@login_required
def hr_recruitment():
    if request.method=="POST":
        no=next_hr_number("recruitment_candidates","CAN")
        execute("""INSERT INTO recruitment_candidates(candidate_no,full_name,nationality,
          phone,email,job_title,department_id,expected_salary,source,stage,
          interview_date,interview_result,notes,created_by,created_at)
          VALUES(:no,:name,:nationality,:phone,:email,:job,:department,:salary,
          :source,:stage,:interview,:result,:notes,:uid,:created)""",
          {"no":no,"name":request.form["full_name"],
           "nationality":request.form.get("nationality",""),
           "phone":request.form.get("phone",""),"email":request.form.get("email",""),
           "job":request.form.get("job_title",""),
           "department":request.form.get("department_id") or None,
           "salary":float(request.form.get("expected_salary") or 0),
           "source":request.form.get("source",""),
           "stage":request.form.get("stage","جديد"),
           "interview":request.form.get("interview_date") or None,
           "result":request.form.get("interview_result",""),
           "notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash(f"تم حفظ المرشح {no}","success")
        return redirect(url_for("hr_recruitment"))
    return render_template("hr_recruitment.html",
      candidates=rows("""SELECT c.*,d.name department_name FROM recruitment_candidates c
                         LEFT JOIN departments d ON d.id=c.department_id
                         ORDER BY c.id DESC"""),
      departments=rows("SELECT id,name FROM departments WHERE active=1 ORDER BY name"))

@app.route("/hr/report")
@login_required
def hr_report():
    data=rows("""SELECT e.id,e.employee_no,e.name,e.job_title,d.name department_name,
      b.name branch_name,e.hire_date,e.contract_end_date,e.iqama_expiry,
      e.passport_expiry,e.medical_insurance_expiry,e.basic_salary,e.active
      FROM employees e
      LEFT JOIN departments d ON d.id=e.department_id
      LEFT JOIN branches b ON b.id=e.branch_id
      ORDER BY e.name""")
    return render_template("hr_report.html",rows=data)

@app.route("/hr/report.xlsx")
@login_required
def hr_report_export():
    data=rows("""SELECT e.employee_no,e.name,e.job_title,d.name department_name,
      b.name branch_name,e.hire_date,e.contract_end_date,e.iqama_expiry,
      e.passport_expiry,e.medical_insurance_expiry,e.basic_salary,e.active
      FROM employees e
      LEFT JOIN departments d ON d.id=e.department_id
      LEFT JOIN branches b ON b.id=e.branch_id
      ORDER BY e.name""")
    return xlsx_response("hr_employees.xlsx","الموارد البشرية",
      ["رقم الموظف","الاسم","الوظيفة","القسم","الفرع","تاريخ التعيين",
       "نهاية العقد","انتهاء الإقامة","انتهاء الجواز","انتهاء التأمين",
       "الراتب الأساسي","الحالة"],
      [[r["employee_no"],r["name"],r["job_title"],r["department_name"],
        r["branch_name"],r["hire_date"],r["contract_end_date"],r["iqama_expiry"],
        r["passport_expiry"],r["medical_insurance_expiry"],r["basic_salary"],
        "نشط" if r["active"] else "موقوف"] for r in data])


@app.route("/budgets")
@login_required
def budgets_center():
    budgets=rows("""SELECT b.*,br.name branch_name,
      COALESCE((SELECT SUM(bl.budget_amount) FROM budget_lines bl
                WHERE bl.budget_id=b.id),0) total_budget,
      COALESCE((SELECT SUM(bl.forecast_amount) FROM budget_lines bl
                WHERE bl.budget_id=b.id),0) total_forecast
      FROM budget_headers b
      LEFT JOIN branches br ON br.id=b.branch_id
      ORDER BY b.fiscal_year DESC,b.id DESC""")
    return render_template("budgets_center.html",budgets=budgets)

@app.route("/budgets/new",methods=["GET","POST"])
@login_required
def budget_new():
    if request.method=="POST":
        year=int(request.form["fiscal_year"])
        no=next_budget_number(year)
        execute("""INSERT INTO budget_headers(budget_no,name,fiscal_year,branch_id,
          status,notes,created_by,created_at)
          VALUES(:no,:name,:year,:branch,'مسودة',:notes,:user,:dt)""",
          {"no":no,"name":request.form["name"],"year":year,
           "branch":request.form.get("branch_id") or None,
           "notes":request.form.get("notes",""),
           "user":session.get("user_id"),"dt":datetime.now()})
        budget_id=row("SELECT id FROM budget_headers WHERE budget_no=:no",{"no":no})["id"]
        flash(f"تم إنشاء الموازنة {no}","success")
        return redirect(url_for("budget_view",budget_id=budget_id))
    return render_template("budget_form.html",
      branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"))

@app.route("/budgets/<int:budget_id>",methods=["GET","POST"])
@login_required
def budget_view(budget_id):
    budget=row("""SELECT b.*,br.name branch_name FROM budget_headers b
                  LEFT JOIN branches br ON br.id=b.branch_id WHERE b.id=:id""",
               {"id":budget_id})
    if not budget:
        return "الموازنة غير موجودة",404
    if request.method=="POST":
        account_id=int(request.form["account_id"])
        cost_center_id=request.form.get("cost_center_id") or None
        month_no=int(request.form["month_no"])
        amount=float(request.form.get("budget_amount") or 0)
        forecast=float(request.form.get("forecast_amount") or amount)
        execute("""INSERT INTO budget_lines(budget_id,account_id,cost_center_id,
          month_no,budget_amount,forecast_amount,notes)
          VALUES(:budget,:account,:cc,:month,:amount,:forecast,:notes)
          ON CONFLICT(budget_id,account_id,cost_center_id,month_no)
          DO UPDATE SET budget_amount=EXCLUDED.budget_amount,
          forecast_amount=EXCLUDED.forecast_amount,notes=EXCLUDED.notes""",
          {"budget":budget_id,"account":account_id,"cc":cost_center_id,
           "month":month_no,"amount":amount,"forecast":forecast,
           "notes":request.form.get("notes","")})
        flash("تم حفظ بند الموازنة","success")
        return redirect(url_for("budget_view",budget_id=budget_id))
    summary=budget_summary_rows(budget_id)
    totals={
        "budget":round(sum(float(x["budget_amount"] or 0) for x in summary),2),
        "forecast":round(sum(float(x["forecast_amount"] or 0) for x in summary),2),
        "actual":round(sum(float(x["actual_amount"] or 0) for x in summary),2),
        "variance":round(sum(float(x["variance_amount"] or 0) for x in summary),2),
    }
    return render_template("budget_view.html",budget=budget,lines=summary,totals=totals,
      accounts=rows("""SELECT id,account_code,account_name_ar,account_type
                       FROM chart_of_accounts
                       WHERE active=1 AND accepts_entries=1
                       ORDER BY account_code"""),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"))

@app.route("/budgets/<int:budget_id>/approve",methods=["POST"])
@login_required
def budget_approve(budget_id):
    execute("UPDATE budget_headers SET status='معتمد' WHERE id=:id",{"id":budget_id})
    audit("APPROVE","BUDGET",f"اعتماد الموازنة رقم {budget_id}")
    flash("تم اعتماد الموازنة","success")
    return redirect(url_for("budget_view",budget_id=budget_id))

@app.route("/budgets/<int:budget_id>/revision",methods=["POST"])
@login_required
def budget_revision(budget_id):
    revision_no=(db.session.execute(text("""SELECT COALESCE(MAX(revision_no),0)+1
      FROM budget_revisions WHERE budget_id=:id"""),{"id":budget_id}).scalar() or 1)
    execute("""INSERT INTO budget_revisions(budget_id,revision_no,revision_date,
      reason,approved_by,created_by,created_at)
      VALUES(:budget,:revision,:dt,:reason,:approved,:created_by,:created)""",
      {"budget":budget_id,"revision":revision_no,"dt":request.form["revision_date"],
       "reason":request.form.get("reason",""),
       "approved":session.get("user_id"),"created_by":session.get("user_id"),
       "created":datetime.now()})
    execute("UPDATE budget_headers SET status='معدلة' WHERE id=:id",{"id":budget_id})
    flash(f"تم تسجيل المراجعة رقم {revision_no}","success")
    return redirect(url_for("budget_view",budget_id=budget_id))

@app.route("/budgets/<int:budget_id>/variance")
@login_required
def budget_variance_report(budget_id):
    budget=row("SELECT * FROM budget_headers WHERE id=:id",{"id":budget_id})
    if not budget:
        return "الموازنة غير موجودة",404
    rows_data=budget_summary_rows(budget_id)
    monthly=[]
    for month_no in range(1,13):
        month_rows=[x for x in rows_data if x["month_no"]==month_no]
        monthly.append({
            "month_no":month_no,
            "budget":round(sum(float(x["budget_amount"] or 0) for x in month_rows),2),
            "forecast":round(sum(float(x["forecast_amount"] or 0) for x in month_rows),2),
            "actual":round(sum(float(x["actual_amount"] or 0) for x in month_rows),2),
            "variance":round(sum(float(x["variance_amount"] or 0) for x in month_rows),2),
        })
    return render_template("budget_variance_report.html",budget=budget,
                           rows=rows_data,monthly=monthly)

@app.route("/budgets/<int:budget_id>/export.xlsx")
@login_required
def budget_export(budget_id):
    budget=row("SELECT * FROM budget_headers WHERE id=:id",{"id":budget_id})
    data=budget_summary_rows(budget_id)
    return xlsx_response(f"budget_{budget['fiscal_year']}.xlsx","الموازنة",
      ["الحساب","اسم الحساب","مركز التكلفة","الشهر","الموازنة","التوقع",
       "الفعلي","الانحراف","نسبة الانحراف"],
      [[r["account_code"],r["account_name_ar"],
        f"{r['cost_center_code'] or ''} {r['cost_center_name'] or ''}".strip(),
        r["month_no"],r["budget_amount"],r["forecast_amount"],
        r["actual_amount"],r["variance_amount"],r["variance_percent"]] for r in data])

@app.route("/budgets/dashboard")
@login_required
def budgets_dashboard():
    latest=row("""SELECT * FROM budget_headers
                  WHERE status IN ('معتمد','معدلة')
                  ORDER BY fiscal_year DESC,id DESC LIMIT 1""")
    if not latest:
        return render_template("budgets_dashboard.html",budget=None,monthly=[],summary={})
    data=budget_summary_rows(latest["id"])
    monthly=[]
    for month_no in range(1,13):
        month_rows=[x for x in data if x["month_no"]==month_no]
        monthly.append({
            "month_no":month_no,
            "budget":round(sum(float(x["budget_amount"] or 0) for x in month_rows),2),
            "actual":round(sum(float(x["actual_amount"] or 0) for x in month_rows),2),
            "variance":round(sum(float(x["variance_amount"] or 0) for x in month_rows),2),
        })
    summary={
        "budget":round(sum(x["budget"] for x in monthly),2),
        "actual":round(sum(x["actual"] for x in monthly),2),
        "variance":round(sum(x["variance"] for x in monthly),2),
    }
    return render_template("budgets_dashboard.html",budget=latest,
                           monthly=monthly,summary=summary)


@app.route("/approvals")
@login_required
def approvals_inbox():
    requests=rows("""SELECT ar.*,aw.name workflow_name,u.username requester,
      s.step_order,s.approver_username,r.name role_name
      FROM approval_requests ar
      JOIN approval_workflows aw ON aw.id=ar.workflow_id
      LEFT JOIN users u ON u.id=ar.requester_user_id
      LEFT JOIN approval_workflow_steps s ON s.workflow_id=ar.workflow_id
        AND s.step_order=ar.current_step_order
      LEFT JOIN system_roles r ON r.id=s.role_id
      WHERE ar.status IN ('قيد الاعتماد','معاد للتعديل')
      ORDER BY ar.submitted_at DESC,ar.id DESC""")
    visible=[]
    for req in requests:
        allowed,_=approval_user_can_act(req)
        if allowed or req["requester_user_id"]==session.get("user_id"):
            visible.append(req)
    history=rows("""SELECT ar.*,u.username requester
                    FROM approval_requests ar
                    LEFT JOIN users u ON u.id=ar.requester_user_id
                    WHERE ar.status IN ('معتمد','مرفوض')
                    ORDER BY ar.completed_at DESC,ar.id DESC LIMIT 200""")
    return render_template("approvals_inbox.html",requests=visible,history=history)

@app.route("/approvals/workflows",methods=["GET","POST"])
@login_required
@permission_required("settings.view")
def approval_workflows():
    if request.method=="POST":
        code=request.form["code"].upper().strip()
        execute("""INSERT INTO approval_workflows(code,name,document_type,min_amount,
          max_amount,active,created_by,created_at)
          VALUES(:code,:name,:doc,:min,:max,1,:user,:dt)
          ON CONFLICT(code) DO UPDATE SET name=EXCLUDED.name,
          document_type=EXCLUDED.document_type,min_amount=EXCLUDED.min_amount,
          max_amount=EXCLUDED.max_amount,active=1""",
          {"code":code,"name":request.form["name"],
           "doc":request.form["document_type"],
           "min":float(request.form.get("min_amount") or 0),
           "max":float(request.form["max_amount"]) if request.form.get("max_amount") else None,
           "user":session.get("user_id"),"dt":datetime.now()})
        workflow_id=row("SELECT id FROM approval_workflows WHERE code=:code",{"code":code})["id"]
        execute("DELETE FROM approval_workflow_steps WHERE workflow_id=:id",
                {"id":workflow_id})
        orders=request.form.getlist("step_order[]")
        roles=request.form.getlist("role_id[]")
        usernames=request.form.getlist("approver_username[]")
        actions=request.form.getlist("action_name[]")
        for idx,order_value in enumerate(orders):
            if not order_value:
                continue
            execute("""INSERT INTO approval_workflow_steps(workflow_id,step_order,
              role_id,approver_username,action_name,is_final)
              VALUES(:workflow,:order,:role,:username,:action,:final)""",
              {"workflow":workflow_id,"order":int(order_value),
               "role":roles[idx] or None,"username":usernames[idx].strip(),
               "action":actions[idx] or "اعتماد",
               "final":1 if idx==len(orders)-1 else 0})
        flash("تم حفظ مسار الاعتماد","success")
        return redirect(url_for("approval_workflows"))
    workflows=rows("""SELECT w.*,COUNT(s.id) steps_count
                      FROM approval_workflows w
                      LEFT JOIN approval_workflow_steps s ON s.workflow_id=w.id
                      GROUP BY w.id ORDER BY w.document_type,w.min_amount""")
    return render_template("approval_workflows.html",workflows=workflows,
      document_types=APPROVAL_DOCUMENT_TYPES.keys(),
      roles=rows("SELECT id,name FROM system_roles WHERE active=1 ORDER BY name"),
      users=rows("SELECT username,full_name FROM users WHERE active=1 ORDER BY username"))

@app.route("/approvals/submit",methods=["POST"])
@login_required
def approval_submit():
    document_type=request.form["document_type"]
    document_id=int(request.form["document_id"])
    document_no=request.form.get("document_no","")
    amount=float(request.form.get("document_amount") or 0)
    try:
        request_id=submit_for_approval(document_type,document_id,document_no,amount,
                                       request.form.get("notes",""))
        flash("تم إرسال المستند للاعتماد","success")
        return redirect(url_for("approval_request_view",request_id=request_id))
    except Exception as exc:
        db.session.rollback()
        flash(str(exc),"danger")
        return redirect(request.referrer or url_for("approvals_inbox"))

@app.route("/approvals/<int:request_id>")
@login_required
def approval_request_view(request_id):
    req=row("""SELECT ar.*,aw.name workflow_name,u.username requester
               FROM approval_requests ar
               JOIN approval_workflows aw ON aw.id=ar.workflow_id
               LEFT JOIN users u ON u.id=ar.requester_user_id
               WHERE ar.id=:id""",{"id":request_id})
    if not req:
        return "طلب الاعتماد غير موجود",404
    actions=rows("""SELECT aa.*,u.username,s.step_order,r.name role_name
                    FROM approval_actions aa
                    LEFT JOIN users u ON u.id=aa.action_by_user_id
                    LEFT JOIN approval_workflow_steps s ON s.id=aa.step_id
                    LEFT JOIN system_roles r ON r.id=s.role_id
                    WHERE aa.request_id=:id ORDER BY aa.action_at,aa.id""",
                 {"id":request_id})
    allowed,step=approval_user_can_act(req)
    return render_template("approval_request_view.html",request=req,
                           actions=actions,can_act=allowed,current_step=step)

@app.route("/approvals/<int:request_id>/action",methods=["POST"])
@login_required
def approval_action(request_id):
    try:
        apply_approval_action(request_id,request.form["action_type"],
                              request.form.get("action_notes",""))
        flash("تم تسجيل إجراء الاعتماد","success")
    except Exception as exc:
        db.session.rollback()
        flash(str(exc),"danger")
    return redirect(url_for("approval_request_view",request_id=request_id))

@app.route("/approvals/report")
@login_required
def approvals_report():
    data=rows("""SELECT ar.request_no,ar.document_type,ar.document_no,
      ar.document_amount,ar.status,ar.submitted_at,ar.completed_at,
      aw.name workflow_name,u.username requester
      FROM approval_requests ar
      JOIN approval_workflows aw ON aw.id=ar.workflow_id
      LEFT JOIN users u ON u.id=ar.requester_user_id
      ORDER BY ar.submitted_at DESC""")
    return render_template("approvals_report.html",rows=data)

@app.route("/approvals/report.xlsx")
@login_required
def approvals_report_export():
    data=rows("""SELECT ar.request_no,ar.document_type,ar.document_no,
      ar.document_amount,ar.status,ar.submitted_at,ar.completed_at,
      aw.name workflow_name,u.username requester
      FROM approval_requests ar
      JOIN approval_workflows aw ON aw.id=ar.workflow_id
      LEFT JOIN users u ON u.id=ar.requester_user_id
      ORDER BY ar.submitted_at DESC""")
    return xlsx_response("approvals_report.xlsx","طلبات الاعتماد",
      ["رقم الطلب","نوع المستند","رقم المستند","القيمة","المسار",
       "مقدم الطلب","الحالة","تاريخ الإرسال","تاريخ الإكمال"],
      [[r["request_no"],r["document_type"],r["document_no"],r["document_amount"],
        r["workflow_name"],r["requester"],r["status"],r["submitted_at"],
        r["completed_at"]] for r in data])


@app.route("/admin/users",methods=["GET","POST"])
@login_required
@permission_required("users.view")
def users_admin():
    if request.method=="POST":
        if not has_permission("users.create"):
            flash("ليس لديك صلاحية إضافة مستخدم","danger")
            return redirect(url_for("users_admin"))
        username=request.form["username"].strip()
        password=request.form["password"]
        if len(password)<8:
            flash("كلمة المرور يجب ألا تقل عن 8 أحرف","danger")
            return redirect(url_for("users_admin"))
        execute("""INSERT INTO users(username,password,role,full_name,email,active,
          must_change_password) VALUES(:username,:password,'user',:full_name,:email,1,1)""",
          {"username":username,"password":generate_password_hash(password),
           "full_name":request.form.get("full_name",""),
           "email":request.form.get("email","")})
        user_id=row("SELECT id FROM users WHERE username=:u",{"u":username})["id"]
        for role_id in request.form.getlist("role_ids"):
            execute("""INSERT INTO user_role_assignments(user_id,role_id)
                       VALUES(:user,:role) ON CONFLICT DO NOTHING""",
                    {"user":user_id,"role":role_id})
        audit("CREATE","USER",f"إنشاء المستخدم {username}")
        flash("تم إنشاء المستخدم","success")
        return redirect(url_for("users_admin"))
    users=rows("""SELECT u.*,STRING_AGG(r.name,', ' ORDER BY r.name) role_names
      FROM users u
      LEFT JOIN user_role_assignments ura ON ura.user_id=u.id
      LEFT JOIN system_roles r ON r.id=ura.role_id
      GROUP BY u.id ORDER BY u.id""")
    return render_template("users_admin.html",users=users,
      roles=rows("SELECT id,code,name FROM system_roles WHERE active=1 ORDER BY name"))

@app.route("/admin/users/<int:user_id>/status",methods=["POST"])
@login_required
@permission_required("users.edit")
def user_status_update(user_id):
    if user_id==session.get("user_id"):
        flash("لا يمكنك إيقاف حسابك الحالي","danger")
        return redirect(url_for("users_admin"))
    execute("UPDATE users SET active=CASE WHEN active=1 THEN 0 ELSE 1 END WHERE id=:id",
            {"id":user_id})
    session.pop("permission_keys",None)
    audit("UPDATE","USER",f"تغيير حالة المستخدم رقم {user_id}")
    flash("تم تحديث حالة المستخدم","success")
    return redirect(url_for("users_admin"))

@app.route("/admin/users/<int:user_id>/reset-password",methods=["POST"])
@login_required
@permission_required("users.edit")
def user_reset_password(user_id):
    password=request.form["new_password"]
    if len(password)<8:
        flash("كلمة المرور يجب ألا تقل عن 8 أحرف","danger")
        return redirect(url_for("users_admin"))
    execute("""UPDATE users SET password=:password,must_change_password=1 WHERE id=:id""",
            {"password":generate_password_hash(password),"id":user_id})
    audit("RESET_PASSWORD","USER",f"إعادة كلمة مرور المستخدم رقم {user_id}")
    flash("تمت إعادة كلمة المرور","success")
    return redirect(url_for("users_admin"))

@app.route("/admin/roles",methods=["GET","POST"])
@login_required
@permission_required("users.view")
def roles_admin():
    if request.method=="POST":
        if not has_permission("users.create"):
            flash("ليس لديك صلاحية إنشاء دور","danger")
            return redirect(url_for("roles_admin"))
        execute("""INSERT INTO system_roles(code,name,description,is_system,active)
                   VALUES(:code,:name,:description,0,1)
                   ON CONFLICT(code) DO UPDATE SET name=EXCLUDED.name,
                   description=EXCLUDED.description""",
                {"code":request.form["code"].upper(),
                 "name":request.form["name"],
                 "description":request.form.get("description","")})
        role_id=row("SELECT id FROM system_roles WHERE code=:code",
                    {"code":request.form["code"].upper()})["id"]
        execute("DELETE FROM role_permissions WHERE role_id=:id",{"id":role_id})
        for permission_id in request.form.getlist("permission_ids"):
            execute("""INSERT INTO role_permissions(role_id,permission_id,allowed)
                       VALUES(:role,:permission,1)""",
                    {"role":role_id,"permission":permission_id})
        audit("CREATE","ROLE",f"حفظ الدور {request.form['code']}")
        flash("تم حفظ الدور والصلاحيات","success")
        return redirect(url_for("roles_admin"))
    permission_rows=rows("""SELECT * FROM system_permissions
                            ORDER BY module_name,action_name""")
    grouped={}
    for p in permission_rows:
        grouped.setdefault(p["module_name"],[]).append(p)
    roles=rows("""SELECT r.*,COUNT(rp.permission_id) permission_count
                  FROM system_roles r LEFT JOIN role_permissions rp ON rp.role_id=r.id
                  GROUP BY r.id ORDER BY r.is_system DESC,r.name""")
    return render_template("roles_admin.html",roles=roles,permissions=grouped,
                           module_names=SECURITY_MODULES)

@app.route("/admin/roles/<int:role_id>/permissions")
@login_required
@permission_required("users.view")
def role_permissions_api(role_id):
    values=db.session.execute(text("""SELECT permission_id FROM role_permissions
                                     WHERE role_id=:id AND allowed=1"""),
                              {"id":role_id}).scalars().all()
    return {"permission_ids":values}

@app.route("/admin/users/<int:user_id>/roles",methods=["POST"])
@login_required
@permission_required("users.edit")
def user_roles_update(user_id):
    execute("DELETE FROM user_role_assignments WHERE user_id=:id",{"id":user_id})
    for role_id in request.form.getlist("role_ids"):
        execute("""INSERT INTO user_role_assignments(user_id,role_id)
                   VALUES(:user,:role) ON CONFLICT DO NOTHING""",
                {"user":user_id,"role":role_id})
    if user_id==session.get("user_id"):
        session["permission_keys"]=list(user_permission_keys(user_id))
    audit("UPDATE","USER_ROLES",f"تحديث أدوار المستخدم رقم {user_id}")
    flash("تم تحديث أدوار المستخدم","success")
    return redirect(url_for("users_admin"))

@app.route("/admin/security-audit")
@login_required
@permission_required("users.view")
def security_audit():
    denied=rows("""SELECT * FROM access_denied_logs
                   ORDER BY created_at DESC,id DESC LIMIT 500""")
    logins=rows("""SELECT * FROM audit_logs WHERE action='LOGIN'
                   ORDER BY created_at DESC,id DESC LIMIT 200""")
    return render_template("security_audit.html",denied=denied,logins=logins)


@app.route("/crm")
@login_required
def crm_center():
    stats={
      "leads":row("SELECT COUNT(*) c FROM crm_leads")["c"],
      "open_opportunities":row("SELECT COUNT(*) c FROM crm_opportunities WHERE status='مفتوحة'")["c"],
      "pipeline":row("SELECT COALESCE(SUM(weighted_value),0) v FROM crm_opportunities WHERE status='مفتوحة'")["v"],
      "due_activities":row("""SELECT COUNT(*) c FROM crm_activities
                              WHERE status='مفتوحة' AND due_date<=CURRENT_TIMESTAMP""")["c"],
    }
    return render_template("crm_center.html",stats=stats,pipeline=crm_pipeline_summary(),
      recent=rows("""SELECT o.*,c.name customer_name,l.company_name lead_company
                     FROM crm_opportunities o
                     LEFT JOIN customers c ON c.id=o.customer_id
                     LEFT JOIN crm_leads l ON l.id=o.lead_id
                     ORDER BY o.id DESC LIMIT 10"""))

@app.route("/crm/leads",methods=["GET","POST"])
@login_required
def crm_leads():
    if request.method=="POST":
        no=next_crm_number("crm_leads","LEAD")
        value=float(request.form.get("estimated_value") or 0)
        probability=float(request.form.get("probability") or 0)
        execute("""INSERT INTO crm_leads(lead_no,company_name,contact_name,phone,email,
          source,industry,city,assigned_to,status,priority,estimated_value,probability,
          expected_close_date,notes,created_by,created_at)
          VALUES(:no,:company,:contact,:phone,:email,:source,:industry,:city,:assigned,
          :status,:priority,:value,:probability,:close,:notes,:uid,:created)""",
          {"no":no,"company":request.form["company_name"],
           "contact":request.form.get("contact_name",""),
           "phone":request.form.get("phone",""),"email":request.form.get("email",""),
           "source":request.form.get("source",""),"industry":request.form.get("industry",""),
           "city":request.form.get("city",""),"assigned":request.form.get("assigned_to",""),
           "status":request.form.get("status","جديد"),
           "priority":request.form.get("priority","متوسطة"),
           "value":value,"probability":probability,
           "close":request.form.get("expected_close_date") or None,
           "notes":request.form.get("notes",""),"uid":session.get("user_id"),
           "created":datetime.now()})
        flash(f"تم إنشاء العميل المحتمل {no}","success")
        return redirect(url_for("crm_leads"))
    q=request.args.get("q","").strip()
    params={};where=""
    if q:
        where="""WHERE company_name ILIKE :q OR contact_name ILIKE :q
                 OR phone ILIKE :q OR email ILIKE :q OR lead_no ILIKE :q"""
        params["q"]=f"%{q}%"
    return render_template("crm_leads.html",q=q,
      leads=rows(f"""SELECT * FROM crm_leads {where} ORDER BY id DESC""",params))

@app.route("/crm/leads/<int:lead_id>")
@login_required
def crm_lead_view(lead_id):
    lead=row("SELECT * FROM crm_leads WHERE id=:id",{"id":lead_id})
    if not lead: return "العميل المحتمل غير موجود",404
    return render_template("crm_lead_view.html",lead=lead,
      activities=rows("""SELECT * FROM crm_activities WHERE lead_id=:id
                         ORDER BY activity_date DESC,id DESC""",{"id":lead_id}),
      opportunities=rows("""SELECT * FROM crm_opportunities WHERE lead_id=:id
                            ORDER BY id DESC""",{"id":lead_id}))

@app.route("/crm/leads/<int:lead_id>/convert",methods=["POST"])
@login_required
def crm_lead_convert(lead_id):
    lead=row("SELECT * FROM crm_leads WHERE id=:id",{"id":lead_id})
    if not lead: return "العميل المحتمل غير موجود",404
    if lead.get("converted_customer_id"):
        return redirect(url_for("customers"))
    execute("""INSERT INTO customers(name,name_en,phone,email,city,active,created_at)
               VALUES(:name,'',:phone,:email,:city,1,:created)""",
            {"name":lead["company_name"],"phone":lead["phone"],
             "email":lead["email"],"city":lead["city"],"created":datetime.now()})
    customer_id=row("SELECT id FROM customers ORDER BY id DESC LIMIT 1")["id"]
    execute("""UPDATE crm_leads SET converted_customer_id=:customer,status='محوّل'
               WHERE id=:id""",{"customer":customer_id,"id":lead_id})
    audit("CONVERT","CRM_LEAD",f"تحويل {lead['lead_no']} إلى عميل")
    flash("تم تحويل العميل المحتمل إلى عميل فعلي","success")
    return redirect(url_for("customers"))

@app.route("/crm/opportunities",methods=["GET","POST"])
@login_required
def crm_opportunities():
    if request.method=="POST":
        no=next_crm_number("crm_opportunities","OPP")
        value=float(request.form.get("estimated_value") or 0)
        probability=float(request.form.get("probability") or 0)
        execute("""INSERT INTO crm_opportunities(opportunity_no,lead_id,customer_id,title,
          stage,estimated_value,probability,weighted_value,expected_close_date,
          sales_person,status,notes,created_by,created_at)
          VALUES(:no,:lead,:customer,:title,:stage,:value,:probability,:weighted,:close,
          :sales_person,:status,:notes,:uid,:created)""",
          {"no":no,"lead":request.form.get("lead_id") or None,
           "customer":request.form.get("customer_id") or None,
           "title":request.form["title"],"stage":request.form.get("stage","تأهيل"),
           "value":value,"probability":probability,
           "weighted":round(value*probability/100,2),
           "close":request.form.get("expected_close_date") or None,
           "sales_person":request.form.get("sales_person",""),
           "status":request.form.get("status","مفتوحة"),
           "notes":request.form.get("notes",""),"uid":session.get("user_id"),
           "created":datetime.now()})
        flash(f"تم إنشاء فرصة البيع {no}","success")
        return redirect(url_for("crm_opportunities"))
    return render_template("crm_opportunities.html",
      opportunities=rows("""SELECT o.*,l.company_name,c.name customer_name
                             FROM crm_opportunities o
                             LEFT JOIN crm_leads l ON l.id=o.lead_id
                             LEFT JOIN customers c ON c.id=o.customer_id
                             ORDER BY o.id DESC"""),
      leads=rows("""SELECT id,lead_no,company_name FROM crm_leads
                    WHERE status<>'محوّل' ORDER BY company_name"""),
      customers=rows("SELECT id,name FROM customers ORDER BY name"))

@app.route("/crm/opportunities/<int:opportunity_id>",methods=["GET","POST"])
@login_required
def crm_opportunity_view(opportunity_id):
    if request.method=="POST":
        value=float(request.form.get("estimated_value") or 0)
        probability=float(request.form.get("probability") or 0)
        execute("""UPDATE crm_opportunities SET stage=:stage,estimated_value=:value,
          probability=:probability,weighted_value=:weighted,expected_close_date=:close,
          sales_person=:sales_person,status=:status,lost_reason=:lost_reason,notes=:notes
          WHERE id=:id""",
          {"stage":request.form.get("stage","تأهيل"),"value":value,
           "probability":probability,"weighted":round(value*probability/100,2),
           "close":request.form.get("expected_close_date") or None,
           "sales_person":request.form.get("sales_person",""),
           "status":request.form.get("status","مفتوحة"),
           "lost_reason":request.form.get("lost_reason",""),
           "notes":request.form.get("notes",""),"id":opportunity_id})
        flash("تم تحديث فرصة البيع","success")
        return redirect(url_for("crm_opportunity_view",opportunity_id=opportunity_id))
    opp=row("""SELECT o.*,l.company_name,c.name customer_name
               FROM crm_opportunities o
               LEFT JOIN crm_leads l ON l.id=o.lead_id
               LEFT JOIN customers c ON c.id=o.customer_id WHERE o.id=:id""",{"id":opportunity_id})
    if not opp: return "فرصة البيع غير موجودة",404
    return render_template("crm_opportunity_view.html",opportunity=opp,
      activities=rows("""SELECT * FROM crm_activities WHERE opportunity_id=:id
                         ORDER BY activity_date DESC,id DESC""",{"id":opportunity_id}))

@app.route("/crm/activities",methods=["GET","POST"])
@login_required
def crm_activities():
    if request.method=="POST":
        execute("""INSERT INTO crm_activities(lead_id,opportunity_id,customer_id,
          activity_type,subject,activity_date,due_date,assigned_to,status,outcome,
          notes,created_by,created_at)
          VALUES(:lead,:opportunity,:customer,:type,:subject,:activity_date,:due,
          :assigned,:status,:outcome,:notes,:uid,:created)""",
          {"lead":request.form.get("lead_id") or None,
           "opportunity":request.form.get("opportunity_id") or None,
           "customer":request.form.get("customer_id") or None,
           "type":request.form["activity_type"],"subject":request.form["subject"],
           "activity_date":request.form["activity_date"],
           "due":request.form.get("due_date") or None,
           "assigned":request.form.get("assigned_to",""),
           "status":request.form.get("status","مفتوحة"),
           "outcome":request.form.get("outcome",""),
           "notes":request.form.get("notes",""),"uid":session.get("user_id"),
           "created":datetime.now()})
        flash("تم حفظ نشاط المتابعة","success")
        return redirect(url_for("crm_activities"))
    return render_template("crm_activities.html",
      activities=rows("""SELECT a.*,l.company_name,o.title opportunity_title,
        c.name customer_name
        FROM crm_activities a
        LEFT JOIN crm_leads l ON l.id=a.lead_id
        LEFT JOIN crm_opportunities o ON o.id=a.opportunity_id
        LEFT JOIN customers c ON c.id=a.customer_id
        ORDER BY a.activity_date DESC,a.id DESC"""),
      leads=rows("SELECT id,lead_no,company_name FROM crm_leads ORDER BY company_name"),
      opportunities=rows("SELECT id,opportunity_no,title FROM crm_opportunities ORDER BY id DESC"),
      customers=rows("SELECT id,name FROM customers ORDER BY name"))

@app.route("/crm/pipeline")
@login_required
def crm_pipeline():
    stages=["تأهيل","عرض سعر","تفاوض","فوز","خسارة"]
    data={stage:rows("""SELECT o.*,l.company_name,c.name customer_name
                       FROM crm_opportunities o
                       LEFT JOIN crm_leads l ON l.id=o.lead_id
                       LEFT JOIN customers c ON c.id=o.customer_id
                       WHERE o.stage=:stage ORDER BY o.expected_close_date,o.id""",
                       {"stage":stage}) for stage in stages}
    return render_template("crm_pipeline.html",stages=stages,data=data)

@app.route("/crm/report")
@login_required
def crm_report():
    data=rows("""SELECT o.opportunity_no,o.title,o.stage,o.status,o.estimated_value,
      o.probability,o.weighted_value,o.expected_close_date,o.sales_person,
      COALESCE(c.name,l.company_name) party_name
      FROM crm_opportunities o
      LEFT JOIN customers c ON c.id=o.customer_id
      LEFT JOIN crm_leads l ON l.id=o.lead_id
      ORDER BY o.id DESC""")
    return render_template("crm_report.html",rows=data)

@app.route("/crm/report.xlsx")
@login_required
def crm_report_export():
    data=rows("""SELECT o.opportunity_no,o.title,o.stage,o.status,o.estimated_value,
      o.probability,o.weighted_value,o.expected_close_date,o.sales_person,
      COALESCE(c.name,l.company_name) party_name
      FROM crm_opportunities o
      LEFT JOIN customers c ON c.id=o.customer_id
      LEFT JOIN crm_leads l ON l.id=o.lead_id
      ORDER BY o.id DESC""")
    return xlsx_response("crm_pipeline.xlsx","فرص البيع",
      ["رقم الفرصة","العنوان","العميل","المرحلة","الحالة","القيمة","الاحتمال",
       "القيمة المرجحة","الإغلاق المتوقع","مندوب المبيعات"],
      [[r["opportunity_no"],r["title"],r["party_name"],r["stage"],r["status"],
        r["estimated_value"],r["probability"],r["weighted_value"],
        r["expected_close_date"],r["sales_person"]] for r in data])


@app.route("/projects")
@login_required
def projects_center():
    data=rows("""SELECT p.*,c.name customer_name,cc.name cost_center_name,
      COALESCE((SELECT SUM(pc.total) FROM progress_certificates pc
                WHERE pc.project_id=p.id AND pc.status IN ('معتمد','مفوتر')),0) revenue,
      COALESCE((SELECT SUM(pe.amount) FROM project_cost_entries pe
                WHERE pe.project_id=p.id),0) cost
      FROM projects p LEFT JOIN customers c ON c.id=p.customer_id
      LEFT JOIN cost_centers cc ON cc.id=p.cost_center_id
      ORDER BY p.id DESC""")
    return render_template("projects_center.html",projects=data)

@app.route("/projects/new",methods=["GET","POST"])
@login_required
def project_new():
    if request.method=="POST":
        no=next_project_number()
        execute("""INSERT INTO projects(project_no,name,name_en,customer_id,branch_id,cost_center_id,
          project_manager,start_date,end_date,contract_value,retention_rate,advance_rate,
          status,location,description,created_by,created_at)
          VALUES(:no,:name,:name_en,:customer,:branch,:cc,:manager,:start,:end,:value,
          :retention,:advance,:status,:location,:description,:uid,:created)""",
          {"no":no,"name":request.form["name"],"name_en":request.form.get("name_en",""),
           "customer":request.form.get("customer_id") or None,
           "branch":request.form.get("branch_id") or None,
           "cc":request.form.get("cost_center_id") or None,
           "manager":request.form.get("project_manager",""),
           "start":request.form.get("start_date") or None,
           "end":request.form.get("end_date") or None,
           "value":float(request.form.get("contract_value") or 0),
           "retention":float(request.form.get("retention_rate") or 0),
           "advance":float(request.form.get("advance_rate") or 0),
           "status":request.form.get("status","نشط"),
           "location":request.form.get("location",""),
           "description":request.form.get("description",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash(f"تم إنشاء المشروع {no}","success")
        return redirect(url_for("projects_center"))
    return render_template("project_form.html",
      customers=rows("SELECT id,name FROM customers ORDER BY name"),
      branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"))

@app.route("/projects/<int:project_id>")
@login_required
def project_view(project_id):
    project=row("""SELECT p.*,c.name customer_name,b.name branch_name,cc.name cost_center_name
                   FROM projects p LEFT JOIN customers c ON c.id=p.customer_id
                   LEFT JOIN branches b ON b.id=p.branch_id
                   LEFT JOIN cost_centers cc ON cc.id=p.cost_center_id
                   WHERE p.id=:id""",{"id":project_id})
    if not project:
        return "المشروع غير موجود",404
    return render_template("project_view.html",project=project,
      summary=project_financial_summary(project_id),
      contracts=rows("SELECT * FROM project_contracts WHERE project_id=:id ORDER BY id DESC",{"id":project_id}),
      boq=rows("SELECT * FROM project_boq_items WHERE project_id=:id ORDER BY id",{"id":project_id}),
      certificates=rows("""SELECT pc.*,i.invoice_no FROM progress_certificates pc
                           LEFT JOIN invoices i ON i.id=pc.invoice_id
                           WHERE pc.project_id=:id ORDER BY pc.id DESC""",{"id":project_id}),
      costs=rows("""SELECT pe.*,s.name supplier_name,e.name employee_name
                    FROM project_cost_entries pe
                    LEFT JOIN suppliers s ON s.id=pe.supplier_id
                    LEFT JOIN employees e ON e.id=pe.employee_id
                    WHERE pe.project_id=:id ORDER BY pe.cost_date DESC,pe.id DESC""",{"id":project_id}))

@app.route("/projects/<int:project_id>/contract",methods=["POST"])
@login_required
def project_contract_add(project_id):
    execute("""INSERT INTO project_contracts(contract_no,project_id,contract_date,contract_type,
      contract_value,retention_rate,advance_amount,tax_rate,status,notes,created_by,created_at)
      VALUES(:no,:project,:dt,:type,:value,:retention,:advance,:tax,:status,:notes,:uid,:created)""",
      {"no":request.form["contract_no"],"project":project_id,
       "dt":request.form["contract_date"],"type":request.form.get("contract_type","مقطوعية"),
       "value":float(request.form.get("contract_value") or 0),
       "retention":float(request.form.get("retention_rate") or 0),
       "advance":float(request.form.get("advance_amount") or 0),
       "tax":float(request.form.get("tax_rate") or 15),
       "status":request.form.get("status","ساري"),
       "notes":request.form.get("notes",""),"uid":session.get("user_id"),
       "created":datetime.now()})
    flash("تم حفظ العقد","success")
    return redirect(url_for("project_view",project_id=project_id))

@app.route("/projects/<int:project_id>/boq",methods=["POST"])
@login_required
def project_boq_add(project_id):
    qty=float(request.form.get("quantity") or 0)
    rate=float(request.form.get("unit_rate") or 0)
    execute("""INSERT INTO project_boq_items(project_id,item_code,description,unit,
      quantity,unit_rate,total_value)
      VALUES(:project,:code,:description,:unit,:qty,:rate,:total)""",
      {"project":project_id,"code":request.form["item_code"],
       "description":request.form["description"],
       "unit":request.form.get("unit","وحدة"),"qty":qty,"rate":rate,
       "total":round(qty*rate,2)})
    flash("تمت إضافة بند الكميات","success")
    return redirect(url_for("project_view",project_id=project_id))

@app.route("/projects/<int:project_id>/cost",methods=["POST"])
@login_required
def project_cost_add(project_id):
    execute("""INSERT INTO project_cost_entries(project_id,cost_date,cost_type,reference,
      description,amount,supplier_id,employee_id,journal_id,created_by,created_at)
      VALUES(:project,:dt,:type,:reference,:description,:amount,:supplier,:employee,
      :journal,:uid,:created)""",
      {"project":project_id,"dt":request.form["cost_date"],
       "type":request.form["cost_type"],"reference":request.form.get("reference",""),
       "description":request.form["description"],
       "amount":float(request.form.get("amount") or 0),
       "supplier":request.form.get("supplier_id") or None,
       "employee":request.form.get("employee_id") or None,
       "journal":request.form.get("journal_id") or None,
       "uid":session.get("user_id"),"created":datetime.now()})
    flash("تم تسجيل تكلفة المشروع","success")
    return redirect(url_for("project_view",project_id=project_id))

@app.route("/projects/<int:project_id>/certificate",methods=["GET","POST"])
@login_required
def project_certificate_new(project_id):
    project=row("SELECT * FROM projects WHERE id=:id",{"id":project_id})
    if request.method=="POST":
        no=next_certificate_number(request.form["certificate_date"])
        item_ids=request.form.getlist("boq_item_id[]")
        current_qtys=request.form.getlist("current_qty[]")
        prepared=[];gross=0
        for idx,item_id in enumerate(item_ids):
            boq=row("SELECT * FROM project_boq_items WHERE id=:id AND project_id=:p",
                    {"id":item_id,"p":project_id})
            if not boq: continue
            current=float(current_qtys[idx] or 0)
            previous=float(boq["cumulative_qty"] or 0)
            cumulative=previous+current
            if cumulative>float(boq["quantity"] or 0)+0.0001:
                flash(f"الكمية التراكمية لبند {boq['item_code']} تجاوزت كمية العقد","danger")
                return redirect(url_for("project_certificate_new",project_id=project_id))
            current_value=round(current*float(boq["unit_rate"] or 0),2)
            cumulative_value=round(cumulative*float(boq["unit_rate"] or 0),2)
            prepared.append((boq,current,previous,cumulative,current_value,cumulative_value))
            gross+=current_value
        retention_rate=float(request.form.get("retention_rate") or project["retention_rate"] or 0)
        retention=round(gross*retention_rate/100,2)
        advance_recovery=float(request.form.get("advance_recovery") or 0)
        other_deductions=float(request.form.get("other_deductions") or 0)
        subtotal=round(gross-retention-advance_recovery-other_deductions,2)
        tax_rate=float(request.form.get("tax_rate") or 15)
        vat=round(subtotal*tax_rate/100,2)
        total=round(subtotal+vat,2)
        execute("""INSERT INTO progress_certificates(certificate_no,project_id,contract_id,
          certificate_date,period_from,period_to,gross_work_value,retention_amount,
          advance_recovery,other_deductions,subtotal,vat,total,status,notes,created_by,created_at)
          VALUES(:no,:project,:contract,:dt,:from,:to,:gross,:retention,:advance,:other,
          :subtotal,:vat,:total,'مسودة',:notes,:uid,:created)""",
          {"no":no,"project":project_id,"contract":request.form.get("contract_id") or None,
           "dt":request.form["certificate_date"],"from":request.form.get("period_from") or None,
           "to":request.form.get("period_to") or None,"gross":round(gross,2),
           "retention":retention,"advance":advance_recovery,"other":other_deductions,
           "subtotal":subtotal,"vat":vat,"total":total,
           "notes":request.form.get("notes",""),"uid":session.get("user_id"),
           "created":datetime.now()})
        cert_id=row("SELECT id FROM progress_certificates WHERE certificate_no=:no",{"no":no})["id"]
        for boq,current,previous,cumulative,current_value,cumulative_value in prepared:
            execute("""INSERT INTO progress_certificate_items(certificate_id,boq_item_id,
              previous_qty,current_qty,cumulative_qty,unit_rate,current_value,cumulative_value)
              VALUES(:cert,:boq,:previous,:current,:cumulative,:rate,:current_value,:cumulative_value)""",
              {"cert":cert_id,"boq":boq["id"],"previous":previous,"current":current,
               "cumulative":cumulative,"rate":boq["unit_rate"],
               "current_value":current_value,"cumulative_value":cumulative_value})
            execute("""UPDATE project_boq_items SET previous_qty=:previous,current_qty=:current,
              cumulative_qty=:cumulative,completion_percent=:pct WHERE id=:id""",
              {"previous":previous,"current":current,"cumulative":cumulative,
               "pct":round(cumulative/float(boq["quantity"] or 1)*100,2),"id":boq["id"]})
        flash(f"تم إنشاء المستخلص {no}","success")
        return redirect(url_for("project_certificate_view",certificate_id=cert_id))
    return render_template("project_certificate_form.html",project=project,
      contracts=rows("SELECT id,contract_no FROM project_contracts WHERE project_id=:id ORDER BY id DESC",{"id":project_id}),
      boq=rows("SELECT * FROM project_boq_items WHERE project_id=:id ORDER BY id",{"id":project_id}))

@app.route("/projects/certificates/<int:certificate_id>")
@login_required
def project_certificate_view(certificate_id):
    cert=row("""SELECT pc.*,p.project_no,p.name project_name,c.name customer_name,
      ct.contract_no,i.invoice_no
      FROM progress_certificates pc JOIN projects p ON p.id=pc.project_id
      LEFT JOIN customers c ON c.id=p.customer_id
      LEFT JOIN project_contracts ct ON ct.id=pc.contract_id
      LEFT JOIN invoices i ON i.id=pc.invoice_id
      WHERE pc.id=:id""",{"id":certificate_id})
    if not cert: return "المستخلص غير موجود",404
    items=rows("""SELECT pci.*,b.item_code,b.description,b.unit,b.quantity contract_qty
                  FROM progress_certificate_items pci
                  JOIN project_boq_items b ON b.id=pci.boq_item_id
                  WHERE pci.certificate_id=:id ORDER BY pci.id""",{"id":certificate_id})
    return render_template("project_certificate_view.html",cert=cert,items=items)

@app.route("/projects/certificates/<int:certificate_id>/approve",methods=["POST"])
@login_required
def project_certificate_approve(certificate_id):
    execute("UPDATE progress_certificates SET status='معتمد' WHERE id=:id",{"id":certificate_id})
    flash("تم اعتماد المستخلص","success")
    return redirect(url_for("project_certificate_view",certificate_id=certificate_id))

@app.route("/projects/report")
@login_required
def projects_report():
    data=rows("""SELECT p.id,p.project_no,p.name,c.name customer_name,p.contract_value,p.status,
      COALESCE((SELECT SUM(pc.total) FROM progress_certificates pc
                WHERE pc.project_id=p.id AND pc.status IN ('معتمد','مفوتر')),0) revenue,
      COALESCE((SELECT SUM(pe.amount) FROM project_cost_entries pe
                WHERE pe.project_id=p.id),0) cost
      FROM projects p LEFT JOIN customers c ON c.id=p.customer_id ORDER BY p.project_no""")
    return render_template("projects_report.html",rows=data)

@app.route("/projects/report.xlsx")
@login_required
def projects_report_export():
    data=rows("""SELECT p.project_no,p.name,c.name customer_name,p.contract_value,p.status,
      COALESCE((SELECT SUM(pc.total) FROM progress_certificates pc
                WHERE pc.project_id=p.id AND pc.status IN ('معتمد','مفوتر')),0) revenue,
      COALESCE((SELECT SUM(pe.amount) FROM project_cost_entries pe
                WHERE pe.project_id=p.id),0) cost
      FROM projects p LEFT JOIN customers c ON c.id=p.customer_id ORDER BY p.project_no""")
    return xlsx_response("projects_report.xlsx","المشاريع",
      ["رقم المشروع","المشروع","العميل","قيمة العقد","الحالة","الإيراد","التكلفة","الربح"],
      [[r["project_no"],r["name"],r["customer_name"],r["contract_value"],r["status"],
        r["revenue"],r["cost"],float(r["revenue"])-float(r["cost"])] for r in data])


@app.route("/hr-payroll")
@login_required
def hr_payroll_center():
    stats={
      "employees":row("SELECT COUNT(*) c FROM employees WHERE active=1")["c"],
      "departments":row("SELECT COUNT(*) c FROM departments WHERE active=1")["c"],
      "runs":row("SELECT COUNT(*) c FROM payroll_runs")["c"],
      "net":row("SELECT COALESCE(SUM(total_net),0) v FROM payroll_runs WHERE status='مرحّل'")["v"],
    }
    return render_template("hr_payroll_center.html",stats=stats)

@app.route("/departments",methods=["GET","POST"])
@login_required
def departments():
    if request.method=="POST":
        execute("""INSERT INTO departments(code,name,active)
                   VALUES(:code,:name,1)
                   ON CONFLICT(code) DO UPDATE SET name=EXCLUDED.name""",
                {"code":request.form["code"],"name":request.form["name"]})
        flash("تم حفظ القسم","success")
        return redirect(url_for("departments"))
    return render_template("departments.html",
      rows=rows("SELECT * FROM departments ORDER BY code"))

@app.route("/payroll/settings",methods=["GET","POST"])
@login_required
def payroll_settings():
    cfg=payroll_configuration()
    if request.method=="POST":
        execute("""UPDATE payroll_settings SET
          salary_expense_account_id=:expense,payroll_payable_account_id=:payable,
          deduction_account_id=:deduction,working_days_per_month=:days,
          working_hours_per_day=:hours,overtime_multiplier=:multiplier WHERE id=1""",
          {"expense":request.form.get("salary_expense_account_id") or None,
           "payable":request.form.get("payroll_payable_account_id") or None,
           "deduction":request.form.get("deduction_account_id") or None,
           "days":float(request.form.get("working_days_per_month") or 30),
           "hours":float(request.form.get("working_hours_per_day") or 8),
           "multiplier":float(request.form.get("overtime_multiplier") or 1.5)})
        flash("تم حفظ إعدادات الرواتب","success")
        return redirect(url_for("payroll_settings"))
    return render_template("payroll_settings.html",config=cfg,
      accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                       WHERE active=1 AND accepts_entries=1 ORDER BY account_code"""))

@app.route("/attendance",methods=["GET","POST"])
@login_required
def attendance():
    if request.method=="POST":
        execute("""INSERT INTO attendance_records(employee_id,attendance_date,status,
          check_in,check_out,overtime_hours,absence_hours,notes,created_by,created_at)
          VALUES(:employee,:dt,:status,:check_in,:check_out,:ot,:absence,:notes,:uid,:created)
          ON CONFLICT(employee_id,attendance_date) DO UPDATE SET
          status=EXCLUDED.status,check_in=EXCLUDED.check_in,check_out=EXCLUDED.check_out,
          overtime_hours=EXCLUDED.overtime_hours,absence_hours=EXCLUDED.absence_hours,
          notes=EXCLUDED.notes""",
          {"employee":request.form["employee_id"],"dt":request.form["attendance_date"],
           "status":request.form.get("status","حاضر"),
           "check_in":request.form.get("check_in") or None,
           "check_out":request.form.get("check_out") or None,
           "ot":float(request.form.get("overtime_hours") or 0),
           "absence":float(request.form.get("absence_hours") or 0),
           "notes":request.form.get("notes",""),"uid":session.get("user_id"),
           "created":datetime.now()})
        flash("تم حفظ سجل الحضور","success")
        return redirect(url_for("attendance"))
    date_from=request.args.get("date_from","")
    date_to=request.args.get("date_to","")
    conditions=[];params={}
    if date_from:
        conditions.append("a.attendance_date>=:date_from");params["date_from"]=date_from
    if date_to:
        conditions.append("a.attendance_date<=:date_to");params["date_to"]=date_to
    where=" WHERE "+" AND ".join(conditions) if conditions else ""
    return render_template("attendance.html",date_from=date_from,date_to=date_to,
      records=rows(f"""SELECT a.*,e.employee_no,e.name FROM attendance_records a
                       JOIN employees e ON e.id=a.employee_id
                       {where} ORDER BY a.attendance_date DESC,e.name""",params),
      employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"))

@app.route("/salary-adjustments",methods=["GET","POST"])
@login_required
def salary_adjustments():
    if request.method=="POST":
        execute("""INSERT INTO employee_salary_adjustments(employee_id,adjustment_date,
          adjustment_type,amount,recurring,description,active,created_by,created_at)
          VALUES(:employee,:dt,:type,:amount,:recurring,:description,1,:uid,:created)""",
          {"employee":request.form["employee_id"],"dt":request.form["adjustment_date"],
           "type":request.form["adjustment_type"],
           "amount":float(request.form["amount"] or 0),
           "recurring":1 if request.form.get("recurring") else 0,
           "description":request.form.get("description",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash("تم حفظ الاستحقاق/الاستقطاع","success")
        return redirect(url_for("salary_adjustments"))
    return render_template("salary_adjustments.html",
      rows=rows("""SELECT a.*,e.name employee_name FROM employee_salary_adjustments a
                   JOIN employees e ON e.id=a.employee_id ORDER BY a.adjustment_date DESC,a.id DESC"""),
      employees=rows("SELECT id,employee_no,name FROM employees WHERE active=1 ORDER BY name"))

@app.route("/payroll/runs",methods=["GET","POST"])
@login_required
def payroll_runs():
    if request.method=="POST":
        start=request.form["period_start"];end=request.form["period_end"]
        no=next_payroll_run_number(end)
        execute("""INSERT INTO payroll_runs(run_no,period_start,period_end,payment_date,
          status,notes,created_by,created_at)
          VALUES(:no,:start,:end,:payment,'مسودة',:notes,:uid,:created)""",
          {"no":no,"start":start,"end":end,
           "payment":request.form.get("payment_date") or end,
           "notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        run_id=row("SELECT id FROM payroll_runs WHERE run_no=:no",{"no":no})["id"]
        total_gross=total_deductions=total_net=0
        employees_data=rows("""SELECT e.*,COALESCE(e.cost_center_id,cc.id) resolved_cc
          FROM employees e LEFT JOIN cost_centers cc ON cc.id=e.cost_center_id
          WHERE e.active=1 ORDER BY e.id""")
        for employee in employees_data:
            calc=calculate_employee_payroll(employee,start,end)
            execute("""INSERT INTO payroll_lines(run_id,employee_id,basic_salary,allowances,
              overtime_hours,overtime_amount,absence_days,absence_deduction,other_deductions,
              gross_salary,total_deductions,net_salary,cost_center_id)
              VALUES(:run,:employee,:basic,:allowances,:ot_hours,:ot_amount,:absence_days,
              :absence_deduction,:other_deductions,:gross,:deductions,:net,:cc)""",
              {"run":run_id,"employee":employee["id"],"basic":calc["basic_salary"],
               "allowances":calc["allowances"],"ot_hours":calc["overtime_hours"],
               "ot_amount":calc["overtime_amount"],"absence_days":calc["absence_days"],
               "absence_deduction":calc["absence_deduction"],
               "other_deductions":calc["other_deductions"],"gross":calc["gross_salary"],
               "deductions":calc["total_deductions"],"net":calc["net_salary"],
               "cc":employee.get("cost_center_id")})
            total_gross+=calc["gross_salary"]
            total_deductions+=calc["total_deductions"]
            total_net+=calc["net_salary"]
        execute("""UPDATE payroll_runs SET total_gross=:gross,total_deductions=:deductions,
                   total_net=:net WHERE id=:id""",
                {"gross":round(total_gross,2),"deductions":round(total_deductions,2),
                 "net":round(total_net,2),"id":run_id})
        flash(f"تم إنشاء مسير الرواتب {no}","success")
        return redirect(url_for("payroll_run_view",run_id=run_id))
    return render_template("payroll_runs.html",
      runs=rows("""SELECT r.*,j.journal_no FROM payroll_runs r
                   LEFT JOIN journal_entries j ON j.id=r.journal_id
                   ORDER BY r.period_end DESC,r.id DESC"""))

@app.route("/payroll/runs/<int:run_id>")
@login_required
def payroll_run_view(run_id):
    run=row("""SELECT r.*,j.journal_no FROM payroll_runs r
               LEFT JOIN journal_entries j ON j.id=r.journal_id WHERE r.id=:id""",{"id":run_id})
    if not run:
        return "مسير الرواتب غير موجود",404
    lines=rows("""SELECT pl.*,e.employee_no,e.name,e.job_title,b.name branch_name
      FROM payroll_lines pl JOIN employees e ON e.id=pl.employee_id
      LEFT JOIN branches b ON b.id=e.branch_id WHERE pl.run_id=:id ORDER BY e.name""",{"id":run_id})
    return render_template("payroll_run_view.html",run=run,lines=lines)

@app.route("/payroll/runs/<int:run_id>/post",methods=["POST"])
@login_required
def payroll_run_post(run_id):
    try:
        jid=post_payroll_run(run_id)
        flash("تم اعتماد وترحيل مسير الرواتب","success")
    except Exception as exc:
        db.session.rollback()
        flash(str(exc),"danger")
    return redirect(url_for("payroll_run_view",run_id=run_id))

@app.route("/payroll/payslip/<int:line_id>")
@login_required
def payroll_payslip(line_id):
    line=row("""SELECT pl.*,pr.run_no,pr.period_start,pr.period_end,pr.payment_date,
      e.employee_no,e.name,e.name_en,e.job_title,e.bank_iban,b.name branch_name
      FROM payroll_lines pl JOIN payroll_runs pr ON pr.id=pl.run_id
      JOIN employees e ON e.id=pl.employee_id
      LEFT JOIN branches b ON b.id=e.branch_id WHERE pl.id=:id""",{"id":line_id})
    if not line:
        return "كشف الراتب غير موجود",404
    return render_template("payroll_payslip.html",line=line)

@app.route("/payroll/report")
@login_required
def payroll_report():
    data=rows("""SELECT pr.run_no,pr.period_start,pr.period_end,e.employee_no,e.name,
      pl.basic_salary,pl.allowances,pl.overtime_amount,pl.gross_salary,
      pl.total_deductions,pl.net_salary
      FROM payroll_lines pl JOIN payroll_runs pr ON pr.id=pl.run_id
      JOIN employees e ON e.id=pl.employee_id
      ORDER BY pr.period_end DESC,e.name""")
    return render_template("payroll_report.html",rows=data)

@app.route("/payroll/report.xlsx")
@login_required
def payroll_report_export():
    data=rows("""SELECT pr.run_no,pr.period_start,pr.period_end,e.employee_no,e.name,
      pl.basic_salary,pl.allowances,pl.overtime_amount,pl.gross_salary,
      pl.total_deductions,pl.net_salary
      FROM payroll_lines pl JOIN payroll_runs pr ON pr.id=pl.run_id
      JOIN employees e ON e.id=pl.employee_id
      ORDER BY pr.period_end DESC,e.name""")
    return xlsx_response("payroll_report.xlsx","الرواتب",
      ["المسير","من","إلى","رقم الموظف","الموظف","الأساسي","البدلات",
       "الإضافي","الإجمالي","الاستقطاعات","الصافي"],
      [[r["run_no"],r["period_start"],r["period_end"],r["employee_no"],r["name"],
        r["basic_salary"],r["allowances"],r["overtime_amount"],r["gross_salary"],
        r["total_deductions"],r["net_salary"]] for r in data])


@app.route("/fixed-assets")
@login_required
def fixed_assets_center():
    stats={
      "assets":row("SELECT COUNT(*) c FROM fixed_assets")["c"],
      "active":row("SELECT COUNT(*) c FROM fixed_assets WHERE status='نشط'")["c"],
      "cost":row("SELECT COALESCE(SUM(purchase_cost),0) v FROM fixed_assets")["v"],
      "nbv":row("SELECT COALESCE(SUM(net_book_value),0) v FROM fixed_assets")["v"],
    }
    return render_template("fixed_assets_center.html",stats=stats)

@app.route("/fixed-assets/categories",methods=["GET","POST"])
@login_required
def asset_categories():
    if request.method=="POST":
        execute("""INSERT INTO asset_categories(code,name,useful_life_months,depreciation_method,
          asset_account_id,accumulated_depr_account_id,depreciation_expense_account_id,active)
          VALUES(:code,:name,:life,:method,:asset,:accum,:expense,1)
          ON CONFLICT(code) DO UPDATE SET name=EXCLUDED.name""",
          {"code":request.form["code"],"name":request.form["name"],
           "life":int(request.form.get("useful_life_months") or 60),
           "method":request.form.get("depreciation_method","القسط الثابت"),
           "asset":request.form.get("asset_account_id") or None,
           "accum":request.form.get("accumulated_depr_account_id") or None,
           "expense":request.form.get("depreciation_expense_account_id") or None})
        flash("تم حفظ فئة الأصل","success")
        return redirect(url_for("asset_categories"))
    return render_template("asset_categories.html",
      rows=rows("""SELECT ac.*,a1.account_name_ar asset_account,
        a2.account_name_ar accumulated_account,a3.account_name_ar expense_account
        FROM asset_categories ac
        LEFT JOIN chart_of_accounts a1 ON a1.id=ac.asset_account_id
        LEFT JOIN chart_of_accounts a2 ON a2.id=ac.accumulated_depr_account_id
        LEFT JOIN chart_of_accounts a3 ON a3.id=ac.depreciation_expense_account_id
        ORDER BY ac.code"""),
      accounts=rows("""SELECT id,account_code,account_name_ar FROM chart_of_accounts
                       WHERE active=1 AND accepts_entries=1 ORDER BY account_code"""))

@app.route("/fixed-assets/assets",methods=["GET","POST"])
@login_required
def fixed_assets():
    if request.method=="POST":
        category=row("SELECT * FROM asset_categories WHERE id=:id",{"id":request.form["category_id"]})
        no=next_asset_number()
        cost=float(request.form.get("purchase_cost") or 0)
        residual=float(request.form.get("residual_value") or 0)
        execute("""INSERT INTO fixed_assets(asset_no,name,name_en,category_id,branch_id,cost_center_id,
          supplier_id,purchase_date,capitalization_date,purchase_cost,residual_value,
          useful_life_months,depreciation_method,accumulated_depreciation,net_book_value,
          serial_number,location,status,notes,created_by,created_at)
          VALUES(:no,:name,:name_en,:category,:branch,:cc,:supplier,:purchase_date,:cap_date,
          :cost,:residual,:life,:method,0,:nbv,:serial,:location,'نشط',:notes,:uid,:created)""",
          {"no":no,"name":request.form["name"],"name_en":request.form.get("name_en",""),
           "category":category["id"],"branch":request.form.get("branch_id") or None,
           "cc":request.form.get("cost_center_id") or None,
           "supplier":request.form.get("supplier_id") or None,
           "purchase_date":request.form["purchase_date"],
           "cap_date":request.form["capitalization_date"],
           "cost":cost,"residual":residual,
           "life":int(request.form.get("useful_life_months") or category["useful_life_months"]),
           "method":request.form.get("depreciation_method") or category["depreciation_method"],
           "nbv":cost,"serial":request.form.get("serial_number",""),
           "location":request.form.get("location",""),"notes":request.form.get("notes",""),
           "uid":session.get("user_id"),"created":datetime.now()})
        flash(f"تم إنشاء الأصل {no}","success")
        return redirect(url_for("fixed_assets"))
    return render_template("fixed_assets.html",
      assets=rows("""SELECT fa.*,ac.name category_name,b.name branch_name,cc.name cost_center_name
                     FROM fixed_assets fa JOIN asset_categories ac ON ac.id=fa.category_id
                     LEFT JOIN branches b ON b.id=fa.branch_id
                     LEFT JOIN cost_centers cc ON cc.id=fa.cost_center_id
                     ORDER BY fa.id DESC"""),
      categories=rows("SELECT * FROM asset_categories WHERE active=1 ORDER BY code"),
      branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"),
      suppliers=rows("SELECT id,name FROM suppliers ORDER BY name"))

@app.route("/fixed-assets/depreciation",methods=["GET","POST"])
@login_required
def asset_depreciation():
    if request.method=="POST":
        period_date=request.form["period_date"]
        no=next_asset_run_number(period_date)
        execute("""INSERT INTO asset_depreciation_runs(run_no,period_date,status,notes,created_by,created_at)
                   VALUES(:no,:dt,'مسودة',:notes,:uid,:created)""",
                {"no":no,"dt":period_date,"notes":request.form.get("notes",""),
                 "uid":session.get("user_id"),"created":datetime.now()})
        run_id=row("SELECT id FROM asset_depreciation_runs WHERE run_no=:no",{"no":no})["id"]
        total=0
        for asset in rows("""SELECT * FROM fixed_assets
                             WHERE status='نشط' AND capitalization_date<=:dt""",{"dt":period_date}):
            amount=calculate_monthly_depreciation(asset)
            if amount<=0: continue
            before=float(asset["accumulated_depreciation"] or 0)
            after=round(before+amount,2)
            nbv=round(float(asset["purchase_cost"])-after,2)
            execute("""INSERT INTO asset_depreciation_lines(run_id,asset_id,depreciation_amount,
              accumulated_before,accumulated_after,net_book_value_after)
              VALUES(:run,:asset,:amount,:before,:after,:nbv)""",
              {"run":run_id,"asset":asset["id"],"amount":amount,
               "before":before,"after":after,"nbv":nbv})
            execute("""UPDATE fixed_assets SET accumulated_depreciation=:after,
                       net_book_value=:nbv WHERE id=:id""",
                    {"after":after,"nbv":nbv,"id":asset["id"]})
            total+=amount
        execute("UPDATE asset_depreciation_runs SET total_amount=:total WHERE id=:id",
                {"total":round(total,2),"id":run_id})
        try:
            post_depreciation_run(run_id)
            flash(f"تم تشغيل وترحيل الإهلاك {no}","success")
        except Exception as exc:
            db.session.rollback()
            flash(str(exc),"danger")
        return redirect(url_for("asset_depreciation"))
    return render_template("asset_depreciation.html",
      runs=rows("""SELECT r.*,j.journal_no FROM asset_depreciation_runs r
                   LEFT JOIN journal_entries j ON j.id=r.journal_id
                   ORDER BY r.period_date DESC,r.id DESC"""))

@app.route("/fixed-assets/report")
@login_required
def fixed_assets_report():
    data=rows("""SELECT fa.asset_no,fa.name,ac.name category_name,fa.purchase_date,
      fa.purchase_cost,fa.accumulated_depreciation,fa.net_book_value,
      fa.status,fa.location
      FROM fixed_assets fa JOIN asset_categories ac ON ac.id=fa.category_id
      ORDER BY ac.code,fa.asset_no""")
    return render_template("fixed_assets_report.html",rows=data)

@app.route("/fixed-assets/report.xlsx")
@login_required
def fixed_assets_report_export():
    data=rows("""SELECT fa.asset_no,fa.name,ac.name category_name,fa.purchase_date,
      fa.purchase_cost,fa.accumulated_depreciation,fa.net_book_value,
      fa.status,fa.location
      FROM fixed_assets fa JOIN asset_categories ac ON ac.id=fa.category_id
      ORDER BY ac.code,fa.asset_no""")
    return xlsx_response("fixed_assets.xlsx","الأصول الثابتة",
      ["رقم الأصل","الأصل","الفئة","تاريخ الشراء","التكلفة","مجمع الإهلاك",
       "صافي القيمة الدفترية","الحالة","الموقع"],
      [[r["asset_no"],r["name"],r["category_name"],r["purchase_date"],r["purchase_cost"],
        r["accumulated_depreciation"],r["net_book_value"],r["status"],r["location"]] for r in data])




@app.route("/api/smart-lookup/<entity>")
@login_required
def smart_lookup(entity):
    cfg=SMART_ENTITY_CONFIG.get(entity)
    if not cfg:
        return {"ok":False,"error":"نوع البحث غير مدعوم"},404
    q=request.args.get("q","").strip()
    params={"q":f"%{q}%"}
    display=cfg["display"]; code=cfg.get("code")
    select_fields=["id",f"{display} AS name"]
    if code: select_fields.append(f"{code} AS code")
    if cfg.get("tax"): select_fields.append(f"{cfg['tax']} AS tax_number")
    if cfg.get("account"): select_fields.append(f"{cfg['account']} AS account_id")
    if entity=="items": select_fields.extend(["unit","unit_cost","quantity"])
    conditions=[f"{display} ILIKE :q"]
    if code: conditions.append(f"COALESCE({code},'') ILIKE :q")
    result=rows(f"""SELECT {','.join(select_fields)} FROM {cfg['table']}
                    WHERE {' OR '.join(conditions)}
                    ORDER BY {display} LIMIT 30""",params)
    return {"ok":True,"results":result}

@app.route("/api/smart-entity/<entity>/<int:entity_id>")
@login_required
def smart_entity_details(entity,entity_id):
    result=smart_entity_row(entity,entity_id)
    if not result:
        return {"ok":False,"error":"السجل غير موجود"},404
    return {"ok":True,"result":result}

@app.route("/api/quick-create/<entity>",methods=["POST"])
@login_required
def quick_create(entity):
    try:
        payload=request.get_json(silent=True) or request.form.to_dict()
        lookup_entity,result=create_quick_entity(entity,payload)
        audit("QUICK_CREATE",entity,f"إنشاء سريع: {result['name']}")
        return {"ok":True,"lookup_entity":lookup_entity,"result":result}
    except Exception as exc:
        db.session.rollback()
        return {"ok":False,"error":str(exc)},400

@app.route("/global-search")
@login_required
def global_search():
    q=request.args.get("q","").strip()
    grouped={}
    if q:
        grouped["العملاء"]=rows("""SELECT id,code,name,'customer' result_type
          FROM customers WHERE name ILIKE :q OR COALESCE(code,'') ILIKE :q
          ORDER BY name LIMIT 20""",{"q":f"%{q}%"})
        grouped["الموردون"]=rows("""SELECT id,code,name,'supplier' result_type
          FROM suppliers WHERE name ILIKE :q OR COALESCE(code,'') ILIKE :q
          ORDER BY name LIMIT 20""",{"q":f"%{q}%"})
        grouped["المواد"]=rows("""SELECT id,code,name,'item' result_type
          FROM inventory WHERE name ILIKE :q OR COALESCE(code,'') ILIKE :q
          ORDER BY name LIMIT 20""",{"q":f"%{q}%"})
        grouped["الفواتير"]=rows("""SELECT i.id,i.invoice_no code,c.name,
          'invoice' result_type FROM invoices i JOIN customers c ON c.id=i.customer_id
          WHERE i.invoice_no ILIKE :q OR c.name ILIKE :q
          ORDER BY i.id DESC LIMIT 20""",{"q":f"%{q}%"})
        grouped["المشاريع"]=rows("""SELECT id,project_no code,name,'project' result_type
          FROM projects WHERE project_no ILIKE :q OR name ILIKE :q
          ORDER BY project_no LIMIT 20""",{"q":f"%{q}%"})
        grouped["الموظفون"]=rows("""SELECT id,employee_no code,name,'employee' result_type
          FROM employees WHERE name ILIKE :q OR COALESCE(employee_no,'') ILIKE :q
          ORDER BY name LIMIT 20""",{"q":f"%{q}%"})
    return render_template("global_search.html",q=q,grouped=grouped)


@app.route("/data-import")
@login_required
def data_import_center():
    return render_template("data_import_center.html",
      modules=EXCEL_IMPORT_DEFINITIONS,
      jobs=rows("""SELECT j.*,u.username FROM data_import_jobs j
                   LEFT JOIN users u ON u.id=j.imported_by
                   ORDER BY j.id DESC LIMIT 100"""))

@app.route("/data-import/template/<module_name>")
@login_required
def data_import_template(module_name):
    if module_name not in EXCEL_IMPORT_DEFINITIONS:
        return "الوحدة غير مدعومة",404
    definition=EXCEL_IMPORT_DEFINITIONS[module_name]
    return xlsx_response(
        f"{module_name}_import_template.xlsx",
        definition["title"],
        definition["headers"],
        [definition["sample"]],
    )

@app.route("/data-import/<module_name>",methods=["GET","POST"])
@login_required
def data_import_upload(module_name):
    if module_name not in EXCEL_IMPORT_DEFINITIONS:
        return "الوحدة غير مدعومة",404
    definition=EXCEL_IMPORT_DEFINITIONS[module_name]
    preview=[]; errors=[]; warnings=[]; mapping={}; mapping_details={}; file_info={}
    quality_score=0; quality={}; simulation={"add":0,"update":0,"skip":0,"errors":0}
    profiles=rows("""SELECT id,profile_name,source_system,mapping_json FROM data_import_profiles
                     WHERE module_name=:module ORDER BY profile_name""",{"module":module_name})
    if request.method=="POST":
        uploaded=request.files.get("excel_file")
        if not uploaded or not uploaded.filename:
            flash("اختر ملف Excel أو CSV أولاً.","danger")
            return redirect(url_for("data_import_upload",module_name=module_name))
        try:
            filename,raw_headers,data_rows=read_import_file(uploaded)
            preferred={}
            profile_id=request.form.get("profile_id",type=int)
            if profile_id:
                saved=row("SELECT mapping_json FROM data_import_profiles WHERE id=:id AND module_name=:module",{"id":profile_id,"module":module_name})
                if saved:
                    try: preferred=json.loads(saved["mapping_json"] or "{}")
                    except ValueError: preferred={}
            mapping,mapping_details=smart_map_headers(raw_headers,definition,module_name,preferred)
            mapped_fields={field for field in mapping.values() if field}
            missing=[h for h in definition["required"] if h not in mapped_fields]
            if missing:
                raise ValueError("تعذر مطابقة الأعمدة المطلوبة: "+", ".join(missing))
            duplicate_keys=set()
            for row_no,values in enumerate(data_rows,start=2):
                normalized_values=[normalize_excel_value(v) for v in values]
                raw_data={raw_headers[i]:normalized_values[i] if i<len(normalized_values) else "" for i in range(len(raw_headers))}
                data={target:clean_import_value(raw_data.get(source,""),target) for source,target in mapping.items() if target}
                if not any(str(v).strip() for v in data.values()): continue
                row_errors,row_warnings=validate_import_row(module_name,data,row_no)
                identity=str(data.get("code") or data.get("employee_no") or data.get("account_code") or data.get("vat_no") or data.get("name") or "").strip().lower()
                duplicate_status=""
                if identity and identity in duplicate_keys:
                    duplicate_status="داخل الملف"; row_warnings.append("السجل مكرر داخل الملف")
                if identity: duplicate_keys.add(identity)
                existing=find_existing_import_record(module_name,data)
                action="update" if existing else "add"
                if existing:
                    duplicate_status="موجود بالنظام"
                    row_warnings.append("يوجد سجل مطابق في النظام؛ سيعتمد الإجراء على وضع الاستيراد")
                if row_errors: action="error"
                preview.append({"row_no":row_no,"data":data,"errors":row_errors,"warnings":row_warnings,
                                "duplicate_status":duplicate_status,"action":action})
                errors.extend([f"الصف {row_no}: {e}" for e in row_errors])
                warnings.extend([f"الصف {row_no}: {w}" for w in row_warnings])
            quality=import_quality_breakdown(preview,definition); quality_score=quality["overall"]
            simulation={"add":sum(1 for x in preview if x["action"]=="add"),
                        "update":sum(1 for x in preview if x["action"]=="update"),
                        "skip":sum(1 for x in preview if x.get("duplicate_status")=="داخل الملف"),
                        "errors":sum(1 for x in preview if x["action"]=="error")}
            file_info={"name":filename,"rows":len(preview),"columns":len(raw_headers)}
            save_import_preview({"module_name":module_name,"file_name":filename,"rows":preview,
                                 "mapping":mapping,"quality":quality,"simulation":simulation})
            flash(f"تمت قراءة {len(preview)} صفًا ومطابقة {len(mapped_fields)} عمودًا تلقائيًا.","warning" if errors else "success")
        except Exception as exc:
            db.session.rollback(); clear_import_preview()
            flash(f"تعذر قراءة الملف: {exc}","danger")
    return render_template("data_import_upload.html",module_name=module_name,definition=definition,
      preview=preview,errors=errors,warnings=warnings,mapping=mapping,mapping_details=mapping_details,
      file_info=file_info,quality_score=quality_score,quality=quality,simulation=simulation,profiles=profiles)


@app.route("/data-import/<module_name>/profile",methods=["POST"])
@login_required
def data_import_save_profile(module_name):
    preview=load_import_preview()
    if preview.get("module_name")!=module_name:
        flash("لا توجد معاينة صالحة لحفظ القالب.","danger")
        return redirect(url_for("data_import_upload",module_name=module_name))
    profile_name=request.form.get("profile_name","").strip()
    if not profile_name:
        flash("اكتب اسم قالب المطابقة.","danger")
        return redirect(url_for("data_import_upload",module_name=module_name))
    execute("""INSERT INTO data_import_profiles(profile_name,module_name,source_system,mapping_json,created_by,created_at,updated_at)
              VALUES(:name,:module,:source,:mapping,:user,:dt,:dt)""",
            {"name":profile_name,"module":module_name,"source":request.form.get("source_system","").strip(),
             "mapping":json.dumps(preview.get("mapping",{}),ensure_ascii=False),"user":session.get("user_id"),"dt":datetime.now()})
    for source,target in preview.get("mapping",{}).items():
        if target:
            try:
                execute("""INSERT INTO data_import_profile_aliases(module_name,source_header,target_field,created_by,created_at)
                          VALUES(:module,:source,:target,:user,:dt)
                          ON CONFLICT(module_name,source_header) DO UPDATE SET target_field=EXCLUDED.target_field""",
                        {"module":module_name,"source":source,"target":target,"user":session.get("user_id"),"dt":datetime.now()})
            except Exception:
                db.session.rollback()
    audit("CREATE","IMPORT_PROFILE",f"حفظ قالب مطابقة {profile_name} لوحدة {module_name}")
    flash("تم حفظ قالب المطابقة وسيظهر في عمليات الاستيراد القادمة.","success")
    return redirect(url_for("data_import_upload",module_name=module_name))


@app.route("/data-import/<module_name>/confirm",methods=["POST"])
@login_required
def data_import_confirm(module_name):
    preview=load_import_preview()
    if preview.get("module_name")!=module_name:
        flash("لا توجد معاينة صالحة للاستيراد.","danger")
        return redirect(url_for("data_import_upload",module_name=module_name))
    import_mode=request.form.get("import_mode","إضافة فقط")
    rows_data=preview.get("rows",[])
    import_no=next_import_number()
    now=datetime.now()
    db.session.execute(text("""INSERT INTO data_import_jobs(import_no,module_name,file_name,import_mode,total_rows,
      success_rows,updated_rows,failed_rows,status,error_details,imported_by,imported_at)
      VALUES(:no,:module,:file,:mode,:total,0,0,0,'قيد الانتظار','',:user,:dt)"""),
      {"no":import_no,"module":module_name,"file":preview.get("file_name",""),"mode":import_mode,
       "total":len(rows_data),"user":session.get("user_id"),"dt":now})
    job_id=db.session.execute(text("SELECT id FROM data_import_jobs WHERE import_no=:no"),{"no":import_no}).scalar_one()
    for item in rows_data:
        initial='فشل' if item.get('errors') else ('متجاهل' if item.get('duplicate_status')=='داخل الملف' else 'قيد الانتظار')
        error=' | '.join(item.get('errors') or [])
        db.session.execute(text("""INSERT INTO data_import_job_rows(job_id,row_no,row_data,row_status,action_type,error_message)
          VALUES(:job,:row_no,:data,:status,:action,:error)"""),
          {"job":job_id,"row_no":item.get("row_no"),"data":json.dumps(item.get("data",{}),ensure_ascii=False,default=str),
           "status":initial,"action":item.get("action",''),"error":error})
    db.session.execute(text("""INSERT INTO data_import_job_events(job_id,event_type,message,created_at)
      VALUES(:job,'QUEUE',:message,:dt)"""),{"job":job_id,"message":f"تمت إضافة {len(rows_data)} صفًا إلى طابور المعالجة", "dt":now})
    db.session.commit()
    clear_import_preview()
    audit("QUEUE","DATA_IMPORT",f"إضافة العملية {import_no} إلى طابور الاستيراد")
    flash("تمت إضافة عملية الاستيراد إلى الطابور. ستبدأ المعالجة من شاشة المتابعة.","success")
    return redirect(url_for("data_import_job_monitor",job_id=job_id))


@app.route("/data-import/job/<int:job_id>")
@login_required
def data_import_job_monitor(job_id):
    job=row("""SELECT j.*,u.username FROM data_import_jobs j LEFT JOIN users u ON u.id=j.imported_by
               WHERE j.id=:id""",{"id":job_id})
    if not job: return "عملية الاستيراد غير موجودة",404
    events=rows("SELECT * FROM data_import_job_events WHERE job_id=:id ORDER BY id DESC LIMIT 30",{"id":job_id})
    return render_template("data_import_job_monitor.html",job=job,events=events)


@app.route("/data-import/job/<int:job_id>/progress")
@login_required
def data_import_job_progress(job_id):
    job=row("SELECT * FROM data_import_jobs WHERE id=:id",{"id":job_id})
    if not job: return jsonify({"error":"not_found"}),404
    counts=row("""SELECT COUNT(*) total,
      SUM(CASE WHEN row_status IN ('نجح','تم التحديث','فشل','متجاهل') THEN 1 ELSE 0 END) processed,
      SUM(CASE WHEN row_status='نجح' THEN 1 ELSE 0 END) success,
      SUM(CASE WHEN row_status='تم التحديث' THEN 1 ELSE 0 END) updated,
      SUM(CASE WHEN row_status='فشل' THEN 1 ELSE 0 END) failed,
      SUM(CASE WHEN row_status='متجاهل' THEN 1 ELSE 0 END) skipped
      FROM data_import_job_rows WHERE job_id=:id""",{"id":job_id})
    total=int(counts.get('total') or 0); processed=int(counts.get('processed') or 0)
    percent=round(processed/total*100,1) if total else 100
    latest=row("SELECT message,created_at FROM data_import_job_events WHERE job_id=:id ORDER BY id DESC LIMIT 1",{"id":job_id})
    return jsonify({"status":job.get("status"),"total":total,"processed":processed,"percent":percent,
      "success":int(counts.get('success') or 0),"updated":int(counts.get('updated') or 0),
      "failed":int(counts.get('failed') or 0),"skipped":int(counts.get('skipped') or 0),
      "message":latest.get('message') if latest else ''})


@app.route("/data-import/job/<int:job_id>/process",methods=["POST"])
@login_required
def data_import_job_process(job_id):
    job=row("SELECT * FROM data_import_jobs WHERE id=:id",{"id":job_id})
    if not job: return jsonify({"error":"not_found"}),404
    if job.get('status') in {'مكتمل','مكتمل مع أخطاء','ملغي'}:
        return jsonify({"done":True,"status":job.get('status')})
    execute("UPDATE data_import_jobs SET status='قيد التنفيذ' WHERE id=:id",{"id":job_id})
    batch=rows("""SELECT * FROM data_import_job_rows WHERE job_id=:id AND row_status='قيد الانتظار'
                  ORDER BY id LIMIT 50""",{"id":job_id})
    for item in batch:
        try:
            data=json.loads(item.get('row_data') or '{}')
            was_updated=import_excel_row(job['module_name'],data,job['import_mode'])
            execute("""UPDATE data_import_job_rows SET row_status=:status,processed_at=:dt WHERE id=:id""",
                    {"status":"تم التحديث" if was_updated else "نجح","dt":datetime.now(),"id":item['id']})
        except Exception as exc:
            db.session.rollback()
            execute("""UPDATE data_import_job_rows SET row_status='فشل',error_message=:error,processed_at=:dt WHERE id=:id""",
                    {"error":str(exc),"dt":datetime.now(),"id":item['id']})
    counts=row("""SELECT COUNT(*) total,
      SUM(CASE WHEN row_status='قيد الانتظار' THEN 1 ELSE 0 END) pending,
      SUM(CASE WHEN row_status='نجح' THEN 1 ELSE 0 END) success,
      SUM(CASE WHEN row_status='تم التحديث' THEN 1 ELSE 0 END) updated,
      SUM(CASE WHEN row_status='فشل' THEN 1 ELSE 0 END) failed
      FROM data_import_job_rows WHERE job_id=:id""",{"id":job_id})
    pending=int(counts.get('pending') or 0); failed=int(counts.get('failed') or 0)
    status=('مكتمل مع أخطاء' if failed else 'مكتمل') if pending==0 else 'قيد التنفيذ'
    error_rows=rows("SELECT row_no,error_message FROM data_import_job_rows WHERE job_id=:id AND row_status='فشل' ORDER BY row_no LIMIT 1000",{"id":job_id})
    errors='\n'.join(f"الصف {x['row_no']}: {x['error_message']}" for x in error_rows)
    execute("""UPDATE data_import_jobs SET success_rows=:success,updated_rows=:updated,failed_rows=:failed,
              status=:status,error_details=:errors WHERE id=:id""",
            {"success":int(counts.get('success') or 0),"updated":int(counts.get('updated') or 0),
             "failed":failed,"status":status,"errors":errors,"id":job_id})
    if batch:
        execute("""INSERT INTO data_import_job_events(job_id,event_type,message,created_at)
                   VALUES(:job,'PROGRESS',:message,:dt)""",
                {"job":job_id,"message":f"تمت معالجة دفعة من {len(batch)} صفًا — المتبقي {pending}","dt":datetime.now()})
    if pending==0:
        audit("IMPORT","BATCH",f"اكتملت عملية {job['import_no']} بالحالة {status}")
    return jsonify({"done":pending==0,"status":status,"processed_batch":len(batch)})


@app.route("/data-import/job/<int:job_id>/cancel",methods=["POST"])
@login_required
def data_import_job_cancel(job_id):
    execute("UPDATE data_import_jobs SET status='ملغي' WHERE id=:id AND status IN ('قيد الانتظار','قيد التنفيذ')",{"id":job_id})
    execute("UPDATE data_import_job_rows SET row_status='متجاهل' WHERE job_id=:id AND row_status='قيد الانتظار'",{"id":job_id})
    execute("""INSERT INTO data_import_job_events(job_id,event_type,message,created_at)
               VALUES(:job,'CANCEL','تم إلغاء العملية بواسطة المستخدم',:dt)""",{"job":job_id,"dt":datetime.now()})
    audit("CANCEL","DATA_IMPORT",f"إلغاء عملية الاستيراد رقم {job_id}")
    return jsonify({"ok":True})


@app.route("/data-import/history/<int:job_id>")
@login_required
def data_import_history_detail(job_id):
    job=row("""SELECT j.*,u.username FROM data_import_jobs j
               LEFT JOIN users u ON u.id=j.imported_by WHERE j.id=:id""",{"id":job_id})
    if not job:
        return "عملية الاستيراد غير موجودة",404
    return render_template("data_import_history_detail.html",job=job,
      errors=[line for line in (job.get("error_details") or "").splitlines() if line])

@app.route("/data-import/history/<int:job_id>/errors.xlsx")
@login_required
def data_import_errors(job_id):
    job=row("SELECT * FROM data_import_jobs WHERE id=:id",{"id":job_id})
    if not job:
        return "عملية الاستيراد غير موجودة",404
    error_lines=[line for line in (job.get("error_details") or "").splitlines() if line]
    records=[]
    for line in error_lines:
        match=re.match(r"الصف\s+(\d+):\s*(.*)",line)
        records.append([match.group(1) if match else "",match.group(2) if match else line])
    return xlsx_response(f"{job['import_no']}_errors.xlsx","Import Errors",["الصف","سبب الخطأ"],records)


@app.route("/financial-statements")
@login_required
def financial_statements_center():
    return render_template("financial_statements_center.html")

@app.route("/financial-statements/income-statement")
@login_required
def income_statement():
    filters=financial_date_filters()
    data=financial_statement_data(filters)
    revenues=[x for x in data if x["account_type"]=="إيراد" and abs(x["balance"])>0.005]
    cost_sales=[x for x in data if x["account_type"]=="مصروف" and
                ("تكلفة" in (x["account_name_ar"] or "") or
                 "مبيعات" in (x["account_name_ar"] or "")) and abs(x["balance"])>0.005]
    expenses=[x for x in data if x["account_type"]=="مصروف" and x not in cost_sales
              and abs(x["balance"])>0.005]
    total_revenue=round(sum(x["balance"] for x in revenues),2)
    total_cost=round(sum(x["balance"] for x in cost_sales),2)
    gross_profit=round(total_revenue-total_cost,2)
    total_expenses=round(sum(x["balance"] for x in expenses),2)
    net_profit=round(gross_profit-total_expenses,2)
    return render_template("income_statement.html",filters=filters,revenues=revenues,
      cost_sales=cost_sales,expenses=expenses,total_revenue=total_revenue,
      total_cost=total_cost,gross_profit=gross_profit,total_expenses=total_expenses,
      net_profit=net_profit,branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"))

@app.route("/financial-statements/balance-sheet")
@login_required
def balance_sheet():
    filters=financial_date_filters()
    data=financial_statement_data(filters)
    assets=[x for x in data if x["account_type"]=="أصل" and abs(x["balance"])>0.005]
    liabilities=[x for x in data if x["account_type"]=="التزام" and abs(x["balance"])>0.005]
    equity=[x for x in data if x["account_type"]=="حقوق ملكية" and abs(x["balance"])>0.005]
    revenue_total=sum(x["balance"] for x in data if x["account_type"]=="إيراد")
    expense_total=sum(x["balance"] for x in data if x["account_type"]=="مصروف")
    retained=round(revenue_total-expense_total,2)
    total_assets=round(sum(x["balance"] for x in assets),2)
    total_liabilities=round(sum(x["balance"] for x in liabilities),2)
    total_equity=round(sum(x["balance"] for x in equity)+retained,2)
    difference=round(total_assets-total_liabilities-total_equity,2)
    return render_template("balance_sheet.html",filters=filters,assets=assets,
      liabilities=liabilities,equity=equity,retained=retained,total_assets=total_assets,
      total_liabilities=total_liabilities,total_equity=total_equity,difference=difference,
      branches=rows("SELECT id,name FROM branches WHERE active=1 ORDER BY name"),
      centers=rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code"))

@app.route("/financial-statements/cash-flow")
@login_required
def cash_flow_statement():
    filters=financial_date_filters()
    conditions=["j.status='مرحّل'"];params={}
    if filters["date_from"]:
        conditions.append("j.journal_date>=:date_from");params["date_from"]=filters["date_from"]
    if filters["date_to"]:
        conditions.append("j.journal_date<=:date_to");params["date_to"]=filters["date_to"]
    data=rows(f"""SELECT a.account_code,a.account_name_ar,
      SUM(l.debit) debit,SUM(l.credit) credit,SUM(l.debit-l.credit) net_movement
      FROM journal_entry_lines l JOIN journal_entries j ON j.id=l.journal_id
      JOIN chart_of_accounts a ON a.id=l.account_id
      WHERE {' AND '.join(conditions)}
      AND a.account_type='أصل'
      AND (a.account_name_ar ILIKE '%صندوق%' OR a.account_name_ar ILIKE '%بنك%'
           OR a.account_name_ar ILIKE '%نقد%')
      GROUP BY a.id,a.account_code,a.account_name_ar ORDER BY a.account_code""",params)
    net_cash=round(sum(float(x["net_movement"] or 0) for x in data),2)
    receipts=row(f"""SELECT COALESCE(SUM(tv.amount),0) amount FROM treasury_vouchers tv
      WHERE tv.voucher_type='قبض' AND tv.posting_status='مرحّل'
      {"AND tv.voucher_date>=:date_from" if filters["date_from"] else ""}
      {"AND tv.voucher_date<=:date_to" if filters["date_to"] else ""}""",params)["amount"]
    payments=row(f"""SELECT COALESCE(SUM(tv.amount),0) amount FROM treasury_vouchers tv
      WHERE tv.voucher_type='صرف' AND tv.posting_status='مرحّل'
      {"AND tv.voucher_date>=:date_from" if filters["date_from"] else ""}
      {"AND tv.voucher_date<=:date_to" if filters["date_to"] else ""}""",params)["amount"]
    return render_template("cash_flow_statement.html",filters=filters,accounts=data,
      receipts=receipts,payments=payments,net_cash=net_cash)

@app.route("/financial-statements/export.xlsx")
@login_required
def financial_statements_export():
    statement=request.args.get("statement","income")
    filters=financial_date_filters()
    data=financial_statement_data(filters)
    if statement=="balance":
        records=[[x["account_code"],x["account_name_ar"],x["account_type"],x["balance"]]
                 for x in data if x["account_type"] in ("أصل","التزام","حقوق ملكية")
                 and abs(x["balance"])>0.005]
        return xlsx_response("balance_sheet.xlsx","الميزانية العمومية",
          ["الكود","الحساب","النوع","الرصيد"],records)
    records=[[x["account_code"],x["account_name_ar"],x["account_type"],x["balance"]]
             for x in data if x["account_type"] in ("إيراد","مصروف")
             and abs(x["balance"])>0.005]
    return xlsx_response("income_statement.xlsx","قائمة الدخل",
      ["الكود","الحساب","النوع","الرصيد"],records)

@app.route("/party-statements")
@login_required
def party_statements():
    party_type=request.args.get("party_type","customer")
    if party_type not in ("customer","supplier"):
        party_type="customer"
    party_id=request.args.get("party_id",type=int)
    date_from=request.args.get("date_from","")
    date_to=request.args.get("date_to","")
    opening=party_opening_balance(party_type,party_id,date_from)
    data=party_statement_rows(party_type,party_id,date_from,date_to,opening) if party_id else []
    parties=rows("SELECT id,name FROM customers ORDER BY name") if party_type=="customer" \
            else rows("SELECT id,name FROM suppliers ORDER BY name")
    party=row(f"SELECT id,name,name_en,vat_number,phone,email FROM {'customers' if party_type=='customer' else 'suppliers'} WHERE id=:id",{"id":party_id}) if party_id else None
    total_debit=round(sum(float(x["debit"] or 0) for x in data),2)
    total_credit=round(sum(float(x["credit"] or 0) for x in data),2)
    closing=round(opening+total_debit-total_credit,2)
    return render_template("party_statement.html",party_type=party_type,party_id=party_id,
      date_from=date_from,date_to=date_to,rows=data,parties=parties,party=party,
      opening_balance=opening,total_debit=total_debit,total_credit=total_credit,closing_balance=closing)

@app.route("/party-statements.xlsx")
@login_required
def party_statements_export():
    party_type=request.args.get("party_type","customer")
    party_id=request.args.get("party_id",type=int)
    date_from=request.args.get("date_from","");date_to=request.args.get("date_to","")
    opening=party_opening_balance(party_type,party_id,date_from)
    data=party_statement_rows(party_type,party_id,date_from,date_to,opening) if party_id else []
    records=[]
    if date_from:
        records.append([date_from,"-","-","الرصيد الافتتاحي","-",0,0,opening])
    records.extend([[r["journal_date"],r["journal_no"],r["reference"],r["line_description"] or r["description"],
        r["invoice_number"],r["debit"],r["credit"],r["running_balance"]] for r in data])
    return xlsx_response("party_statement.xlsx","كشف الحساب",
      ["التاريخ","القيد","المرجع","البيان","رقم الفاتورة","مدين","دائن","الرصيد"],
      records)

@app.route("/payables-aging")
@login_required
def payables_aging():
    data=rows("""SELECT s.id,s.name,
      SUM(CASE WHEN CURRENT_DATE-si.invoice_date<=30 THEN si.total ELSE 0 END) bucket_0_30,
      SUM(CASE WHEN CURRENT_DATE-si.invoice_date BETWEEN 31 AND 60 THEN si.total ELSE 0 END) bucket_31_60,
      SUM(CASE WHEN CURRENT_DATE-si.invoice_date BETWEEN 61 AND 90 THEN si.total ELSE 0 END) bucket_61_90,
      SUM(CASE WHEN CURRENT_DATE-si.invoice_date>90 THEN si.total ELSE 0 END) bucket_over_90,
      SUM(si.total) total_outstanding
      FROM supplier_invoices si JOIN suppliers s ON s.id=si.supplier_id
      WHERE si.status='معتمدة'
      GROUP BY s.id,s.name ORDER BY total_outstanding DESC""")
    return render_template("payables_aging.html",rows=data)

@app.route("/executive-dashboard")
@login_required
def executive_dashboard():
    sales=row("""SELECT COALESCE(SUM(total),0) amount,COUNT(*) count
                 FROM invoices WHERE status='معتمدة'""")
    purchases=row("""SELECT COALESCE(SUM(total),0) amount,COUNT(*) count
                     FROM supplier_invoices WHERE status='معتمدة'""")
    expenses=row("""SELECT COALESCE(SUM(amount),0) amount,COUNT(*) count
                    FROM expenses WHERE status='معتمدة'""")
    receivables=row("""SELECT COALESCE(SUM(i.total-COALESCE(
      (SELECT SUM(a.allocated_amount) FROM invoice_payment_allocations a
       WHERE a.invoice_id=i.id),0)),0) amount
      FROM invoices i WHERE i.status='معتمدة'""")
    payables=row("""SELECT COALESCE(SUM(total),0) amount FROM supplier_invoices
                    WHERE status='معتمدة'""")
    inventory_value=row("""SELECT COALESCE(SUM(quantity*cost),0) amount FROM inventory
                           WHERE active=1""")
    cash=row("""SELECT COALESCE(SUM(CASE WHEN l.debit>0 THEN l.debit ELSE -l.credit END),0) amount
      FROM journal_entry_lines l JOIN journal_entries j ON j.id=l.journal_id
      JOIN chart_of_accounts a ON a.id=l.account_id
      WHERE j.status='مرحّل' AND a.account_type='أصل'
      AND (a.account_name_ar ILIKE '%صندوق%' OR a.account_name_ar ILIKE '%بنك%'
           OR a.account_name_ar ILIKE '%نقد%')""")
    top_customers=rows("""SELECT c.name,SUM(i.total) total FROM invoices i
      JOIN customers c ON c.id=i.customer_id WHERE i.status='معتمدة'
      GROUP BY c.id,c.name ORDER BY total DESC LIMIT 10""")
    top_suppliers=rows("""SELECT s.name,SUM(si.total) total FROM supplier_invoices si
      JOIN suppliers s ON s.id=si.supplier_id WHERE si.status='معتمدة'
      GROUP BY s.id,s.name ORDER BY total DESC LIMIT 10""")
    low_stock=rows("""SELECT sku,name,quantity,reorder_level FROM inventory
      WHERE active=1 AND quantity<=reorder_level ORDER BY quantity LIMIT 10""")
    monthly=rows("""SELECT TO_CHAR(invoice_date,'YYYY-MM') month,SUM(total) total
      FROM invoices WHERE status='معتمدة'
      GROUP BY TO_CHAR(invoice_date,'YYYY-MM') ORDER BY month DESC LIMIT 12""")
    net_profit=round(float(sales["amount"] or 0)-float(expenses["amount"] or 0),2)
    return render_template("executive_dashboard.html",sales=sales,purchases=purchases,
      expenses=expenses,receivables=receivables,payables=payables,
      inventory_value=inventory_value,cash=cash,net_profit=net_profit,
      top_customers=top_customers,top_suppliers=top_suppliers,
      low_stock=low_stock,monthly=monthly)


@app.route("/reports")
@login_required
def reports():
    grouped = {}
    for slug, definition in ACCOUNTING_REPORTS.items():
        grouped.setdefault(definition["group"], []).append({"slug": slug, **definition})
    return render_template("reports_center.html", grouped=grouped)

@app.route("/reports/<slug>")
@login_required
def accounting_report(slug):
    if slug not in ACCOUNTING_REPORTS:
        return "التقرير غير موجود", 404
    filters = report_filters()
    context = report_context(slug, filters)
    accounts = rows("SELECT id,account_code,account_name_ar FROM chart_of_accounts WHERE active=1 ORDER BY account_code")
    centers = rows("SELECT id,code,name FROM cost_centers WHERE active=1 ORDER BY code")
    company = row("SELECT * FROM settings WHERE id=1")
    audit("VIEW", "REPORT", f"عرض تقرير {context['definition']['title']}")
    return render_template("accounting_report.html", slug=slug, filters=filters, accounts=accounts,
                           centers=centers, company=company, **context)

@app.route("/reports/<slug>/export.xlsx")
@login_required
def accounting_report_export(slug):
    if slug not in ACCOUNTING_REPORTS:
        return "التقرير غير موجود", 404
    filters = report_filters()
    context = report_context(slug, filters)
    data = context["data"]
    keys_by_slug = {
        "coa-report":["account_code","account_name_ar","account_name_en","account_type","parent_name","level","normal_balance","statement_type","accepts_entries","active"],
        "ledger-summary":["account_code","account_name_ar","total_debit","total_credit","debit_balance","credit_balance","movement_count"],
        "account-balances":["account_code","account_name_ar","total_debit","total_credit","debit_balance","credit_balance","movement_count"],
        "trial-balance":["account_code","account_name_ar","total_debit","total_credit","debit_balance","credit_balance","movement_count"],
        "ledger-detail":["journal_date","journal_no","reference","line_description","invoice_number","invoice_date","debit","credit","running_balance","status"],
        "journal-report":["journal_date","journal_no","status","reference","description","account_code","account_name_ar","debit","credit","party_name","tax_number","invoice_number","invoice_date","cost_center_name"],
        "cost-center-movements-center":["cost_center_name","journal_date","journal_no","account_code","account_name_ar","line_description","debit","credit","running_balance","status"],
        "cc-ledger-detail":["cost_center_name","journal_date","journal_no","account_code","account_name_ar","line_description","debit","credit","running_balance","status"],
        "cost-center-movements-account":["account_code","account_name_ar","cost_center_name","journal_date","journal_no","line_description","debit","credit","status"],
        "cc-ledger-summary":["cost_center_code","cost_center_name","total_debit","total_credit","debit_balance","credit_balance","movement_count"],
        "cc-trial-balance":["cost_center_code","cost_center_name","total_debit","total_credit","debit_balance","credit_balance","movement_count"],
        "cc-balances":["cost_center_code","cost_center_name","total_debit","total_credit","debit_balance","credit_balance","movement_count"],
        "accounts-with-centers":["account_code","account_name_ar","cost_center_code","cost_center_name","total_debit","total_credit","balance"],
        "centers-with-accounts":["cost_center_code","cost_center_name","account_code","account_name_ar","total_debit","total_credit","balance"],
        "vat-report":["journal_date","journal_no","invoice_date","invoice_number","party_name","tax_number","tax_direction","amount_before_tax","vat_amount","total_amount","compliance_status"],
    }
    keys = keys_by_slug[slug]
    records = [[item.get(key) for key in keys] for item in data]
    audit("EXPORT", "REPORT", f"تصدير تقرير {context['definition']['title']}")
    return xlsx_response(f"{slug}.xlsx", context["definition"]["title"][:31], context["headers"], records)


def xlsx_response(filename, sheet_name, headers, records):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for record in records:
        ws.append(list(record))
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 3, 45)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def csv_response(filename, headers, records):
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(records)
    return Response(
        output.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.route("/export/invoices.xlsx")
@login_required
def export_invoices():
    data = rows("""
        SELECT i.invoice_no, i.invoice_date, c.name customer_name,
               i.subtotal, i.vat, i.total, i.status
        FROM invoices i
        JOIN customers c ON c.id=i.customer_id
        ORDER BY i.invoice_date DESC, i.id DESC
    """)
    records = [
        [r["invoice_no"], r["invoice_date"], r["customer_name"],
         r["subtotal"], r["vat"], r["total"], r["status"]]
        for r in data
    ]
    return xlsx_response(
        "invoices.xlsx",
        "الفواتير",
        ["رقم الفاتورة", "التاريخ", "العميل", "قبل الضريبة",
         "الضريبة", "الإجمالي", "الحالة"],
        records
    )


@app.route("/export/expenses.xlsx")
@login_required
def export_expenses():
    data = rows("""
        SELECT expense_date, category, description, amount, vat, total
        FROM expenses ORDER BY expense_date DESC, id DESC
    """)
    records = [
        [r["expense_date"], r["category"], r["description"],
         r["amount"], r["vat"], r["total"]]
        for r in data
    ]
    return xlsx_response(
        "expenses.xlsx",
        "المصروفات",
        ["التاريخ", "التصنيف", "الوصف", "قبل الضريبة", "الضريبة", "الإجمالي"],
        records
    )


@app.route("/export/customers.xlsx")
@login_required
def export_customers():
    data = rows("""
        SELECT name, vat_number, phone, email, balance
        FROM customers ORDER BY name
    """)
    records = [
        [r["name"], r["vat_number"], r["phone"], r["email"], r["balance"]]
        for r in data
    ]
    return xlsx_response(
        "customers.xlsx",
        "العملاء",
        ["اسم العميل", "الرقم الضريبي", "الهاتف", "البريد", "الرصيد"],
        records
    )


@app.errorhandler(404)
def handle_not_found(error):
    return render_template("error_page.html",
                           error_title="الصفحة غير موجودة",
                           error_message="تعذر العثور على الصفحة المطلوبة.",
                           error_code=404), 404

@app.errorhandler(500)
def handle_internal_error(error):
    db.session.rollback()
    app.logger.exception("Unhandled application error")
    return render_template("error_page.html",
                           error_title="حدث خطأ في النظام",
                           error_message="تم تسجيل الخطأ. يرجى العودة للوحة التحكم والمحاولة مرة أخرى.",
                           error_code=500), 500

if __name__ == "__main__":
    with app.app_context():
        init_db()
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", "5000")))
